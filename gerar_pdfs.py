"""
Gerador de PDFs - Fluxo de Renda Fixa
Gera um PDF por assessor com os eventos de pagamento dos seus clientes.
Padrao Somus Capital.
"""

import os
from datetime import datetime
from collections import defaultdict, OrderedDict
import openpyxl
from fpdf import FPDF

# === CONFIGURAÇÕES ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENTRADA_FILE = os.path.join(BASE_DIR, "Mesa Produtos", "Fluxo RF", "ENTRADA", "Agenda de eventos.xlsx")
BASE_FILE = os.path.join(BASE_DIR, "BASE", "BASE EMAILS.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Fluxo RF", "PDFs")
LOGO_PATH = os.path.join(BASE_DIR, "assets", "logo_somus.png")

# Cores Somus Capital
GREEN = (0, 77, 51)
GREEN_LIGHT = (0, 102, 68)
BLUE = (24, 99, 220)
BLUE_LIGHT = (230, 240, 255)
WHITE = (255, 255, 255)
LIGHT_GRAY = (246, 247, 249)
DARK_GRAY = (45, 45, 45)
MID_GRAY = (180, 180, 180)
TEXT_GRAY = (80, 80, 80)

# Mapeamento de dia da semana
DIAS_SEMANA = {
    0: "Segunda-feira", 1: "Terca-feira", 2: "Quarta-feira",
    3: "Quinta-feira", 4: "Sexta-feira", 5: "Sabado", 6: "Domingo",
}
MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Marco", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}


def sanitize_text(text):
    if text is None:
        return ""
    text = str(text)
    replacements = {
        '\u2013': '-', '\u2014': '-', '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2026': '...', '\u2022': '-',
        '\u00b2': '2', '\u00b3': '3', '\u00b9': '1', '\u2070': '0',
        '\u00ba': 'o', '\u00aa': 'a', '\u20ac': 'EUR', '\u00a0': ' ',
        '\ufeff': '', '\u200b': '', '\u200e': '', '\u200f': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    try:
        text.encode('latin-1')
    except UnicodeEncodeError:
        cleaned = []
        for ch in text:
            try:
                ch.encode('latin-1')
                cleaned.append(ch)
            except UnicodeEncodeError:
                cleaned.append('?')
        text = ''.join(cleaned)
    return text


def fmt_currency(value):
    if value is None or value == 0:
        return "R$ 0,00"
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "R$ 0,00"
    neg = v < 0
    v = abs(v)
    inteiro = int(v)
    centavos = round((v - inteiro) * 100)
    if centavos >= 100:
        inteiro += 1
        centavos = 0
    s_int = f"{inteiro:,}".replace(",", ".")
    result = f"R$ {s_int},{centavos:02d}"
    if neg:
        result = f"-{result}"
    return result


def fmt_date_short(dt):
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def fmt_date_full(dt):
    if isinstance(dt, datetime):
        dia_semana = DIAS_SEMANA.get(dt.weekday(), "")
        mes = MESES.get(dt.month, "")
        return f"{dia_semana}, {dt.day:02d} de {mes} de {dt.year}"
    return str(dt)


def load_base_emails(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    assessores = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        código = str(row[0]).strip()
        if not código.startswith("A"):
            continue
        assessores[código] = {
            "nome": sanitize_text(row[1]) if row[1] else "",
            "email": sanitize_text(row[2]) if row[2] else "",
            "assistente": sanitize_text(row[3]) if row[3] else "-",
            "email_assistente": sanitize_text(row[4]) if row[4] else "-",
        }
    wb.close()
    return assessores


def load_eventos(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    eventos_por_assessor = defaultdict(list)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cod_assessor = row[1]
        if not cod_assessor or not isinstance(cod_assessor, str) or not cod_assessor.startswith("A"):
            continue
        data = row[0]
        if not isinstance(data, datetime):
            continue
        evento = {
            "data": data,
            "cod_conta": str(row[3]).strip() if row[3] else "",
            "risco": sanitize_text(row[4]) if row[4] else "",
            "qtd_posição": row[5] if row[5] else 0,
            "valor_estimado": row[6] if row[6] else 0,
            "total_estimado": row[7] if row[7] else 0,
            "código_ativo": sanitize_text(row[8]) if row[8] else "",
            "tipo_evento": sanitize_text(row[9]) if row[9] else "",
        }
        eventos_por_assessor[cod_assessor].append(evento)
    wb.close()
    return eventos_por_assessor


class SomusPDF(FPDF):

    def __init__(self, assessor_nome, assessor_código):
        super().__init__(orientation="L", unit="mm", format="A4")
        self.assessor_nome = sanitize_text(assessor_nome)
        self.assessor_código = sanitize_text(assessor_código)
        self.set_auto_page_break(auto=True, margin=18)
        self.logo_path = LOGO_PATH if os.path.exists(LOGO_PATH) else None

    def header(self):
        # Barra verde superior
        self.set_fill_color(*GREEN)
        self.rect(0, 0, 297, 16, "F")

        # Logo
        if self.logo_path:
            try:
                self.image(self.logo_path, x=8, y=2.5, h=11)
            except Exception:
                self.set_font("Helvetica", "B", 11)
                self.set_text_color(*WHITE)
                self.set_xy(8, 3)
                self.cell(40, 10, "SOMUS CAPITAL")

        # Título central
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*WHITE)
        self.set_xy(0, 3)
        self.cell(297, 10, "RELATÓRIO DE FLUXO DE RENDA FIXA", align="C")

        # Assessor no canto direito
        self.set_font("Helvetica", "", 7.5)
        self.set_xy(200, 3)
        self.cell(90, 5, f"{self.assessor_nome}", align="R")
        self.set_xy(200, 7.5)
        self.set_font("Helvetica", "", 6.5)
        self.cell(90, 5, f"Cod. {self.assessor_código}", align="R")

        # Filete azul
        self.set_fill_color(*BLUE)
        self.rect(0, 16, 297, 1, "F")
        self.set_y(20)

    def footer(self):
        self.set_y(-13)
        # Linha fina
        self.set_draw_color(*MID_GRAY)
        self.line(10, self.get_y(), 287, self.get_y())
        self.ln(1.5)
        self.set_font("Helvetica", "", 6)
        self.set_text_color(140, 140, 140)
        self.cell(0, 5, f"Somus Capital  |  Fluxo de Renda Fixa  |  {self.assessor_nome} ({self.assessor_código})", align="L")
        self.cell(0, 5, f"Página {self.page_no()}/{{nb}}", align="R")

    # ================================================================
    #  CAPA
    # ================================================================
    def add_cover(self, assessor_info, eventos):
        """Página de capa com resumo executivo."""
        self.ln(8)

        # Título grande
        self.set_font("Helvetica", "B", 22)
        self.set_text_color(*GREEN)
        self.cell(0, 12, "Fluxo de Renda Fixa", new_x="LMARGIN", new_y="NEXT")

        # Subtítulo
        self.set_font("Helvetica", "", 11)
        self.set_text_color(*TEXT_GRAY)
        self.cell(0, 7, "Relatório de eventos previstos  -  Somus Capital", new_x="LMARGIN", new_y="NEXT")

        self.ln(6)

        # Card do assessor
        y = self.get_y()
        self.set_fill_color(*LIGHT_GRAY)
        self.rect(10, y, 277, 28, "F")
        # Borda esquerda verde
        self.set_fill_color(*GREEN)
        self.rect(10, y, 2.5, 28, "F")

        self.set_xy(17, y + 3)
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(*GREEN)
        self.cell(130, 7, self.assessor_nome)

        self.set_xy(17, y + 10)
        self.set_font("Helvetica", "", 8.5)
        self.set_text_color(*DARK_GRAY)
        self.cell(130, 6, f"Código: {self.assessor_código}        Email: {assessor_info.get('email', '-')}")

        if assessor_info.get("assistente", "-") != "-":
            self.set_xy(17, y + 17)
            self.cell(130, 6, f"Assistente: {assessor_info['assistente']}        Email: {assessor_info.get('email_assistente', '-')}")

        # Metricas no lado direito do card
        if eventos:
            total_valor = sum(e["total_estimado"] for e in eventos)
            datas_unicas = sorted(set(e["data"] for e in eventos))
            ativos_únicos = len(set(e["risco"] for e in eventos))

            self.set_xy(170, y + 3)
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(*BLUE)
            self.cell(50, 6, f"{len(eventos)} eventos", align="R")

            self.set_xy(230, y + 3)
            self.cell(50, 6, f"{ativos_únicos} ativos", align="R")

            self.set_xy(170, y + 10)
            self.set_font("Helvetica", "", 8)
            self.set_text_color(*DARK_GRAY)
            self.cell(50, 6, f"De {fmt_date_short(datas_unicas[0])}", align="R")

            self.set_xy(230, y + 10)
            self.cell(50, 6, f"Ate {fmt_date_short(datas_unicas[-1])}", align="R")

            self.set_xy(170, y + 17)
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(*GREEN)
            self.cell(110, 6, f"Total estimado: {fmt_currency(total_valor)}", align="R")

        self.set_y(y + 33)

        if not eventos:
            return

        # ---- Mini resumo por tipo ----
        self.ln(2)
        por_tipo = defaultdict(lambda: {"count": 0, "total": 0.0})
        for ev in eventos:
            por_tipo[ev["tipo_evento"]]["count"] += 1
            por_tipo[ev["tipo_evento"]]["total"] += ev["total_estimado"]

        y2 = self.get_y()
        card_w = 66
        card_h = 18
        gap = 4.7
        tipos_display = {
            "PAGAMENTO DE JUROS": "Pagamento de Juros",
            "AMORTIZACAO": "Amortizacao",
            "INCORPORACAO": "Incorporacao",
            "PREMIO": "Premio",
        }
        cores_tipo = {
            "PAGAMENTO DE JUROS": BLUE,
            "AMORTIZACAO": GREEN,
            "INCORPORACAO": (180, 120, 0),
            "PREMIO": (140, 40, 140),
        }

        x = 10
        for tipo_key in ["PAGAMENTO DE JUROS", "AMORTIZACAO", "INCORPORACAO", "PREMIO"]:
            data = por_tipo.get(tipo_key)
            if not data:
                continue
            cor = cores_tipo.get(tipo_key, BLUE)
            # Card
            self.set_fill_color(*LIGHT_GRAY)
            self.rect(x, y2, card_w, card_h, "F")
            self.set_fill_color(*cor)
            self.rect(x, y2, card_w, 2, "F")

            self.set_xy(x + 3, y2 + 3)
            self.set_font("Helvetica", "B", 7.5)
            self.set_text_color(*cor)
            self.cell(card_w - 6, 5, tipos_display.get(tipo_key, tipo_key))

            self.set_xy(x + 3, y2 + 9)
            self.set_font("Helvetica", "", 7)
            self.set_text_color(*DARK_GRAY)
            self.cell(card_w - 6, 5, f"{data['count']} eventos  |  {fmt_currency(data['total'])}")

            x += card_w + gap

        self.set_y(y2 + card_h + 6)

        # ---- Top 3 Pagadoras ----
        por_risco_top = defaultdict(lambda: {"total": 0.0, "count": 0})
        for ev in eventos:
            por_risco_top[ev["risco"]]["total"] += ev["total_estimado"]
            por_risco_top[ev["risco"]]["count"] += 1
        top3 = sorted(por_risco_top.items(), key=lambda x: -x[1]["total"])[:3]

        if top3:
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(*GREEN)
            self.cell(0, 6, "Top 3 Pagadoras", new_x="LMARGIN", new_y="NEXT")
            self.ln(1)

            yt = self.get_y()
            tw = 89
            tgap = 5
            tx = 10
            medals = ["1o", "2o", "3o"]
            medal_colors = [BLUE, GREEN_LIGHT, (120, 120, 120)]

            for i, (risco, info) in enumerate(top3):
                cor = medal_colors[i]
                # Card
                self.set_fill_color(*LIGHT_GRAY)
                self.rect(tx, yt, tw, 16, "F")
                # Borda superior colorida
                self.set_fill_color(*cor)
                self.rect(tx, yt, tw, 2, "F")

                # Posicao
                self.set_xy(tx + 3, yt + 3)
                self.set_font("Helvetica", "B", 7)
                self.set_text_color(*cor)
                self.cell(8, 5, medals[i])

                # Nome do ativo
                risco_short = risco[:32] if len(risco) > 32 else risco
                self.set_xy(tx + 11, yt + 3)
                self.set_font("Helvetica", "B", 7)
                self.set_text_color(*DARK_GRAY)
                self.cell(tw - 14, 5, risco_short)

                # Valor total
                self.set_xy(tx + 3, yt + 9)
                self.set_font("Helvetica", "", 7)
                self.set_text_color(*cor)
                self.cell(tw - 6, 5, f"{fmt_currency(info['total'])}  ({info['count']} eventos)")

                tx += tw + tgap

            self.set_y(yt + 22)

        # ---- Calendario: resumo por data ----
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*GREEN)
        self.cell(0, 7, "Calendario de Eventos", new_x="LMARGIN", new_y="NEXT")
        self.ln(1)

        por_data = defaultdict(lambda: {"count": 0, "total": 0.0})
        for ev in eventos:
            por_data[ev["data"]]["count"] += 1
            por_data[ev["data"]]["total"] += ev["total_estimado"]

        # Header
        self._draw_table_header_calendario()

        self.set_font("Helvetica", "", 7.5)
        for idx, (data, info) in enumerate(sorted(por_data.items())):
            if self.get_y() > 185:
                self.add_page()
                self._draw_table_header_calendario()
                self.set_font("Helvetica", "", 7.5)

            if idx % 2 == 0:
                self.set_fill_color(*LIGHT_GRAY)
            else:
                self.set_fill_color(*WHITE)

            self.set_text_color(*DARK_GRAY)
            self.set_x(10)
            self.cell(45, 6, f"  {fmt_date_short(data)}", fill=True)
            self.cell(90, 6, f"  {fmt_date_full(data)}", fill=True)
            self.cell(40, 6, f"{info['count']} eventos", fill=True, align="C")
            self.cell(80, 6, f"{fmt_currency(info['total'])}  ", fill=True, align="R")
            self.ln(6)

    def _draw_table_header_calendario(self):
        self.set_fill_color(*GREEN)
        self.set_font("Helvetica", "B", 7)
        self.set_text_color(*WHITE)
        self.set_x(10)
        self.cell(45, 6, "  DATA", fill=True)
        self.cell(90, 6, "  DIA DA SEMANA", fill=True)
        self.cell(40, 6, "  EVENTOS", fill=True, align="C")
        self.cell(80, 6, "  TOTAL ESTIMADO", fill=True, align="R")
        self.ln(6)

    # ================================================================
    #  DETALHAMENTO POR DATA
    # ================================================================
    def add_detail_pages(self, eventos):
        """Páginas de detalhe: cada data com tabela de todos os clientes."""
        self.add_page()

        # Título da secao
        self.set_font("Helvetica", "B", 13)
        self.set_text_color(*GREEN)
        self.cell(0, 9, "Detalhamento por Data", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 8)
        self.set_text_color(*TEXT_GRAY)
        self.cell(0, 5, "Cada linha representa um fluxo individual de um cliente. Valores em Reais (R$).", new_x="LMARGIN", new_y="NEXT")
        self.ln(4)

        # Agrupar por data
        por_data = defaultdict(list)
        for ev in eventos:
            por_data[ev["data"]].append(ev)

        for data in sorted(por_data.keys()):
            evts = por_data[data]
            evts.sort(key=lambda e: (e["risco"], e["tipo_evento"], -e["total_estimado"]))
            total_data = sum(e["total_estimado"] for e in evts)

            # Verifica espaco (header data + header tabela + pelo menos 2 linhas)
            if self.get_y() > 165:
                self.add_page()

            # Banner da data
            y = self.get_y()
            self.set_fill_color(*GREEN)
            self.rect(10, y, 277, 8, "F")
            self.set_xy(12, y + 0.5)
            self.set_font("Helvetica", "B", 8.5)
            self.set_text_color(*WHITE)
            self.cell(140, 7, f"  {fmt_date_short(data)}   -   {fmt_date_full(data)}")
            self.set_font("Helvetica", "", 8)
            self.cell(125, 7, f"{len(evts)} fluxos  |  Total: {fmt_currency(total_data)}  ", align="R")
            self.ln(8)

            # Header da tabela
            self._draw_table_header_detail()

            # Linhas
            self.set_font("Helvetica", "", 7)
            cliente_num = 0
            for idx, ev in enumerate(evts):
                if self.get_y() > 185:
                    self.add_page()
                    # Repetir banner resumido
                    yb = self.get_y()
                    self.set_fill_color(*GREEN_LIGHT)
                    self.rect(10, yb, 277, 6, "F")
                    self.set_xy(12, yb)
                    self.set_font("Helvetica", "B", 7)
                    self.set_text_color(*WHITE)
                    self.cell(265, 6, f"  {fmt_date_short(data)}  -  {fmt_date_full(data)}  (continuação)")
                    self.ln(6)
                    self._draw_table_header_detail()
                    self.set_font("Helvetica", "", 7)

                cliente_num += 1

                if idx % 2 == 0:
                    self.set_fill_color(*LIGHT_GRAY)
                else:
                    self.set_fill_color(*WHITE)

                tipo_short = ev["tipo_evento"]
                if tipo_short == "PAGAMENTO DE JUROS":
                    tipo_short = "Pgto. Juros"
                elif tipo_short == "AMORTIZACAO":
                    tipo_short = "Amortizacao"
                elif tipo_short == "INCORPORACAO":
                    tipo_short = "Incorporacao"
                elif tipo_short == "PREMIO":
                    tipo_short = "Premio"

                risco = ev["risco"][:30] if len(ev["risco"]) > 30 else ev["risco"]

                self.set_text_color(*DARK_GRAY)
                self.set_x(10)
                # #  | Conta | Ativo | Cod.Ativo | Tipo | Qtd.Pos | Val.Unit | Total
                self.cell(12, 5.5, f"  {cliente_num}", fill=True, align="C")
                self.cell(22, 5.5, f"  {ev.get('cod_conta', '')[:10]}", fill=True)
                self.cell(62, 5.5, f"  {risco}", fill=True)
                self.cell(32, 5.5, f"  {ev['código_ativo'][:16]}", fill=True)
                self.cell(34, 5.5, f"  {tipo_short}", fill=True)
                self.cell(30, 5.5, f"{ev['qtd_posição']:,.0f}".replace(",", ".") + "  ", fill=True, align="R")
                self.cell(34, 5.5, f"{fmt_currency(ev['valor_estimado'])}  ", fill=True, align="R")
                self.cell(46, 5.5, f"{fmt_currency(ev['total_estimado'])}  ", fill=True, align="R")

                # Indicador de cor por tipo
                cor_tipo = BLUE if "JUROS" in ev["tipo_evento"] else GREEN
                if "INCORP" in ev["tipo_evento"]:
                    cor_tipo = (180, 120, 0)
                elif "PREMIO" in ev["tipo_evento"]:
                    cor_tipo = (140, 40, 140)
                self.set_fill_color(*cor_tipo)
                self.rect(10, self.get_y(), 1.2, 5.5, "F")

                self.ln(5.5)

            # Linha de total da data
            self.set_fill_color(*GREEN)
            self.set_text_color(*WHITE)
            self.set_font("Helvetica", "B", 7)
            self.set_x(10)
            self.cell(226, 5.5, f"  TOTAL {fmt_date_short(data)}  ", fill=True, align="R")
            self.cell(46, 5.5, f"{fmt_currency(total_data)}  ", fill=True, align="R")
            self.ln(5.5)

            self.set_font("Helvetica", "", 7)
            self.ln(5)

    def _draw_table_header_detail(self):
        self.set_fill_color(*BLUE)
        self.set_font("Helvetica", "B", 6.5)
        self.set_text_color(*WHITE)
        self.set_x(10)
        self.cell(12, 6, "  #", fill=True, align="C")
        self.cell(22, 6, "  CONTA", fill=True)
        self.cell(62, 6, "  ATIVO / RISCO", fill=True)
        self.cell(32, 6, "  COD. ATIVO", fill=True)
        self.cell(34, 6, "  TIPO EVENTO", fill=True)
        self.cell(30, 6, "  QTD. POSICAO", fill=True, align="R")
        self.cell(34, 6, "  VAL. UNITARIO", fill=True, align="R")
        self.cell(46, 6, "  TOTAL ESTIMADO", fill=True, align="R")
        self.ln(6)

    # ================================================================
    #  RESUMO POR ATIVO
    # ================================================================
    def add_asset_summary(self, eventos):
        """Página de resumo por ativo."""
        self.add_page()

        self.set_font("Helvetica", "B", 13)
        self.set_text_color(*GREEN)
        self.cell(0, 9, "Resumo por Ativo", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 8)
        self.set_text_color(*TEXT_GRAY)
        self.cell(0, 5, "Totais agrupados por ativo/risco, ordenados pelo maior valor total.", new_x="LMARGIN", new_y="NEXT")
        self.ln(3)

        por_risco = defaultdict(lambda: {"count": 0, "total": 0.0, "datas": set()})
        for ev in eventos:
            r = ev["risco"]
            por_risco[r]["count"] += 1
            por_risco[r]["total"] += ev["total_estimado"]
            por_risco[r]["datas"].add(ev["data"])

        # Header
        self.set_fill_color(*GREEN)
        self.set_font("Helvetica", "B", 7)
        self.set_text_color(*WHITE)
        self.set_x(10)
        self.cell(14, 6, "  #", fill=True, align="C")
        self.cell(85, 6, "  ATIVO / RISCO", fill=True)
        self.cell(35, 6, "  EVENTOS", fill=True, align="C")
        self.cell(60, 6, "  DATAS", fill=True)
        self.cell(73, 6, "  TOTAL ESTIMADO", fill=True, align="R")
        self.ln(6)

        sorted_riscos = sorted(por_risco.items(), key=lambda x: -x[1]["total"])
        grand_total = 0.0
        self.set_font("Helvetica", "", 7)

        for idx, (risco, info) in enumerate(sorted_riscos):
            if self.get_y() > 185:
                self.add_page()
                self.set_fill_color(*GREEN)
                self.set_font("Helvetica", "B", 7)
                self.set_text_color(*WHITE)
                self.set_x(10)
                self.cell(14, 6, "  #", fill=True, align="C")
                self.cell(85, 6, "  ATIVO / RISCO", fill=True)
                self.cell(35, 6, "  EVENTOS", fill=True, align="C")
                self.cell(60, 6, "  DATAS", fill=True)
                self.cell(73, 6, "  TOTAL ESTIMADO", fill=True, align="R")
                self.ln(6)
                self.set_font("Helvetica", "", 7)

            if idx % 2 == 0:
                self.set_fill_color(*LIGHT_GRAY)
            else:
                self.set_fill_color(*WHITE)

            datas_str = ", ".join(fmt_date_short(d) for d in sorted(info["datas"]))
            if len(datas_str) > 38:
                datas_str = datas_str[:35] + "..."

            risco_text = risco[:45] if len(risco) > 45 else risco

            self.set_text_color(*DARK_GRAY)
            self.set_x(10)
            self.cell(14, 5.5, f"{idx + 1}", fill=True, align="C")
            self.cell(85, 5.5, f"  {risco_text}", fill=True)
            self.cell(35, 5.5, f"{info['count']}", fill=True, align="C")
            self.cell(60, 5.5, f"  {datas_str}", fill=True)
            self.cell(73, 5.5, f"{fmt_currency(info['total'])}  ", fill=True, align="R")
            self.ln(5.5)
            grand_total += info["total"]

        # Total geral
        self.set_fill_color(*GREEN)
        self.set_text_color(*WHITE)
        self.set_font("Helvetica", "B", 7.5)
        self.set_x(10)
        self.cell(194, 6, f"  TOTAL GERAL  ({len(sorted_riscos)} ativos)", fill=True, align="R")
        self.cell(73, 6, f"{fmt_currency(grand_total)}  ", fill=True, align="R")
        self.ln(6)

    # ================================================================
    #  PAGINA SEM EVENTOS
    # ================================================================
    def add_no_events(self, assessor_info):
        self.ln(30)
        self.set_font("Helvetica", "", 14)
        self.set_text_color(*TEXT_GRAY)
        self.cell(0, 10, "Nenhum evento de renda fixa encontrado", new_x="LMARGIN", new_y="NEXT", align="C")
        self.set_font("Helvetica", "", 10)
        self.cell(0, 8, "para este assessor no período analisado.", new_x="LMARGIN", new_y="NEXT", align="C")

    # ================================================================
    #  DISCLAIMER
    # ================================================================
    def add_disclaimer(self):
        self.ln(6)
        self.set_draw_color(*MID_GRAY)
        self.line(10, self.get_y(), 287, self.get_y())
        self.ln(3)
        self.set_font("Helvetica", "I", 6)
        self.set_text_color(150, 150, 150)
        self.multi_cell(
            277, 3.5,
            "Este relatório e gerado automaticamente com base nos dados de eventos previstos. "
            "Os valores são estimativas e podem sofrer alterações. "
            "Para informações oficiais, consulte a plataforma XP Investimentos. "
            f"Gerado em {datetime.now().strftime('%d/%m/%Y as %H:%M')}.",
            align="C",
        )


def generate_pdf(assessor_código, assessor_info, eventos, output_dir):
    pdf = SomusPDF(assessor_info["nome"], assessor_código)
    pdf.alias_nb_pages()
    pdf.add_page()

    if not eventos:
        pdf.add_cover(assessor_info, eventos)
        pdf.add_no_events(assessor_info)
        pdf.add_disclaimer()
    else:
        eventos.sort(key=lambda e: (e["data"], e["risco"], e["tipo_evento"]))

        # Página 1: Capa com resumo
        pdf.add_cover(assessor_info, eventos)

        # Páginas de detalhe por data
        pdf.add_detail_pages(eventos)

        # Resumo por ativo
        pdf.add_asset_summary(eventos)

        # Disclaimer
        pdf.add_disclaimer()

    nome_limpo = assessor_info["nome"].replace(" ", "_").replace("/", "_")
    filename = f"Fluxo_RF_{assessor_código}_{nome_limpo}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


def main():
    print("=" * 60)
    print("  GERADOR DE PDFs - FLUXO DE RENDA FIXA")
    print("  Somus Capital")
    print("=" * 60)
    print()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("[1/3] Carregando base de assessores...")
    assessores = load_base_emails(BASE_FILE)
    print(f"       {len(assessores)} assessores encontrados na base.")

    print("[2/3] Carregando eventos de renda fixa...")
    eventos = load_eventos(ENTRADA_FILE)
    total_eventos = sum(len(v) for v in eventos.values())
    print(f"       {total_eventos} eventos carregados para {len(eventos)} assessores.")

    print("[3/3] Gerando PDFs...")
    print()

    gerados = 0
    erros = 0

    for código in sorted(assessores.keys()):
        info = assessores[código]
        evts = eventos.get(código, [])

        try:
            filepath = generate_pdf(código, info, evts, OUTPUT_DIR)
            status = f"{len(evts)} eventos" if evts else "SEM EVENTOS"
            total = fmt_currency(sum(e["total_estimado"] for e in evts)) if evts else "-"
            print(f"  [OK] {código} - {info['nome']:<25} | {status:<15} | {total}")
            gerados += 1
        except Exception as e:
            print(f"  [ERRO] {código} - {info['nome']}: {e}")
            import traceback
            traceback.print_exc()
            erros += 1

    print()
    print("=" * 60)
    print(f"  Concluído! {gerados} PDFs gerados, {erros} erros.")
    print(f"  Pasta de saida: {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
