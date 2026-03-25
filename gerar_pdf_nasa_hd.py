"""
Gerador de PDFs e Excel - NASA HD Engine
Gera 6 relatorios PDF profissionais + 1 workbook Excel completo
a partir dos resultados do motor NASA HD VPL.

Somus Capital - Mesa de Produtos
"""

import os
from datetime import datetime
from dataclasses import dataclass, field
from typing import Optional

from fpdf import FPDF

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "assets", "logo_somus.png")

# =====================================================================
# CORES - Paleta Somus Capital
# =====================================================================
GREEN = (0, 77, 51)
GREEN_LIGHT = (0, 92, 61)
BLUE = (24, 99, 220)
WHITE = (255, 255, 255)
LIGHT_GRAY = (246, 247, 249)
DARK_GRAY = (45, 45, 45)
MID_GRAY = (180, 180, 180)
TEXT_GRAY = (80, 80, 80)
ORANGE = (230, 131, 42)
PURPLE = (140, 40, 140)
RED = (180, 40, 40)
TEAL = (0, 130, 90)
RED_BG = (255, 230, 230)
GREEN_BG = (230, 255, 240)


# =====================================================================
# DATA CLASS - Container de Operacao Completa
# =====================================================================

@dataclass
class OperacaoCompleta:
    """Container with ALL data from a NASA HD calculation."""
    # Inputs
    params: dict = field(default_factory=dict)
    config: dict = field(default_factory=dict)

    # Core results
    fluxo: dict = field(default_factory=dict)
    vpl: dict = field(default_factory=dict)
    resumo: dict = field(default_factory=dict)
    parcelas: list = field(default_factory=list)

    # Optional results (may be None)
    financiamento: dict = None
    comparativo: dict = None
    venda: dict = None
    credito_lance: dict = None
    custo_combinado: dict = None
    credito_equivalente: float = None
    consolidacao: dict = None

    # Metadata
    cliente_nome: str = ""
    assessor: str = ""
    administradora: str = ""
    data_geracao: str = ""


# =====================================================================
# HELPERS
# =====================================================================

def sanitize_text(text):
    """Remove caracteres incompativeis com latin-1 (fpdf2)."""
    if text is None:
        return ""
    text = str(text)
    replacements = {
        '\u2013': '-', '\u2014': '-', '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2026': '...', '\u2022': '-',
        '\u00ba': 'o', '\u00aa': 'a', '\u00a0': ' ',
        '\ufeff': '', '\u200b': '', '\u200e': '', '\u200f': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    try:
        text.encode('latin-1')
    except UnicodeEncodeError:
        cleaned = []
        for ch in text:
            try:
                ch.encode('latin-1')
                cleaned.append(ch)
            except UnicodeEncodeError:
                cleaned.append('?')
        text = ''.join(cleaned)
    return text


def fmt_currency(value):
    """Formata valor monetario: R$ 1.234.567,89"""
    if value is None or value == 0:
        return "R$ 0,00"
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "R$ 0,00"
    neg = v < 0
    v = abs(v)
    inteiro = int(v)
    centavos = round((v - inteiro) * 100)
    if centavos >= 100:
        inteiro += 1
        centavos = 0
    s_int = f"{inteiro:,}".replace(",", ".")
    result = f"R$ {s_int},{centavos:02d}"
    return f"-{result}" if neg else result


def fmt_pct(value):
    """Formata percentual: 12,50%"""
    if value is None:
        return "0,00%"
    return f"{float(value):.2f}%".replace(".", ",")


def fmt_pct4(value):
    """Formata percentual com 4 casas: 0,0123%"""
    if value is None:
        return "0,0000%"
    return f"{float(value):.4f}%".replace(".", ",")


def fmt_num(value, decimals=2):
    """Formata numero generico com separador brasileiro."""
    if value is None:
        return "0"
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "0"
    formatted = f"{v:,.{decimals}f}"
    # Swap . and , for Brazilian format
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    return formatted


def _safe_get(d, *keys, default=0):
    """Busca segura em dicts aninhados."""
    for k in keys:
        if isinstance(d, dict):
            d = d.get(k, default)
        else:
            return default
    return d if d is not None else default


# =====================================================================
# BASE PDF CLASS - Estilo Somus Capital
# =====================================================================

class SomusPDF(FPDF):
    """Classe base para todos os PDFs Somus Capital (NASA HD)."""

    def __init__(self, orientation="P", title_text="RELATORIO NASA HD"):
        super().__init__(orientation=orientation, unit="mm",
                         format="A4")
        self.title_text = title_text
        self.set_auto_page_break(auto=True, margin=20)
        self.logo_path = LOGO_PATH if os.path.exists(LOGO_PATH) else None

    def header(self):
        page_w = self.w  # 210 portrait, 297 landscape
        self.set_fill_color(*GREEN)
        self.rect(0, 0, page_w, 18, "F")

        if self.logo_path:
            try:
                self.image(self.logo_path, x=8, y=3, h=12)
            except Exception:
                self.set_font("Helvetica", "B", 12)
                self.set_text_color(*WHITE)
                self.set_xy(8, 4)
                self.cell(50, 10, "SOMUS CAPITAL")

        self.set_font("Helvetica", "B", 11)
        self.set_text_color(*WHITE)
        self.set_xy(0, 4)
        self.cell(page_w, 10, sanitize_text(self.title_text), align="C")

        self.set_fill_color(*BLUE)
        self.rect(0, 18, page_w, 1.2, "F")
        self.set_y(23)

    def footer(self):
        page_w = self.w
        self.set_y(-14)
        self.set_draw_color(*MID_GRAY)
        self.line(10, self.get_y(), page_w - 10, self.get_y())
        self.ln(1.5)
        self.set_font("Helvetica", "", 6)
        self.set_text_color(140, 140, 140)
        self.cell(0, 5, "Somus Capital  |  Credito Estruturado - NASA HD", align="L")
        self.cell(0, 5, f"Pagina {self.page_no()}/{{nb}}", align="R")

    # -- Helpers de layout --

    def section_title(self, title):
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(*GREEN)
        self.cell(0, 8, sanitize_text(title), new_x="LMARGIN", new_y="NEXT")

    def section_subtitle(self, title):
        self.set_font("Helvetica", "B", 9)
        self.set_text_color(*BLUE)
        self.cell(0, 7, sanitize_text(title), new_x="LMARGIN", new_y="NEXT")

    def param_row(self, label, value, idx):
        """Linha parametro/valor com zebra."""
        self.set_fill_color(*(LIGHT_GRAY if idx % 2 == 0 else WHITE))
        self.set_x(10)
        self.set_font("Helvetica", "B", 8)
        self.set_text_color(*TEXT_GRAY)
        self.cell(80, 7, f"  {sanitize_text(label)}", fill=True)
        self.set_font("Helvetica", "", 8.5)
        self.set_text_color(*DARK_GRAY)
        self.cell(110, 7, f"  {sanitize_text(str(value))}", fill=True)
        self.ln(7)

    def table_header(self, cols, widths, row_h=6):
        """Cabecalho de tabela com fundo azul."""
        self.set_fill_color(*BLUE)
        self.set_font("Helvetica", "B", 6)
        self.set_text_color(*WHITE)
        self.set_x(10)
        for col, w in zip(cols, widths):
            self.cell(w, row_h, sanitize_text(col), fill=True, align="C")
        self.ln(row_h)

    def table_row(self, values, widths, idx, row_h=5.5, aligns=None,
                  bold_mask=None, font_size=6.5):
        """Linha de tabela com zebra."""
        self.set_fill_color(*(LIGHT_GRAY if idx % 2 == 0 else WHITE))
        self.set_font("Helvetica", "", font_size)
        self.set_text_color(*DARK_GRAY)
        self.set_x(10)
        if aligns is None:
            aligns = ["C"] + ["R"] * (len(values) - 1)
        if bold_mask is None:
            bold_mask = [False] * len(values)
        for val, w, al, bld in zip(values, widths, aligns, bold_mask):
            if bld:
                self.set_font("Helvetica", "B", font_size)
            else:
                self.set_font("Helvetica", "", font_size)
            self.cell(w, row_h, sanitize_text(str(val)), fill=True, align=al)
        self.ln(row_h)

    def table_total_row(self, values, widths, row_h=6, aligns=None, font_size=6.5):
        """Linha de total com fundo verde."""
        self.set_fill_color(*GREEN)
        self.set_font("Helvetica", "B", font_size)
        self.set_text_color(*WHITE)
        self.set_x(10)
        if aligns is None:
            aligns = ["C"] + ["R"] * (len(values) - 1)
        for val, w, al in zip(values, widths, aligns):
            self.cell(w, row_h, sanitize_text(str(val)), fill=True, align=al)
        self.ln(row_h)

    def kpi_box(self, x, y, w, h, label, value, color=GREEN):
        """Caixa de KPI destacada."""
        self.set_fill_color(*LIGHT_GRAY)
        self.rect(x, y, w, h, "F")
        self.set_fill_color(*color)
        self.rect(x, y, w, 8, "F")
        self.set_xy(x + 2, y + 1)
        self.set_font("Helvetica", "B", 7)
        self.set_text_color(*WHITE)
        self.cell(w - 4, 6, sanitize_text(label), align="C")
        self.set_xy(x + 2, y + 11)
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*color)
        self.cell(w - 4, 6, sanitize_text(str(value)), align="C")

    def disclaimer(self):
        """Rodape legal."""
        self.ln(6)
        self.set_draw_color(*MID_GRAY)
        self.line(10, self.get_y(), self.w - 10, self.get_y())
        self.ln(3)
        self.set_font("Helvetica", "I", 6)
        self.set_text_color(150, 150, 150)
        self.multi_cell(
            self.w - 20, 3.5,
            "Esta analise e meramente ilustrativa e nao constitui oferta ou proposta formal. "
            "Os valores apresentados sao estimativas baseadas nos parametros informados e podem "
            "sofrer alteracoes conforme regras da administradora do consorcio. "
            "As parcelas podem ser reajustadas conforme indice previsto em contrato. "
            "Para informacoes oficiais, consulte a administradora ou seu assessor. "
            f"Relatorio gerado em {datetime.now().strftime('%d/%m/%Y as %H:%M')}.",
            align="C",
        )

    def check_page_break(self, needed_h=15):
        """Verifica se precisa quebra de pagina."""
        if self.get_y() + needed_h > self.h - 20:
            self.add_page()
            return True
        return False


# =====================================================================
# REPORT 1: RESUMO EXECUTIVO
# =====================================================================

def gerar_pdf_resumo_executivo(op: OperacaoCompleta, output_dir: str) -> str:
    """Gera PDF de resumo executivo (client-facing)."""
    os.makedirs(output_dir, exist_ok=True)
    pdf = SomusPDF(title_text="RELATORIO DE CREDITO ESTRUTURADO")
    pdf.alias_nb_pages()

    resumo = op.resumo or {}
    fluxo = op.fluxo or {}
    vpl_data = op.vpl or {}
    params = op.params or {}
    totais = fluxo.get("totais", {})
    metricas = fluxo.get("metricas", {})

    credito = params.get("valor_credito", resumo.get("valor_credito", 0))
    prazo = params.get("prazo_meses", resumo.get("prazo_meses", 0))
    contemp = params.get("momento_contemplacao", resumo.get("momento_contemplacao", 0))
    carta_liq = totais.get("carta_liquida", resumo.get("carta_liquida", 0))
    lance_emb = totais.get("lance_embutido_valor", resumo.get("lance_embutido", 0))
    lance_livre = totais.get("lance_livre_valor", resumo.get("lance_livre", 0))
    lance_total = lance_emb + lance_livre
    lance_emb_pct = (lance_emb / credito * 100) if credito > 0 else 0
    lance_livre_pct = (lance_livre / credito * 100) if credito > 0 else 0

    total_pago = totais.get("total_pago", resumo.get("total_pago", 0))
    total_fc = totais.get("total_fundo_comum", resumo.get("total_fundo_comum", 0))
    total_ta = totais.get("total_taxa_adm", resumo.get("total_taxa_adm", 0))
    total_fr = totais.get("total_fundo_reserva", resumo.get("total_fundo_reserva", 0))
    total_seg = totais.get("total_seguro", resumo.get("total_seguro", 0))
    total_custos_ac = totais.get("total_custos_acessorios", resumo.get("total_custos_acessorios", 0))

    tir_m = vpl_data.get("tir_mensal", metricas.get("tir_mensal", 0))
    tir_a = vpl_data.get("tir_anual", metricas.get("tir_anual", 0))
    cet_a = vpl_data.get("cet_anual", metricas.get("cet_anual", 0))
    delta_vpl = vpl_data.get("delta_vpl", resumo.get("delta_vpl", 0))
    cria_valor = vpl_data.get("cria_valor", resumo.get("cria_valor", False))
    be_lance = vpl_data.get("break_even_lance", resumo.get("break_even_lance", 0))
    custo_total_pct = metricas.get("custo_total_pct", resumo.get("custo_total_pct", 0))

    # ---- CAPA ----
    pdf.add_page()
    pdf.ln(15)

    pdf.set_font("Helvetica", "B", 24)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 14, "Relatorio de Credito Estruturado", new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(*TEXT_GRAY)
    pdf.cell(0, 7, "Analise completa da operacao de consorcio  -  Somus Capital",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    # Card do cliente
    y = pdf.get_y()
    pdf.set_fill_color(*LIGHT_GRAY)
    pdf.rect(10, y, 190, 30, "F")
    pdf.set_fill_color(*GREEN)
    pdf.rect(10, y, 3, 30, "F")

    pdf.set_xy(18, y + 4)
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 8, sanitize_text(op.cliente_nome or "Cliente"))

    pdf.set_xy(18, y + 14)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*DARK_GRAY)
    assessor_txt = f"Assessor: {sanitize_text(op.assessor)}" if op.assessor else ""
    adm_txt = f"Administradora: {sanitize_text(op.administradora)}" if op.administradora else ""
    data_txt = f"Data: {op.data_geracao or datetime.now().strftime('%d/%m/%Y')}"
    pdf.cell(0, 6, f"{assessor_txt}     {adm_txt}")
    pdf.set_xy(18, y + 22)
    pdf.cell(0, 6, data_txt)

    pdf.set_y(y + 36)

    # Destaques
    pdf.ln(4)
    y = pdf.get_y()
    box_w = 60
    pdf.kpi_box(10, y, box_w, 26, "VALOR DO CREDITO", fmt_currency(credito), GREEN)
    pdf.kpi_box(75, y, box_w, 26, "CARTA LIQUIDA", fmt_currency(carta_liq), BLUE)
    pdf.kpi_box(140, y, box_w, 26, "PRAZO TOTAL", f"{prazo} meses", ORANGE)
    pdf.set_y(y + 32)

    y = pdf.get_y()
    pdf.kpi_box(10, y, box_w, 26, "CET ANUAL", fmt_pct(cet_a * 100), GREEN)
    valor_label = "CRIA VALOR" if cria_valor else "DESTROI VALOR"
    valor_color = GREEN if cria_valor else RED
    pdf.kpi_box(75, y, box_w, 26, valor_label, fmt_currency(delta_vpl), valor_color)
    pdf.kpi_box(140, y, box_w, 26, "TIR ANUAL", fmt_pct(tir_a * 100), BLUE)
    pdf.set_y(y + 32)

    pdf.disclaimer()

    # ---- PAGINA 2: RESUMO DA OPERACAO ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Resumo da Operacao", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Dados Principais
    pdf.section_title("Dados Principais")
    pdf.ln(1)

    dados_principais = [
        ("Valor do Credito", fmt_currency(credito)),
        ("Carta Liquida", fmt_currency(carta_liq)),
        ("Prazo Total", f"{prazo} meses"),
        ("Contemplacao", f"{contemp}o mes"),
        ("Lance Embutido", f"{fmt_pct(lance_emb_pct)} ({fmt_currency(lance_emb)})"),
        ("Lance Livre", f"{fmt_pct(lance_livre_pct)} ({fmt_currency(lance_livre)})"),
        ("Total Lances", fmt_currency(lance_total)),
    ]
    for i, (label, value) in enumerate(dados_principais):
        pdf.param_row(label, value, i)

    pdf.ln(4)

    # Custo da Operacao
    pdf.section_title("Custo da Operacao")
    pdf.ln(1)

    # Header
    cols_custo = ["Componente", "Total", "Mensal"]
    w_custo = [80, 55, 55]
    pdf.table_header(cols_custo, w_custo)

    custo_mensal_fc = total_fc / prazo if prazo > 0 else 0
    custo_mensal_ta = total_ta / prazo if prazo > 0 else 0
    custo_mensal_fr = total_fr / prazo if prazo > 0 else 0
    custo_mensal_seg = total_seg / prazo if prazo > 0 else 0
    custo_mensal_ac = total_custos_ac / prazo if prazo > 0 else 0
    total_custo = total_fc + total_ta + total_fr + total_seg + total_custos_ac
    total_custo_mensal = total_custo / prazo if prazo > 0 else 0

    custo_rows = [
        ("Fundo Comum", fmt_currency(total_fc), fmt_currency(custo_mensal_fc)),
        ("Taxa Administracao", fmt_currency(total_ta), fmt_currency(custo_mensal_ta)),
        ("Fundo Reserva", fmt_currency(total_fr), fmt_currency(custo_mensal_fr)),
        ("Seguro", fmt_currency(total_seg), fmt_currency(custo_mensal_seg)),
        ("Custos Acessorios", fmt_currency(total_custos_ac), fmt_currency(custo_mensal_ac)),
    ]
    for i, row in enumerate(custo_rows):
        pdf.table_row(list(row), w_custo, i, aligns=["L", "R", "R"])

    pdf.table_total_row(["TOTAL", fmt_currency(total_custo), fmt_currency(total_custo_mensal)],
                        w_custo, aligns=["L", "R", "R"])

    pdf.ln(4)

    # Indicadores
    pdf.section_title("Indicadores")
    pdf.ln(1)

    cet_m = vpl_data.get("tir_mensal", metricas.get("tir_mensal", 0))

    y = pdf.get_y()
    box_w2 = 46
    gap = 3
    x0 = 10
    pdf.kpi_box(x0, y, box_w2, 26, "CET Anual", fmt_pct(cet_a * 100), GREEN)
    pdf.kpi_box(x0 + box_w2 + gap, y, box_w2, 26, "CET Mensal", fmt_pct(cet_m * 100), GREEN)
    pdf.kpi_box(x0 + 2 * (box_w2 + gap), y, box_w2, 26, "TIR Anual", fmt_pct(tir_a * 100), BLUE)
    pdf.kpi_box(x0 + 3 * (box_w2 + gap), y, box_w2 - 3, 26, "Custo Total",
                fmt_pct(custo_total_pct), ORANGE)
    pdf.set_y(y + 30)

    y = pdf.get_y()
    valor_color2 = GREEN if cria_valor else RED
    valor_label2 = "CRIA VALOR" if cria_valor else "DESTROI VALOR"
    pdf.kpi_box(x0, y, 93, 26, f"Delta VPL ({valor_label2})", fmt_currency(delta_vpl), valor_color2)
    pdf.kpi_box(x0 + 96, y, 94, 26, "Break-even Lance", fmt_pct(be_lance), BLUE)
    pdf.set_y(y + 30)

    # ---- PAGINA 3: PARCELAS RESUMO ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Resumo de Parcelas", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    primeira_parcela = resumo.get("primeira_parcela", 0)
    ultima_parcela = resumo.get("ultima_parcela", 0)
    parcela_media = metricas.get("parcela_media", resumo.get("parcela_media", 0))
    parcela_max = metricas.get("parcela_maxima", resumo.get("parcela_maxima", 0))
    parcela_min = metricas.get("parcela_minima", resumo.get("parcela_minima", 0))

    # Parcelas do fluxo
    fluxo_mensal = fluxo.get("fluxo", fluxo.get("fluxo_mensal", []))
    if fluxo_mensal and len(fluxo_mensal) > 1:
        # Parcela base fase 1 e fase 2
        f1_row = fluxo_mensal[1] if len(fluxo_mensal) > 1 else {}
        parcela_f1 = f1_row.get("valor_parcela", 0)
        # Encontrar primeira parcela fase 2
        parcela_f2 = 0
        for r in fluxo_mensal:
            if r.get("mes", 0) > contemp:
                parcela_f2 = r.get("valor_parcela", 0)
                break
    else:
        parcela_f1 = 0
        parcela_f2 = 0

    pdf.section_title("Estatisticas de Parcelas")
    pdf.ln(1)

    parc_stats = [
        ("Primeira Parcela", fmt_currency(primeira_parcela)),
        ("Ultima Parcela", fmt_currency(ultima_parcela)),
        ("Parcela Media", fmt_currency(parcela_media)),
        ("Parcela Maxima", fmt_currency(parcela_max)),
        ("Parcela Minima", fmt_currency(parcela_min)),
        ("Parcela Base Fase 1 (pre-contemplacao)", fmt_currency(parcela_f1)),
        ("Parcela Base Fase 2 (pos-contemplacao)", fmt_currency(parcela_f2)),
    ]
    for i, (label, value) in enumerate(parc_stats):
        pdf.param_row(label, value, i)

    pdf.ln(4)

    # Reajuste
    pdf.section_title("Informacoes de Reajuste")
    pdf.ln(1)

    reaj_pre = params.get("reajuste_pre_pct", resumo.get("reajuste_pre_pct", 0))
    reaj_pos = params.get("reajuste_pos_pct", resumo.get("reajuste_pos_pct", 0))
    freq_pre = params.get("reajuste_pre_freq", resumo.get("reajuste_pre_freq", "Anual"))
    freq_pos = params.get("reajuste_pos_freq", resumo.get("reajuste_pos_freq", "Anual"))

    reaj_rows = [
        ("Reajuste Pre-Contemplacao", f"{fmt_pct(reaj_pre)} {freq_pre}"),
        ("Reajuste Pos-Contemplacao", f"{fmt_pct(reaj_pos)} {freq_pos}"),
    ]
    for i, (label, value) in enumerate(reaj_rows):
        pdf.param_row(label, value, i)

    pdf.disclaimer()

    # Output
    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Resumo_Executivo_{nome}_{ts}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


# =====================================================================
# REPORT 2: FLUXO FINANCEIRO COMPLETO (Landscape)
# =====================================================================

def gerar_pdf_fluxo_financeiro(op: OperacaoCompleta, output_dir: str) -> str:
    """Gera PDF com fluxo financeiro completo em landscape."""
    os.makedirs(output_dir, exist_ok=True)

    fluxo = op.fluxo or {}
    fluxo_mensal = fluxo.get("fluxo", fluxo.get("fluxo_mensal", []))
    totais = fluxo.get("totais", {})
    cliente = sanitize_text(op.cliente_nome or "Cliente")

    # Divide em 4 grupos de colunas
    # Group A: Mes | Meses Rest | FC Base | Lance Emb | Lance Livre | Base Final | Amortiz | Saldo Principal
    # Group B: Mes | %FC | %FC Acum | Tx Adm | %TA | %TA Acum | Fdo Res | %FR | %FR Acum
    # Group C: Mes | Parcela Base | Saldo Dev | Peso | %Reaj | %Reaj Acum | Parcela Reaj | Saldo Dev Reaj
    # Group D: Mes | Seguro | Parc c/ Seguro | Outros Custos | Carta Cred | Carta Reaj | Fluxo Caixa | Fluxo TIR

    groups = [
        {
            "title": f"Fluxo Financeiro - Grupo A: Amortizacao - {cliente}",
            "cols": ["Mes", "Rest.", "FC Base", "Lance Emb.", "Lance Livre",
                     "Base Final", "Amortiz.", "Saldo Princ."],
            "widths": [18, 18, 38, 38, 38, 38, 38, 51],
            "keys": ["mes", "meses_restantes", "valor_base_fc", "lance_embutido",
                     "lance_livre", "valor_base_final", "amortizacao", "saldo_principal"],
            "formats": ["int", "int", "cur", "cur", "cur", "cur", "cur", "cur"],
        },
        {
            "title": f"Fluxo Financeiro - Grupo B: Taxas - {cliente}",
            "cols": ["Mes", "% FC", "% FC Ac.", "Tx Adm.", "% TA", "% TA Ac.",
                     "Fdo Res.", "% FR", "% FR Ac."],
            "widths": [18, 30, 30, 35, 30, 30, 35, 30, 39],
            "keys": ["mes", "pct_mensal_fc", "pct_acum_fc", "valor_parcela_ta",
                     "pct_ta_mensal", "pct_ta_acum", "valor_fundo_reserva",
                     "pct_fr_mensal", "pct_fr_acum"],
            "formats": ["int", "pct4", "pct4", "cur", "pct4", "pct4", "cur", "pct4", "pct4"],
        },
        {
            "title": f"Fluxo Financeiro - Grupo C: Reajuste - {cliente}",
            "cols": ["Mes", "Parc. Base", "Saldo Dev.", "Peso",
                     "% Reaj.", "% Reaj. Ac.", "Parc. Reaj.", "Saldo Dev. Reaj."],
            "widths": [18, 35, 40, 25, 30, 35, 37, 57],
            "keys": ["mes", "valor_parcela", "saldo_devedor", "peso_parcela",
                     "pct_reajuste", "pct_reajuste_acum", "parcela_apos_reajuste",
                     "saldo_devedor_reajustado"],
            "formats": ["int", "cur", "cur", "pct4", "pct4", "pct4", "cur", "cur"],
        },
        {
            "title": f"Fluxo Financeiro - Grupo D: Fluxo de Caixa - {cliente}",
            "cols": ["Mes", "Seguro", "Parc.+Seguro", "Outros Cust.",
                     "Carta Cred.", "Carta Reaj.", "Fluxo Caixa", "Fluxo TIR"],
            "widths": [18, 33, 40, 33, 38, 38, 38, 39],
            "keys": ["mes", "seguro_vida", "parcela_com_seguro", "outros_custos",
                     "carta_credito_original", "carta_credito_reajustada",
                     "fluxo_caixa", "fluxo_caixa_tir"],
            "formats": ["int", "cur", "cur", "cur", "cur", "cur", "cur", "cur"],
        },
    ]

    pdf = SomusPDF(orientation="L", title_text="FLUXO FINANCEIRO DO CONSORCIO")
    pdf.alias_nb_pages()

    for grp in groups:
        pdf.add_page()
        pdf.ln(1)
        pdf.section_title(grp["title"])
        pdf.ln(1)
        pdf.table_header(grp["cols"], grp["widths"])

        for idx, row in enumerate(fluxo_mensal):
            if pdf.get_y() > 185:
                pdf.add_page()
                pdf.ln(1)
                pdf.section_subtitle(grp["title"] + " (cont.)")
                pdf.ln(1)
                pdf.table_header(grp["cols"], grp["widths"])

            vals = []
            for key, fmt_type in zip(grp["keys"], grp["formats"]):
                v = row.get(key, 0)
                if fmt_type == "int":
                    vals.append(str(int(v)) if v is not None else "0")
                elif fmt_type == "cur":
                    vals.append(fmt_currency(v))
                elif fmt_type == "pct4":
                    vals.append(fmt_pct4(v * 100 if v is not None else 0))
                elif fmt_type == "pct":
                    vals.append(fmt_pct(v * 100 if v is not None else 0))
                else:
                    vals.append(str(v))

            pdf.table_row(vals, grp["widths"], idx, font_size=6, row_h=5)

        # Totals row
        if fluxo_mensal:
            tot_vals = []
            for key, fmt_type in zip(grp["keys"], grp["formats"]):
                if key == "mes":
                    tot_vals.append("TOTAL")
                elif fmt_type == "cur":
                    total_val = sum(r.get(key, 0) for r in fluxo_mensal)
                    tot_vals.append(fmt_currency(total_val))
                elif fmt_type in ("pct4", "pct"):
                    tot_vals.append("")
                else:
                    tot_vals.append("")
            pdf.table_total_row(tot_vals, grp["widths"], font_size=6)

    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Fluxo_Financeiro_{nome}_{ts}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


# =====================================================================
# REPORT 3: ANALISE VPL
# =====================================================================

def gerar_pdf_analise_vpl(op: OperacaoCompleta, output_dir: str) -> str:
    """Gera PDF com analise VPL HD."""
    os.makedirs(output_dir, exist_ok=True)

    pdf = SomusPDF(title_text="ANALISE VPL HD")
    pdf.alias_nb_pages()

    vpl_data = op.vpl or {}
    params = op.params or {}
    resumo = op.resumo or {}
    fluxo = op.fluxo or {}
    totais = fluxo.get("totais", {})

    credito = params.get("valor_credito", 0)
    carta_liq = totais.get("carta_liquida", resumo.get("carta_liquida", 0))
    b0 = vpl_data.get("b0", 0)
    h0 = vpl_data.get("h0", 0)
    d0 = vpl_data.get("d0", 0)
    pv_pos_t = vpl_data.get("pv_pos_t", 0)
    delta_vpl = vpl_data.get("delta_vpl", 0)
    cria_valor = vpl_data.get("cria_valor", False)
    be_lance = vpl_data.get("break_even_lance", 0)
    tir_m = vpl_data.get("tir_mensal", 0)
    tir_a = vpl_data.get("tir_anual", 0)
    vpl_total = vpl_data.get("vpl_total", 0)

    alm_a = params.get("alm_anual", 12.0)
    hurdle_a = params.get("hurdle_anual", 12.0)

    # ---- PAGINA 1 ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Analise de VPL", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Banner resultado
    y = pdf.get_y()
    if cria_valor:
        banner_color = GREEN
        banner_text = "Esta operacao CRIA VALOR"
    else:
        banner_color = RED
        banner_text = "Esta operacao DESTROI VALOR"

    pdf.set_fill_color(*banner_color)
    pdf.rect(10, y, 190, 22, "F")
    pdf.set_xy(15, y + 3)
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*WHITE)
    pdf.cell(0, 5, "RESULTADO DA ANALISE VPL HD")
    pdf.set_xy(15, y + 10)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(100, 8, sanitize_text(banner_text))
    pdf.set_xy(140, y + 10)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(55, 8, fmt_currency(delta_vpl), align="R")
    pdf.set_y(y + 26)

    # Decomposicao VPL
    pdf.ln(2)
    pdf.section_title("Decomposicao VPL")
    pdf.ln(1)

    cols_vpl = ["Componente", "Valor", "% do Credito"]
    w_vpl = [90, 50, 50]
    pdf.table_header(cols_vpl, w_vpl)

    cheque_liq_pct = (carta_liq / credito * 100) if credito > 0 else 100
    b0_pct = (b0 / credito * 100) if credito > 0 else 0
    h0_pct = (h0 / credito * 100) if credito > 0 else 0
    d0_pct = (d0 / credito * 100) if credito > 0 else 0
    pv_pos_pct = (pv_pos_t / credito * 100) if credito > 0 else 0
    delta_pct = (delta_vpl / credito * 100) if credito > 0 else 0

    decomp_rows = [
        ("Cheque liquido em T", fmt_currency(carta_liq), fmt_pct(cheque_liq_pct)),
        ("B0 (VP Credito @ ALM)", fmt_currency(b0), fmt_pct(b0_pct)),
        ("H0 (VP Pagtos pre-T)", fmt_currency(h0), fmt_pct(h0_pct)),
        ("D0 (Valor criado)", fmt_currency(d0), fmt_pct(d0_pct)),
        ("PV pos-T (@ Hurdle)", fmt_currency(pv_pos_t), fmt_pct(pv_pos_pct)),
    ]
    for i, row in enumerate(decomp_rows):
        pdf.table_row(list(row), w_vpl, i, aligns=["L", "R", "R"])

    # Total row
    pdf.table_total_row(
        ["Delta VPL", fmt_currency(delta_vpl), fmt_pct(delta_pct)],
        w_vpl, aligns=["L", "R", "R"]
    )

    pdf.ln(4)

    # Parametros VPL
    pdf.section_title("Parametros VPL")
    pdf.ln(1)

    vpl_params = [
        ("ALM/CDI", f"{fmt_pct(alm_a)} a.a."),
        ("Hurdle", f"{fmt_pct(hurdle_a)} a.a."),
        ("Break-even Lance", fmt_pct(be_lance)),
        ("TIR Mensal", fmt_pct(tir_m * 100)),
        ("TIR Anual", fmt_pct(tir_a * 100)),
        ("VPL Total (@ ALM)", fmt_currency(vpl_total)),
    ]
    for i, (label, value) in enumerate(vpl_params):
        pdf.param_row(label, value, i)

    # ---- PAGINA 2+: VP por Mes ----
    pv_pre = vpl_data.get("pv_pre_t_detail", [])
    pv_pos = vpl_data.get("pv_pos_t_detail", [])

    if pv_pre or pv_pos:
        pdf.add_page()
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*GREEN)
        pdf.cell(0, 10, "VP por Mes", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Pre-contemplacao
        if pv_pre:
            pdf.section_title("Pagamentos Pre-Contemplacao (descontados @ ALM)")
            pdf.ln(1)

            cols_pre = ["Mes", "Pagamento", "VP @ ALM"]
            w_pre = [30, 80, 80]
            pdf.table_header(cols_pre, w_pre)

            for i, item in enumerate(pv_pre):
                if pdf.get_y() > 265:
                    pdf.add_page()
                    pdf.section_subtitle("VP Pre-Contemplacao (cont.)")
                    pdf.ln(1)
                    pdf.table_header(cols_pre, w_pre)
                pdf.table_row(
                    [str(item.get("mes", 0)),
                     fmt_currency(item.get("valor", 0)),
                     fmt_currency(item.get("pv", 0))],
                    w_pre, i, aligns=["C", "R", "R"]
                )

            total_pre_val = sum(it.get("valor", 0) for it in pv_pre)
            total_pre_pv = sum(it.get("pv", 0) for it in pv_pre)
            pdf.table_total_row(
                ["TOTAL", fmt_currency(total_pre_val), fmt_currency(total_pre_pv)],
                w_pre, aligns=["C", "R", "R"]
            )
            pdf.ln(4)

        # Pos-contemplacao
        if pv_pos:
            pdf.check_page_break(30)
            pdf.section_title("Pagamentos Pos-Contemplacao (descontados @ Hurdle)")
            pdf.ln(1)

            cols_pos = ["Mes", "Pagamento", "VP @ Hurdle"]
            w_pos = [30, 80, 80]
            pdf.table_header(cols_pos, w_pos)

            for i, item in enumerate(pv_pos):
                if pdf.get_y() > 265:
                    pdf.add_page()
                    pdf.section_subtitle("VP Pos-Contemplacao (cont.)")
                    pdf.ln(1)
                    pdf.table_header(cols_pos, w_pos)
                pdf.table_row(
                    [str(item.get("mes", 0)),
                     fmt_currency(item.get("valor", 0)),
                     fmt_currency(item.get("pv", 0))],
                    w_pos, i, aligns=["C", "R", "R"]
                )

            total_pos_val = sum(it.get("valor", 0) for it in pv_pos)
            total_pos_pv = sum(it.get("pv", 0) for it in pv_pos)
            pdf.table_total_row(
                ["TOTAL", fmt_currency(total_pos_val), fmt_currency(total_pos_pv)],
                w_pos, aligns=["C", "R", "R"]
            )

    pdf.disclaimer()

    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Analise_VPL_{nome}_{ts}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


# =====================================================================
# REPORT 4: COMPARATIVO CONSORCIO VS FINANCIAMENTO
# =====================================================================

def gerar_pdf_comparativo(op: OperacaoCompleta, output_dir: str) -> str:
    """Gera PDF comparativo consorcio vs financiamento."""
    os.makedirs(output_dir, exist_ok=True)

    if not op.comparativo:
        return ""

    pdf = SomusPDF(title_text="COMPARATIVO: CONSORCIO VS FINANCIAMENTO")
    pdf.alias_nb_pages()

    comp = op.comparativo
    params = op.params or {}
    fluxo = op.fluxo or {}

    total_cons = comp.get("total_pago_consorcio", 0)
    total_fin = comp.get("total_pago_financiamento", 0)
    econ_nominal = comp.get("economia_nominal", 0)
    vpl_cons = comp.get("vpl_consorcio", 0)
    vpl_fin = comp.get("vpl_financiamento", 0)
    econ_vpl = comp.get("economia_vpl", 0)
    tir_c_a = comp.get("tir_consorcio_anual", 0)
    tir_f_a = comp.get("tir_financ_anual", 0)
    razao_c = comp.get("razao_vpl_consorcio", 0)
    razao_f = comp.get("razao_vpl_financ", 0)

    # ---- PAGINA 1: RESUMO COMPARATIVO ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Resumo Comparativo", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Banner vantagem
    if econ_vpl > 0:
        vant_text = f"Consorcio e {fmt_currency(abs(econ_vpl))} mais economico em termos de VPL"
        vant_color = GREEN
    else:
        vant_text = f"Financiamento e {fmt_currency(abs(econ_vpl))} mais economico em termos de VPL"
        vant_color = BLUE

    y = pdf.get_y()
    pdf.set_fill_color(*vant_color)
    pdf.rect(10, y, 190, 14, "F")
    pdf.set_xy(15, y + 3)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*WHITE)
    pdf.cell(180, 8, sanitize_text(vant_text), align="C")
    pdf.set_y(y + 18)

    # Side-by-side KPIs
    pdf.ln(2)
    cols_comp = ["Indicador", "Consorcio", "Financiamento", "Diferenca"]
    w_comp = [55, 45, 45, 45]
    pdf.table_header(cols_comp, w_comp)

    comp_rows = [
        ("Total Pago", fmt_currency(total_cons), fmt_currency(total_fin),
         fmt_currency(total_cons - total_fin)),
        ("VPL", fmt_currency(vpl_cons), fmt_currency(vpl_fin),
         fmt_currency(econ_vpl)),
        ("TIR Anual", fmt_pct(tir_c_a * 100), fmt_pct(tir_f_a * 100), ""),
        ("Custo/Credito", fmt_num(razao_c, 2) + "x", fmt_num(razao_f, 2) + "x", ""),
        ("Economia Nominal", "", "", fmt_currency(econ_nominal)),
    ]
    for i, row in enumerate(comp_rows):
        pdf.table_row(list(row), w_comp, i, aligns=["L", "R", "R", "R"])

    # ---- PAGINA 2: FLUXOS NOMINAIS ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Fluxos Nominais", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    cf_cons = comp.get("consorcio", {}).get("cashflow", [])
    cf_fin_result = comp.get("financiamento", {})
    cf_fin = cf_fin_result.get("cashflow", [])
    max_len = max(len(cf_cons), len(cf_fin))

    cols_nom = ["Mes", "Cons. Parcela", "Cons. Acum.", "Financ. Parcela", "Financ. Acum.", "Diferenca"]
    w_nom = [18, 34, 34, 34, 34, 36]
    pdf.table_header(cols_nom, w_nom)

    acum_c = 0.0
    acum_f = 0.0
    for m in range(max_len):
        if pdf.get_y() > 265:
            pdf.add_page()
            pdf.section_subtitle("Fluxos Nominais (cont.)")
            pdf.ln(1)
            pdf.table_header(cols_nom, w_nom)

        vc = abs(cf_cons[m]) if m < len(cf_cons) else 0
        vf = abs(cf_fin[m]) if m < len(cf_fin) else 0
        # Skip month 0 accumulation for credit received
        if m > 0:
            acum_c += vc
            acum_f += vf

        pdf.table_row(
            [str(m), fmt_currency(vc), fmt_currency(acum_c),
             fmt_currency(vf), fmt_currency(acum_f), fmt_currency(vc - vf)],
            w_nom, m, font_size=6, row_h=5, aligns=["C", "R", "R", "R", "R", "R"]
        )

    # ---- PAGINA 3: FLUXOS EM VALOR PRESENTE ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Fluxos em Valor Presente", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pv_cons = comp.get("pv_consorcio", [])
    pv_fin = comp.get("pv_financiamento", [])
    max_pv_len = max(len(pv_cons), len(pv_fin))

    cols_pv = ["Mes", "VP Cons.", "VP Acum. Cons.", "VP Financ.", "VP Acum. Financ."]
    w_pv = [18, 40, 42, 40, 50]
    pdf.table_header(cols_pv, w_pv)

    acum_pvc = 0.0
    acum_pvf = 0.0
    for m in range(max_pv_len):
        if pdf.get_y() > 265:
            pdf.add_page()
            pdf.section_subtitle("Fluxos em VP (cont.)")
            pdf.ln(1)
            pdf.table_header(cols_pv, w_pv)

        vpc = pv_cons[m] if m < len(pv_cons) else 0
        vpf = pv_fin[m] if m < len(pv_fin) else 0
        acum_pvc += vpc
        acum_pvf += vpf

        pdf.table_row(
            [str(m), fmt_currency(vpc), fmt_currency(acum_pvc),
             fmt_currency(vpf), fmt_currency(acum_pvf)],
            w_pv, m, font_size=6, row_h=5, aligns=["C", "R", "R", "R", "R"]
        )

    # ---- PAGINA 4: TABELA AMORTIZACAO FINANCIAMENTO ----
    fin_parcelas = cf_fin_result.get("parcelas", [])
    if fin_parcelas:
        pdf.add_page()
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*GREEN)
        pdf.cell(0, 10, "Tabela de Amortizacao - Financiamento", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        cols_amort = ["Mes", "Parcela", "Juros", "Amortizacao", "Saldo"]
        w_amort = [20, 42, 42, 42, 44]
        pdf.table_header(cols_amort, w_amort)

        for i, p in enumerate(fin_parcelas):
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.section_subtitle("Amortizacao Financiamento (cont.)")
                pdf.ln(1)
                pdf.table_header(cols_amort, w_amort)
            pdf.table_row(
                [str(p.get("mes", 0)), fmt_currency(p.get("parcela", 0)),
                 fmt_currency(p.get("juros", 0)), fmt_currency(p.get("amortizacao", 0)),
                 fmt_currency(p.get("saldo", 0))],
                w_amort, i, aligns=["C", "R", "R", "R", "R"]
            )

        total_parc = sum(p.get("parcela", 0) for p in fin_parcelas)
        total_juros = sum(p.get("juros", 0) for p in fin_parcelas)
        total_amort_val = sum(p.get("amortizacao", 0) for p in fin_parcelas)
        pdf.table_total_row(
            ["TOTAL", fmt_currency(total_parc), fmt_currency(total_juros),
             fmt_currency(total_amort_val), ""],
            w_amort, aligns=["C", "R", "R", "R", "R"]
        )

    pdf.disclaimer()

    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Comparativo_{nome}_{ts}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


# =====================================================================
# REPORT 5: VENDA DA OPERACAO
# =====================================================================

def gerar_pdf_venda_operacao(op: OperacaoCompleta, output_dir: str) -> str:
    """Gera PDF com analise de venda da operacao."""
    os.makedirs(output_dir, exist_ok=True)

    if not op.venda:
        return ""

    pdf = SomusPDF(title_text="ANALISE DE VENDA DA OPERACAO")
    pdf.alias_nb_pages()

    venda = op.venda
    params = op.params or {}

    vpl_vendedor = venda.get("vpl_vendedor", 0)
    vpl_comprador = venda.get("vpl_comprador", 0)
    ganho_nominal = venda.get("ganho_nominal", 0)
    ganho_pct = venda.get("ganho_pct", 0)
    total_investido = venda.get("total_investido", 0)
    valor_venda = venda.get("valor_venda", 0)
    prazo_medio = venda.get("prazo_medio", 0)
    ganho_mensal = venda.get("ganho_mensal", 0)
    margem_mensal = venda.get("margem_mensal_pct", 0)
    tir_v_m = venda.get("tir_vendedor_mensal", 0)
    tir_v_a = venda.get("tir_vendedor_anual", 0)
    tir_c_m = venda.get("tir_comprador_mensal", 0)
    tir_c_a = venda.get("tir_comprador_anual", 0)

    # ---- PAGINA 1: ANALISE DE VENDA ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Analise de Venda", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # KPIs
    y = pdf.get_y()
    pdf.kpi_box(10, y, 60, 26, "VPL Venda", fmt_currency(vpl_vendedor), GREEN)
    pdf.kpi_box(75, y, 60, 26, "Ganho %", fmt_pct(ganho_pct), BLUE)
    pdf.kpi_box(140, y, 60, 26, "Prazo Medio", f"{fmt_num(prazo_medio, 1)} meses", ORANGE)
    pdf.set_y(y + 30)

    y = pdf.get_y()
    pdf.kpi_box(10, y, 60, 26, "Ganho Mensal", fmt_currency(ganho_mensal), GREEN)
    pdf.kpi_box(75, y, 60, 26, "Margem Mensal", fmt_pct(margem_mensal), BLUE)
    pdf.kpi_box(140, y, 60, 26, "TIR Comprador", fmt_pct(tir_c_a * 100), ORANGE)
    pdf.set_y(y + 30)

    pdf.ln(2)

    # Resumo do vendedor
    pdf.section_title("Resumo do Vendedor")
    pdf.ln(1)

    vend_rows = [
        ("Total Investido", fmt_currency(total_investido)),
        ("Valor de Venda", fmt_currency(valor_venda)),
        ("Ganho Nominal", fmt_currency(ganho_nominal)),
        ("Ganho Percentual", fmt_pct(ganho_pct)),
        ("Ganho Mensal Medio", fmt_currency(ganho_mensal)),
        ("Margem Mensal", fmt_pct(margem_mensal)),
        ("TIR Vendedor Mensal", fmt_pct(tir_v_m * 100)),
        ("TIR Vendedor Anual", fmt_pct(tir_v_a * 100)),
        ("VPL Vendedor", fmt_currency(vpl_vendedor)),
    ]
    for i, (label, value) in enumerate(vend_rows):
        pdf.param_row(label, value, i)

    pdf.ln(2)

    pdf.section_title("Custo do Comprador")
    pdf.ln(1)

    comp_rows = [
        ("Valor Pago na Compra", fmt_currency(valor_venda)),
        ("TIR Comprador Mensal", fmt_pct(tir_c_m * 100)),
        ("TIR Comprador Anual", fmt_pct(tir_c_a * 100)),
        ("VPL Comprador", fmt_currency(vpl_comprador)),
    ]
    for i, (label, value) in enumerate(comp_rows):
        pdf.param_row(label, value, i)

    # ---- PAGINA 2: FLUXOS ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Fluxos da Venda", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Fluxo do Vendedor
    cf_vendedor = venda.get("cashflow_vendedor", [])
    if cf_vendedor:
        pdf.section_title("Fluxo do Vendedor")
        pdf.ln(1)

        cols_v = ["Mes", "Pagamento", "Tipo"]
        w_v = [30, 80, 80]
        pdf.table_header(cols_v, w_v)

        for i, cf in enumerate(cf_vendedor):
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.section_subtitle("Fluxo Vendedor (cont.)")
                pdf.ln(1)
                pdf.table_header(cols_v, w_v)

            if i == len(cf_vendedor) - 1 and cf > 0:
                tipo = "Venda"
            elif cf < 0:
                tipo = "Parcela"
            elif cf > 0:
                tipo = "Credito"
            else:
                tipo = "-"

            pdf.table_row(
                [str(i), fmt_currency(cf), tipo],
                w_v, i, aligns=["C", "R", "L"]
            )

    pdf.ln(4)

    # Fluxo do Comprador
    cf_comprador = venda.get("cashflow_comprador", [])
    if cf_comprador:
        pdf.check_page_break(30)
        pdf.section_title("Fluxo do Comprador")
        pdf.ln(1)

        cols_c = ["Mes", "Pagamento", "Tipo"]
        w_c = [30, 80, 80]
        pdf.table_header(cols_c, w_c)

        for i, cf in enumerate(cf_comprador):
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.section_subtitle("Fluxo Comprador (cont.)")
                pdf.ln(1)
                pdf.table_header(cols_c, w_c)

            if i == 0:
                tipo = "Compra"
            elif cf < 0:
                tipo = "Parcela"
            elif cf > 0:
                tipo = "Credito"
            else:
                tipo = "-"

            pdf.table_row(
                [str(i), fmt_currency(cf), tipo],
                w_c, i, aligns=["C", "R", "L"]
            )

    pdf.disclaimer()

    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Venda_Operacao_{nome}_{ts}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


# =====================================================================
# REPORT 6: CREDITO LANCE + CUSTO COMBINADO
# =====================================================================

def gerar_pdf_credito_lance(op: OperacaoCompleta, output_dir: str) -> str:
    """Gera PDF com analise do credito para lance e custo combinado."""
    os.makedirs(output_dir, exist_ok=True)

    if not op.credito_lance:
        return ""

    pdf = SomusPDF(title_text="OPERACAO DE CREDITO PARA LANCE")
    pdf.alias_nb_pages()

    cl = op.credito_lance
    comb = op.custo_combinado
    params = op.params or {}

    valor_lance = cl.get("valor", 0)
    total_pago = cl.get("total_pago", 0)
    total_juros = cl.get("total_juros", 0)
    iof = cl.get("iof", 0)
    custos_iniciais = cl.get("custos_iniciais", 0)
    cet_total = cl.get("custo_efetivo_total", 0)
    tir_m = cl.get("tir_mensal", 0)
    tir_a = cl.get("tir_anual", 0)
    cet_a = cl.get("cet_anual", 0)
    parcelas_lance = cl.get("parcelas", [])

    # ---- PAGINA 1: FINANCIAMENTO DO LANCE ----
    pdf.add_page()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(*GREEN)
    pdf.cell(0, 10, "Financiamento do Lance", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    # Dados
    pdf.section_title("Dados do Financiamento")
    pdf.ln(1)

    dados_lance = [
        ("Valor Financiado", fmt_currency(valor_lance)),
        ("Total Pago", fmt_currency(total_pago)),
        ("Total Juros", fmt_currency(total_juros)),
    ]
    for i, (label, value) in enumerate(dados_lance):
        pdf.param_row(label, value, i)

    pdf.ln(2)

    # Custos
    pdf.section_title("Custos Adicionais")
    pdf.ln(1)

    custos_rows = [
        ("IOF", fmt_currency(iof)),
        ("Custos Iniciais (TAC + Aval. + Comissao)", fmt_currency(custos_iniciais)),
        ("Custo Efetivo Total", fmt_currency(cet_total)),
    ]
    for i, (label, value) in enumerate(custos_rows):
        pdf.param_row(label, value, i)

    pdf.ln(2)

    # KPIs
    pdf.section_title("Indicadores")
    pdf.ln(1)

    y = pdf.get_y()
    pdf.kpi_box(10, y, 60, 26, "Total Pago", fmt_currency(total_pago), GREEN)
    pdf.kpi_box(75, y, 60, 26, "Total Juros", fmt_currency(total_juros), BLUE)
    pdf.kpi_box(140, y, 60, 26, "CET Anual", fmt_pct(cet_a * 100), ORANGE)
    pdf.set_y(y + 30)

    # ---- PAGINA 2: TABELA AMORTIZACAO ----
    if parcelas_lance:
        pdf.add_page()
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*GREEN)
        pdf.cell(0, 10, "Tabela de Amortizacao - Credito Lance", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        cols_al = ["Mes", "Parcela", "Juros", "Amortizacao", "Saldo"]
        w_al = [20, 42, 42, 42, 44]
        pdf.table_header(cols_al, w_al)

        for i, p in enumerate(parcelas_lance):
            if pdf.get_y() > 265:
                pdf.add_page()
                pdf.section_subtitle("Amortizacao Lance (cont.)")
                pdf.ln(1)
                pdf.table_header(cols_al, w_al)
            pdf.table_row(
                [str(p.get("mes", 0)), fmt_currency(p.get("parcela", 0)),
                 fmt_currency(p.get("juros", 0)), fmt_currency(p.get("amortizacao", 0)),
                 fmt_currency(p.get("saldo", 0))],
                w_al, i, aligns=["C", "R", "R", "R", "R"]
            )

        t_parc = sum(p.get("parcela", 0) for p in parcelas_lance)
        t_juros = sum(p.get("juros", 0) for p in parcelas_lance)
        t_amort = sum(p.get("amortizacao", 0) for p in parcelas_lance)
        pdf.table_total_row(
            ["TOTAL", fmt_currency(t_parc), fmt_currency(t_juros),
             fmt_currency(t_amort), ""],
            w_al, aligns=["C", "R", "R", "R", "R"]
        )

    # ---- PAGINA 3: CUSTO COMBINADO ----
    if comb:
        pdf.add_page()
        pdf.ln(2)
        pdf.set_font("Helvetica", "B", 16)
        pdf.set_text_color(*GREEN)
        pdf.cell(0, 10, "Custo Combinado", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*TEXT_GRAY)
        pdf.cell(0, 6, "Consorcio + Financiamento do Lance", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

        total_cons = comb.get("total_pago_consorcio", 0)
        total_lance = comb.get("total_pago_lance", 0)
        total_comb = comb.get("total_pago_combinado", 0)
        tir_comb_m = comb.get("tir_mensal_combinado", 0)
        tir_comb_a = comb.get("tir_anual_combinado", 0)
        cet_comb_a = comb.get("cet_anual_combinado", 0)

        # KPIs
        y = pdf.get_y()
        pdf.kpi_box(10, y, 60, 26, "Total Consorcio", fmt_currency(total_cons), GREEN)
        pdf.kpi_box(75, y, 60, 26, "Total Lance", fmt_currency(total_lance), BLUE)
        pdf.kpi_box(140, y, 60, 26, "Total Combinado", fmt_currency(total_comb), ORANGE)
        pdf.set_y(y + 30)

        y = pdf.get_y()
        pdf.kpi_box(10, y, 93, 26, "TIR Combinado Anual", fmt_pct(tir_comb_a * 100), GREEN)
        pdf.kpi_box(107, y, 93, 26, "CET Combinado Anual", fmt_pct(cet_comb_a * 100), BLUE)
        pdf.set_y(y + 30)

        # Tabela resumo
        pdf.ln(2)
        pdf.section_title("Composicao do Custo Combinado")
        pdf.ln(1)

        comb_detail = [
            ("Total Pago - Consorcio", fmt_currency(total_cons)),
            ("Total Pago - Financiamento Lance", fmt_currency(total_lance)),
            ("Total Pago - Combinado", fmt_currency(total_comb)),
            ("TIR Mensal Combinado", fmt_pct(tir_comb_m * 100)),
            ("TIR Anual Combinado", fmt_pct(tir_comb_a * 100)),
            ("CET Anual Combinado", fmt_pct(cet_comb_a * 100)),
        ]
        for i, (label, value) in enumerate(comb_detail):
            pdf.param_row(label, value, i)

    pdf.disclaimer()

    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Credito_Lance_{nome}_{ts}.pdf"
    filepath = os.path.join(output_dir, filename)
    pdf.output(filepath)
    return filepath


# =====================================================================
# MAIN ENTRY POINT - ALL PDFs
# =====================================================================

def gerar_relatorio_completo(op: OperacaoCompleta, output_dir: str) -> dict:
    """Generate ALL reports for an operation. Returns dict of {report_name: filepath}."""
    results = {}
    results['resumo_executivo'] = gerar_pdf_resumo_executivo(op, output_dir)
    results['fluxo_financeiro'] = gerar_pdf_fluxo_financeiro(op, output_dir)
    results['analise_vpl'] = gerar_pdf_analise_vpl(op, output_dir)
    if op.comparativo:
        results['comparativo'] = gerar_pdf_comparativo(op, output_dir)
    if op.venda:
        results['venda_operacao'] = gerar_pdf_venda_operacao(op, output_dir)
    if op.credito_lance:
        results['credito_lance'] = gerar_pdf_credito_lance(op, output_dir)
    return results


# =====================================================================
# EXCEL GENERATOR
# =====================================================================

def gerar_excel_completo(op: OperacaoCompleta, output_dir: str) -> str:
    """
    Generate a complete Excel workbook mirroring the NASA HD spreadsheet.
    Uses openpyxl. Returns filepath.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl e necessario para gerar Excel. Instale com: pip install openpyxl")

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()

    # Estilos
    header_fill = PatternFill(start_color="004D33", end_color="004D33", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    alt_fill_1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    alt_fill_2 = PatternFill(start_color="F6F7F9", end_color="F6F7F9", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", bold=True, size=10)
    currency_fmt = '#.##0,00'
    pct_fmt = '0,00%'
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    def style_header_row(ws, row_num, num_cols):
        """Aplica estilo ao cabecalho."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    def style_data_cell(ws, row_num, col_num, is_alt=False, is_currency=False,
                        is_pct=False, is_bold=False):
        """Aplica estilo a celula de dados."""
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = alt_fill_2 if is_alt else alt_fill_1
        cell.font = bold_font if is_bold else data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right' if (is_currency or is_pct) else 'center',
                                   vertical='center')
        if is_currency:
            cell.number_format = currency_fmt
        elif is_pct:
            cell.number_format = pct_fmt

    def auto_width(ws, min_width=10, max_width=25):
        """Ajusta largura das colunas."""
        for col in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)) + 2)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len, max_width)

    def freeze_pane(ws, row=2, col=1):
        """Congela paineis."""
        ws.freeze_panes = ws.cell(row=row, column=col)

    params = op.params or {}
    fluxo = op.fluxo or {}
    resumo = op.resumo or {}
    vpl_data = op.vpl or {}
    parcelas = op.parcelas or []
    fluxo_mensal = fluxo.get("fluxo", fluxo.get("fluxo_mensal", []))
    totais = fluxo.get("totais", {})
    metricas = fluxo.get("metricas", {})

    # ================================================================
    # ABA 1: Dados do Consorcio
    # ================================================================
    ws1 = wb.active
    ws1.title = "Dados do Consorcio"

    ws1.cell(row=1, column=1, value="Parametro")
    ws1.cell(row=1, column=2, value="Valor")
    style_header_row(ws1, 1, 2)

    param_rows = [
        ("Cliente", op.cliente_nome),
        ("Assessor", op.assessor),
        ("Administradora", op.administradora),
        ("Data Geracao", op.data_geracao or datetime.now().strftime("%d/%m/%Y")),
        ("", ""),
        ("Valor do Credito", params.get("valor_credito", 0)),
        ("Prazo (meses)", params.get("prazo_meses", 0)),
        ("Momento Contemplacao", params.get("momento_contemplacao", 0)),
        ("Taxa Adm. (%)", params.get("taxa_adm_pct", 0)),
        ("Fundo Reserva (%)", params.get("fundo_reserva_pct", 0)),
        ("Seguro Vida (% mensal)", params.get("seguro_vida_pct", 0)),
        ("Lance Embutido (%)", params.get("lance_embutido_pct", 0)),
        ("Lance Livre (%)", params.get("lance_livre_pct", 0)),
        ("Reajuste Pre (%)", params.get("reajuste_pre_pct", 0)),
        ("Reajuste Pre Freq.", params.get("reajuste_pre_freq", "Anual")),
        ("Reajuste Pos (%)", params.get("reajuste_pos_pct", 0)),
        ("Reajuste Pos Freq.", params.get("reajuste_pos_freq", "Anual")),
        ("ALM Anual (%)", params.get("alm_anual", 12.0)),
        ("Hurdle Anual (%)", params.get("hurdle_anual", 12.0)),
        ("TMA (mensal)", params.get("tma", 0.01)),
        ("Antecipacao TA (%)", params.get("antecipacao_ta_pct", 0)),
        ("Antecipacao TA Parcelas", params.get("antecipacao_ta_parcelas", 1)),
        ("", ""),
        ("--- RESULTADOS ---", ""),
        ("Carta Liquida", totais.get("carta_liquida", 0)),
        ("Total Pago", totais.get("total_pago", 0)),
        ("Total Fundo Comum", totais.get("total_fundo_comum", 0)),
        ("Total Taxa Adm", totais.get("total_taxa_adm", 0)),
        ("Total Fundo Reserva", totais.get("total_fundo_reserva", 0)),
        ("Total Seguro", totais.get("total_seguro", 0)),
        ("Total Custos Acessorios", totais.get("total_custos_acessorios", 0)),
        ("Lance Embutido (R$)", totais.get("lance_embutido_valor", 0)),
        ("Lance Livre (R$)", totais.get("lance_livre_valor", 0)),
        ("TIR Mensal", metricas.get("tir_mensal", 0)),
        ("TIR Anual", metricas.get("tir_anual", 0)),
        ("Parcela Media", metricas.get("parcela_media", 0)),
        ("Parcela Maxima", metricas.get("parcela_maxima", 0)),
        ("Parcela Minima", metricas.get("parcela_minima", 0)),
        ("Custo Total %", metricas.get("custo_total_pct", 0)),
    ]

    for i, (param, val) in enumerate(param_rows, start=2):
        ws1.cell(row=i, column=1, value=param)
        ws1.cell(row=i, column=2, value=val)
        is_alt = (i % 2 == 0)
        is_cur = isinstance(val, (int, float)) and val != 0 and param not in ("Prazo (meses)",
                 "Momento Contemplacao", "Antecipacao TA Parcelas")
        style_data_cell(ws1, i, 1, is_alt)
        style_data_cell(ws1, i, 2, is_alt, is_currency=is_cur)

    auto_width(ws1)
    freeze_pane(ws1)

    # ================================================================
    # ABA 2: Fluxo Financeiro (35 colunas)
    # ================================================================
    ws2 = wb.create_sheet("Fluxo Financeiro")

    fluxo_cols = [
        ("Mes", "mes", False, False),
        ("Meses Rest.", "meses_restantes", False, False),
        ("FC Base", "valor_base_fc", True, False),
        ("Lance Emb.", "lance_embutido", True, False),
        ("Lance Livre", "lance_livre", True, False),
        ("Base Final", "valor_base_final", True, False),
        ("% FC Mensal", "pct_mensal_fc", False, True),
        ("% FC Acum.", "pct_acum_fc", False, True),
        ("Amortizacao", "amortizacao", True, False),
        ("Saldo Principal", "saldo_principal", True, False),
        ("TA Antecip.", "taxa_adm_antecipada", True, False),
        ("% TA Mensal", "pct_ta_mensal", False, True),
        ("% TA Acum.", "pct_ta_acum", False, True),
        ("Valor TA", "valor_parcela_ta", True, False),
        ("% FR Mensal", "pct_fr_mensal", False, True),
        ("% FR Acum.", "pct_fr_acum", False, True),
        ("Valor FR", "valor_fundo_reserva", True, False),
        ("Parcela Base", "valor_parcela", True, False),
        ("Saldo Devedor", "saldo_devedor", True, False),
        ("Peso Parcela", "peso_parcela", False, True),
        ("% Reajuste", "pct_reajuste", False, True),
        ("% Reaj. Acum.", "pct_reajuste_acum", False, True),
        ("Fator Reajuste", "fator_reajuste", False, False),
        ("Parc. Reajustada", "parcela_apos_reajuste", True, False),
        ("Saldo Dev. Reaj.", "saldo_devedor_reajustado", True, False),
        ("Seguro Vida", "seguro_vida", True, False),
        ("Parc.+Seguro", "parcela_com_seguro", True, False),
        ("Outros Custos", "outros_custos", True, False),
        ("Carta Credito", "carta_credito_original", True, False),
        ("Carta Reajust.", "carta_credito_reajustada", True, False),
        ("Credito Receb.", "credito_recebido", True, False),
        ("Fluxo Caixa", "fluxo_caixa", True, False),
        ("Fluxo TIR", "fluxo_caixa_tir", True, False),
    ]

    for col_idx, (col_name, _, _, _) in enumerate(fluxo_cols, start=1):
        ws2.cell(row=1, column=col_idx, value=col_name)
    style_header_row(ws2, 1, len(fluxo_cols))

    for row_idx, row_data in enumerate(fluxo_mensal, start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, (_, key, is_cur, is_pct) in enumerate(fluxo_cols, start=1):
            val = row_data.get(key, 0)
            if val is None:
                val = 0
            ws2.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(ws2, row_idx, col_idx, is_alt, is_currency=is_cur, is_pct=is_pct)

    auto_width(ws2, max_width=18)
    freeze_pane(ws2)

    # ================================================================
    # ABA 3: Parcelas
    # ================================================================
    ws3 = wb.create_sheet("Parcelas")

    parc_cols = [
        ("Mes", "mes", False),
        ("Fundo Comum", "fundo_comum", True),
        ("Taxa Adm.", "taxa_adm", True),
        ("Fundo Reserva", "fundo_reserva", True),
        ("Parcela Base", "parcela_base", True),
        ("Fator Reajuste", "reajuste", False),
        ("Parcela Reajustada", "parcela_reajustada", True),
        ("Seguro", "seguro", True),
        ("Parcela Total", "parcela_total", True),
        ("Outros Custos", "outros_custos", True),
        ("Desembolso Total", "desembolso_total", True),
        ("Saldo Devedor", "saldo_devedor", True),
        ("Lance Embutido", "lance_embutido", True),
        ("Lance Livre", "lance_livre", True),
        ("Credito Recebido", "credito_recebido", True),
    ]

    for col_idx, (col_name, _, _) in enumerate(parc_cols, start=1):
        ws3.cell(row=1, column=col_idx, value=col_name)
    style_header_row(ws3, 1, len(parc_cols))

    for row_idx, p in enumerate(parcelas, start=2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, (_, key, is_cur) in enumerate(parc_cols, start=1):
            val = p.get(key, 0)
            if val is None:
                val = 0
            ws3.cell(row=row_idx, column=col_idx, value=val)
            style_data_cell(ws3, row_idx, col_idx, is_alt, is_currency=is_cur)

    auto_width(ws3, max_width=18)
    freeze_pane(ws3)

    # ================================================================
    # ABA 4: Comparativo VPL
    # ================================================================
    ws4 = wb.create_sheet("Comparativo VPL")

    # Resumo VPL
    vpl_summary = [
        ("Componente", "Valor", "% do Credito"),
    ]
    credito = params.get("valor_credito", 0)
    carta_liq = totais.get("carta_liquida", 0)

    ws4.cell(row=1, column=1, value="Componente")
    ws4.cell(row=1, column=2, value="Valor")
    ws4.cell(row=1, column=3, value="% do Credito")
    ws4.cell(row=1, column=4, value="")
    ws4.cell(row=1, column=5, value="Parametro")
    ws4.cell(row=1, column=6, value="Valor")
    style_header_row(ws4, 1, 6)

    b0 = vpl_data.get("b0", 0)
    h0 = vpl_data.get("h0", 0)
    d0 = vpl_data.get("d0", 0)
    pv_pos_t = vpl_data.get("pv_pos_t", 0)
    delta_vpl = vpl_data.get("delta_vpl", 0)

    vpl_rows = [
        ("Cheque liquido em T", carta_liq, (carta_liq / credito * 100) if credito > 0 else 0),
        ("B0 (VP Credito @ ALM)", b0, (b0 / credito * 100) if credito > 0 else 0),
        ("H0 (VP Pagtos pre-T)", h0, (h0 / credito * 100) if credito > 0 else 0),
        ("D0 (Valor criado)", d0, (d0 / credito * 100) if credito > 0 else 0),
        ("PV pos-T (@ Hurdle)", pv_pos_t, (pv_pos_t / credito * 100) if credito > 0 else 0),
        ("Delta VPL", delta_vpl, (delta_vpl / credito * 100) if credito > 0 else 0),
    ]
    for i, (comp, val, pct) in enumerate(vpl_rows, start=2):
        ws4.cell(row=i, column=1, value=comp)
        ws4.cell(row=i, column=2, value=val)
        ws4.cell(row=i, column=3, value=pct / 100)
        is_alt = (i % 2 == 0)
        style_data_cell(ws4, i, 1, is_alt)
        style_data_cell(ws4, i, 2, is_alt, is_currency=True)
        style_data_cell(ws4, i, 3, is_alt, is_pct=True)

    # Parametros na coluna E-F
    vpl_params_data = [
        ("ALM/CDI (% a.a.)", params.get("alm_anual", 12.0)),
        ("Hurdle (% a.a.)", params.get("hurdle_anual", 12.0)),
        ("Break-even Lance (%)", vpl_data.get("break_even_lance", 0)),
        ("TIR Mensal", vpl_data.get("tir_mensal", 0)),
        ("TIR Anual", vpl_data.get("tir_anual", 0)),
        ("Cria Valor?", "SIM" if vpl_data.get("cria_valor", False) else "NAO"),
    ]
    for i, (param, val) in enumerate(vpl_params_data, start=2):
        ws4.cell(row=i, column=5, value=param)
        ws4.cell(row=i, column=6, value=val)
        is_alt = (i % 2 == 0)
        style_data_cell(ws4, i, 5, is_alt)
        style_data_cell(ws4, i, 6, is_alt)

    # VP por mes
    pv_pre_detail = vpl_data.get("pv_pre_t_detail", [])
    pv_pos_detail = vpl_data.get("pv_pos_t_detail", [])
    all_pv = pv_pre_detail + pv_pos_detail

    if all_pv:
        start_row = len(vpl_rows) + 4
        ws4.cell(row=start_row, column=1, value="Mes")
        ws4.cell(row=start_row, column=2, value="Pagamento")
        ws4.cell(row=start_row, column=3, value="VP")
        ws4.cell(row=start_row, column=4, value="Periodo")
        style_header_row(ws4, start_row, 4)

        for i, item in enumerate(all_pv, start=start_row + 1):
            mes = item.get("mes", 0)
            is_pre = item in pv_pre_detail
            ws4.cell(row=i, column=1, value=mes)
            ws4.cell(row=i, column=2, value=item.get("valor", 0))
            ws4.cell(row=i, column=3, value=item.get("pv", 0))
            ws4.cell(row=i, column=4, value="Pre-T" if is_pre else "Pos-T")
            is_alt = (i % 2 == 0)
            style_data_cell(ws4, i, 1, is_alt)
            style_data_cell(ws4, i, 2, is_alt, is_currency=True)
            style_data_cell(ws4, i, 3, is_alt, is_currency=True)
            style_data_cell(ws4, i, 4, is_alt)

    auto_width(ws4, max_width=20)
    freeze_pane(ws4)

    # ================================================================
    # ABA 5: Cenarios (placeholder)
    # ================================================================
    ws5 = wb.create_sheet("Cenarios")
    ws5.cell(row=1, column=1, value="Cenario")
    ws5.cell(row=1, column=2, value="Valor Credito")
    ws5.cell(row=1, column=3, value="Prazo")
    ws5.cell(row=1, column=4, value="Total Pago")
    ws5.cell(row=1, column=5, value="TIR Mensal")
    ws5.cell(row=1, column=6, value="TIR Anual")
    ws5.cell(row=1, column=7, value="Delta VPL")
    ws5.cell(row=1, column=8, value="Parcela Media")
    style_header_row(ws5, 1, 8)

    # Se consolidacao tiver fluxos individuais (cenarios multi-cota)
    if op.consolidacao:
        cons = op.consolidacao
        fluxos_ind = cons.get("fluxos_individuais", [])
        for i, fl in enumerate(fluxos_ind, start=2):
            m = fl.get("metricas", {})
            t = fl.get("totais", {})
            ws5.cell(row=i, column=1, value=f"Cota {i-1}")
            ws5.cell(row=i, column=2, value=t.get("carta_liquida", 0))
            ws5.cell(row=i, column=3, value=len(fl.get("cashflow", [])) - 1)
            ws5.cell(row=i, column=4, value=t.get("total_pago", 0))
            ws5.cell(row=i, column=5, value=m.get("tir_mensal", 0))
            ws5.cell(row=i, column=6, value=m.get("tir_anual", 0))
            ws5.cell(row=i, column=7, value=0)
            ws5.cell(row=i, column=8, value=m.get("parcela_media", 0))
            is_alt = (i % 2 == 0)
            for c in range(1, 9):
                style_data_cell(ws5, i, c, is_alt, is_currency=(c in (2, 4, 7, 8)))
    else:
        ws5.cell(row=2, column=1, value="Cenario Base")
        ws5.cell(row=2, column=2, value=params.get("valor_credito", 0))
        ws5.cell(row=2, column=3, value=params.get("prazo_meses", 0))
        ws5.cell(row=2, column=4, value=totais.get("total_pago", 0))
        ws5.cell(row=2, column=5, value=metricas.get("tir_mensal", 0))
        ws5.cell(row=2, column=6, value=metricas.get("tir_anual", 0))
        ws5.cell(row=2, column=7, value=vpl_data.get("delta_vpl", 0))
        ws5.cell(row=2, column=8, value=metricas.get("parcela_media", 0))
        for c in range(1, 9):
            style_data_cell(ws5, 2, c, False, is_currency=(c in (2, 4, 7, 8)))

    auto_width(ws5, max_width=18)
    freeze_pane(ws5)

    # ================================================================
    # ABA 6: Venda da Operacao
    # ================================================================
    ws6 = wb.create_sheet("Venda da Operacao")

    if op.venda:
        v = op.venda
        ws6.cell(row=1, column=1, value="Indicador")
        ws6.cell(row=1, column=2, value="Valor")
        style_header_row(ws6, 1, 2)

        venda_rows = [
            ("Total Investido", v.get("total_investido", 0)),
            ("Valor de Venda", v.get("valor_venda", 0)),
            ("Ganho Nominal", v.get("ganho_nominal", 0)),
            ("Ganho %", v.get("ganho_pct", 0)),
            ("Prazo Medio", v.get("prazo_medio", 0)),
            ("Ganho Mensal", v.get("ganho_mensal", 0)),
            ("Margem Mensal %", v.get("margem_mensal_pct", 0)),
            ("TIR Vendedor Mensal", v.get("tir_vendedor_mensal", 0)),
            ("TIR Vendedor Anual", v.get("tir_vendedor_anual", 0)),
            ("VPL Vendedor", v.get("vpl_vendedor", 0)),
            ("TIR Comprador Mensal", v.get("tir_comprador_mensal", 0)),
            ("TIR Comprador Anual", v.get("tir_comprador_anual", 0)),
            ("VPL Comprador", v.get("vpl_comprador", 0)),
        ]
        for i, (ind, val) in enumerate(venda_rows, start=2):
            ws6.cell(row=i, column=1, value=ind)
            ws6.cell(row=i, column=2, value=val)
            is_alt = (i % 2 == 0)
            style_data_cell(ws6, i, 1, is_alt)
            style_data_cell(ws6, i, 2, is_alt, is_currency=True)

        # Fluxo vendedor
        cf_v = v.get("cashflow_vendedor", [])
        if cf_v:
            start = len(venda_rows) + 4
            ws6.cell(row=start, column=1, value="Mes")
            ws6.cell(row=start, column=2, value="Fluxo Vendedor")
            style_header_row(ws6, start, 2)
            for j, cf in enumerate(cf_v, start=start + 1):
                ws6.cell(row=j, column=1, value=j - start - 1)
                ws6.cell(row=j, column=2, value=cf)
                is_alt = (j % 2 == 0)
                style_data_cell(ws6, j, 1, is_alt)
                style_data_cell(ws6, j, 2, is_alt, is_currency=True)

        # Fluxo comprador (coluna D-E)
        cf_c = v.get("cashflow_comprador", [])
        if cf_c:
            start = len(venda_rows) + 4
            ws6.cell(row=start, column=4, value="Mes")
            ws6.cell(row=start, column=5, value="Fluxo Comprador")
            style_header_row(ws6, start, 5)
            # Re-style only cols 4-5
            for cc in (4, 5):
                cell = ws6.cell(row=start, column=cc)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            for j, cf in enumerate(cf_c, start=start + 1):
                ws6.cell(row=j, column=4, value=j - start - 1)
                ws6.cell(row=j, column=5, value=cf)
                is_alt = (j % 2 == 0)
                style_data_cell(ws6, j, 4, is_alt)
                style_data_cell(ws6, j, 5, is_alt, is_currency=True)
    else:
        ws6.cell(row=1, column=1, value="Sem dados de venda disponiveis")

    auto_width(ws6, max_width=18)
    freeze_pane(ws6)

    # ================================================================
    # ABA 7: Op. Credito Lance
    # ================================================================
    ws7 = wb.create_sheet("Op. Credito Lance")

    if op.credito_lance:
        cl = op.credito_lance
        ws7.cell(row=1, column=1, value="Indicador")
        ws7.cell(row=1, column=2, value="Valor")
        style_header_row(ws7, 1, 2)

        cl_rows = [
            ("Valor Financiado", cl.get("valor", 0)),
            ("Total Pago", cl.get("total_pago", 0)),
            ("Total Juros", cl.get("total_juros", 0)),
            ("IOF", cl.get("iof", 0)),
            ("Custos Iniciais", cl.get("custos_iniciais", 0)),
            ("Custo Efetivo Total", cl.get("custo_efetivo_total", 0)),
            ("TIR Mensal", cl.get("tir_mensal", 0)),
            ("TIR Anual", cl.get("tir_anual", 0)),
            ("CET Anual", cl.get("cet_anual", 0)),
        ]
        for i, (ind, val) in enumerate(cl_rows, start=2):
            ws7.cell(row=i, column=1, value=ind)
            ws7.cell(row=i, column=2, value=val)
            is_alt = (i % 2 == 0)
            style_data_cell(ws7, i, 1, is_alt)
            style_data_cell(ws7, i, 2, is_alt, is_currency=True)

        # Tabela amortizacao
        cl_parcelas = cl.get("parcelas", [])
        if cl_parcelas:
            start = len(cl_rows) + 4
            cols_cl = ["Mes", "Parcela", "Juros", "Amortizacao", "Saldo"]
            for c_idx, col_name in enumerate(cols_cl, start=1):
                ws7.cell(row=start, column=c_idx, value=col_name)
            style_header_row(ws7, start, len(cols_cl))

            for j, p in enumerate(cl_parcelas, start=start + 1):
                ws7.cell(row=j, column=1, value=p.get("mes", 0))
                ws7.cell(row=j, column=2, value=p.get("parcela", 0))
                ws7.cell(row=j, column=3, value=p.get("juros", 0))
                ws7.cell(row=j, column=4, value=p.get("amortizacao", 0))
                ws7.cell(row=j, column=5, value=p.get("saldo", 0))
                is_alt = (j % 2 == 0)
                style_data_cell(ws7, j, 1, is_alt)
                for cc in range(2, 6):
                    style_data_cell(ws7, j, cc, is_alt, is_currency=True)
    else:
        ws7.cell(row=1, column=1, value="Sem dados de credito lance disponiveis")

    auto_width(ws7, max_width=18)
    freeze_pane(ws7)

    # ================================================================
    # ABA 8: Custo Combinado
    # ================================================================
    ws8 = wb.create_sheet("Custo Combinado")

    if op.custo_combinado:
        comb = op.custo_combinado
        ws8.cell(row=1, column=1, value="Indicador")
        ws8.cell(row=1, column=2, value="Valor")
        style_header_row(ws8, 1, 2)

        comb_rows = [
            ("Total Pago Consorcio", comb.get("total_pago_consorcio", 0)),
            ("Total Pago Lance", comb.get("total_pago_lance", 0)),
            ("Total Pago Combinado", comb.get("total_pago_combinado", 0)),
            ("TIR Mensal Combinado", comb.get("tir_mensal_combinado", 0)),
            ("TIR Anual Combinado", comb.get("tir_anual_combinado", 0)),
            ("CET Anual Combinado", comb.get("cet_anual_combinado", 0)),
        ]
        for i, (ind, val) in enumerate(comb_rows, start=2):
            ws8.cell(row=i, column=1, value=ind)
            ws8.cell(row=i, column=2, value=val)
            is_alt = (i % 2 == 0)
            style_data_cell(ws8, i, 1, is_alt)
            style_data_cell(ws8, i, 2, is_alt, is_currency=True)

        # Cashflow combinado
        cf_comb = comb.get("cashflow_combinado", [])
        if cf_comb:
            start = len(comb_rows) + 4
            ws8.cell(row=start, column=1, value="Mes")
            ws8.cell(row=start, column=2, value="Fluxo Combinado")
            style_header_row(ws8, start, 2)
            for j, cf in enumerate(cf_comb, start=start + 1):
                ws8.cell(row=j, column=1, value=j - start - 1)
                ws8.cell(row=j, column=2, value=cf)
                is_alt = (j % 2 == 0)
                style_data_cell(ws8, j, 1, is_alt)
                style_data_cell(ws8, j, 2, is_alt, is_currency=True)
    else:
        ws8.cell(row=1, column=1, value="Sem dados de custo combinado disponiveis")

    auto_width(ws8, max_width=18)
    freeze_pane(ws8)

    # ================================================================
    # ABA 9: Consorcio X Financ.
    # ================================================================
    ws9 = wb.create_sheet("Consorcio X Financ.")

    if op.comparativo:
        comp = op.comparativo
        ws9.cell(row=1, column=1, value="Indicador")
        ws9.cell(row=1, column=2, value="Consorcio")
        ws9.cell(row=1, column=3, value="Financiamento")
        ws9.cell(row=1, column=4, value="Diferenca")
        style_header_row(ws9, 1, 4)

        total_c = comp.get("total_pago_consorcio", 0)
        total_f = comp.get("total_pago_financiamento", 0)
        vpl_c = comp.get("vpl_consorcio", 0)
        vpl_f = comp.get("vpl_financiamento", 0)
        tir_c = comp.get("tir_consorcio_anual", 0)
        tir_f = comp.get("tir_financ_anual", 0)

        comp_rows_data = [
            ("Total Pago", total_c, total_f, total_c - total_f),
            ("VPL", vpl_c, vpl_f, comp.get("economia_vpl", 0)),
            ("VPL (TMA)", comp.get("vpl_consorcio_tma", 0),
             comp.get("vpl_financiamento_tma", 0), comp.get("economia_vpl_tma", 0)),
            ("TIR Anual", tir_c, tir_f, tir_c - tir_f),
            ("Razao Custo/Credito", comp.get("razao_vpl_consorcio", 0),
             comp.get("razao_vpl_financ", 0), 0),
            ("Economia Nominal", "", "", comp.get("economia_nominal", 0)),
        ]

        for i, (ind, vc, vf, diff) in enumerate(comp_rows_data, start=2):
            ws9.cell(row=i, column=1, value=ind)
            ws9.cell(row=i, column=2, value=vc)
            ws9.cell(row=i, column=3, value=vf)
            ws9.cell(row=i, column=4, value=diff)
            is_alt = (i % 2 == 0)
            style_data_cell(ws9, i, 1, is_alt)
            for cc in range(2, 5):
                is_cur = isinstance(ws9.cell(row=i, column=cc).value, (int, float))
                style_data_cell(ws9, i, cc, is_alt, is_currency=is_cur)

        # Fluxos lado a lado
        cf_cons = comp.get("consorcio", {}).get("cashflow", [])
        cf_fin = comp.get("financiamento", {}).get("cashflow", [])
        pv_cons = comp.get("pv_consorcio", [])
        pv_fin = comp.get("pv_financiamento", [])
        max_len = max(len(cf_cons), len(cf_fin), len(pv_cons), len(pv_fin))

        if max_len > 0:
            start = len(comp_rows_data) + 4
            flow_cols = ["Mes", "CF Consorcio", "CF Financ.", "PV Consorcio", "PV Financ."]
            for c_idx, col_name in enumerate(flow_cols, start=1):
                ws9.cell(row=start, column=c_idx, value=col_name)
            style_header_row(ws9, start, len(flow_cols))

            for m in range(max_len):
                r = start + 1 + m
                ws9.cell(row=r, column=1, value=m)
                ws9.cell(row=r, column=2, value=cf_cons[m] if m < len(cf_cons) else 0)
                ws9.cell(row=r, column=3, value=cf_fin[m] if m < len(cf_fin) else 0)
                ws9.cell(row=r, column=4, value=pv_cons[m] if m < len(pv_cons) else 0)
                ws9.cell(row=r, column=5, value=pv_fin[m] if m < len(pv_fin) else 0)
                is_alt = (r % 2 == 0)
                style_data_cell(ws9, r, 1, is_alt)
                for cc in range(2, 6):
                    style_data_cell(ws9, r, cc, is_alt, is_currency=True)
    else:
        ws9.cell(row=1, column=1, value="Sem dados comparativos disponiveis")

    auto_width(ws9, max_width=18)
    freeze_pane(ws9)

    # ================================================================
    # ABA 10: Resumo Cliente
    # ================================================================
    ws10 = wb.create_sheet("Resumo Cliente")

    ws10.cell(row=1, column=1, value="Indicador")
    ws10.cell(row=1, column=2, value="Valor")
    style_header_row(ws10, 1, 2)

    resumo_rows = [
        ("Cliente", op.cliente_nome),
        ("Assessor", op.assessor),
        ("Administradora", op.administradora),
        ("Data", op.data_geracao or datetime.now().strftime("%d/%m/%Y")),
        ("", ""),
        ("--- OPERACAO ---", ""),
        ("Valor do Credito", resumo.get("valor_credito", credito)),
        ("Prazo (meses)", resumo.get("prazo_meses", params.get("prazo_meses", 0))),
        ("Momento Contemplacao", resumo.get("momento_contemplacao", 0)),
        ("Carta Liquida", resumo.get("carta_liquida", 0)),
        ("Lance Embutido", resumo.get("lance_embutido", 0)),
        ("Lance Embutido %", resumo.get("lance_embutido_pct", 0)),
        ("Lance Livre", resumo.get("lance_livre", 0)),
        ("Lance Livre %", resumo.get("lance_livre_pct", 0)),
        ("Lance Total", resumo.get("lance_total", 0)),
        ("", ""),
        ("--- PARCELAS ---", ""),
        ("Primeira Parcela", resumo.get("primeira_parcela", 0)),
        ("Ultima Parcela", resumo.get("ultima_parcela", 0)),
        ("Parcela Media", resumo.get("parcela_media", 0)),
        ("Parcela Maxima", resumo.get("parcela_maxima", 0)),
        ("Parcela Minima", resumo.get("parcela_minima", 0)),
        ("", ""),
        ("--- TOTAIS ---", ""),
        ("Total Pago", resumo.get("total_pago", 0)),
        ("Total Fundo Comum", resumo.get("total_fundo_comum", 0)),
        ("Total Taxa Adm.", resumo.get("total_taxa_adm", 0)),
        ("Total Fundo Reserva", resumo.get("total_fundo_reserva", 0)),
        ("Total Seguro", resumo.get("total_seguro", 0)),
        ("Total Custos Acessorios", resumo.get("total_custos_acessorios", 0)),
        ("Custo Total %", resumo.get("custo_total_pct", 0)),
        ("", ""),
        ("--- TAXAS ---", ""),
        ("Taxa Adm. %", resumo.get("taxa_adm_pct", 0)),
        ("Fundo Reserva %", resumo.get("fundo_reserva_pct", 0)),
        ("Seguro %", resumo.get("seguro_pct", 0)),
        ("", ""),
        ("--- VPL ---", ""),
        ("TIR Mensal", resumo.get("tir_mensal", 0)),
        ("TIR Anual", resumo.get("tir_anual", 0)),
        ("CET Anual", resumo.get("cet_anual", 0)),
        ("B0", resumo.get("b0", 0)),
        ("H0", resumo.get("h0", 0)),
        ("D0", resumo.get("d0", 0)),
        ("PV pos-T", resumo.get("pv_pos_t", 0)),
        ("Delta VPL", resumo.get("delta_vpl", 0)),
        ("Cria Valor?", "SIM" if resumo.get("cria_valor", False) else "NAO"),
        ("VPL Total", resumo.get("vpl_total", 0)),
        ("Break-even Lance %", resumo.get("break_even_lance", 0)),
        ("", ""),
        ("--- REAJUSTE ---", ""),
        ("Reajuste Pre %", resumo.get("reajuste_pre_pct", 0)),
        ("Reajuste Pre Freq.", resumo.get("reajuste_pre_freq", "Anual")),
        ("Reajuste Pos %", resumo.get("reajuste_pos_pct", 0)),
        ("Reajuste Pos Freq.", resumo.get("reajuste_pos_freq", "Anual")),
    ]

    if op.credito_equivalente:
        resumo_rows.append(("", ""))
        resumo_rows.append(("--- CREDITO EQUIVALENTE ---", ""))
        resumo_rows.append(("Credito Equivalente", op.credito_equivalente))

    for i, (ind, val) in enumerate(resumo_rows, start=2):
        ws10.cell(row=i, column=1, value=ind)
        ws10.cell(row=i, column=2, value=val)
        is_alt = (i % 2 == 0)
        is_cur = isinstance(val, (int, float)) and val != 0 and "%" not in str(ind) and \
                 "meses" not in str(ind).lower() and "Freq" not in str(ind)
        style_data_cell(ws10, i, 1, is_alt)
        style_data_cell(ws10, i, 2, is_alt, is_currency=is_cur)

    auto_width(ws10, max_width=25)
    freeze_pane(ws10)

    # ================================================================
    # SALVAR
    # ================================================================
    nome = sanitize_text(op.cliente_nome or "Cliente").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"NASA_HD_Completo_{nome}_{ts}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    return filepath
