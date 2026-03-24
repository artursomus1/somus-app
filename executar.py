"""
Somus Capital - Fluxo de Renda Fixa
App empresarial moderno.
"""

import os
import re
import base64
import threading
from datetime import datetime
from version import VERSION
from updater import verificar_atualizacao, baixar_e_instalar, reiniciar_app
from nasa_engine import (calcular_fluxo_consorcio, calcular_vpl_hd,
                         calcular_financiamento, comparar_consorcio_financiamento,
                         _annual_from_monthly, _monthly_from_annual)
from collections import defaultdict

import customtkinter as ctk
from tkinter import messagebox, filedialog
from PIL import Image
import openpyxl

import matplotlib
matplotlib.use('Agg')
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from gerar_pdf_consorcio import generate_consorcio_pdf, generate_relatorio_consorcio

# Import dos módulos internos
import sys
import tempfile
_SALDOS_MODULE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Mesa Produtos", "Envio Saldos")
if _SALDOS_MODULE not in sys.path:
    sys.path.insert(0, _SALDOS_MODULE)
_SEGUROS_MODULE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Seguros", "Renovacoes Anuais")
if _SEGUROS_MODULE not in sys.path:
    sys.path.insert(0, _SEGUROS_MODULE)

# Import do consolidador (projeto agente investimentos - na pasta Documents)
_DOCS_DIR = os.path.join(os.path.expanduser("~"), "Documents")
_AGENTE_PROJECT = os.path.join(_DOCS_DIR, "Projeto - AGENTE INVESTIMENTOS")
if os.path.isdir(_AGENTE_PROJECT) and _AGENTE_PROJECT not in sys.path:
    sys.path.insert(0, _AGENTE_PROJECT)

# === PATHS ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Shared
LOGO_PATH = os.path.join(BASE_DIR, "assets", "logo_somus.png")
BASE_FILE = os.path.join(BASE_DIR, "BASE", "BASE EMAILS.xlsx")

# Mesa Produtos
ENTRADA_FILE = os.path.join(BASE_DIR, "Mesa Produtos", "Fluxo RF", "ENTRADA", "Agenda de eventos.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Fluxo RF", "PDFs")
AGIO_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Info Agio", "Agio")
RECEITA_FILE = os.path.join(BASE_DIR, "Mesa Produtos", "Ctrl Receita", "CONTROLE RECEITA TOTAL.xlsx")
ORGANIZADOR_OUTPUT_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Organizador", "ORGANIZADO")
CONSOLIDADOR_OUTPUT_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Consolidador", "PDFs")
SALDOS_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Envio Saldos")
ANIVERSARIOS_DIR = os.path.join(BASE_DIR, "Mesa Produtos", "Envio Aniversarios")

# Corporate
CONSÓRCIO_OUTPUT_DIR = os.path.join(BASE_DIR, "Corporate", "Simulador", "PDFs")

# Seguros
SEGUROS_DIR = os.path.join(os.path.expanduser("~"), "SOMUS CAPITAL ASSESSOR DE INVESTIMENTOS LTDA", "Somus Capital - Seguros - Documentos")

# === PALETA SOMUS ===
BG_PRIMARY = "#f7f8fa"
BG_SECONDARY = "#ffffff"
BG_SIDEBAR = "#004d33"
BG_SIDEBAR_HOVER = "#006644"
BG_SIDEBAR_ACTIVE = "#00835a"
BG_HEADER = "#ffffff"
BG_CARD = "#ffffff"
BG_LOG = "#1a1f2e"
BG_LOG_HEADER = "#141824"
BG_PROGRESS_TRACK = "#e4e7ec"
BG_INPUT = "#f2f4f7"

ACCENT_GREEN = "#004d33"
ACCENT_BLUE = "#1863DC"
ACCENT_TEAL = "#00a86b"
ACCENT_ORANGE = "#e6832a"
ACCENT_PURPLE = "#7c5cbf"
ACCENT_RED = "#dc3545"

TEXT_PRIMARY = "#111827"
TEXT_SECONDARY = "#4b5563"
TEXT_TERTIARY = "#9ca3af"
TEXT_WHITE = "#ffffff"
TEXT_SIDEBAR = "#c8e6d8"
TEXT_SIDEBAR_ACTIVE = "#ffffff"
TEXT_LOG = "#c8d0dc"

BORDER_LIGHT = "#e5e7eb"
BORDER_CARD = "#e8ecf0"
SHADOW_CARD = "#d1d5db"


# === CTRL RECEITA: DETAIL SHEETS MAPPING ===
CTRL_DETAIL_SHEETS = [
    {"produto": "Sec RF", "sheet": "SECUND\u00c1RIO RF", "col_assessor": 3, "col_receita": 14},
    {"produto": "Prim RF", "sheet": "PRIMARIO RF", "col_assessor": 2, "col_receita": None},
    {"produto": "RV Corretagem", "sheet": "RENDA VARIAVEL | CORRETAGEM", "col_assessor": 4, "col_receita": 8},
    {"produto": "RV Estruturadas", "sheet": "RENDA VARIAVEL | ESTRUTURADAS", "col_assessor": 9, "col_receita": 7},
    {"produto": "Cambio", "sheet": "C\u00c2MBIO", "col_assessor": 7, "col_receita": 15},
    {"produto": "Fundos Fechados", "sheet": "FUNDOS FECHADOS", "col_assessor": 10, "col_receita": 9},
    {"produto": "Sec Fundos", "sheet": "SECUNDARIO FUNDOS", "col_assessor": 4, "col_receita": 15},
    {"produto": "COE", "sheet": "COE", "col_assessor": 2, "col_receita": 11},
]


# === DATA HELPERS ===

def sanitize_text(text):
    if text is None:
        return ""
    text = str(text)
    replacements = {
        '\u2013': '-', '\u2014': '-', '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2026': '...', '\u2022': '-',
        '\u00ba': 'o', '\u00aa': 'a', '\u00a0': ' ',
        '\ufeff': '', '\u200b': '', '\u200e': '', '\u200f': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def load_base_emails():
    wb = openpyxl.load_workbook(BASE_FILE)
    ws = wb[wb.sheetnames[0]]
    assessores = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        código = str(row[0]).strip()
        if not código.startswith("A"):
            continue
        assessores[código] = {
            "nome": sanitize_text(row[1]) if row[1] else "",
            "email": str(row[2]).strip() if row[2] else "",
            "assistente": sanitize_text(row[3]) if row[3] else "-",
            "email_assistente": str(row[4]).strip() if row[4] else "-",
        }
    wb.close()
    return assessores


def load_equipe_mapping():
    """Carrega mapeamento codigo_assessor -> equipe da planilha CONTROLE RECEITA."""
    mapping = {}
    try:
        if not os.path.exists(RECEITA_FILE):
            return mapping
        wb = openpyxl.load_workbook(RECEITA_FILE, data_only=True)
        ws = wb["CONSOLIDADO"]
        for r in range(7, ws.max_row + 1):
            cod = ws.cell(r, 2).value
            if not cod:
                continue
            equipe = sanitize_text(ws.cell(r, 3).value)
            nome = sanitize_text(ws.cell(r, 4).value).upper()
            # Agrupar PRODUTOS e CORPORATE dentro de SP
            if equipe in ("PRODUTOS", "CORPORATE"):
                equipe = "SP"
            # Reagrupar BACKOFFICE individualmente
            if equipe == "BACKOFFICE":
                if "JULIA" in nome:
                    equipe = "LEBLON"
                elif "GUILHERME" in nome and "PRAXEDES" in nome:
                    equipe = "SP"
                else:
                    equipe = "SP"  # fallback para outros backoffice
            cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
            mapping[f"A{cod_str}"] = equipe
        wb.close()
    except Exception:
        pass
    return mapping


def load_eventos_summary():
    wb = openpyxl.load_workbook(ENTRADA_FILE)
    ws = wb[wb.sheetnames[0]]
    total_eventos = 0
    assessores_set = set()
    ativos = set()
    valor_total = 0.0
    datas = set()
    tipos = defaultdict(lambda: {"count": 0, "total": 0.0})
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cod = row[1]
        if not cod or not isinstance(cod, str) or not cod.startswith("A"):
            continue
        data = row[0]
        if not isinstance(data, datetime):
            continue
        total_eventos += 1
        assessores_set.add(cod)
        if row[8]:
            ativos.add(row[8])
        val = float(row[7]) if row[7] else 0
        valor_total += val
        datas.add(data)
        tipo = sanitize_text(row[9]) if row[9] else "OUTRO"
        tipos[tipo]["count"] += 1
        tipos[tipo]["total"] += val
    wb.close()
    datas_sorted = sorted(datas) if datas else []
    return {
        "total_eventos": total_eventos,
        "assessores": len(assessores_set),
        "ativos": len(ativos),
        "valor_total": valor_total,
        "data_min": datas_sorted[0] if datas_sorted else None,
        "data_max": datas_sorted[-1] if datas_sorted else None,
        "datas_unicas": len(datas_sorted),
        "tipos": dict(tipos),
    }


def find_pdf_for_assessor(código, nome):
    nome_limpo = nome.replace(" ", "_").replace("/", "_")
    expected = f"Fluxo_RF_{código}_{nome_limpo}.pdf"
    path = os.path.join(OUTPUT_DIR, expected)
    if os.path.exists(path):
        return path
    if os.path.exists(OUTPUT_DIR):
        for f in os.listdir(OUTPUT_DIR):
            if f.startswith(f"Fluxo_RF_{código}_") and f.endswith(".pdf"):
                return os.path.join(OUTPUT_DIR, f)
    return None


def fmt_currency(value):
    if value is None or value == 0:
        return "R$ 0,00"
    try:
        v = float(value)
    except (ValueError, TypeError):
        return "R$ 0,00"
    neg = v < 0
    v = abs(v)
    inteiro = int(v)
    centavos = round((v - inteiro) * 100)
    if centavos >= 100:
        inteiro += 1
        centavos = 0
    s_int = f"{inteiro:,}".replace(",", ".")
    result = f"R$ {s_int},{centavos:02d}"
    return f"-{result}" if neg else result


def fmt_pct(value):
    return f"{value:.2f}%".replace(".", ",")


def _load_logo_b64():
    """Carrega logo Somus como base64 para embedding inline no email."""
    if not os.path.exists(LOGO_PATH):
        return ""
    with open(LOGO_PATH, "rb") as f:
        return base64.b64encode(f.read()).decode()


# Tag HTML referenciando a logo via CID (funciona no Outlook)
LOGO_CID = "somus_logo"
LOGO_TAG_CID = (
    f'<img src="cid:{LOGO_CID}" width="155" height="38"'
    f' style="vertical-align:middle;margin-right:16px;" alt="Somus Capital">'
)


def _attach_logo_cid(mail):
    """Anexa logo como imagem oculta com CID para exibir no corpo HTML."""
    if not os.path.exists(LOGO_PATH):
        return
    att = mail.Attachments.Add(os.path.abspath(LOGO_PATH))
    # PR_ATTACH_CONTENT_ID (DASL property)
    att.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
        LOGO_CID)
    # PR_ATTACHMENT_HIDDEN — esconde da lista de anexos
    att.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x7FFE000B",
        True)


def build_email_body(nome_assessor):
    hoje = datetime.now().strftime("%d/%m/%Y")
    primeiro_nome = nome_assessor.split()[0] if nome_assessor else "Assessor"

    logo_tag = LOGO_TAG_CID if os.path.exists(LOGO_PATH) else ""

    return f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">

<!-- Header -->
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>

<!-- Corpo -->
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">

    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#4b5563;margin-top:0;">
      Segue em anexo o relat&oacute;rio de <b>Fluxo de Renda Fixa</b> com os eventos previstos para os pr&oacute;ximos dias.
    </p>

  </td></tr>

  <!-- Secao: Conteudo do relatório -->
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="background-color:#00b876;width:4px;border-radius:4px;">&nbsp;</td>
      <td style="padding-left:12px;">
        <span style="font-family:Calibri,Arial,sans-serif;font-size:12.5pt;color:#004d33;font-weight:bold;letter-spacing:0.3px;">Conte&uacute;do do Relat&oacute;rio</span>
      </td>
    </tr></table>
  </td></tr>

  <tr><td style="padding:0 4px;">
    <table cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;font-size:9.5pt;margin-bottom:6px;">
      <tr style="background:#f7faf9;">
        <td style="padding:5px 12px;color:#00785a;font-weight:bold;border-bottom:1.5px solid #00785a;white-space:nowrap;">&#8226;</td>
        <td style="padding:5px 12px;color:#1a1a2e;border-bottom:1.5px solid #00785a;">Calend&aacute;rio com todas as datas de eventos</td>
      </tr>
      <tr style="background:#ffffff;">
        <td style="padding:5px 12px;color:#00785a;font-weight:bold;border-bottom:1px solid #eef1ef;">&#8226;</td>
        <td style="padding:5px 12px;color:#1a1a2e;border-bottom:1px solid #eef1ef;">Detalhamento por data com ativo, tipo, quantidade e valor</td>
      </tr>
      <tr style="background:#f7faf9;">
        <td style="padding:5px 12px;color:#00785a;font-weight:bold;border-bottom:1px solid #eef1ef;">&#8226;</td>
        <td style="padding:5px 12px;color:#1a1a2e;border-bottom:1px solid #eef1ef;">Resumo consolidado por ativo</td>
      </tr>
    </table>
  </td></tr>

  <tr><td style="padding:14px 4px 0 4px;">
    <p style="font-size:10.5pt;color:#4b5563;margin:0;">
      Em caso de d&uacute;vidas, estamos &agrave; disposi&ccedil;&atilde;o.
    </p>
  </td></tr>

  <!-- Footer -->
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">
            Fluxo de Renda Fixa &middot; {hoje}
          </span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">
            E-mail gerado automaticamente
          </span>
        </td>
      </tr>
    </table>
  </td></tr>

</table>
</div>
"""


def build_email_consorcio(nome_cliente, dados_resumo):
    hoje = datetime.now().strftime("%d/%m/%Y")
    return f"""<html><head><style>
body{{font-family:Calibri,Arial,sans-serif;font-size:11pt;color:#2d2d2d}}
.hb{{background:#004d33;padding:12px 20px}}.hb span{{color:#fff;font-size:14pt;font-weight:bold}}
.ct{{padding:15px 20px}}.ft{{padding:10px 20px;font-size:9pt;color:#888;border-top:1px solid #e0e0e0;margin-top:20px}}
.hl{{color:#004d33;font-weight:bold}}
table{{border-collapse:collapse;width:100%;margin:10px 0}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:10pt}}
td.lb{{color:#666;width:55%}}td.vl{{font-weight:bold;color:#2d2d2d}}
.tag{{display:inline-block;background:#004d33;color:#fff;padding:3px 10px;border-radius:4px;font-size:9pt;margin-bottom:8px}}
</style></head><body>
<div class="hb"><span>Somus Capital</span></div>
<div class="ct">
<p>Prezado(a) <span class="hl">{nome_cliente}</span>,</p>
<p>Segue em anexo a <b>simula&ccedil;&atilde;o de cons&oacute;rcio</b> elaborada pela Somus Capital.</p>
<span class="tag">Cons&oacute;rcio - {dados_resumo.get('tipo_bem', '')}</span>
<table>
<tr><td class="lb">Carta de Cr&eacute;dito</td><td class="vl">{dados_resumo.get('valor_carta_fmt', '')}</td></tr>
<tr><td class="lb">Administradora</td><td class="vl">{dados_resumo.get('administradora', '')}</td></tr>
<tr><td class="lb">Prazo do Grupo</td><td class="vl">{dados_resumo.get('prazo', '')} meses</td></tr>
<tr><td class="lb">Parcela Fase 1</td><td class="vl">{dados_resumo.get('parcela_f1', '')}/m&ecirc;s</td></tr>
<tr><td class="lb">Custo Efetivo</td><td class="vl">{dados_resumo.get('ce_anual', '')} a.a.</td></tr>
</table>
<p>Os documentos anexos cont&ecirc;m:</p>
<ul>
<li><b>Proposta de Simula&ccedil;&atilde;o</b> - valores detalhados, cronograma e resumo financeiro</li>
<li><b>Relat&oacute;rio Explicativo</b> - métodologia de c&aacute;lculo passo a passo</li>
</ul>
<p>Fico &agrave; disposi&ccedil;&atilde;o para esclarecer qualquer d&uacute;vida.</p>
<p>Atenciosamente,<br><span class="hl">Equipe Somus Capital</span><br>Cons&oacute;rcios</p>
</div>
<div class="ft">Simula&ccedil;&atilde;o gerada em {hoje}. Valores estimados, sujeitos a altera&ccedil;&atilde;o pela administradora.</div>
</body></html>"""


# =====================================================================
#  APP
# =====================================================================

class App(ctk.CTk):
    def __init__(self, role="admin"):
        super().__init__()
        self.role = role
        self.title("Somus Capital")
        self.geometry("1100x720")
        self.minsize(1000, 650)
        self.configure(fg_color=BG_PRIMARY)
        self._set_icon()

        self.current_page = "corp_dashboard" if role == "corporate" else "dashboard"
        self.current_module = "Corporate" if role == "corporate" else "Mesa Produtos"
        self.sidebar_buttons = {}
        self.mp_nav_keys = []
        self.cs_nav_keys = []
        self.sg_nav_keys = []
        self.pages = {}

        # Grid: sidebar | content
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self._build_sidebar()
        self._build_pages()

        if role == "corporate":
            self._show_page("corp_dashboard")
            self._load_data_async()
            self.after(50, lambda: self._on_module_change("Corporate"))
        else:
            self._show_page("dashboard")
            self._load_data_async()
            self.after(50, lambda: self._on_module_change("Mesa Produtos"))

        # Verificar atualizacoes em background (4s de delay)
        self.after(4000, self._checar_atualizacao)

    def _set_icon(self):
        try:
            ico_path = os.path.join(BASE_DIR, "assets", "icon_somus.ico")
            if os.path.exists(ico_path):
                self.iconbitmap(ico_path)
        except Exception:
            pass

    # =================================================================
    #  AUTO UPDATE
    # =================================================================

    def _checar_atualizacao(self):
        def _on_disponivel(versao, url):
            self.after(0, lambda: self._mostrar_update_dialog(versao, url))
        verificar_atualizacao(_on_disponivel)

    def _mostrar_update_dialog(self, versao, url):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Atualizacao disponivel")
        dialog.geometry("420x220")
        dialog.resizable(False, False)
        dialog.grab_set()
        dialog.configure(fg_color=BG_SECONDARY)
        dialog.after(200, dialog.lift)

        ctk.CTkLabel(
            dialog, text=f"Nova versao disponivel: v{versao}",
            font=ctk.CTkFont(size=15, weight="bold"), text_color=ACCENT_GREEN
        ).pack(pady=(28, 4))
        ctk.CTkLabel(
            dialog, text="Deseja atualizar agora? O app sera reiniciado.",
            font=ctk.CTkFont(size=12), text_color=TEXT_SECONDARY
        ).pack(pady=(0, 16))

        barra = ctk.CTkProgressBar(dialog, width=340)
        barra.set(0)

        label_status = ctk.CTkLabel(dialog, text="", font=ctk.CTkFont(size=11), text_color=TEXT_TERTIARY)

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(pady=4)

        def _iniciar():
            btn_sim.configure(state="disabled")
            btn_nao.configure(state="disabled")
            barra.pack(pady=(4, 2))
            label_status.pack()

            def _progresso(pct, msg):
                self.after(0, lambda: barra.set(pct / 100))
                self.after(0, lambda: label_status.configure(text=msg))

            def _concluido():
                self.after(0, lambda: label_status.configure(text="Reiniciando..."))
                self.after(800, reiniciar_app)

            def _erro(msg):
                self.after(0, lambda: label_status.configure(
                    text=f"Erro: {msg}", text_color=ACCENT_RED
                ))
                self.after(0, lambda: btn_nao.configure(state="normal", text="Fechar"))

            baixar_e_instalar(url, _progresso, _concluido, _erro)

        btn_sim = ctk.CTkButton(
            btn_frame, text="Atualizar agora", width=160,
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            command=_iniciar
        )
        btn_sim.pack(side="left", padx=8)

        btn_nao = ctk.CTkButton(
            btn_frame, text="Agora nao", width=120,
            fg_color=BG_INPUT, hover_color=BORDER_LIGHT,
            text_color=TEXT_PRIMARY, command=dialog.destroy
        )
        btn_nao.pack(side="left", padx=8)

    # =================================================================
    #  SIDEBAR
    # =================================================================
    def _build_sidebar(self):
        sidebar = ctk.CTkFrame(self, fg_color=BG_SIDEBAR, width=245, corner_radius=0)
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        self._sidebar = sidebar  # prevent GC

        # === HEADER FIXO (topo) ===
        header = ctk.CTkFrame(sidebar, fg_color="transparent")
        header.pack(fill="x", side="top")

        # Logo
        logo_frame = ctk.CTkFrame(header, fg_color="transparent", height=78)
        logo_frame.pack(fill="x")
        logo_frame.pack_propagate(False)
        self._logo_frame = logo_frame  # prevent GC

        try:
            logo_img = Image.open(LOGO_PATH)
            self._logo_ctk = ctk.CTkImage(light_image=logo_img, dark_image=logo_img, size=(155, 38))
            ctk.CTkLabel(logo_frame, image=self._logo_ctk, text="").place(relx=0.5, rely=0.5, anchor="center")
        except Exception:
            ctk.CTkLabel(
                logo_frame, text="SOMUS CAPITAL",
                font=("Segoe UI", 16, "bold"), text_color=TEXT_WHITE
            ).place(relx=0.5, rely=0.5, anchor="center")

        # Separador
        ctk.CTkFrame(header, fg_color="#006644", height=1, corner_radius=0).pack(fill="x", padx=18)

        # Seletor de modulo
        selector_frame = ctk.CTkFrame(header, fg_color="transparent")
        selector_frame.pack(fill="x", padx=10, pady=(12, 10))
        self._selector_frame = selector_frame  # prevent GC

        ctk.CTkLabel(
            selector_frame, text="MÓDULO",
            font=("Segoe UI", 9, "bold"), text_color="#5a9a7a", anchor="w"
        ).pack(fill="x", padx=4, pady=(0, 4))

        _module_values = ["Corporate"] if self.role == "corporate" else ["Mesa Produtos", "Corporate", "Seguros"]
        self.module_selector = ctk.CTkOptionMenu(
            selector_frame,
            values=_module_values,
            font=("Segoe UI", 12, "bold"),
            fg_color="#003d2b",
            button_color=ACCENT_BLUE,
            button_hover_color="#1555bb",
            dropdown_fg_color="#002a1a",
            dropdown_hover_color=BG_SIDEBAR_HOVER,
            dropdown_text_color=TEXT_WHITE,
            text_color=TEXT_WHITE,
            corner_radius=8,
            height=36,
            anchor="w",
            command=self._on_module_change,
        )
        self.module_selector.pack(fill="x")
        self.module_selector.set("Corporate" if self.role == "corporate" else "Mesa Produtos")

        # Separador 2
        ctk.CTkFrame(header, fg_color="#006644", height=1, corner_radius=0).pack(fill="x", padx=18)

        # Label MENU
        ctk.CTkLabel(
            header, text="MENU", font=("Segoe UI", 9, "bold"),
            text_color="#5a9a7a", anchor="w"
        ).pack(fill="x", padx=24, pady=(12, 6))

        # === FOOTER FIXO (base) — declarar ANTES do scroll para pack order ===
        footer = ctk.CTkFrame(sidebar, fg_color="transparent")
        footer.pack(fill="x", side="bottom")

        ctk.CTkFrame(footer, fg_color="#006644", height=1, corner_radius=0).pack(fill="x", padx=18)

        ctk.CTkButton(
            footer,
            text="  \u26a0    Reportar Erro",
            font=("Segoe UI", 11),
            fg_color="transparent",
            hover_color="#4d1a1a",
            text_color="#e07070",
            anchor="w",
            height=36,
            corner_radius=8,
            command=self._on_reportar_erro,
        ).pack(fill="x", padx=12, pady=(8, 4))

        ctk.CTkLabel(
            footer, text="SomusApp - BETA",
            font=("Segoe UI", 8), text_color="#3a6a50"
        ).pack(pady=(4, 14))

        # === AREA SCROLLAVEL para botoes de navegacao ===
        nav_scroll = ctk.CTkScrollableFrame(
            sidebar, fg_color="transparent",
            scrollbar_button_color=BG_SIDEBAR_HOVER,
            scrollbar_button_hover_color=BG_SIDEBAR_ACTIVE,
        )
        nav_scroll.pack(fill="both", expand=True, padx=0, pady=0)
        self._nav_scroll = nav_scroll  # prevent GC

        # === MESA PRODUTOS nav buttons — oculto para corporate ===
        mp_nav_items = [] if self.role == "corporate" else [
            ("dashboard", "Dashboard", "\u25a3"),
            ("fluxo_rf", "FLUXO - RF", "\u2913"),
            ("informativo", "Informativo", "\u2709"),
            ("info_agio", "Info - Agio", "\u25b2"),
            ("envio_ordens", "Envio de Ordens", "\u2191"),
            ("ordem_massa", "Ordem MASSA", "\u21c8"),
            ("ctrl_receita", "Ctrl Receita", "$"),
            ("organizador", "Organizador", "\u25a6"),
            ("consolidador", "Consolidador", "\u229a"),
            ("saldos", "Envio Saldos", "\u21c6"),
            ("envio_mesa", "Envio Mesa", "\u2709"),
            ("envio_aniversarios", "Envio Aniversários", "\u2605"),
            ("tarefas", "Tarefas", "\u2611"),
            ("top_picks", "Top Picks", "\u2605"),
        ] if self.role != "corporate" else []
        for key, label, icon in mp_nav_items:
            btn = ctk.CTkButton(
                nav_scroll,
                text=f"  {icon}    {label}",
                font=("Segoe UI", 13),
                fg_color="transparent",
                hover_color=BG_SIDEBAR_HOVER,
                text_color=TEXT_SIDEBAR,
                anchor="w",
                height=42,
                corner_radius=8,
                command=lambda k=key: self._on_nav(k),
            )
            btn.pack(fill="x", padx=12, pady=2)
            self.sidebar_buttons[key] = btn
            self.mp_nav_keys.append(key)

        # === CORPORATE nav buttons (ocultos inicialmente) ===
        cs_nav_items = [
            ("corp_dashboard", "Dashboard", "\u25a3"),
            ("simulador", "Simulador", "\u2630"),
            ("comparativo_vpl", "Comparativo VPL", "\u25b2"),
            ("consorcio_vs_financ", "Cons. vs Financ.", "\u21c4"),
            ("fluxo_receitas", "Fluxo de Receitas", "\u25b6"),
        ]
        for key, label, icon in cs_nav_items:
            btn = ctk.CTkButton(
                nav_scroll,
                text=f"  {icon}    {label}",
                font=("Segoe UI", 13),
                fg_color="transparent",
                hover_color=BG_SIDEBAR_HOVER,
                text_color=TEXT_SIDEBAR,
                anchor="w",
                height=42,
                corner_radius=8,
                command=lambda k=key: self._on_nav(k),
            )
            btn.pack(fill="x", padx=12, pady=2)
            btn.pack_forget()
            self.sidebar_buttons[key] = btn
            self.cs_nav_keys.append(key)

        # === SEGUROS nav buttons (ocultos para corporate) ===
        sg_nav_items = [] if self.role == "corporate" else [
            ("seg_renovacoes", "Renovações Anuais", "\u26e8"),
        ]
        for key, label, icon in sg_nav_items:
            btn = ctk.CTkButton(
                nav_scroll,
                text=f"  {icon}    {label}",
                font=("Segoe UI", 13),
                fg_color="transparent",
                hover_color=BG_SIDEBAR_HOVER,
                text_color=TEXT_SIDEBAR,
                anchor="w",
                height=42,
                corner_radius=8,
                command=lambda k=key: self._on_nav(k),
            )
            btn.pack(fill="x", padx=12, pady=2)
            btn.pack_forget()
            self.sidebar_buttons[key] = btn
            self.sg_nav_keys.append(key)

    def _on_nav(self, page_key):
        if page_key == "simulador":
            self._show_page("consorcio")
        elif page_key == "corp_dashboard":
            self._show_page("corp_dashboard")
        elif page_key == "seg_renovacoes":
            self._show_page("seg_renovacoes")
        else:
            self._show_page(page_key)

    def _on_module_change(self, value):
        self.current_module = value
        all_nav = [
            (self.mp_nav_keys, "Mesa Produtos", "dashboard"),
            (self.cs_nav_keys, "Corporate", "corp_dashboard"),
            (self.sg_nav_keys, "Seguros", "seg_renovacoes"),
        ]
        for keys, module, default_page in all_nav:
            if value == module:
                for key in keys:
                    self.sidebar_buttons[key].pack(fill="x", padx=12, pady=2)
            else:
                for key in keys:
                    self.sidebar_buttons[key].pack_forget()
        default_pages = {"Mesa Produtos": "dashboard", "Corporate": "corp_dashboard", "Seguros": "seg_renovacoes"}
        self._show_page(default_pages.get(value, "dashboard"))

    def _update_sidebar_active(self, active_key):
        for key, btn in self.sidebar_buttons.items():
            if key == active_key:
                btn.configure(fg_color=BG_SIDEBAR_ACTIVE, text_color=TEXT_WHITE)
            else:
                btn.configure(fg_color="transparent", text_color=TEXT_SIDEBAR)

    # =================================================================
    #  TOPBAR HELPER
    # =================================================================
    def _make_topbar(self, parent, title, subtitle="", show_date=True,
                     back_btn=False, back_cmd=None):
        """Cria topbar profissional padronizado com logo no canto direito."""
        wrapper = ctk.CTkFrame(parent, fg_color="transparent", corner_radius=0)
        wrapper.pack(fill="x")

        topbar = ctk.CTkFrame(wrapper, fg_color=BG_HEADER, height=68, corner_radius=0)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)

        # --- Lado esquerdo: título + subtítulo ---
        left = ctk.CTkFrame(topbar, fg_color="transparent")
        left.pack(side="left", fill="y", padx=28)

        title_lbl = ctk.CTkLabel(
            left, text=title,
            font=("Segoe UI", 21, "bold"), text_color=TEXT_PRIMARY
        )
        title_lbl.pack(side="left", anchor="w", pady=0)

        if subtitle:
            sep_lbl = ctk.CTkLabel(
                left, text="  |  ",
                font=("Segoe UI", 13), text_color=BORDER_CARD
            )
            sep_lbl.pack(side="left", anchor="w")
            sub_lbl = ctk.CTkLabel(
                left, text=subtitle,
                font=("Segoe UI", 12), text_color=TEXT_TERTIARY
            )
            sub_lbl.pack(side="left", anchor="w")

        # --- Lado direito: logo + data ---
        right = ctk.CTkFrame(topbar, fg_color="transparent")
        right.pack(side="right", fill="y", padx=24)

        try:
            logo_img = Image.open(LOGO_PATH)
            logo_topbar = ctk.CTkImage(light_image=logo_img, dark_image=logo_img, size=(140, 34))
            logo_lbl = ctk.CTkLabel(right, image=logo_topbar, text="")
            logo_lbl.pack(side="right", anchor="e", padx=(16, 4))
            logo_lbl._topbar_logo = logo_topbar  # manter referencia
        except Exception:
            pass

        if show_date:
            date_lbl = ctk.CTkLabel(
                right, text=datetime.now().strftime("%d/%m/%Y   %H:%M"),
                font=("Segoe UI", 10), text_color=TEXT_TERTIARY
            )
            date_lbl.pack(side="right", anchor="e", padx=(0, 8))

        if back_btn and back_cmd:
            ctk.CTkButton(
                right, text="\u2190  Voltar",
                font=("Segoe UI", 10), fg_color=BG_INPUT,
                hover_color=BORDER_CARD, text_color=TEXT_SECONDARY,
                height=30, corner_radius=8, width=90,
                command=back_cmd
            ).pack(side="right", anchor="e", padx=(0, 12))

        # --- Filete verde embaixo ---
        accent = ctk.CTkFrame(wrapper, fg_color=ACCENT_GREEN, height=3, corner_radius=0)
        accent.pack(fill="x")

        return wrapper, title_lbl, date_lbl if show_date else None

    # =================================================================
    #  PAGES
    # =================================================================
    def _build_pages(self):
        # Corporate: apenas modulo Corporate
        self.pages["consorcio"] = self._build_consorcio_page()
        self.pages["corp_dashboard"] = self._build_corp_dashboard_page()
        self.pages["comparativo_vpl"] = self._build_comparativo_vpl_page()
        self.pages["consorcio_vs_financ"] = self._build_consorcio_vs_financ_page()
        self.pages["fluxo_receitas"] = self._build_fluxo_receitas_page()

        if self.role == "corporate":
            return

        # Admin: todos os modulos
        self.pages["dashboard"] = self._build_receita_dashboard_page()
        self.pages["fluxo_rf"] = self._build_dashboard_page()
        self.pages["operations"] = self._build_operations_page()
        self.pages["informativo"] = self._build_informativo_page()
        self.pages["info_agio"] = self._build_info_agio_page()
        self.pages["envio_ordens"] = self._build_envio_ordens_page()
        self.pages["ordem_massa"] = self._build_ordem_massa_page()
        self.pages["ctrl_receita"] = self._build_ctrl_receita_page()
        self.pages["organizador"] = self._build_organizador_page()
        self.pages["consolidador"] = self._build_consolidador_page()
        self.pages["saldos"] = self._build_saldos_page()
        self.pages["envio_mesa"] = self._build_envio_mesa_page()
        self.pages["envio_aniversarios"] = self._build_envio_aniversarios_page()
        self.pages["tarefas"] = self._build_tarefas_page()
        self.pages["seg_renovacoes"] = self._build_renovacoes_page()
        self.pages["top_picks"] = self._build_top_picks_page()

    def _show_page(self, page_key):
        for key, frame in self.pages.items():
            frame.grid_remove()
        self.pages[page_key].grid(row=0, column=1, sticky="nsew")
        self.current_page = page_key

        # Lazy load ctrl_receita data on first visit
        if page_key == "ctrl_receita" and not self._cr_loading and self._cr_data is None:
            self._cr_loading = True
            threading.Thread(target=self._load_ctrl_receita_data, daemon=True).start()

        # Lazy load info_agio data on first visit
        if page_key == "info_agio" and not getattr(self, "_ag_loaded", False):
            self._ag_loaded = True
            threading.Thread(target=self._ag_load_data_thread, daemon=True).start()

        nav_map = {
            "dashboard": "dashboard",
            "fluxo_rf": "fluxo_rf",
            "ctrl_receita": "ctrl_receita",
            "operations": "fluxo_rf",
            "corp_dashboard": "corp_dashboard",
            "consorcio": "simulador",
            "comparativo_vpl": "comparativo_vpl",
            "consorcio_vs_financ": "consorcio_vs_financ",
            "fluxo_receitas": "fluxo_receitas",
            "seg_renovacoes": "seg_renovacoes",
        }
        self._update_sidebar_active(nav_map.get(page_key, page_key))

    # -----------------------------------------------------------------
    #  PAGE: RECEITA DASHBOARD
    # -----------------------------------------------------------------
    def _build_receita_dashboard_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        _, _, _ = self._make_topbar(page, "Dashboard", subtitle="Mesa de Produtos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Status banner
        self.rc_banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=44)
        self.rc_banner.pack(fill="x", pady=(0, 18))
        self.rc_banner.pack_propagate(False)

        self.rc_banner_text = ctk.CTkLabel(
            self.rc_banner, text="  Carregando dados de receita...",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        )
        self.rc_banner_text.pack(side="left", padx=18, pady=10)

        self.rc_banner_date = ctk.CTkLabel(
            self.rc_banner, text="",
            font=("Segoe UI", 11), text_color="#80c0a0"
        )
        self.rc_banner_date.pack(side="right", padx=18)

        # RECEITA TOTAL destaque
        total_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        total_card.pack(fill="x", pady=(0, 18))
        total_inner = ctk.CTkFrame(total_card, fg_color="transparent")
        total_inner.pack(fill="x", padx=24, pady=16)

        ctk.CTkLabel(
            total_inner, text="RECEITA TOTAL",
            font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY
        ).pack(anchor="w")

        self.rc_total_label = ctk.CTkLabel(
            total_inner, text="Carregando...",
            font=("Segoe UI", 34, "bold"), text_color=ACCENT_GREEN
        )
        self.rc_total_label.pack(anchor="w", pady=(2, 0))

        self.rc_total_sub = ctk.CTkLabel(
            total_inner, text="",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY
        )
        self.rc_total_sub.pack(anchor="w")

        # KPI cards - 4 maiores categorias
        kpi_frame = ctk.CTkFrame(content, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=(0, 8))
        kpi_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self.rc_kpi_rf = self._make_kpi_card(kpi_frame, "Renda Fixa", "...", ACCENT_GREEN, 0)
        self.rc_kpi_rv_corr = self._make_kpi_card(kpi_frame, "RV Corretagem", "...", ACCENT_BLUE, 1)
        self.rc_kpi_fundos = self._make_kpi_card(kpi_frame, "Fundos Fechados", "...", ACCENT_PURPLE, 2)
        self.rc_kpi_rv_pe = self._make_kpi_card(kpi_frame, "RV Estruturadas", "...", ACCENT_ORANGE, 3)

        # KPI row 2 - menores categorias
        kpi_frame2 = ctk.CTkFrame(content, fg_color="transparent")
        kpi_frame2.pack(fill="x", pady=(0, 18))
        kpi_frame2.columnconfigure((0, 1, 2), weight=1)

        self.rc_kpi_cambio = self._make_kpi_card(kpi_frame2, "Cambio", "...", ACCENT_TEAL, 0)
        self.rc_kpi_coe = self._make_kpi_card(kpi_frame2, "COE", "...", ACCENT_RED, 1)
        self.rc_kpi_assessores = self._make_kpi_card(kpi_frame2, "Assessores", "...", "#6b7280", 2)

        # Distribuição por Produto (barra horizontal visual)
        ctk.CTkLabel(
            content, text="Distribuição por Produto",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        self.rc_bars_frame = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        self.rc_bars_frame.pack(fill="x", pady=(0, 18))

        self.rc_bars_inner = ctk.CTkFrame(self.rc_bars_frame, fg_color="transparent")
        self.rc_bars_inner.pack(fill="x", padx=20, pady=16)

        # Ranking Assessores
        ctk.CTkLabel(
            content, text="Ranking por Assessor",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        self.rc_ranking_frame = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        self.rc_ranking_frame.pack(fill="x", pady=(0, 18))

        self.rc_ranking_inner = ctk.CTkFrame(self.rc_ranking_frame, fg_color="transparent")
        self.rc_ranking_inner.pack(fill="x", padx=20, pady=16)

        # Placeholder
        self.rc_ranking_placeholder = ctk.CTkLabel(
            self.rc_ranking_inner, text="Carregando ranking...",
            font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.rc_ranking_placeholder.pack(anchor="w")

        # Ranking por Equipe
        ctk.CTkLabel(
            content, text="Receita por Equipe",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        self.rc_equipe_frame = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        self.rc_equipe_frame.pack(fill="x", pady=(0, 18))

        self.rc_equipe_inner = ctk.CTkFrame(self.rc_equipe_frame, fg_color="transparent")
        self.rc_equipe_inner.pack(fill="x", padx=20, pady=16)

        # Botao Atualizar
        self._make_atualizar_btn(content, RECEITA_FILE)

        # Load data async (delay to ensure mainloop is running)
        self.after(300, lambda: threading.Thread(target=self._load_receita_data, daemon=True).start())

        return page

    def _load_receita_data(self):
        """Carrega dados de CONTROLE RECEITA TOTAL.xlsx e atualiza o dashboard."""
        try:
            if not os.path.exists(RECEITA_FILE):
                self.after(0, lambda: self.rc_banner_text.configure(
                    text="  Arquivo CONTROLE RECEITA TOTAL.xlsx não encontrado"))
                return

            wb = openpyxl.load_workbook(RECEITA_FILE, data_only=True)
            ws = wb["CONSOLIDADO"]

            # Data base
            data_base = ws.cell(3, 2).value
            if data_base and hasattr(data_base, "strftime"):
                data_str = data_base.strftime("%d/%m/%Y")
            else:
                data_str = str(data_base) if data_base else "N/A"

            # Headers na row 6 (B=2..K=11)
            # B=Assessor, C=Equipe, D=Nome, E=RF, F=RV Corr, G=RV PE, H=Cambio, I=Fundos, J=COE, K=Receita Geral
            assessores = []
            for r in range(7, ws.max_row + 1):
                cod = ws.cell(r, 2).value
                if not cod:
                    continue
                nome = sanitize_text(ws.cell(r, 4).value)
                equipe = sanitize_text(ws.cell(r, 3).value)
                rf = float(ws.cell(r, 5).value or 0)
                rv_corr = float(ws.cell(r, 6).value or 0)
                rv_pe = float(ws.cell(r, 7).value or 0)
                cambio = float(ws.cell(r, 8).value or 0)
                fundos = float(ws.cell(r, 9).value or 0)
                coe = float(ws.cell(r, 10).value or 0)
                receita = float(ws.cell(r, 11).value or 0)
                assessores.append({
                    "cod": str(cod), "nome": nome, "equipe": equipe,
                    "rf": rf, "rv_corr": rv_corr, "rv_pe": rv_pe,
                    "cambio": cambio, "fundos": fundos, "coe": coe,
                    "receita": receita,
                })
            wb.close()

            # Totais por categoria
            total_rf = sum(a["rf"] for a in assessores)
            total_rv_corr = sum(a["rv_corr"] for a in assessores)
            total_rv_pe = sum(a["rv_pe"] for a in assessores)
            total_cambio = sum(a["cambio"] for a in assessores)
            total_fundos = sum(a["fundos"] for a in assessores)
            total_coe = sum(a["coe"] for a in assessores)
            total_geral = sum(a["receita"] for a in assessores)
            n_assessores = len([a for a in assessores if a["receita"] > 0])

            # Equipes
            equipes = defaultdict(float)
            for a in assessores:
                if a["receita"] > 0:
                    equipes[a["equipe"]] += a["receita"]

            # Ranking (top 10 assessores por receita)
            ranking = sorted(assessores, key=lambda x: x["receita"], reverse=True)[:10]

            # Categorias para barras
            categorias = [
                ("Renda Fixa", total_rf, ACCENT_GREEN),
                ("RV Corretagem", total_rv_corr, ACCENT_BLUE),
                ("Fundos Fechados", total_fundos, ACCENT_PURPLE),
                ("RV Estruturadas", total_rv_pe, ACCENT_ORANGE),
                ("Cambio", total_cambio, ACCENT_TEAL),
                ("COE", total_coe, ACCENT_RED),
            ]

            # Update UI on main thread
            def update_ui():
                self.rc_banner_text.configure(text=f"  Receita consolidada - {len(assessores)} assessores")
                self.rc_banner_date.configure(text=f"Base: {data_str}")

                self.rc_total_label.configure(text=fmt_currency(total_geral))
                self.rc_total_sub.configure(text=f"{n_assessores} assessores com receita ativa")

                self.rc_kpi_rf.configure(text=fmt_currency(total_rf))
                self.rc_kpi_rv_corr.configure(text=fmt_currency(total_rv_corr))
                self.rc_kpi_fundos.configure(text=fmt_currency(total_fundos))
                self.rc_kpi_rv_pe.configure(text=fmt_currency(total_rv_pe))
                self.rc_kpi_cambio.configure(text=fmt_currency(total_cambio))
                self.rc_kpi_coe.configure(text=fmt_currency(total_coe))
                self.rc_kpi_assessores.configure(text=str(n_assessores))

                # Barras de distribuição
                for w in self.rc_bars_inner.winfo_children():
                    w.destroy()

                max_val = max(v for _, v, _ in categorias) if categorias else 1
                for cat_name, cat_val, cat_color in categorias:
                    row_frame = ctk.CTkFrame(self.rc_bars_inner, fg_color="transparent")
                    row_frame.pack(fill="x", pady=3)

                    pct = (cat_val / total_geral * 100) if total_geral > 0 else 0

                    lbl = ctk.CTkLabel(
                        row_frame, text=cat_name, font=("Segoe UI", 11),
                        text_color=TEXT_PRIMARY, width=120, anchor="w"
                    )
                    lbl.pack(side="left")

                    bar_track = ctk.CTkFrame(row_frame, fg_color=BG_PROGRESS_TRACK, height=18, corner_radius=4)
                    bar_track.pack(side="left", fill="x", expand=True, padx=(8, 8))
                    bar_track.pack_propagate(False)

                    bar_width = max(cat_val / max_val, 0.02) if max_val > 0 else 0.02
                    bar_fill = ctk.CTkFrame(bar_track, fg_color=cat_color, corner_radius=4)
                    bar_fill.place(relx=0, rely=0, relwidth=bar_width, relheight=1.0)

                    val_lbl = ctk.CTkLabel(
                        row_frame, text=f"{fmt_currency(cat_val)}  ({pct:.1f}%)",
                        font=("Segoe UI", 10, "bold"), text_color=cat_color, width=180, anchor="e"
                    )
                    val_lbl.pack(side="right")

                # Ranking assessores
                self.rc_ranking_placeholder.destroy()
                for w in self.rc_ranking_inner.winfo_children():
                    w.destroy()

                # Header
                hdr = ctk.CTkFrame(self.rc_ranking_inner, fg_color="transparent")
                hdr.pack(fill="x", pady=(0, 8))
                ctk.CTkLabel(hdr, text="#", font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY, width=30, anchor="w").pack(side="left")
                ctk.CTkLabel(hdr, text="Assessor", font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY, width=200, anchor="w").pack(side="left")
                ctk.CTkLabel(hdr, text="Equipe", font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY, width=100, anchor="w").pack(side="left")
                ctk.CTkLabel(hdr, text="Receita Total", font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY, anchor="e").pack(side="right")

                sep = ctk.CTkFrame(self.rc_ranking_inner, fg_color=BORDER_LIGHT, height=1)
                sep.pack(fill="x", pady=(0, 4))

                medal_colors = [("#FFD700", "#B8860B"), ("#C0C0C0", "#808080"), ("#CD7F32", "#8B5A2B")]

                for idx, a in enumerate(ranking):
                    if a["receita"] <= 0:
                        continue
                    row = ctk.CTkFrame(self.rc_ranking_inner, fg_color="transparent")
                    row.pack(fill="x", pady=2)

                    pos = idx + 1
                    if idx < 3:
                        pos_color = medal_colors[idx][1]
                        pos_font = ("Segoe UI", 11, "bold")
                    else:
                        pos_color = TEXT_TERTIARY
                        pos_font = ("Segoe UI", 10)

                    ctk.CTkLabel(row, text=str(pos), font=pos_font, text_color=pos_color, width=30, anchor="w").pack(side="left")
                    ctk.CTkLabel(row, text=a["nome"], font=("Segoe UI", 11), text_color=TEXT_PRIMARY, width=200, anchor="w").pack(side="left")
                    ctk.CTkLabel(row, text=a["equipe"], font=("Segoe UI", 10), text_color=TEXT_SECONDARY, width=100, anchor="w").pack(side="left")

                    val_color = ACCENT_GREEN if idx < 3 else TEXT_PRIMARY
                    ctk.CTkLabel(row, text=fmt_currency(a["receita"]), font=("Segoe UI", 11, "bold"), text_color=val_color, anchor="e").pack(side="right")

                    if idx < len(ranking) - 1 and ranking[idx + 1]["receita"] > 0:
                        ctk.CTkFrame(self.rc_ranking_inner, fg_color=BORDER_LIGHT, height=1).pack(fill="x")

                # Equipes
                for w in self.rc_equipe_inner.winfo_children():
                    w.destroy()

                equipe_sorted = sorted(equipes.items(), key=lambda x: x[1], reverse=True)
                equipe_colors = [ACCENT_GREEN, ACCENT_BLUE, ACCENT_ORANGE, ACCENT_PURPLE, ACCENT_TEAL, ACCENT_RED, "#6b7280", "#9b59b6"]
                max_equipe = equipe_sorted[0][1] if equipe_sorted else 1

                for ei, (eq_name, eq_val) in enumerate(equipe_sorted):
                    row_frame = ctk.CTkFrame(self.rc_equipe_inner, fg_color="transparent")
                    row_frame.pack(fill="x", pady=4)

                    eq_color = equipe_colors[ei % len(equipe_colors)]
                    pct = (eq_val / total_geral * 100) if total_geral > 0 else 0

                    ctk.CTkLabel(
                        row_frame, text=eq_name, font=("Segoe UI", 11, "bold"),
                        text_color=TEXT_PRIMARY, width=120, anchor="w"
                    ).pack(side="left")

                    bar_track = ctk.CTkFrame(row_frame, fg_color=BG_PROGRESS_TRACK, height=22, corner_radius=5)
                    bar_track.pack(side="left", fill="x", expand=True, padx=(8, 8))
                    bar_track.pack_propagate(False)

                    bar_width = max(eq_val / max_equipe, 0.02) if max_equipe > 0 else 0.02
                    bar_fill = ctk.CTkFrame(bar_track, fg_color=eq_color, corner_radius=5)
                    bar_fill.place(relx=0, rely=0, relwidth=bar_width, relheight=1.0)

                    ctk.CTkLabel(
                        row_frame, text=f"{fmt_currency(eq_val)}  ({pct:.1f}%)",
                        font=("Segoe UI", 10, "bold"), text_color=eq_color, width=180, anchor="e"
                    ).pack(side="right")

            self.after(0, update_ui)

        except Exception as e:
            import traceback; traceback.print_exc()
            err_msg = f"  Erro ao carregar receita: {str(e)[:60]}"
            self.after(0, lambda m=err_msg: self.rc_banner_text.configure(text=m))

    # -----------------------------------------------------------------
    #  PAGE: CTRL RECEITA (BI Dashboard)
    # -----------------------------------------------------------------
    def _build_ctrl_receita_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        _, _, _ = self._make_topbar(page, "Controle Receita", subtitle="Dashboard BI")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Banner
        self.cr_banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=44)
        self.cr_banner.pack(fill="x", pady=(0, 14))
        self.cr_banner.pack_propagate(False)

        self.cr_banner_text = ctk.CTkLabel(
            self.cr_banner, text="  Carregando dados...",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        )
        self.cr_banner_text.pack(side="left", padx=18, pady=10)

        self.cr_banner_date = ctk.CTkLabel(
            self.cr_banner, text="",
            font=("Segoe UI", 11), text_color="#80c0a0"
        )
        self.cr_banner_date.pack(side="right", padx=18)

        # Filter
        filter_frame = ctk.CTkFrame(content, fg_color="transparent")
        filter_frame.pack(fill="x", pady=(0, 14))

        ctk.CTkLabel(
            filter_frame, text="Assessor:",
            font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY
        ).pack(side="left", padx=(0, 8))

        self.cr_assessor_var = ctk.StringVar(value="Todos")
        self.cr_assessor_menu = ctk.CTkOptionMenu(
            filter_frame, variable=self.cr_assessor_var,
            values=["Todos"], width=300,
            font=("Segoe UI", 12),
            fg_color=BG_INPUT, button_color=ACCENT_GREEN,
            button_hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_PRIMARY,
            command=self._on_ctrl_assessor_change,
        )
        self.cr_assessor_menu.pack(side="left")

        # === OVERVIEW FRAME ===
        self.cr_overview_frame = ctk.CTkFrame(content, fg_color="transparent")
        self.cr_overview_frame.pack(fill="x")

        # RECEITA TOTAL card
        cr_total_card = ctk.CTkFrame(
            self.cr_overview_frame, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER_CARD
        )
        cr_total_card.pack(fill="x", pady=(0, 14))
        cr_total_inner = ctk.CTkFrame(cr_total_card, fg_color="transparent")
        cr_total_inner.pack(fill="x", padx=24, pady=16)

        ctk.CTkLabel(
            cr_total_inner, text="RECEITA TOTAL",
            font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY
        ).pack(anchor="w")

        self.cr_total_label = ctk.CTkLabel(
            cr_total_inner, text="...",
            font=("Segoe UI", 34, "bold"), text_color=ACCENT_GREEN
        )
        self.cr_total_label.pack(anchor="w", pady=(2, 0))

        self.cr_total_sub = ctk.CTkLabel(
            cr_total_inner, text="",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY
        )
        self.cr_total_sub.pack(anchor="w")

        # KPI Row 1
        cr_kpi1 = ctk.CTkFrame(self.cr_overview_frame, fg_color="transparent")
        cr_kpi1.pack(fill="x", pady=(0, 8))
        cr_kpi1.columnconfigure((0, 1, 2, 3), weight=1)

        self.cr_kpi_rf = self._make_kpi_card(cr_kpi1, "Renda Fixa", "...", ACCENT_GREEN, 0)
        self.cr_kpi_rv_corr = self._make_kpi_card(cr_kpi1, "RV Corretagem", "...", ACCENT_BLUE, 1)
        self.cr_kpi_fundos = self._make_kpi_card(cr_kpi1, "Fundos", "...", ACCENT_PURPLE, 2)
        self.cr_kpi_rv_pe = self._make_kpi_card(cr_kpi1, "RV Estruturadas", "...", ACCENT_ORANGE, 3)

        # KPI Row 2
        cr_kpi2 = ctk.CTkFrame(self.cr_overview_frame, fg_color="transparent")
        cr_kpi2.pack(fill="x", pady=(0, 14))
        cr_kpi2.columnconfigure((0, 1, 2), weight=1)

        self.cr_kpi_cambio = self._make_kpi_card(cr_kpi2, "Cambio", "...", ACCENT_TEAL, 0)
        self.cr_kpi_coe = self._make_kpi_card(cr_kpi2, "COE", "...", ACCENT_RED, 1)
        self.cr_kpi_n_assessores = self._make_kpi_card(cr_kpi2, "Assessores", "...", "#6b7280", 2)

        # Charts row (pie + bar)
        cr_charts = ctk.CTkFrame(self.cr_overview_frame, fg_color="transparent")
        cr_charts.pack(fill="x", pady=(0, 14))
        cr_charts.columnconfigure((0, 1), weight=1)

        pie_card, self.cr_ov_pie_fig, self.cr_ov_pie_ax, self.cr_ov_pie_canvas = \
            self._make_chart_card(cr_charts, "Distribuição por Produto", 4.5, 3)
        pie_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        bar_card, self.cr_ov_bar_fig, self.cr_ov_bar_ax, self.cr_ov_bar_canvas = \
            self._make_chart_card(cr_charts, "Top 10 Assessores", 4.5, 3)
        bar_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        # Distribuição por Produto (barras)
        ctk.CTkLabel(
            self.cr_overview_frame, text="Distribuição por Produto",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        self.cr_ov_bars_card = ctk.CTkFrame(
            self.cr_overview_frame, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER_CARD
        )
        self.cr_ov_bars_card.pack(fill="x", pady=(0, 14))
        self.cr_ov_bars_inner = ctk.CTkFrame(self.cr_ov_bars_card, fg_color="transparent")
        self.cr_ov_bars_inner.pack(fill="x", padx=20, pady=16)

        # Receita por Equipe
        ctk.CTkLabel(
            self.cr_overview_frame, text="Receita por Equipe",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        self.cr_ov_equipe_card = ctk.CTkFrame(
            self.cr_overview_frame, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER_CARD
        )
        self.cr_ov_equipe_card.pack(fill="x", pady=(0, 14))
        self.cr_ov_equipe_inner = ctk.CTkFrame(self.cr_ov_equipe_card, fg_color="transparent")
        self.cr_ov_equipe_inner.pack(fill="x", padx=20, pady=16)

        # === DEEPDIVE FRAME (hidden initially) ===
        self.cr_deepdive_frame = ctk.CTkFrame(content, fg_color="transparent")

        # Header card
        self.cr_dd_header = ctk.CTkFrame(
            self.cr_deepdive_frame, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER_CARD
        )
        self.cr_dd_header.pack(fill="x", pady=(0, 14))

        dd_hdr_inner = ctk.CTkFrame(self.cr_dd_header, fg_color="transparent")
        dd_hdr_inner.pack(fill="x", padx=24, pady=16)

        self.cr_dd_nome = ctk.CTkLabel(
            dd_hdr_inner, text="",
            font=("Segoe UI", 22, "bold"), text_color=TEXT_PRIMARY
        )
        self.cr_dd_nome.pack(anchor="w")

        self.cr_dd_info = ctk.CTkLabel(
            dd_hdr_inner, text="",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.cr_dd_info.pack(anchor="w", pady=(2, 0))

        self.cr_dd_total = ctk.CTkLabel(
            dd_hdr_inner, text="",
            font=("Segoe UI", 28, "bold"), text_color=ACCENT_GREEN
        )
        self.cr_dd_total.pack(anchor="w", pady=(8, 0))

        self.cr_dd_rank = ctk.CTkLabel(
            dd_hdr_inner, text="",
            font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.cr_dd_rank.pack(anchor="w")

        # KPI Row 1 (deepdive)
        dd_kpi1 = ctk.CTkFrame(self.cr_deepdive_frame, fg_color="transparent")
        dd_kpi1.pack(fill="x", pady=(0, 8))
        dd_kpi1.columnconfigure((0, 1, 2, 3), weight=1)

        self.cr_dd_kpi_rf = self._make_kpi_card(dd_kpi1, "Renda Fixa", "...", ACCENT_GREEN, 0)
        self.cr_dd_kpi_rv_corr = self._make_kpi_card(dd_kpi1, "RV Corretagem", "...", ACCENT_BLUE, 1)
        self.cr_dd_kpi_fundos = self._make_kpi_card(dd_kpi1, "Fundos", "...", ACCENT_PURPLE, 2)
        self.cr_dd_kpi_rv_pe = self._make_kpi_card(dd_kpi1, "RV Estruturadas", "...", ACCENT_ORANGE, 3)

        # KPI Row 2 (deepdive)
        dd_kpi2 = ctk.CTkFrame(self.cr_deepdive_frame, fg_color="transparent")
        dd_kpi2.pack(fill="x", pady=(0, 14))
        dd_kpi2.columnconfigure((0, 1, 2, 3), weight=1)

        self.cr_dd_kpi_cambio = self._make_kpi_card(dd_kpi2, "Cambio", "...", ACCENT_TEAL, 0)
        self.cr_dd_kpi_coe = self._make_kpi_card(dd_kpi2, "COE", "...", ACCENT_RED, 1)
        self.cr_dd_kpi_pct = self._make_kpi_card(dd_kpi2, "% do Total", "...", "#6b7280", 2)
        self.cr_dd_kpi_ranking = self._make_kpi_card(dd_kpi2, "Ranking", "...", "#6b7280", 3)

        # Charts (deepdive)
        dd_charts = ctk.CTkFrame(self.cr_deepdive_frame, fg_color="transparent")
        dd_charts.pack(fill="x", pady=(0, 14))
        dd_charts.columnconfigure((0, 1), weight=1)

        dd_pie_card, self.cr_dd_pie_fig, self.cr_dd_pie_ax, self.cr_dd_pie_canvas = \
            self._make_chart_card(dd_charts, "Mix Receita", 4.5, 3)
        dd_pie_card.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        dd_bar_card, self.cr_dd_bar_fig, self.cr_dd_bar_ax, self.cr_dd_bar_canvas = \
            self._make_chart_card(dd_charts, "Receita por Produto", 4.5, 3)
        dd_bar_card.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        # Detail table
        ctk.CTkLabel(
            self.cr_deepdive_frame, text="Operações Detalhadas",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        self.cr_dd_table_card = ctk.CTkFrame(
            self.cr_deepdive_frame, fg_color=BG_CARD,
            corner_radius=12, border_width=1, border_color=BORDER_CARD
        )
        self.cr_dd_table_card.pack(fill="x", pady=(0, 14))

        self.cr_dd_table_inner = ctk.CTkFrame(self.cr_dd_table_card, fg_color="transparent")
        self.cr_dd_table_inner.pack(fill="x", padx=16, pady=12)

        ctk.CTkLabel(
            self.cr_dd_table_inner, text="Selecione um assessor para ver operações",
            font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        ).pack(anchor="w")

        # Botao Atualizar
        self._make_atualizar_btn(content, RECEITA_FILE)

        # Data cache
        self._cr_data = None
        self._cr_detail_cache = {}
        self._cr_loading = False

        return page

    # -- Chart helpers --
    def _make_chart_card(self, parent, title, fig_width=4, fig_height=3):
        card = ctk.CTkFrame(
            parent, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        ctk.CTkLabel(
            card, text=title, font=("Segoe UI", 12, "bold"),
            text_color=TEXT_PRIMARY
        ).pack(anchor="w", padx=16, pady=(12, 4))

        fig = Figure(figsize=(fig_width, fig_height), dpi=90, facecolor='white')
        ax = fig.add_subplot(111)
        canvas = FigureCanvasTkAgg(fig, master=card)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=8, pady=(0, 8))

        return card, fig, ax, canvas

    def _update_pie_chart(self, fig, ax, canvas, labels, values, colors):
        ax.clear()
        if not values or sum(values) == 0:
            ax.text(0.5, 0.5, 'Sem dados', ha='center', va='center',
                    fontsize=12, color='#9ca3af')
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            canvas.draw()
            return
        wedges, texts, autotexts = ax.pie(
            values, labels=None, colors=colors,
            autopct='%1.1f%%', startangle=90,
            pctdistance=0.78, wedgeprops=dict(width=0.42)
        )
        for t in autotexts:
            t.set_fontsize(7)
            t.set_color('#333333')
        ax.legend(labels, loc='center left', bbox_to_anchor=(-0.15, 0.5),
                  fontsize=7, frameon=False)
        fig.tight_layout()
        canvas.draw()

    def _update_bar_chart(self, fig, ax, canvas, names, values, colors):
        ax.clear()
        if not values:
            ax.text(0.5, 0.5, 'Sem dados', ha='center', va='center',
                    fontsize=12, color='#9ca3af')
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            canvas.draw()
            return
        y_pos = range(len(names))
        bars = ax.barh(y_pos, values, color=colors, height=0.6)
        ax.set_yticks(list(y_pos))
        ax.set_yticklabels(names, fontsize=7)
        ax.invert_yaxis()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.tick_params(axis='x', labelsize=7)
        max_v = max(values) if values else 1
        for bar, val in zip(bars, values):
            ax.text(bar.get_width() + max_v * 0.02,
                    bar.get_y() + bar.get_height() / 2,
                    fmt_currency(val), va='center', fontsize=6)
        fig.tight_layout()
        canvas.draw()

    # -- Ctrl Receita data loading --
    def _load_ctrl_receita_data(self):
        try:
            if not os.path.exists(RECEITA_FILE):
                self.after(0, lambda: self.cr_banner_text.configure(
                    text="  Arquivo CONTROLE RECEITA TOTAL.xlsx não encontrado"))
                return

            wb = openpyxl.load_workbook(RECEITA_FILE, data_only=True)
            ws = wb["CONSOLIDADO"]

            data_base = ws.cell(3, 2).value
            if data_base and hasattr(data_base, "strftime"):
                data_str = data_base.strftime("%d/%m/%Y")
            else:
                data_str = str(data_base) if data_base else "N/A"

            assessores = []
            for r in range(7, ws.max_row + 1):
                cod = ws.cell(r, 2).value
                if not cod:
                    continue
                nome = sanitize_text(ws.cell(r, 4).value)
                equipe = sanitize_text(ws.cell(r, 3).value)
                rf = float(ws.cell(r, 5).value or 0)
                rv_corr = float(ws.cell(r, 6).value or 0)
                rv_pe = float(ws.cell(r, 7).value or 0)
                cambio = float(ws.cell(r, 8).value or 0)
                fundos = float(ws.cell(r, 9).value or 0)
                coe = float(ws.cell(r, 10).value or 0)
                receita = float(ws.cell(r, 11).value or 0)
                assessores.append({
                    "cod": str(cod), "nome": nome, "equipe": equipe,
                    "rf": rf, "rv_corr": rv_corr, "rv_pe": rv_pe,
                    "cambio": cambio, "fundos": fundos, "coe": coe,
                    "receita": receita,
                })
            wb.close()

            total_geral = sum(a["receita"] for a in assessores)
            ranking = sorted(assessores, key=lambda x: x["receita"], reverse=True)

            self._cr_data = {
                "assessores": assessores,
                "data_str": data_str,
                "total_geral": total_geral,
                "ranking": ranking,
            }

            names = ["Todos"] + [f'{a["cod"]} - {a["nome"]}' for a in ranking]

            def update_ui():
                self.cr_assessor_menu.configure(values=names)
                self.cr_banner_text.configure(
                    text=f"  Receita consolidada - {len(assessores)} assessores")
                self.cr_banner_date.configure(text=f"Base: {data_str}")
                self._update_ctrl_overview()

            self.after(0, update_ui)

        except Exception as e:
            import traceback; traceback.print_exc()
            err_msg = f"  Erro: {str(e)[:60]}"
            self.after(0, lambda m=err_msg: self.cr_banner_text.configure(text=m))

    def _on_ctrl_assessor_change(self, value):
        if value == "Todos":
            self.cr_deepdive_frame.pack_forget()
            self.cr_overview_frame.pack(fill="x")
            self._update_ctrl_overview()
        else:
            self.cr_overview_frame.pack_forget()
            self.cr_deepdive_frame.pack(fill="x")
            cod = value.split(" - ")[0].strip()
            self._update_ctrl_deepdive(cod)

    def _update_ctrl_overview(self):
        if not self._cr_data:
            return

        data = self._cr_data
        assessores = data["assessores"]
        total_geral = data["total_geral"]
        ranking = data["ranking"]

        total_rf = sum(a["rf"] for a in assessores)
        total_rv_corr = sum(a["rv_corr"] for a in assessores)
        total_rv_pe = sum(a["rv_pe"] for a in assessores)
        total_cambio = sum(a["cambio"] for a in assessores)
        total_fundos = sum(a["fundos"] for a in assessores)
        total_coe = sum(a["coe"] for a in assessores)
        n_assessores = len([a for a in assessores if a["receita"] > 0])

        self.cr_total_label.configure(text=fmt_currency(total_geral))
        self.cr_total_sub.configure(text=f"{n_assessores} assessores com receita ativa")

        self.cr_kpi_rf.configure(text=fmt_currency(total_rf))
        self.cr_kpi_rv_corr.configure(text=fmt_currency(total_rv_corr))
        self.cr_kpi_fundos.configure(text=fmt_currency(total_fundos))
        self.cr_kpi_rv_pe.configure(text=fmt_currency(total_rv_pe))
        self.cr_kpi_cambio.configure(text=fmt_currency(total_cambio))
        self.cr_kpi_coe.configure(text=fmt_currency(total_coe))
        self.cr_kpi_n_assessores.configure(text=str(n_assessores))

        # Pie chart
        cat_labels = ["RF", "RV Corr", "RV Estrut", "Fundos", "Cambio", "COE"]
        cat_values = [total_rf, total_rv_corr, total_rv_pe, total_fundos,
                      total_cambio, total_coe]
        cat_colors = [ACCENT_GREEN, ACCENT_BLUE, ACCENT_ORANGE, ACCENT_PURPLE,
                      ACCENT_TEAL, ACCENT_RED]

        pie_data = [(l, v, c) for l, v, c in zip(cat_labels, cat_values, cat_colors) if v > 0]
        if pie_data:
            self._update_pie_chart(
                self.cr_ov_pie_fig, self.cr_ov_pie_ax, self.cr_ov_pie_canvas,
                [d[0] for d in pie_data], [d[1] for d in pie_data],
                [d[2] for d in pie_data]
            )

        # Bar chart - top 10
        top10 = [a for a in ranking[:10] if a["receita"] > 0]
        top10_names = [a["nome"][:20] for a in top10]
        top10_values = [a["receita"] for a in top10]
        top10_colors = [ACCENT_GREEN] * len(top10_names)

        self._update_bar_chart(
            self.cr_ov_bar_fig, self.cr_ov_bar_ax, self.cr_ov_bar_canvas,
            top10_names, top10_values, top10_colors
        )

        # Barras distribuição produto
        categorias = [
            ("Renda Fixa", total_rf, ACCENT_GREEN),
            ("RV Corretagem", total_rv_corr, ACCENT_BLUE),
            ("Fundos", total_fundos, ACCENT_PURPLE),
            ("RV Estruturadas", total_rv_pe, ACCENT_ORANGE),
            ("Cambio", total_cambio, ACCENT_TEAL),
            ("COE", total_coe, ACCENT_RED),
        ]

        for w in self.cr_ov_bars_inner.winfo_children():
            w.destroy()

        max_val = max(v for _, v, _ in categorias) if categorias else 1
        for cat_name, cat_val, cat_color in categorias:
            row_frame = ctk.CTkFrame(self.cr_ov_bars_inner, fg_color="transparent")
            row_frame.pack(fill="x", pady=3)

            pct = (cat_val / total_geral * 100) if total_geral > 0 else 0

            ctk.CTkLabel(
                row_frame, text=cat_name, font=("Segoe UI", 11),
                text_color=TEXT_PRIMARY, width=120, anchor="w"
            ).pack(side="left")

            bar_track = ctk.CTkFrame(
                row_frame, fg_color=BG_PROGRESS_TRACK, height=18, corner_radius=4
            )
            bar_track.pack(side="left", fill="x", expand=True, padx=(8, 8))
            bar_track.pack_propagate(False)

            bar_w = max(cat_val / max_val, 0.02) if max_val > 0 else 0.02
            bar_fill = ctk.CTkFrame(bar_track, fg_color=cat_color, corner_radius=4)
            bar_fill.place(relx=0, rely=0, relwidth=bar_w, relheight=1.0)

            ctk.CTkLabel(
                row_frame, text=f"{fmt_currency(cat_val)}  ({pct:.1f}%)",
                font=("Segoe UI", 10, "bold"), text_color=cat_color,
                width=180, anchor="e"
            ).pack(side="right")

        # Equipes
        equipes = defaultdict(float)
        for a in assessores:
            if a["receita"] > 0:
                equipes[a["equipe"]] += a["receita"]

        for w in self.cr_ov_equipe_inner.winfo_children():
            w.destroy()

        equipe_sorted = sorted(equipes.items(), key=lambda x: x[1], reverse=True)
        equipe_colors = [ACCENT_GREEN, ACCENT_BLUE, ACCENT_ORANGE, ACCENT_PURPLE,
                         ACCENT_TEAL, ACCENT_RED, "#6b7280"]
        max_equipe = equipe_sorted[0][1] if equipe_sorted else 1

        for ei, (eq_name, eq_val) in enumerate(equipe_sorted):
            row_frame = ctk.CTkFrame(self.cr_ov_equipe_inner, fg_color="transparent")
            row_frame.pack(fill="x", pady=4)

            eq_color = equipe_colors[ei % len(equipe_colors)]
            pct = (eq_val / total_geral * 100) if total_geral > 0 else 0

            ctk.CTkLabel(
                row_frame, text=eq_name, font=("Segoe UI", 11, "bold"),
                text_color=TEXT_PRIMARY, width=120, anchor="w"
            ).pack(side="left")

            bar_track = ctk.CTkFrame(
                row_frame, fg_color=BG_PROGRESS_TRACK, height=22, corner_radius=5
            )
            bar_track.pack(side="left", fill="x", expand=True, padx=(8, 8))
            bar_track.pack_propagate(False)

            bar_w = max(eq_val / max_equipe, 0.02) if max_equipe > 0 else 0.02
            bar_fill = ctk.CTkFrame(bar_track, fg_color=eq_color, corner_radius=5)
            bar_fill.place(relx=0, rely=0, relwidth=bar_w, relheight=1.0)

            ctk.CTkLabel(
                row_frame, text=f"{fmt_currency(eq_val)}  ({pct:.1f}%)",
                font=("Segoe UI", 10, "bold"), text_color=eq_color,
                width=180, anchor="e"
            ).pack(side="right")

    def _update_ctrl_deepdive(self, cod):
        if not self._cr_data:
            return

        data = self._cr_data
        ranking = data["ranking"]
        total_geral = data["total_geral"]

        assessor = None
        rank_pos = 0
        for i, a in enumerate(ranking):
            if a["cod"] == cod:
                assessor = a
                rank_pos = i + 1
                break
        if not assessor:
            return

        # Header
        self.cr_dd_nome.configure(text=assessor["nome"])
        self.cr_dd_info.configure(
            text=f'Equipe: {assessor["equipe"]}  |  Código: {assessor["cod"]}'
        )
        self.cr_dd_total.configure(text=fmt_currency(assessor["receita"]))
        self.cr_dd_rank.configure(text=f'Ranking #{rank_pos} de {len(ranking)}')

        # KPIs
        self.cr_dd_kpi_rf.configure(text=fmt_currency(assessor["rf"]))
        self.cr_dd_kpi_rv_corr.configure(text=fmt_currency(assessor["rv_corr"]))
        self.cr_dd_kpi_fundos.configure(text=fmt_currency(assessor["fundos"]))
        self.cr_dd_kpi_rv_pe.configure(text=fmt_currency(assessor["rv_pe"]))
        self.cr_dd_kpi_cambio.configure(text=fmt_currency(assessor["cambio"]))
        self.cr_dd_kpi_coe.configure(text=fmt_currency(assessor["coe"]))

        pct = (assessor["receita"] / total_geral * 100) if total_geral > 0 else 0
        self.cr_dd_kpi_pct.configure(text=f"{pct:.1f}%")
        self.cr_dd_kpi_ranking.configure(text=f"#{rank_pos}")

        # Pie chart
        labels = ["RF", "RV Corr", "RV Estrut", "Fundos", "Cambio", "COE"]
        values = [assessor["rf"], assessor["rv_corr"], assessor["rv_pe"],
                  assessor["fundos"], assessor["cambio"], assessor["coe"]]
        colors = [ACCENT_GREEN, ACCENT_BLUE, ACCENT_ORANGE,
                  ACCENT_PURPLE, ACCENT_TEAL, ACCENT_RED]

        pie_data = [(l, v, c) for l, v, c in zip(labels, values, colors) if v > 0]
        if pie_data:
            self._update_pie_chart(
                self.cr_dd_pie_fig, self.cr_dd_pie_ax, self.cr_dd_pie_canvas,
                [d[0] for d in pie_data], [d[1] for d in pie_data],
                [d[2] for d in pie_data]
            )
        else:
            self.cr_dd_pie_ax.clear()
            self.cr_dd_pie_ax.text(
                0.5, 0.5, 'Sem dados', ha='center', va='center',
                fontsize=12, color='#9ca3af'
            )
            self.cr_dd_pie_ax.axis('off')
            self.cr_dd_pie_canvas.draw()

        # Bar chart
        bar_data = [(l, v, c) for l, v, c in zip(labels, values, colors) if v > 0]
        if bar_data:
            self._update_bar_chart(
                self.cr_dd_bar_fig, self.cr_dd_bar_ax, self.cr_dd_bar_canvas,
                [d[0] for d in bar_data], [d[1] for d in bar_data],
                [d[2] for d in bar_data]
            )
        else:
            self.cr_dd_bar_ax.clear()
            self.cr_dd_bar_ax.text(
                0.5, 0.5, 'Sem dados', ha='center', va='center',
                fontsize=12, color='#9ca3af'
            )
            self.cr_dd_bar_ax.axis('off')
            self.cr_dd_bar_canvas.draw()

        # Load detail table async
        threading.Thread(
            target=self._load_ctrl_assessor_details, args=(cod,), daemon=True
        ).start()

    def _load_ctrl_assessor_details(self, cod):
        if cod in self._cr_detail_cache:
            self.after(0, lambda: self._update_ctrl_detail_table(cod))
            return

        try:
            self.after(0, lambda: self._set_ctrl_detail_loading("Carregando operações..."))

            wb = openpyxl.load_workbook(RECEITA_FILE, data_only=True)
            details = []

            for sheet_info in CTRL_DETAIL_SHEETS:
                sheet_name = sheet_info["sheet"]
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]
                col_a = sheet_info["col_assessor"]
                col_r = sheet_info["col_receita"]

                for r in range(2, ws.max_row + 1):
                    cell_val = ws.cell(r, col_a).value
                    if not cell_val:
                        continue
                    cell_str = str(cell_val).strip()
                    if cell_str != cod and cod not in cell_str:
                        continue

                    # Grab first columns for description
                    desc_parts = []
                    for c in range(1, min(ws.max_column + 1, 8)):
                        v = ws.cell(r, c).value
                        if v is not None and c != col_a:
                            desc_parts.append(sanitize_text(str(v))[:30])

                    receita_val = None
                    if col_r:
                        try:
                            receita_val = float(ws.cell(r, col_r).value or 0)
                        except (ValueError, TypeError):
                            receita_val = 0

                    details.append({
                        "produto": sheet_info["produto"],
                        "desc": " | ".join(desc_parts[:3]),
                        "receita": receita_val,
                    })

            wb.close()

            self._cr_detail_cache[cod] = details
            self.after(0, lambda: self._update_ctrl_detail_table(cod))

        except Exception as e:
            self.after(0, lambda: self._set_ctrl_detail_loading(
                f"Erro: {str(e)[:50]}"))

    def _set_ctrl_detail_loading(self, text):
        for w in self.cr_dd_table_inner.winfo_children():
            w.destroy()
        ctk.CTkLabel(
            self.cr_dd_table_inner, text=text,
            font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        ).pack(anchor="w")

    def _update_ctrl_detail_table(self, cod):
        details = self._cr_detail_cache.get(cod, [])

        for w in self.cr_dd_table_inner.winfo_children():
            w.destroy()

        if not details:
            ctk.CTkLabel(
                self.cr_dd_table_inner, text="Nenhuma operação encontrada",
                font=("Segoe UI", 11), text_color=TEXT_TERTIARY
            ).pack(anchor="w")
            return

        # Header
        hdr = ctk.CTkFrame(self.cr_dd_table_inner, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 6))
        ctk.CTkLabel(
            hdr, text="Produto", font=("Segoe UI", 10, "bold"),
            text_color=TEXT_TERTIARY, width=120, anchor="w"
        ).pack(side="left")
        ctk.CTkLabel(
            hdr, text="Descrição", font=("Segoe UI", 10, "bold"),
            text_color=TEXT_TERTIARY, anchor="w"
        ).pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(
            hdr, text="Receita", font=("Segoe UI", 10, "bold"),
            text_color=TEXT_TERTIARY, width=120, anchor="e"
        ).pack(side="right")

        ctk.CTkFrame(
            self.cr_dd_table_inner, fg_color=BORDER_LIGHT, height=1
        ).pack(fill="x", pady=(0, 4))

        prod_colors = {
            "Sec RF": ACCENT_GREEN, "Prim RF": ACCENT_GREEN,
            "RV Corretagem": ACCENT_BLUE, "RV Estruturadas": ACCENT_ORANGE,
            "Cambio": ACCENT_TEAL, "Fundos Fechados": ACCENT_PURPLE,
            "Sec Fundos": ACCENT_PURPLE, "COE": ACCENT_RED,
        }

        for d in details[:50]:
            row = ctk.CTkFrame(self.cr_dd_table_inner, fg_color="transparent")
            row.pack(fill="x", pady=1)

            color = prod_colors.get(d["produto"], TEXT_PRIMARY)
            ctk.CTkLabel(
                row, text=d["produto"], font=("Segoe UI", 10),
                text_color=color, width=120, anchor="w"
            ).pack(side="left")
            ctk.CTkLabel(
                row, text=d["desc"][:60], font=("Segoe UI", 9),
                text_color=TEXT_SECONDARY, anchor="w"
            ).pack(side="left", fill="x", expand=True)

            rec_text = fmt_currency(d["receita"]) if d["receita"] is not None else "-"
            ctk.CTkLabel(
                row, text=rec_text, font=("Segoe UI", 10, "bold"),
                text_color=TEXT_PRIMARY, width=120, anchor="e"
            ).pack(side="right")

        if len(details) > 50:
            ctk.CTkLabel(
                self.cr_dd_table_inner,
                text=f"... e mais {len(details) - 50} operações",
                font=("Segoe UI", 10), text_color=TEXT_TERTIARY
            ).pack(anchor="w", pady=(4, 0))

    # -----------------------------------------------------------------
    #  PAGE: FLUXO - RF
    # -----------------------------------------------------------------
    def _build_dashboard_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        # Top bar profissional
        _, _, self.topbar_date = self._make_topbar(
            page, "FLUXO - RF", subtitle="Mesa de Produtos"
        )

        # Scroll area
        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Periodo banner
        self.período_banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=44)
        self.período_banner.pack(fill="x", pady=(0, 18))
        self.período_banner.pack_propagate(False)

        self.período_text = ctk.CTkLabel(
            self.período_banner,
            text="  Carregando dados...",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        )
        self.período_text.pack(side="left", padx=18, pady=10)

        self.período_pdfs = ctk.CTkLabel(
            self.período_banner, text="",
            font=("Segoe UI", 11), text_color="#80c0a0"
        )
        self.período_pdfs.pack(side="right", padx=18)

        # KPI Cards row
        kpi_frame = ctk.CTkFrame(content, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=(0, 18))
        kpi_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self.kpi_assessores = self._make_kpi_card(kpi_frame, "Assessores", "...", ACCENT_GREEN, 0)
        self.kpi_eventos = self._make_kpi_card(kpi_frame, "Eventos", "...", ACCENT_BLUE, 1)
        self.kpi_ativos = self._make_kpi_card(kpi_frame, "Ativos Distintos", "...", ACCENT_ORANGE, 2)
        self.kpi_datas = self._make_kpi_card(kpi_frame, "Datas com Eventos", "...", ACCENT_PURPLE, 3)

        # Total destaque
        total_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        total_card.pack(fill="x", pady=(0, 18))

        total_inner = ctk.CTkFrame(total_card, fg_color="transparent")
        total_inner.pack(fill="x", padx=24, pady=16)

        ctk.CTkLabel(
            total_inner, text="VALOR TOTAL ESTIMADO",
            font=("Segoe UI", 10, "bold"), text_color=TEXT_TERTIARY
        ).pack(anchor="w")

        self.total_value_label = ctk.CTkLabel(
            total_inner, text="Carregando...",
            font=("Segoe UI", 32, "bold"), text_color=ACCENT_GREEN
        )
        self.total_value_label.pack(anchor="w", pady=(2, 0))

        self.total_sub_label = ctk.CTkLabel(
            total_inner, text="",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY
        )
        self.total_sub_label.pack(anchor="w")

        # Breakdown por tipo
        tipo_label = ctk.CTkLabel(
            content, text="Distribuição por Tipo de Evento",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        tipo_label.pack(fill="x", pady=(0, 10))

        self.tipo_frame = ctk.CTkFrame(content, fg_color="transparent")
        self.tipo_frame.pack(fill="x", pady=(0, 18))
        self.tipo_frame.columnconfigure((0, 1), weight=1)

        # Placeholder - será preenchido quando os dados carregarem
        self.tipo_cards = {}

        # Ações rapidas
        act_label = ctk.CTkLabel(
            content, text="Ações Rápidas",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        act_label.pack(fill="x", pady=(0, 10))

        act_frame = ctk.CTkFrame(content, fg_color="transparent")
        act_frame.pack(fill="x", pady=(0, 10))
        act_frame.columnconfigure((0, 1, 2), weight=1)

        self._make_action_card(
            act_frame, "\u2913", "Gerar PDFs",
            "Gera um relatório PDF para cada assessor com todos os fluxos de renda fixa detalhados por data.",
            ACCENT_GREEN, 0, self._go_gerar_pdfs
        )
        self._make_action_card(
            act_frame, "\u2709", "Gerar Emails",
            "Cria rascunhos no Outlook com o PDF anexado, prontos para revisão e envio.",
            ACCENT_BLUE, 1, self._go_gerar_emails
        )
        self._make_action_card(
            act_frame, "\u2750", "Abrir Pasta",
            "Abre a pasta de PDFs gerados no Windows Explorer.",
            "#6b7280", 2, self._on_abrir_pasta
        )

        # Botao Atualizar
        self._make_atualizar_btn(content, ENTRADA_FILE)

        return page

    def _make_kpi_card(self, parent, label, value, color, col):
        card = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        card.grid(row=0, column=col, padx=(0 if col == 0 else 6, 0 if col == 3 else 6), sticky="nsew")

        # Indicador de cor
        bar = ctk.CTkFrame(card, fg_color=color, height=4, corner_radius=2)
        bar.pack(fill="x", padx=14, pady=(14, 0))

        ctk.CTkLabel(
            card, text=label, font=("Segoe UI", 10),
            text_color=TEXT_SECONDARY
        ).pack(anchor="w", padx=16, pady=(10, 0))

        val_label = ctk.CTkLabel(
            card, text=value, font=("Segoe UI", 26, "bold"),
            text_color=TEXT_PRIMARY
        )
        val_label.pack(anchor="w", padx=16, pady=(0, 14))

        return val_label

    def _make_action_card(self, parent, icon, title, desc, color, col, command):
        card = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        card.grid(row=0, column=col, padx=(0 if col == 0 else 5, 0 if col == 2 else 5), sticky="nsew")

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=18, pady=16)

        header = ctk.CTkFrame(inner, fg_color="transparent")
        header.pack(fill="x")

        ctk.CTkLabel(
            header, text=icon, font=("Segoe UI", 22),
            text_color=color
        ).pack(side="left")

        ctk.CTkLabel(
            header, text=title, font=("Segoe UI", 14, "bold"),
            text_color=TEXT_PRIMARY
        ).pack(side="left", padx=(10, 0))

        ctk.CTkLabel(
            inner, text=desc, font=("Segoe UI", 10),
            text_color=TEXT_SECONDARY, wraplength=250, justify="left", anchor="w"
        ).pack(fill="x", pady=(8, 12))

        ctk.CTkButton(
            inner, text=f"Executar", font=("Segoe UI", 11, "bold"),
            fg_color=color, hover_color=self._darken(color),
            height=36, corner_radius=8, command=command
        ).pack(anchor="w")

    def _make_tipo_card(self, parent, tipo_name, count, total, color, row, col):
        card = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=10, border_width=1, border_color=BORDER_CARD)
        card.grid(row=row, column=col, padx=(0 if col == 0 else 4, 0 if col == 1 else 4), pady=4, sticky="nsew")

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=16, pady=12)

        # Dot + name
        header = ctk.CTkFrame(inner, fg_color="transparent")
        header.pack(fill="x")

        ctk.CTkLabel(header, text="\u25cf", font=("Segoe UI", 10), text_color=color).pack(side="left")
        ctk.CTkLabel(
            header, text=f"  {tipo_name}", font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY
        ).pack(side="left")
        ctk.CTkLabel(
            header, text=f"{count} eventos", font=("Segoe UI", 10), text_color=TEXT_TERTIARY
        ).pack(side="right")

        # Value
        ctk.CTkLabel(
            inner, text=fmt_currency(total),
            font=("Segoe UI", 16, "bold"), text_color=color, anchor="w"
        ).pack(fill="x", pady=(4, 0))

    def _make_atualizar_btn(self, parent, file_path):
        """Botao verde 'Atualizar' que abre a pasta da planilha fonte."""
        folder = os.path.dirname(os.path.abspath(file_path))
        btn_frame = ctk.CTkFrame(parent, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(24, 8))
        ctk.CTkButton(
            btn_frame, text="\u21BB  Atualizar",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_WHITE, height=42, corner_radius=10,
            command=lambda: os.startfile(folder)
        ).pack(fill="x")

    # -----------------------------------------------------------------
    #  PAGE: OPERATIONS
    # -----------------------------------------------------------------
    def _build_operations_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        # Top bar profissional
        _, self.ops_title, _ = self._make_topbar(
            page, "Operações", subtitle="Geração de PDFs e Emails",
            back_btn=True, back_cmd=lambda: self._show_page("fluxo_rf")
        )

        # Content
        content = ctk.CTkFrame(page, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=28, pady=20)

        # Status card
        status_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12, border_width=1, border_color=BORDER_CARD)
        status_card.pack(fill="x", pady=(0, 14))

        status_inner = ctk.CTkFrame(status_card, fg_color="transparent")
        status_inner.pack(fill="x", padx=24, pady=18)

        self.ops_status_dot = ctk.CTkLabel(
            status_inner, text="\u25cf", font=("Segoe UI", 14), text_color=TEXT_TERTIARY
        )
        self.ops_status_dot.pack(side="left")

        self.ops_status_text = ctk.CTkLabel(
            status_inner, text="  Aguardando...",
            font=("Segoe UI", 13), text_color=TEXT_PRIMARY
        )
        self.ops_status_text.pack(side="left")

        self.ops_counter = ctk.CTkLabel(
            status_inner, text="",
            font=("Segoe UI", 12, "bold"), text_color=ACCENT_GREEN
        )
        self.ops_counter.pack(side="right")

        # Progress
        self.ops_progress = ctk.CTkProgressBar(
            content, fg_color=BG_PROGRESS_TRACK, progress_color=ACCENT_GREEN,
            height=8, corner_radius=4
        )
        self.ops_progress.pack(fill="x", pady=(0, 4))
        self.ops_progress.set(0)

        self.ops_progress_text = ctk.CTkLabel(
            content, text="", font=("Segoe UI", 9), text_color=TEXT_TERTIARY, anchor="w"
        )
        self.ops_progress_text.pack(fill="x", pady=(0, 12))

        # Log
        log_card = ctk.CTkFrame(content, fg_color=BG_LOG, corner_radius=12, border_width=1, border_color="#2a3040")
        log_card.pack(fill="both", expand=True)

        log_header = ctk.CTkFrame(log_card, fg_color=BG_LOG_HEADER, corner_radius=0)
        log_header.pack(fill="x")

        ctk.CTkLabel(
            log_header, text="  Terminal",
            font=("Consolas", 10, "bold"), text_color="#5a6a7a"
        ).pack(side="left", padx=12, pady=6)

        ctk.CTkButton(
            log_header, text="Limpar", font=("Segoe UI", 9),
            fg_color="transparent", hover_color="#2a3040",
            text_color="#5a6a7a", width=50, height=22,
            corner_radius=4, command=self._clear_ops_log
        ).pack(side="right", padx=8, pady=4)

        self.ops_log = ctk.CTkTextbox(
            log_card, font=("Consolas", 10), fg_color=BG_LOG,
            text_color=TEXT_LOG, corner_radius=0, wrap="word",
            border_width=0
        )
        self.ops_log.pack(fill="both", expand=True, padx=2, pady=(0, 2))
        self.ops_log.configure(state="disabled")

        return page

    # -----------------------------------------------------------------
    #  PAGE: INFORMATIVO - E-MAIL
    # -----------------------------------------------------------------
    def _build_informativo_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Informativo", subtitle="E-mail")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Banner
        banner = ctk.CTkFrame(content, fg_color=ACCENT_BLUE, corner_radius=10, height=44)
        banner.pack(fill="x", pady=(0, 18))
        banner.pack_propagate(False)

        ctk.CTkLabel(
            banner, text="  Compose e envie informativos por e-mail aos assessores",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        ).pack(side="left", padx=18, pady=10)

        # === ASSUNTO ===
        ctk.CTkLabel(
            content, text="Assunto",
            font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.info_assunto = ctk.CTkEntry(
            content, font=("Segoe UI", 12), height=40,
            corner_radius=8, border_color=BORDER_CARD,
            placeholder_text="Ex: Informativo Semanal - Mesa de Produtos"
        )
        self.info_assunto.pack(fill="x", pady=(0, 14))

        # === CORPO DO EMAIL ===
        ctk.CTkLabel(
            content, text="Corpo do E-mail",
            font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.info_corpo = ctk.CTkTextbox(
            content, font=("Segoe UI", 11), height=220,
            corner_radius=8, border_width=1, border_color=BORDER_CARD,
            fg_color=BG_CARD, text_color=TEXT_PRIMARY, wrap="word"
        )
        self.info_corpo.pack(fill="x", pady=(0, 14))

        # === CC - CÓPIA ===
        ctk.CTkLabel(
            content, text="CC - Cópia",
            font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        cc_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=10,
                               border_width=1, border_color=BORDER_CARD)
        cc_card.pack(fill="x", pady=(0, 14))

        cc_inner = ctk.CTkFrame(cc_card, fg_color="transparent")
        cc_inner.pack(fill="x", padx=16, pady=12)

        # Backoffice - assistente da base por assessor (checkbox pré-marcado)
        self.info_cc_backoffice_var = ctk.IntVar(value=1)
        ctk.CTkCheckBox(
            cc_inner, text="Backoffice / Assistente (da planilha BASE EMAILS)",
            font=("Segoe UI", 11), text_color=TEXT_PRIMARY,
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            variable=self.info_cc_backoffice_var,
        ).pack(anchor="w", pady=(0, 8))

        # Separador
        ctk.CTkFrame(cc_inner, fg_color=BORDER_CARD, height=1).pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(
            cc_inner, text="Pessoas em cópia:",
            font=("Segoe UI", 10, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(anchor="w", pady=(0, 6))

        # Lista de pessoas selecionáveis
        self._info_cc_pessoas = [
            ("Leonardo Dellatorre", "Leonardo.dellatorre@somuscapital.com.br"),
            ("João Morais", "joao.morais@somuscapital.com.br"),
            ("Alexandre Achui", "alexandre.achui@somuscapital.com.br"),
            ("Alessandra Lima", "alessandra.lima@somuscapital.com.br"),
        ]
        self._info_cc_vars = []

        cc_grid = ctk.CTkFrame(cc_inner, fg_color="transparent")
        cc_grid.pack(fill="x", pady=(0, 8))

        for i, (nome, email_addr) in enumerate(self._info_cc_pessoas):
            var = ctk.IntVar(value=0)
            self._info_cc_vars.append(var)
            cb = ctk.CTkCheckBox(
                cc_grid, text=f"{nome}",
                font=("Segoe UI", 11), text_color=TEXT_PRIMARY,
                fg_color=ACCENT_BLUE, hover_color="#1555bb",
                variable=var,
            )
            row = i // 2
            col = i % 2
            cb.grid(row=row, column=col, sticky="w", padx=(0, 24), pady=2)

        cc_grid.columnconfigure(0, weight=1)
        cc_grid.columnconfigure(1, weight=1)

        # Separador
        ctk.CTkFrame(cc_inner, fg_color=BORDER_CARD, height=1).pack(fill="x", pady=(4, 8))

        # Campo manual para CC adicional
        ctk.CTkLabel(
            cc_inner, text="Outros (separar com ponto e vírgula):",
            font=("Segoe UI", 10), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(anchor="w", pady=(0, 4))

        self.info_cc_manual = ctk.CTkEntry(
            cc_inner, font=("Segoe UI", 11), height=36,
            corner_radius=8, border_color=BORDER_CARD,
            placeholder_text="Ex: nome@email.com; outro@email.com"
        )
        self.info_cc_manual.pack(fill="x")

        # === ANEXO ===
        ctk.CTkLabel(
            content, text="Anexo (opcional)",
            font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        anexo_frame = ctk.CTkFrame(content, fg_color="transparent")
        anexo_frame.pack(fill="x", pady=(0, 14))

        self.info_anexo_label = ctk.CTkLabel(
            anexo_frame, text="Nenhum arquivo selecionado",
            font=("Segoe UI", 11), text_color=TEXT_TERTIARY, anchor="w"
        )
        self.info_anexo_label.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(
            anexo_frame, text="Selecionar",
            font=("Segoe UI", 11, "bold"),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=34, corner_radius=8, width=110,
            command=self._on_info_browse_anexo,
        ).pack(side="right", padx=(8, 0))

        ctk.CTkButton(
            anexo_frame, text="Limpar",
            font=("Segoe UI", 10),
            fg_color="transparent", hover_color=BORDER_CARD,
            text_color=TEXT_TERTIARY, height=34, corner_radius=8, width=70,
            command=self._on_info_clear_anexo,
        ).pack(side="right")

        # === DESTINATARIOS ===
        ctk.CTkLabel(
            content, text="Destinatarios",
            font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        dest_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=10,
                                 border_width=1, border_color=BORDER_CARD)
        dest_card.pack(fill="x", pady=(0, 14))

        dest_inner = ctk.CTkFrame(dest_card, fg_color="transparent")
        dest_inner.pack(fill="x", padx=16, pady=12)

        # Seletor de Time/Equipe
        team_row = ctk.CTkFrame(dest_inner, fg_color="transparent")
        team_row.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(
            team_row, text="Enviar para:",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(side="left", padx=(0, 10))

        # Carregar equipes disponiveis da base
        self._info_equipe_map = load_equipe_mapping()
        equipes_disponiveis = sorted(set(self._info_equipe_map.values()))
        opcoes_time = ["Todos"] + equipes_disponiveis

        self.info_team_var = ctk.StringVar(value="Todos")
        self.info_team_selector = ctk.CTkSegmentedButton(
            team_row,
            values=opcoes_time,
            variable=self.info_team_var,
            font=("Segoe UI", 11),
            selected_color=ACCENT_GREEN,
            selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_WHITE,
            corner_radius=8,
            command=self._on_info_team_changed,
        )
        self.info_team_selector.pack(side="left", fill="x", expand=True)

        self.info_dest_count = ctk.CTkLabel(
            dest_inner, text="",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w"
        )
        self.info_dest_count.pack(anchor="w", pady=(4, 0))

        # Lista de assessores do time selecionado
        self.info_team_list = ctk.CTkLabel(
            dest_inner, text="",
            font=("Segoe UI", 9), text_color=TEXT_SECONDARY, anchor="w",
            justify="left", wraplength=600
        )
        self.info_team_list.pack(anchor="w", pady=(2, 0))

        # Carregar contagem inicial
        self._info_assessores_base = {}
        try:
            self._info_assessores_base = load_base_emails()
            self._on_info_team_changed("Todos")
        except Exception:
            self.info_dest_count.configure(text="Erro ao carregar base")

        # === AÇÕES ===
        act_frame = ctk.CTkFrame(content, fg_color="transparent")
        act_frame.pack(fill="x", pady=(6, 14))

        self.info_enviar_btn = ctk.CTkButton(
            act_frame, text="\u2709  Gerar Rascunhos no Outlook",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=44, corner_radius=10,
            command=self._on_info_gerar_emails,
        )
        self.info_enviar_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            act_frame, text="Limpar Tudo",
            font=("Segoe UI", 11),
            fg_color="#6b7280", hover_color="#555d6a",
            height=44, corner_radius=10, width=120,
            command=self._on_info_limpar,
        ).pack(side="left")

        # === STATUS ===
        self.info_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.info_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.info_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.info_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.info_status_dot.pack(side="left")

        self.info_status_text = ctk.CTkLabel(
            si, text="  Preencha os campos e clique em Gerar Rascunhos",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.info_status_text.pack(side="left")

        # Botao Atualizar
        self._make_atualizar_btn(content, BASE_FILE)

        # Internal state
        self._info_anexo_path = None

        return page

    # -----------------------------------------------------------------
    #  INFORMATIVO: Ações
    # -----------------------------------------------------------------
    def _on_info_team_changed(self, selected):
        """Atualiza contagem e lista de assessores quando o time muda."""
        assessores = self._info_assessores_base
        equipe_map = self._info_equipe_map
        filtered = self._get_info_filtered_assessores()
        n = len(filtered)
        if selected == "Todos":
            self.info_dest_count.configure(text=f"{n} assessores com e-mail cadastrado")
            self.info_team_list.configure(text="")
        else:
            nomes = [filtered[c]["nome"] for c in sorted(filtered.keys())]
            self.info_dest_count.configure(text=f"{n} assessores na equipe {selected}")
            self.info_team_list.configure(text=", ".join(nomes))

    def _get_info_filtered_assessores(self):
        """Retorna assessores filtrados pelo time selecionado."""
        assessores = self._info_assessores_base
        selected = self.info_team_var.get()
        filtered = {}
        for cod, info in assessores.items():
            email = info.get("email", "")
            if not email or email == "-":
                continue
            if selected == "Todos":
                filtered[cod] = info
            else:
                equipe = self._info_equipe_map.get(cod, "")
                if equipe == selected:
                    filtered[cod] = info
        return filtered

    def _on_info_browse_anexo(self):
        path = filedialog.askopenfilename(
            title="Selecionar anexo",
            filetypes=[("Todos", "*.*"), ("PDF", "*.pdf"), ("Excel", "*.xlsx"), ("Imagem", "*.png *.jpg")]
        )
        if path:
            self._info_anexo_path = path
            self.info_anexo_label.configure(
                text=os.path.basename(path), text_color=ACCENT_GREEN
            )

    def _on_info_clear_anexo(self):
        self._info_anexo_path = None
        self.info_anexo_label.configure(
            text="Nenhum arquivo selecionado", text_color=TEXT_TERTIARY
        )

    def _on_info_limpar(self):
        self.info_assunto.delete(0, "end")
        self.info_corpo.delete("1.0", "end")
        self.info_cc_manual.delete(0, "end")
        self.info_cc_backoffice_var.set(1)  # Backoffice sempre marcado por padrão
        for var in self._info_cc_vars:
            var.set(0)
        self._on_info_clear_anexo()
        self.info_team_var.set("Todos")
        self._on_info_team_changed("Todos")
        self.info_status_dot.configure(text_color=TEXT_TERTIARY)
        self.info_status_text.configure(text="  Preencha os campos e clique em Gerar Rascunhos")

    def _on_info_gerar_emails(self):
        assunto = self.info_assunto.get().strip()
        corpo = self.info_corpo.get("1.0", "end").strip()

        if not assunto:
            messagebox.showwarning("Campo obrigatorio", "Preencha o assunto do e-mail.")
            return
        if not corpo:
            messagebox.showwarning("Campo obrigatorio", "Preencha o corpo do e-mail.")
            return

        filtered = self._get_info_filtered_assessores()
        if not filtered:
            messagebox.showwarning("Sem destinatarios", "Nenhum assessor encontrado para o time selecionado.")
            return

        team_sel = self.info_team_var.get()
        team_label = f"equipe {team_sel}" if team_sel != "Todos" else "todos os assessores"

        resp = messagebox.askyesno(
            "Gerar Rascunhos",
            f"Sera criado um rascunho no Outlook para {len(filtered)} assessor(es) ({team_label}).\n\n"
            "Os e-mails NAO serao enviados, apenas salvos como rascunho.\n\n"
            "O Outlook precisa estar aberto. Deseja continuar?"
        )
        if not resp:
            return

        # Montar lista de CC extra (checkboxes + manual)
        include_backoffice = bool(self.info_cc_backoffice_var.get())
        cc_extras = []
        for i, var in enumerate(self._info_cc_vars):
            if var.get():
                cc_extras.append(self._info_cc_pessoas[i][1])
        cc_manual = self.info_cc_manual.get().strip()
        if cc_manual:
            cc_extras.append(cc_manual)
        cc_extra_str = "; ".join(cc_extras)

        self.info_enviar_btn.configure(state="disabled")
        self.info_status_dot.configure(text_color=ACCENT_ORANGE)
        self.info_status_text.configure(text=f"  Gerando rascunhos para {team_label}...")
        threading.Thread(target=self._run_info_emails, args=(assunto, corpo, filtered, cc_extra_str, include_backoffice), daemon=True).start()

    def _run_info_emails(self, assunto, corpo, filtered_assessores, cc_manual="", include_backoffice=True):
        try:
            import win32com.client as win32
            import pythoncom
            pythoncom.CoInitialize()

            # Recarregar base de emails para pegar alterações recentes
            try:
                self._info_assessores_base = load_base_emails()
            except Exception:
                pass

            outlook = win32.Dispatch("Outlook.Application")

            hoje = datetime.now().strftime("%d/%m/%Y")
            logo_tag = LOGO_TAG_CID if os.path.exists(LOGO_PATH) else ""

            # Converter quebras de linha em <br>
            corpo_html = corpo.replace("\n", "<br>")

            criados = 0
            erros = 0
            codes = sorted(filtered_assessores.keys())

            for código in codes:
                info = filtered_assessores[código]
                email = info.get("email", "")
                if not email or email == "-":
                    continue

                nome = info["nome"]
                primeiro_nome = nome.split()[0] if nome else "Assessor"

                html_body = f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">
    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#000000;margin-top:0;">
      {corpo_html}
    </p>
  </td></tr>
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">Informativo &middot; {hoje}</span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">E-mail gerado automaticamente</span>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</div>
"""
                try:
                    mail = outlook.CreateItem(0)
                    mail.To = email

                    # Montar CC: backoffice/assistente da base (se marcado) + CC extras
                    cc_parts = []
                    if include_backoffice:
                        email_assist = info.get("email_assistente", "-")
                        if email_assist and email_assist != "-":
                            cc_parts.append(email_assist)
                    if cc_manual:
                        cc_parts.append(cc_manual)
                    if cc_parts:
                        mail.CC = "; ".join(cc_parts)

                    mail.Subject = assunto
                    mail.HTMLBody = html_body
                    _attach_logo_cid(mail)
                    if self._info_anexo_path and os.path.exists(self._info_anexo_path):
                        mail.Attachments.Add(os.path.abspath(self._info_anexo_path))
                    mail.Save()
                    criados += 1
                except Exception:
                    erros += 1

            pythoncom.CoUninitialize()

            def _done():
                self.info_enviar_btn.configure(state="normal")
                self.info_status_dot.configure(text_color="#00a86b")
                self.info_status_text.configure(
                    text=f"  Concluído - {criados} rascunhos criados, {erros} erros"
                )
                messagebox.showinfo(
                    "Rascunhos Criados",
                    f"{criados} rascunhos criados no Outlook!\n{erros} erros."
                )
            self.after(0, _done)

        except Exception as e:
            def _err():
                self.info_enviar_btn.configure(state="normal")
                self.info_status_dot.configure(text_color=ACCENT_RED)
                self.info_status_text.configure(text=f"  Erro: {e}")
                messagebox.showerror("Erro", str(e))
            self.after(0, _err)

    # -----------------------------------------------------------------
    #  PAGE: INFORMATIVO - AGIO
    # -----------------------------------------------------------------
    def _build_info_agio_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Informativo - Agio", subtitle="Mesa de Produtos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Banner
        banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=44)
        banner.pack(fill="x", pady=(0, 18))
        banner.pack_propagate(False)

        ctk.CTkLabel(
            banner, text="  Informativo de Agio/Desagio - Dashboard e Envio",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        ).pack(side="left", padx=18, pady=10)

        ctk.CTkButton(
            banner, text="\u21BB  Recarregar",
            font=("Segoe UI", 10, "bold"),
            fg_color=BG_SIDEBAR_HOVER, hover_color=BG_SIDEBAR_ACTIVE,
            text_color=TEXT_WHITE, height=28, corner_radius=6, width=110,
            command=self._ag_load_data,
        ).pack(side="right", padx=18)

        # ======== KPI CARDS ========
        kpi_frame = ctk.CTkFrame(content, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=(0, 8))
        kpi_frame.columnconfigure((0, 1, 2, 3), weight=1)

        # KPI 1: Total R$ em Agio
        kpi1 = ctk.CTkFrame(kpi_frame, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color=BORDER_CARD)
        kpi1.grid(row=0, column=0, padx=(0, 6), sticky="nsew")
        ctk.CTkFrame(kpi1, fg_color=ACCENT_GREEN, height=4,
                     corner_radius=2).pack(fill="x", padx=14, pady=(14, 0))
        ctk.CTkLabel(kpi1, text="Total Agio (R$)", font=("Segoe UI", 10),
                     text_color=TEXT_SECONDARY).pack(anchor="w", padx=16, pady=(10, 0))
        self.ag_kpi_total = ctk.CTkLabel(
            kpi1, text="...", font=("Segoe UI", 22, "bold"), text_color=TEXT_PRIMARY)
        self.ag_kpi_total.pack(anchor="w", padx=16, pady=(0, 14))

        # KPI 2: Assessores
        kpi2 = ctk.CTkFrame(kpi_frame, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color=BORDER_CARD)
        kpi2.grid(row=0, column=1, padx=6, sticky="nsew")
        ctk.CTkFrame(kpi2, fg_color=ACCENT_BLUE, height=4,
                     corner_radius=2).pack(fill="x", padx=14, pady=(14, 0))
        ctk.CTkLabel(kpi2, text="Assessores", font=("Segoe UI", 10),
                     text_color=TEXT_SECONDARY).pack(anchor="w", padx=16, pady=(10, 0))
        self.ag_kpi_assessores = ctk.CTkLabel(
            kpi2, text="...", font=("Segoe UI", 22, "bold"), text_color=TEXT_PRIMARY)
        self.ag_kpi_assessores.pack(anchor="w", padx=16, pady=(0, 14))

        # KPI 3: Papeis em Agio
        kpi3 = ctk.CTkFrame(kpi_frame, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color=BORDER_CARD)
        kpi3.grid(row=0, column=2, padx=6, sticky="nsew")
        ctk.CTkFrame(kpi3, fg_color=ACCENT_ORANGE, height=4,
                     corner_radius=2).pack(fill="x", padx=14, pady=(14, 0))
        ctk.CTkLabel(kpi3, text="Papeis em Agio", font=("Segoe UI", 10),
                     text_color=TEXT_SECONDARY).pack(anchor="w", padx=16, pady=(10, 0))
        self.ag_kpi_papeis = ctk.CTkLabel(
            kpi3, text="...", font=("Segoe UI", 22, "bold"), text_color=TEXT_PRIMARY)
        self.ag_kpi_papeis.pack(anchor="w", padx=16, pady=(0, 14))

        # KPI 4: Principal Indexador
        kpi4 = ctk.CTkFrame(kpi_frame, fg_color=BG_CARD, corner_radius=12,
                             border_width=1, border_color=BORDER_CARD)
        kpi4.grid(row=0, column=3, padx=(6, 0), sticky="nsew")
        ctk.CTkFrame(kpi4, fg_color=ACCENT_PURPLE, height=4,
                     corner_radius=2).pack(fill="x", padx=14, pady=(14, 0))
        ctk.CTkLabel(kpi4, text="Principal Indexador", font=("Segoe UI", 10),
                     text_color=TEXT_SECONDARY).pack(anchor="w", padx=16, pady=(10, 0))
        self.ag_kpi_indexador = ctk.CTkLabel(
            kpi4, text="...", font=("Segoe UI", 22, "bold"), text_color=TEXT_PRIMARY)
        self.ag_kpi_indexador.pack(anchor="w", padx=16, pady=(0, 14))

        # ======== TOTAL DESTAQUE ========
        total_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                  border_width=1, border_color=BORDER_CARD)
        total_card.pack(fill="x", pady=(0, 18))
        total_inner = ctk.CTkFrame(total_card, fg_color="transparent")
        total_inner.pack(fill="x", padx=24, pady=16)
        ctk.CTkLabel(total_inner, text="VALOR TOTAL ESTIMADO EM AGIO",
                     font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_TERTIARY).pack(anchor="w")
        self.ag_total_value = ctk.CTkLabel(
            total_inner, text="Carregando...",
            font=("Segoe UI", 32, "bold"), text_color=ACCENT_GREEN)
        self.ag_total_value.pack(anchor="w", pady=(2, 0))
        self.ag_total_sub = ctk.CTkLabel(
            total_inner, text="",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY)
        self.ag_total_sub.pack(anchor="w")

        # ======== TOP ASSESSORES ========
        ctk.CTkLabel(
            content, text="Top Assessores - Papeis em Agio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        ranking_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                    border_width=1, border_color=BORDER_CARD)
        ranking_card.pack(fill="x", pady=(0, 18))
        self.ag_ranking_inner = ctk.CTkFrame(ranking_card, fg_color="transparent")
        self.ag_ranking_inner.pack(fill="x", padx=20, pady=16)
        ctk.CTkLabel(self.ag_ranking_inner, text="Carregando...",
                     font=("Segoe UI", 11), text_color=TEXT_TERTIARY).pack(anchor="w")

        # ======== DISTRIBUICAO POR INDEXADOR ========
        ctk.CTkLabel(
            content, text="Distribuição por Indexador",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 10))

        dist_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                 border_width=1, border_color=BORDER_CARD)
        dist_card.pack(fill="x", pady=(0, 18))
        self.ag_dist_inner = ctk.CTkFrame(dist_card, fg_color="transparent")
        self.ag_dist_inner.pack(fill="x", padx=20, pady=16)
        ctk.CTkLabel(self.ag_dist_inner, text="Carregando...",
                     font=("Segoe UI", 11), text_color=TEXT_TERTIARY).pack(anchor="w")

        # ======== ASSUNTO ========
        ctk.CTkLabel(
            content, text="Configurações do Envio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        config_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                   border_width=1, border_color=BORDER_CARD)
        config_card.pack(fill="x", pady=(0, 16))

        ci = ctk.CTkFrame(config_card, fg_color="transparent")
        ci.pack(fill="x", padx=20, pady=16)

        ctk.CTkLabel(
            ci, text="Tipo de Ativo",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 6))

        self.ag_tipo_ativo = ctk.CTkSegmentedButton(
            ci,
            values=["Tesouro Direto", "Crédito Privado", "Título Publico"],
            font=("Segoe UI", 11, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_GREEN,
            selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY,
            corner_radius=8,
            height=38,
        )
        self.ag_tipo_ativo.pack(fill="x")
        self.ag_tipo_ativo.set("Crédito Privado")

        # ======== AÇÕES ========
        act_frame = ctk.CTkFrame(content, fg_color="transparent")
        act_frame.pack(fill="x", pady=(14, 14))

        self.ag_enviar_btn = ctk.CTkButton(
            act_frame, text="\u2709  Gerar Rascunhos no Outlook",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            height=44, corner_radius=10,
            command=self._on_ag_gerar_emails,
        )
        self.ag_enviar_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            act_frame, text="\u2702  Limpar Planilha",
            font=("Segoe UI", 11, "bold"),
            fg_color=ACCENT_RED, hover_color="#b02a37",
            height=44, corner_radius=10, width=160,
            command=self._on_ag_limpar_xlsx,
        ).pack(side="left")

        # ======== STATUS ========
        self.ag_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.ag_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.ag_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.ag_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.ag_status_dot.pack(side="left")

        self.ag_status_text = ctk.CTkLabel(
            si, text="  Preencha o assunto e gere os rascunhos",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.ag_status_text.pack(side="left")

        # Atualizar (abre pasta da planilha)
        os.makedirs(AGIO_DIR, exist_ok=True)
        self._make_atualizar_btn(content, os.path.join(AGIO_DIR, "agio.xlsx"))

        # Internal state
        self._ag_data = []
        self._ag_headers = []
        self._ag_loaded = False

        return page

    # -----------------------------------------------------------------
    #  INFORMATIVO AGIO: Carregamento e Dashboard
    # -----------------------------------------------------------------
    def _ag_load_data_thread(self):
        """Le XLSX em background thread (read_only) e pre-computa dashboard."""
        from collections import Counter, defaultdict

        os.makedirs(AGIO_DIR, exist_ok=True)
        xlsx_files = [f for f in os.listdir(AGIO_DIR)
                      if f.lower().endswith(".xlsx") and not f.startswith("~")]
        if not xlsx_files:
            self._ag_data = []
            self._ag_headers = []
            self.after(0, lambda: self._ag_update_dashboard(None))
            return

        path = os.path.join(AGIO_DIR, xlsx_files[0])
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)

            first_row = next(rows_iter, None)
            if not first_row:
                wb.close()
                self._ag_data = []
                self._ag_headers = []
                self.after(0, lambda: self._ag_update_dashboard(None))
                return

            headers = [str(h) if h else f"Col{i}"
                       for i, h in enumerate(first_row, 1)]
            rows = []
            for row_vals in rows_iter:
                if any(v is not None for v in row_vals):
                    rows.append(dict(zip(headers, row_vals)))
            wb.close()

            self._ag_data = rows
            self._ag_headers = headers

            # --- Pre-computar dados do dashboard na thread ---
            def _find_col(*keywords):
                for h in headers:
                    hl = h.lower()
                    if all(k in hl for k in keywords):
                        return h
                return None

            col_assessor = _find_col("assessor")
            col_agio_pct = _find_col("gio", "%")
            col_agio_rs = _find_col("gio", "r$")
            col_index = _find_col("indexador")

            filtered = []
            for row in rows:
                val = row.get(col_agio_pct, 0) if col_agio_pct else 0
                try:
                    num = float(val) if val is not None else 0
                except (ValueError, TypeError):
                    continue
                if num > 0.005:
                    filtered.append(row)

            total_rs = 0.0
            for r in filtered:
                try:
                    total_rs += float(r.get(col_agio_rs, 0) or 0)
                except (ValueError, TypeError):
                    pass

            assessores_set = set()
            for r in filtered:
                cod = str(r.get(col_assessor, "")).strip() if col_assessor else ""
                if cod:
                    assessores_set.add(cod)

            idx_counts = Counter()
            if col_index:
                for r in filtered:
                    idx = str(r.get(col_index, "")).strip()
                    if idx:
                        idx_counts[idx] += 1
            top_idx = idx_counts.most_common(1)[0][0] if idx_counts else "--"

            by_assessor = defaultdict(int)
            for r in filtered:
                cod = str(r.get(col_assessor, "")).strip() if col_assessor else ""
                if cod:
                    by_assessor[cod] += 1

            info = {
                "path": path,
                "total_rs": total_rs,
                "n_assessores": len(assessores_set),
                "n_papeis": len(filtered),
                "top_idx": top_idx,
                "top5": sorted(by_assessor.items(), key=lambda x: -x[1])[:5],
                "top_idxs": idx_counts.most_common(6),
                "n_total": len(rows),
            }
            self.after(0, lambda: self._ag_update_dashboard(info))

        except Exception as e:
            self._ag_data = []
            self._ag_headers = []
            self.after(0, lambda err=e: self._ag_update_dashboard({"error": err}))

    def _ag_load_data(self):
        """Carrega dados (sync) — chamado pelo botao Recarregar e Gerar Emails."""
        self._ag_loaded = True
        threading.Thread(target=self._ag_load_data_thread, daemon=True).start()

    def _ag_update_dashboard(self, info):
        """Atualiza widgets do dashboard com dados pre-computados (main thread)."""
        if info is None:
            self.ag_kpi_total.configure(text="--")
            self.ag_kpi_assessores.configure(text="--")
            self.ag_kpi_papeis.configure(text="--")
            self.ag_kpi_indexador.configure(text="--")
            self.ag_total_value.configure(text="Sem dados")
            self.ag_total_sub.configure(
                text="Insira planilha .xlsx via botao Atualizar")
            for w in self.ag_ranking_inner.winfo_children():
                w.destroy()
            ctk.CTkLabel(self.ag_ranking_inner, text="Sem dados",
                         font=("Segoe UI", 11),
                         text_color=TEXT_TERTIARY).pack(anchor="w")
            for w in self.ag_dist_inner.winfo_children():
                w.destroy()
            ctk.CTkLabel(self.ag_dist_inner, text="Sem dados",
                         font=("Segoe UI", 11),
                         text_color=TEXT_TERTIARY).pack(anchor="w")
            self.ag_status_dot.configure(text_color=ACCENT_RED)
            self.ag_status_text.configure(
                text="  Nenhuma planilha encontrada em Mesa Produtos/Info Agio/Agio/")
            return

        if "error" in info:
            self.ag_status_dot.configure(text_color=ACCENT_RED)
            self.ag_status_text.configure(
                text=f"  Erro ao carregar: {info['error']}")
            return

        total_rs = info["total_rs"]
        n_assessores = info["n_assessores"]
        n_papeis = info["n_papeis"]
        top_idx = info["top_idx"]
        top5 = info["top5"]
        top_idxs = info["top_idxs"]
        n_total = info["n_total"]
        path = info["path"]

        # --- KPI cards ---
        fmt_total = f"R$ {abs(total_rs):,.0f}".replace(",", ".")
        self.ag_kpi_total.configure(text=fmt_total)
        self.ag_kpi_assessores.configure(text=str(n_assessores))
        self.ag_kpi_papeis.configure(text=str(n_papeis))
        self.ag_kpi_indexador.configure(text=top_idx)

        # --- Total destaque ---
        fmt_full = f"R$ {abs(total_rs):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        self.ag_total_value.configure(text=fmt_full)
        self.ag_total_sub.configure(
            text=f"{n_papeis} papeis com agio > 0,50%  |  "
                 f"{n_assessores} assessores  |  "
                 f"{n_total} registros totais")

        # --- Top Assessores ranking ---
        for w in self.ag_ranking_inner.winfo_children():
            w.destroy()
        max_count = top5[0][1] if top5 else 1
        for i, (cod, count) in enumerate(top5):
            row_f = ctk.CTkFrame(self.ag_ranking_inner, fg_color="transparent")
            row_f.pack(fill="x", pady=3)
            ctk.CTkLabel(
                row_f, text=f"{i + 1}.", font=("Segoe UI", 11, "bold"),
                text_color=TEXT_TERTIARY, width=24
            ).pack(side="left")
            ctk.CTkLabel(
                row_f, text=cod, font=("Segoe UI", 11, "bold"),
                text_color=TEXT_PRIMARY, width=80
            ).pack(side="left", padx=(0, 12))
            bar_bg = ctk.CTkFrame(
                row_f, fg_color=BG_PROGRESS_TRACK, height=18, corner_radius=4)
            bar_bg.pack(side="left", fill="x", expand=True, padx=(0, 12))
            bar_bg.pack_propagate(False)
            pct = count / max_count
            bar_fill = ctk.CTkFrame(
                bar_bg, fg_color=ACCENT_GREEN, corner_radius=4)
            bar_fill.place(relx=0, rely=0, relwidth=pct, relheight=1)
            ctk.CTkLabel(
                row_f, text=str(count), font=("Segoe UI", 11, "bold"),
                text_color=ACCENT_GREEN, width=40
            ).pack(side="right")

        # --- Distribuição por Indexador ---
        for w in self.ag_dist_inner.winfo_children():
            w.destroy()
        max_idx = top_idxs[0][1] if top_idxs else 1
        colors = [ACCENT_GREEN, ACCENT_BLUE, ACCENT_ORANGE,
                  ACCENT_PURPLE, ACCENT_TEAL, ACCENT_RED]
        for i, (idx_name, count) in enumerate(top_idxs):
            row_f = ctk.CTkFrame(self.ag_dist_inner, fg_color="transparent")
            row_f.pack(fill="x", pady=3)
            ctk.CTkLabel(
                row_f, text=sanitize_text(idx_name),
                font=("Segoe UI", 11), text_color=TEXT_PRIMARY,
                width=120, anchor="w"
            ).pack(side="left", padx=(0, 12))
            bar_bg = ctk.CTkFrame(
                row_f, fg_color=BG_PROGRESS_TRACK, height=18, corner_radius=4)
            bar_bg.pack(side="left", fill="x", expand=True, padx=(0, 12))
            bar_bg.pack_propagate(False)
            pct = count / max_idx
            c = colors[i % len(colors)]
            bar_fill = ctk.CTkFrame(bar_bg, fg_color=c, corner_radius=4)
            bar_fill.place(relx=0, rely=0, relwidth=pct, relheight=1)
            ctk.CTkLabel(
                row_f, text=str(count), font=("Segoe UI", 11, "bold"),
                text_color=c, width=40
            ).pack(side="right")

        # --- Status ---
        self.ag_status_dot.configure(text_color=ACCENT_GREEN)
        self.ag_status_text.configure(
            text=f"  Dados carregados - {os.path.basename(path)}"
                 f"  |  {n_total} registros  |  {n_papeis} com agio > 0,50%")

    def _on_ag_limpar_xlsx(self):
        """Remove do XLSX todas as linhas com |agio %| < 0,50%."""
        os.makedirs(AGIO_DIR, exist_ok=True)
        xlsx_files = [f for f in os.listdir(AGIO_DIR)
                      if f.lower().endswith(".xlsx") and not f.startswith("~")]
        if not xlsx_files:
            messagebox.showwarning(
                "Sem planilha",
                "Nenhuma planilha .xlsx encontrada em Mesa Produtos/Info Agio/Agio/.\n"
                "Clique em 'Atualizar' para inserir o arquivo.")
            return

        path = os.path.join(AGIO_DIR, xlsx_files[0])
        resp = messagebox.askyesno(
            "Limpar Planilha",
            f"Isso vai REMOVER do arquivo todas as linhas com\n"
            f"|agio| inferior a 0,50%.\n\n"
            f"Arquivo: {os.path.basename(path)}\n\n"
            "Esta acao não pode ser desfeita. Continuar?")
        if not resp:
            return

        self.ag_status_dot.configure(text_color=ACCENT_ORANGE)
        self.ag_status_text.configure(text="  Limpando planilha...")

        threading.Thread(
            target=self._run_limpar_xlsx, args=(path,), daemon=True
        ).start()

    def _run_limpar_xlsx(self, path):
        """Thread: reescreve XLSX mantendo so linhas com agio >= 0,50%."""
        try:
            from openpyxl import Workbook
            wb_src = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws_src = wb_src[wb_src.sheetnames[0]]
            rows_iter = ws_src.iter_rows(values_only=True)

            header_row = next(rows_iter, None)
            if not header_row:
                wb_src.close()
                self.after(0, lambda: messagebox.showerror(
                    "Erro", "Planilha vazia."))
                return

            # Encontrar coluna de agio %
            agio_col = None
            for i, val in enumerate(header_row):
                if val and "gio" in str(val).lower() and "%" in str(val):
                    agio_col = i
                    break

            if agio_col is None:
                wb_src.close()
                self.after(0, lambda: messagebox.showerror(
                    "Erro", "Coluna de Agio (%) não encontrada."))
                return

            # Ler todas as linhas e filtrar
            kept_rows = []
            total_src = 0
            for row_vals in rows_iter:
                if not any(v is not None for v in row_vals):
                    continue
                total_src += 1
                val = row_vals[agio_col]
                try:
                    num = float(val) if val is not None else 0
                except (ValueError, TypeError):
                    num = 0
                if num > 0.005:
                    kept_rows.append(row_vals)

            wb_src.close()

            # Criar novo workbook com linhas filtradas
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.title = "Export"
            for c, val in enumerate(header_row, 1):
                ws_new.cell(1, c, val)
            for r, row_vals in enumerate(kept_rows, 2):
                for c, val in enumerate(row_vals, 1):
                    ws_new.cell(r, c, val)

            kept = len(kept_rows)
            removed = total_src - kept

            wb_new.save(path)
            wb_new.close()

            def _done():
                self.ag_status_dot.configure(text_color=ACCENT_GREEN)
                self.ag_status_text.configure(
                    text=f"  Limpeza concluída - {removed} removidas"
                         f"  |  {kept} restantes")
                messagebox.showinfo(
                    "Limpeza Concluida",
                    f"{removed} linhas com |agio| < 0,50% removidas.\n"
                    f"{kept} linhas restantes no arquivo.")
                self._ag_load_data()
            self.after(0, _done)

        except Exception as e:
            self.after(0, lambda: messagebox.showerror(
                "Erro ao limpar", str(e)))
            self.after(0, lambda: self.ag_status_dot.configure(
                text_color=ACCENT_RED))
            self.after(0, lambda: self.ag_status_text.configure(
                text=f"  Erro: {e}"))

    def _on_ag_gerar_emails(self):
        if not self._ag_data:
            messagebox.showwarning(
                "Dados não carregados",
                "Aguarde o carregamento ou clique em Recarregar.")
            return

        tipo_ativo = self.ag_tipo_ativo.get()

        # Contar registros filtrados e assessores (dados ja em memoria)
        headers = getattr(self, "_ag_headers", [])
        col_agio_pct = None
        col_assessor = None
        for h in headers:
            hl = h.lower()
            if "gio" in hl and "%" in hl:
                col_agio_pct = h
            if "assessor" in hl:
                col_assessor = h

        n_filtrados = 0
        assessores_set = set()
        for row in self._ag_data:
            agio_val = row.get(col_agio_pct, 0) if col_agio_pct else 0
            try:
                agio_num = float(agio_val) if agio_val is not None else 0
            except (ValueError, TypeError):
                continue
            if agio_num > 0.005:
                n_filtrados += 1
                cod = str(row.get(col_assessor, "")).strip() if col_assessor else ""
                if cod:
                    assessores_set.add(cod)

        if n_filtrados == 0:
            messagebox.showinfo("Sem dados", "Nenhum registro com agio acima de 0,50%.")
            return

        resp = messagebox.askyesno(
            "Gerar Rascunhos",
            f"Serao criados rascunhos no Outlook:\n\n"
            f"Tipo: {tipo_ativo}\n"
            f"{n_filtrados} registros com agio > 0,50%\n"
            f"{len(assessores_set)} assessores\n\n"
            "O Outlook precisa estar aberto. Continuar?"
        )
        if not resp:
            return

        self.ag_enviar_btn.configure(state="disabled")
        self.ag_status_dot.configure(text_color=ACCENT_ORANGE)
        self.ag_status_text.configure(text="  Gerando rascunhos...")
        threading.Thread(
            target=self._run_ag_emails,
            args=(tipo_ativo,),
            daemon=True
        ).start()

    def _run_ag_emails(self, tipo_ativo):
        try:
            import win32com.client as win32
            import pythoncom
            from collections import defaultdict
            pythoncom.CoInitialize()

            outlook = win32.Dispatch("Outlook.Application")
            assessores = load_base_emails()

            hoje = datetime.now().strftime("%d/%m/%Y")
            logo_tag = LOGO_TAG_CID if os.path.exists(LOGO_PATH) else ""

            headers = self._ag_headers

            # --- Localizar colunas por correspondencia parcial ---
            def _find_col(*keywords):
                for h in headers:
                    hl = h.lower()
                    if all(k in hl for k in keywords):
                        return h
                return None

            col_assessor = _find_col("assessor") or headers[1]
            col_agio_pct = _find_col("gio", "%") or headers[-2]
            col_agio_rs = _find_col("gio", "r$") or headers[-1]
            col_conta = _find_col("conta")
            col_ativo = _find_col("nome", "ativo") or _find_col("ativo")
            col_ticker = _find_col("ticker")
            col_venc = _find_col("vencimento")
            col_index = _find_col("indexador")
            col_valor = _find_col("valor", "aplic")
            col_posição = _find_col("posi")

            # Colunas exibidas no email (chave_original, label_email)
            email_cols = []
            if col_conta:    email_cols.append((col_conta, "Conta"))
            if col_ativo:    email_cols.append((col_ativo, "Ativo"))
            if col_ticker:   email_cols.append((col_ticker, "Ticker"))
            if col_venc:     email_cols.append((col_venc, "Vencimento"))
            if col_index:    email_cols.append((col_index, "Indexador"))
            if col_valor:    email_cols.append((col_valor, "Valor Aplicado"))
            if col_posição:  email_cols.append((col_posição, "Pos. Atual"))
            if col_agio_pct: email_cols.append((col_agio_pct, "Agio (%)"))
            if col_agio_rs:  email_cols.append((col_agio_rs, "Agio (R$)"))

            money_cols = {col_valor, col_posição, col_agio_rs}
            pct_cols = {col_agio_pct}
            date_cols = {col_venc}

            # --- Filtrar: |agio %| > 0,50% (0.005 em decimal) ---
            filtered = []
            for row in self._ag_data:
                agio_val = row.get(col_agio_pct, 0)
                try:
                    agio_num = float(agio_val) if agio_val is not None else 0
                except (ValueError, TypeError):
                    continue
                if agio_num > 0.005:
                    filtered.append(row)

            # --- Agrupar por assessor e ordenar por financeiro (Agio R$) desc ---
            by_assessor = defaultdict(list)
            for row in filtered:
                cod = str(row.get(col_assessor, "")).strip()
                if cod:
                    by_assessor[cod].append(row)
            for cod in by_assessor:
                by_assessor[cod].sort(
                    key=lambda r: abs(float(r.get(col_agio_rs, 0) or 0)),
                    reverse=True)

            # --- Montar tabela HTML para um assessor ---
            def _build_table(rows):
                hdr_html = "".join(
                    f'<td style="padding:4px 10px;font-weight:bold;color:#00785a;'
                    f'font-size:8.5pt;border-bottom:1.5px solid #00785a;white-space:nowrap;">'
                    f'{sanitize_text(label)}</td>'
                    for _col_key, label in email_cols
                )
                rows_html = ""
                for i, row in enumerate(rows):
                    bg = "#f7faf9" if i % 2 == 0 else "#ffffff"
                    cells = ""
                    for col_key, _label in email_cols:
                        val = row.get(col_key, "")
                        color = "#1a1a2e"
                        if col_key in date_cols and val is not None:
                            try:
                                if hasattr(val, "strftime"):
                                    val = val.strftime("%d/%m/%y")
                                else:
                                    s = str(val).strip()[:10]
                                    dt = datetime.strptime(s, "%Y-%m-%d")
                                    val = dt.strftime("%d/%m/%y")
                            except (ValueError, TypeError):
                                val = str(val) if val else ""
                        elif col_key in pct_cols and val is not None:
                            try:
                                v = float(val)
                                val = f"{v * 100:.2f}%"
                                if v < 0:
                                    color = "#c0392b"
                            except (ValueError, TypeError):
                                val = str(val) if val else ""
                        elif col_key in money_cols and val is not None:
                            try:
                                v = float(val)
                                val = f"R$ {v:,.2f}"
                                if v < 0:
                                    color = "#c0392b"
                            except (ValueError, TypeError):
                                val = str(val) if val else ""
                        else:
                            val = str(val) if val is not None else ""
                        cells += (
                            f'<td style="padding:3px 10px;color:{color};'
                            f'border-bottom:1px solid #eef1ef;font-size:8.5pt;'
                            f'white-space:nowrap;">{sanitize_text(val)}</td>'
                        )
                    rows_html += f'<tr style="background:{bg};">{cells}</tr>\n'
                return (
                    f'<table cellpadding="0" cellspacing="0" border="0" '
                    f'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;'
                    f'font-size:9pt;margin-bottom:6px;">'
                    f'<tr>{hdr_html}</tr>\n{rows_html}</table>'
                )

            # --- HTML completo do email (padrao Saldos Diários) ---
            def _build_html(primeiro_nome, data_table):
                return f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">
    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#4b5563;margin-top:0;">
      Segue abaixo o informativo de &aacute;gio/des&aacute;gio dos seus clientes.
    </p>
  </td></tr>
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="background-color:#00b876;width:4px;border-radius:4px;">&nbsp;</td>
      <td style="padding-left:12px;">
        <span style="font-size:12.5pt;color:#004d33;font-weight:bold;letter-spacing:0.3px;">
          Ativos com &Aacute;gio/Des&aacute;gio acima de 0,50%
        </span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:0 4px;">
    {data_table}
  </td></tr>
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">Informativo &Aacute;gio &middot; {hoje}</span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">E-mail gerado automaticamente</span>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</div>
"""

            criados = 0
            erros = 0
            ignorados = 0

            for cod_assess, rows in sorted(by_assessor.items()):
                info = assessores.get(cod_assess)
                if not info:
                    ignorados += 1
                    continue
                email = info.get("email", "")
                if not email or email == "-":
                    ignorados += 1
                    continue

                nome = info["nome"]
                primeiro = nome.split()[0] if nome else "Assessor"
                tabela = _build_table(rows)

                try:
                    mail = outlook.CreateItem(0)
                    mail.To = email
                    email_assist = info.get("email_assistente", "-")
                    if email_assist and email_assist != "-":
                        mail.CC = email_assist
                    mail.Subject = (
                        f"{cod_assess} | Segue a sua relação de agio de "
                        f"{tipo_ativo}")
                    mail.HTMLBody = _build_html(primeiro, tabela)
                    _attach_logo_cid(mail)
                    mail.Save()
                    criados += 1
                except Exception:
                    erros += 1

            pythoncom.CoUninitialize()

            def _done():
                self.ag_enviar_btn.configure(state="normal")
                self.ag_status_dot.configure(text_color="#00a86b")
                self.ag_status_text.configure(
                    text=f"  Concluído - {criados} rascunhos, {erros} erros, {ignorados} sem email"
                )
                messagebox.showinfo("Rascunhos Criados",
                    f"{criados} rascunhos criados no Outlook!\n"
                    f"{erros} erros\n{ignorados} assessores sem email cadastrado")
            self.after(0, _done)

        except Exception as e:
            def _err():
                self.ag_enviar_btn.configure(state="normal")
                self.ag_status_dot.configure(text_color=ACCENT_RED)
                self.ag_status_text.configure(text=f"  Erro: {e}")
                messagebox.showerror("Erro", str(e))
            self.after(0, _err)

    # -----------------------------------------------------------------
    #  PAGE: ENVIO DE ORDENS
    # -----------------------------------------------------------------
    def _build_envio_ordens_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Envio de Ordens", subtitle="Mesa de Produtos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Banner
        banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=44)
        banner.pack(fill="x", pady=(0, 18))
        banner.pack_propagate(False)

        ctk.CTkLabel(
            banner, text="  Envio de ordens por e-mail a partir de planilha",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        ).pack(side="left", padx=18, pady=10)

        # ======== DESTINATÁRIOS ========
        ctk.CTkLabel(
            content, text="Destinatarios",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 6))

        self.eo_lista_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=12,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.eo_lista_card.pack(fill="x", pady=(0, 16))

        lista_inner = ctk.CTkFrame(self.eo_lista_card, fg_color="transparent")
        lista_inner.pack(fill="x", padx=20, pady=16)

        lista_left = ctk.CTkFrame(lista_inner, fg_color="transparent")
        lista_left.pack(side="left", fill="x", expand=True)

        self.eo_lista_icon = ctk.CTkLabel(
            lista_left, text="\u2709",
            font=("Segoe UI", 20), text_color=ACCENT_BLUE
        )
        self.eo_lista_icon.pack(side="left")

        self.eo_lista_info = ctk.CTkLabel(
            lista_left, text="  Nenhuma lista carregada",
            font=("Segoe UI", 12), text_color=TEXT_SECONDARY, anchor="w"
        )
        self.eo_lista_info.pack(side="left", padx=(6, 0))

        ctk.CTkButton(
            lista_inner, text="  Inserir Lista de Emails",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=38, corner_radius=8, width=200,
            command=self._on_eo_browse_lista,
        ).pack(side="right")

        # Preview da lista carregada
        self.eo_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.eo_preview_frame.pack(fill="x", pady=(0, 16))
        self.eo_preview_frame.pack_forget()

        self.eo_preview_text = ctk.CTkTextbox(
            self.eo_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=140, wrap="none"
        )
        self.eo_preview_text.pack(fill="x", padx=4, pady=4)
        self.eo_preview_text.configure(state="disabled")

        # --- Email manual (caso a lista não seja carregada) ---
        manual_card = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        manual_card.pack(fill="x", pady=(0, 16))

        manual_inner = ctk.CTkFrame(manual_card, fg_color="transparent")
        manual_inner.pack(fill="x", padx=20, pady=14)

        ctk.CTkLabel(
            manual_inner, text="Ou digite manualmente:",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        manual_row = ctk.CTkFrame(manual_inner, fg_color="transparent")
        manual_row.pack(fill="x")
        manual_row.columnconfigure((0, 1), weight=1)

        man_nome_f = ctk.CTkFrame(manual_row, fg_color="transparent")
        man_nome_f.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ctk.CTkLabel(man_nome_f, text="Nome", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_manual_nome = ctk.CTkEntry(man_nome_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Ex: Joao Silva")
        self.eo_manual_nome.pack(fill="x", pady=(2, 0))

        man_email_f = ctk.CTkFrame(manual_row, fg_color="transparent")
        man_email_f.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        ctk.CTkLabel(man_email_f, text="E-mail", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_manual_email = ctk.CTkEntry(man_email_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Ex: joao@email.com")
        self.eo_manual_email.pack(fill="x", pady=(2, 0))

        # --- Campo CC ---
        cc_frame = ctk.CTkFrame(manual_inner, fg_color="transparent")
        cc_frame.pack(fill="x", pady=(10, 0))
        ctk.CTkLabel(cc_frame, text="CC (opcional - separe multiplos com ;)", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_cc = ctk.CTkEntry(cc_frame, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Ex: gestor@email.com; compliance@email.com")
        self.eo_cc.pack(fill="x", pady=(2, 0))

        # ======== CONFIGURAÇÕES ========
        ctk.CTkLabel(
            content, text="Configurações do Envio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        config_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                   border_width=1, border_color=BORDER_CARD)
        config_card.pack(fill="x", pady=(0, 16))

        ci = ctk.CTkFrame(config_card, fg_color="transparent")
        ci.pack(fill="x", padx=20, pady=16)

        # Row 1: Modo de envio + Tipo de ativo
        row1 = ctk.CTkFrame(ci, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 14))
        row1.columnconfigure((0, 1), weight=1)

        # -- Modo de envio --
        modo_frame = ctk.CTkFrame(row1, fg_color="transparent")
        modo_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        ctk.CTkLabel(
            modo_frame, text="Modo de Envio",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 6))

        self.eo_modo = ctk.CTkSegmentedButton(
            modo_frame,
            values=["Grupo", "Individual"],
            font=("Segoe UI", 11, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_GREEN,
            selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY,
            corner_radius=8,
            height=38,
        )
        self.eo_modo.pack(fill="x")
        self.eo_modo.set("Individual")

        # -- Tipo de ativo --
        tipo_frame = ctk.CTkFrame(row1, fg_color="transparent")
        tipo_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        ctk.CTkLabel(
            tipo_frame, text="Tipo de Ativo",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 6))

        self.eo_tipo_ativo = ctk.CTkSegmentedButton(
            tipo_frame,
            values=["Fundos", "Renda Variavel", "Renda Fixa", "Alternativo"],
            font=("Segoe UI", 11, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_BLUE,
            selected_hover_color="#1555bb",
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY,
            corner_radius=8,
            height=38,
        )
        self.eo_tipo_ativo.pack(fill="x")
        self.eo_tipo_ativo.set("Fundos")

        # === SELETOR TIPO DE ORDEM ===
        ctk.CTkLabel(
            ci, text="Tipo de Ordem",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(14, 6))

        self.eo_tipo_ordem = ctk.CTkSegmentedButton(
            ci,
            values=["Normal", "OFERTA P\u00daBLICA - RF"],
            font=("Segoe UI", 11, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_PURPLE,
            selected_hover_color="#6a4db0",
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY,
            corner_radius=8,
            height=38,
            command=self._on_eo_tipo_ordem_change,
        )
        self.eo_tipo_ordem.pack(fill="x", pady=(0, 14))
        self.eo_tipo_ordem.set("Normal")

        # ===== FRAME NORMAL (Dados da Ordem dinâmicos + Venda) =====
        self._eo_normal_frame = ctk.CTkFrame(ci, fg_color="transparent")
        self._eo_normal_frame.pack(fill="x")

        # Container para múltiplos blocos de ordem
        self._eo_ordens_container = ctk.CTkFrame(self._eo_normal_frame, fg_color="transparent")
        self._eo_ordens_container.pack(fill="x")

        # Lista que armazena referências aos widgets de cada bloco de ordem
        self._eo_ordens_blocks = []

        # Criar o primeiro bloco de ordem
        self._eo_add_ordem_block()

        # Botão "Adicionar Ativo"
        self._eo_add_btn_frame = ctk.CTkFrame(self._eo_normal_frame, fg_color="transparent")
        self._eo_add_btn_frame.pack(fill="x", pady=(4, 10))

        ctk.CTkButton(
            self._eo_add_btn_frame, text="+  Adicionar Ativo",
            font=("Segoe UI", 11, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=34, corner_radius=8, width=180,
            command=self._eo_add_ordem_block,
        ).pack(side="left")

        # Venda
        venda_sep = ctk.CTkFrame(self._eo_normal_frame, fg_color="#fff3e8", corner_radius=8, border_width=1, border_color=ACCENT_ORANGE)
        venda_sep.pack(fill="x", pady=(8, 10))
        venda_sep_inner = ctk.CTkFrame(venda_sep, fg_color="transparent")
        venda_sep_inner.pack(fill="x", padx=14, pady=10)
        ctk.CTkLabel(venda_sep_inner, text="\u2193  Venda \u2014 Ativo de Sa\u00edda", font=("Segoe UI", 11, "bold"), text_color=ACCENT_ORANGE, anchor="w").pack(side="left")
        ctk.CTkLabel(venda_sep_inner, text="(opcional \u2014 preencha caso o cliente precise sair de algum ativo)", font=("Segoe UI", 9), text_color=TEXT_TERTIARY, anchor="w").pack(side="left", padx=(10, 0))

        venda_row = ctk.CTkFrame(self._eo_normal_frame, fg_color="transparent")
        venda_row.pack(fill="x", pady=(0, 4))
        venda_row.columnconfigure((0, 1, 2, 3), weight=1)

        vativo_frame = ctk.CTkFrame(venda_row, fg_color="transparent")
        vativo_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ctk.CTkLabel(vativo_frame, text="Ativo", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_venda_ativo = ctk.CTkEntry(vativo_frame, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_ORANGE, placeholder_text="Ex: FIXA2028")
        self.eo_venda_ativo.pack(fill="x", pady=(2, 0))

        vqtd_frame = ctk.CTkFrame(venda_row, fg_color="transparent")
        vqtd_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 6))
        ctk.CTkLabel(vqtd_frame, text="Quantidade", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_venda_quantidade = ctk.CTkEntry(vqtd_frame, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_ORANGE, placeholder_text="Ex: 100")
        self.eo_venda_quantidade.pack(fill="x", pady=(2, 0))

        vfin_frame = ctk.CTkFrame(venda_row, fg_color="transparent")
        vfin_frame.grid(row=0, column=2, sticky="nsew", padx=(6, 6))
        ctk.CTkLabel(vfin_frame, text="Financeiro", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_venda_financeiro = ctk.CTkEntry(vfin_frame, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_ORANGE, placeholder_text="Ex: 10000.00")
        self.eo_venda_financeiro.pack(fill="x", pady=(2, 0))

        vcot_frame = ctk.CTkFrame(venda_row, fg_color="transparent")
        vcot_frame.grid(row=0, column=3, sticky="nsew", padx=(6, 0))
        ctk.CTkLabel(vcot_frame, text="Cota\u00e7\u00e3o", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_venda_cotacao = ctk.CTkEntry(vcot_frame, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_ORANGE, placeholder_text="Opcional")
        self.eo_venda_cotacao.pack(fill="x", pady=(2, 0))

        # ===== FRAME OFERTA PUBLICA RF (oculto inicialmente) =====
        self._eo_oferta_frame = ctk.CTkFrame(ci, fg_color="transparent")
        # Não faz pack — fica oculto até selecionar

        of_banner = ctk.CTkFrame(self._eo_oferta_frame, fg_color="#f3eeff", corner_radius=8,
                                 border_width=1, border_color=ACCENT_PURPLE)
        of_banner.pack(fill="x", pady=(0, 14))
        ctk.CTkLabel(
            of_banner, text="  Reserva de Oferta P\u00fablica - Renda Fixa",
            font=("Segoe UI", 11, "bold"), text_color=ACCENT_PURPLE, anchor="w"
        ).pack(fill="x", padx=14, pady=10)

        # Row 1: Oferta + Prospecto
        of_row1 = ctk.CTkFrame(self._eo_oferta_frame, fg_color="transparent")
        of_row1.pack(fill="x", pady=(0, 10))
        of_row1.columnconfigure((0, 1), weight=1)

        of_oferta_f = ctk.CTkFrame(of_row1, fg_color="transparent")
        of_oferta_f.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ctk.CTkLabel(of_oferta_f, text="Oferta", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_oferta = ctk.CTkEntry(of_oferta_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="Ex: CRI Direcional")
        self.eo_of_oferta.pack(fill="x", pady=(2, 0))

        of_prosp_f = ctk.CTkFrame(of_row1, fg_color="transparent")
        of_prosp_f.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        ctk.CTkLabel(of_prosp_f, text="Link do Prospecto", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_prospecto = ctk.CTkEntry(of_prosp_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="https://...")
        self.eo_of_prospecto.pack(fill="x", pady=(2, 0))

        # Row 2: Serie
        of_serie_f = ctk.CTkFrame(self._eo_oferta_frame, fg_color="transparent")
        of_serie_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(of_serie_f, text="S\u00e9rie", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_serie = ctk.CTkEntry(of_serie_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="Ex: 2\u00aa S\u00e9rie")
        self.eo_of_serie.pack(fill="x", pady=(2, 0))

        # Row 3: Valor + Taxa + Duration + Inv. Vinculado
        of_row3 = ctk.CTkFrame(self._eo_oferta_frame, fg_color="transparent")
        of_row3.pack(fill="x", pady=(0, 10))
        of_row3.columnconfigure((0, 1, 2, 3), weight=1)

        of_valor_f = ctk.CTkFrame(of_row3, fg_color="transparent")
        of_valor_f.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ctk.CTkLabel(of_valor_f, text="Valor da Reserva", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_valor = ctk.CTkEntry(of_valor_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="Ex: R$ 1.000.000,00")
        self.eo_of_valor.pack(fill="x", pady=(2, 0))

        of_taxa_f = ctk.CTkFrame(of_row3, fg_color="transparent")
        of_taxa_f.grid(row=0, column=1, sticky="nsew", padx=(6, 6))
        ctk.CTkLabel(of_taxa_f, text="Taxa", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_taxa = ctk.CTkEntry(of_taxa_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="Ex: TETO (DI Jan/31)")
        self.eo_of_taxa.pack(fill="x", pady=(2, 0))

        of_dur_f = ctk.CTkFrame(of_row3, fg_color="transparent")
        of_dur_f.grid(row=0, column=2, sticky="nsew", padx=(6, 6))
        ctk.CTkLabel(of_dur_f, text="Duration", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_duration = ctk.CTkEntry(of_dur_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="Ex: 4,8 anos")
        self.eo_of_duration.pack(fill="x", pady=(2, 0))

        of_vinc_f = ctk.CTkFrame(of_row3, fg_color="transparent")
        of_vinc_f.grid(row=0, column=3, sticky="nsew", padx=(6, 0))
        ctk.CTkLabel(of_vinc_f, text="Investidor Vinculado", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_vinculado = ctk.CTkSegmentedButton(
            of_vinc_f, values=["N\u00c3O", "SIM"],
            font=("Segoe UI", 11, "bold"), fg_color=BG_INPUT,
            selected_color=ACCENT_PURPLE, selected_hover_color="#6a4db0",
            unselected_color=BG_INPUT, unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY, corner_radius=8, height=38,
        )
        self.eo_of_vinculado.pack(fill="x", pady=(2, 0))
        self.eo_of_vinculado.set("N\u00c3O")

        # Row 4: Codigo do cliente
        of_cod_f = ctk.CTkFrame(self._eo_oferta_frame, fg_color="transparent")
        of_cod_f.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(of_cod_f, text="C\u00f3digo do Cliente", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        self.eo_of_codigo = ctk.CTkEntry(of_cod_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=ACCENT_PURPLE, placeholder_text="Ex: 5714703")
        self.eo_of_codigo.pack(fill="x", pady=(2, 0))

        # Row 3: Assunto
        ctk.CTkLabel(
            ci, text="Assunto do E-mail",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.eo_assunto = ctk.CTkEntry(
            ci, font=("Segoe UI", 12), height=40,
            corner_radius=8, border_color=BORDER_CARD,
            placeholder_text="Ex: Ordem de Compra - Fundos"
        )
        self.eo_assunto.pack(fill="x", pady=(0, 14))

        # Row 3: Corpo
        ctk.CTkLabel(
            ci, text="Corpo do E-mail",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.eo_corpo = ctk.CTkTextbox(
            ci, font=("Segoe UI", 11), height=160,
            corner_radius=8, border_width=1, border_color=BORDER_CARD,
            fg_color=BG_PRIMARY, text_color=TEXT_PRIMARY, wrap="word"
        )
        self.eo_corpo.pack(fill="x", pady=(0, 14))

        # Row 4: Anexo
        anexo_row = ctk.CTkFrame(ci, fg_color="transparent")
        anexo_row.pack(fill="x", pady=(0, 0))

        ctk.CTkLabel(
            anexo_row, text="Anexo (opcional)",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(side="left")

        self.eo_anexo_label = ctk.CTkLabel(
            anexo_row, text="Nenhum arquivo",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="e"
        )
        self.eo_anexo_label.pack(side="right", padx=(0, 8))

        ctk.CTkButton(
            anexo_row, text="Selecionar",
            font=("Segoe UI", 10), fg_color=BG_INPUT,
            hover_color=BORDER_CARD, text_color=TEXT_SECONDARY,
            height=30, corner_radius=6, width=90,
            command=self._on_eo_browse_anexo,
        ).pack(side="right", padx=(0, 4))

        ctk.CTkButton(
            anexo_row, text="Limpar",
            font=("Segoe UI", 10), fg_color="transparent",
            hover_color=BORDER_CARD, text_color=TEXT_TERTIARY,
            height=30, corner_radius=6, width=60,
            command=self._on_eo_clear_anexo,
        ).pack(side="right")

        # ======== AÇÕES ========
        act_frame = ctk.CTkFrame(content, fg_color="transparent")
        act_frame.pack(fill="x", pady=(6, 14))

        self.eo_enviar_btn = ctk.CTkButton(
            act_frame, text="\u2191  Gerar Rascunhos no Outlook",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            height=44, corner_radius=10,
            command=self._on_eo_gerar_emails,
        )
        self.eo_enviar_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            act_frame, text="Limpar Tudo",
            font=("Segoe UI", 11),
            fg_color="#6b7280", hover_color="#555d6a",
            height=44, corner_radius=10, width=120,
            command=self._on_eo_limpar,
        ).pack(side="left")

        # ======== STATUS ========
        self.eo_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.eo_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.eo_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.eo_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.eo_status_dot.pack(side="left")

        self.eo_status_text = ctk.CTkLabel(
            si, text="  Carregue a lista de emails e configure o envio",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.eo_status_text.pack(side="left")

        # Internal state
        self._eo_lista_path = None
        self._eo_destinatarios = []
        self._eo_anexo_path = None

        # Compatibilidade — referências ao primeiro bloco de ordem
        if self._eo_ordens_blocks:
            b0 = self._eo_ordens_blocks[0]
            self.eo_ativo = b0["ativo"]
            self.eo_quantidade = b0["quantidade"]
            self.eo_financeiro = b0["financeiro"]
            self.eo_cotação = b0["cotacao"]

        return page

    # -----------------------------------------------------------------
    #  ENVIO DE ORDENS: Blocos de ordem dinâmicos
    # -----------------------------------------------------------------
    def _eo_add_ordem_block(self):
        """Adiciona um bloco de 'Dados da Ordem' ao container."""
        idx = len(self._eo_ordens_blocks)

        block_frame = ctk.CTkFrame(self._eo_ordens_container, fg_color="transparent")
        block_frame.pack(fill="x", pady=(0, 4))

        # Header com label e botão remover (exceto o primeiro)
        header = ctk.CTkFrame(block_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(
            header, text=f"Dados da Ordem  #{idx + 1}",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(side="left")

        remove_btn = None
        if idx > 0:
            remove_btn = ctk.CTkButton(
                header, text="Remover",
                font=("Segoe UI", 9), fg_color=ACCENT_RED, hover_color="#c0392b",
                height=24, corner_radius=6, width=70,
            )
            remove_btn.pack(side="right")

        # Campos: Ativo, Quantidade, Financeiro, Cotação
        ordem_row = ctk.CTkFrame(block_frame, fg_color="transparent")
        ordem_row.pack(fill="x", pady=(0, 6))
        ordem_row.columnconfigure((0, 1, 2, 3), weight=1)

        ativo_f = ctk.CTkFrame(ordem_row, fg_color="transparent")
        ativo_f.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        ctk.CTkLabel(ativo_f, text="Ativo", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        ativo_entry = ctk.CTkEntry(ativo_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Ex: XPML11")
        ativo_entry.pack(fill="x", pady=(2, 0))

        qtd_f = ctk.CTkFrame(ordem_row, fg_color="transparent")
        qtd_f.grid(row=0, column=1, sticky="nsew", padx=(6, 6))
        ctk.CTkLabel(qtd_f, text="Quantidade", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        qtd_entry = ctk.CTkEntry(qtd_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Ex: 100")
        qtd_entry.pack(fill="x", pady=(2, 0))

        fin_f = ctk.CTkFrame(ordem_row, fg_color="transparent")
        fin_f.grid(row=0, column=2, sticky="nsew", padx=(6, 6))
        ctk.CTkLabel(fin_f, text="Financeiro", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        fin_entry = ctk.CTkEntry(fin_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Ex: 10000.00")
        fin_entry.pack(fill="x", pady=(2, 0))

        cot_f = ctk.CTkFrame(ordem_row, fg_color="transparent")
        cot_f.grid(row=0, column=3, sticky="nsew", padx=(6, 0))
        ctk.CTkLabel(cot_f, text="Cota\u00e7\u00e3o", font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w").pack(fill="x")
        cot_entry = ctk.CTkEntry(cot_f, font=("Segoe UI", 12), height=38, corner_radius=8, border_color=BORDER_CARD, placeholder_text="Opcional")
        cot_entry.pack(fill="x", pady=(2, 0))

        block_data = {
            "frame": block_frame,
            "ativo": ativo_entry,
            "quantidade": qtd_entry,
            "financeiro": fin_entry,
            "cotacao": cot_entry,
        }
        self._eo_ordens_blocks.append(block_data)

        if remove_btn:
            remove_btn.configure(command=lambda bd=block_data: self._eo_remove_ordem_block(bd))

    def _eo_remove_ordem_block(self, block_data):
        """Remove um bloco de ordem do container."""
        if block_data in self._eo_ordens_blocks:
            block_data["frame"].destroy()
            self._eo_ordens_blocks.remove(block_data)
            # Renumerar headers
            for i, b in enumerate(self._eo_ordens_blocks):
                for widget in b["frame"].winfo_children():
                    if isinstance(widget, ctk.CTkFrame):
                        for child in widget.winfo_children():
                            if isinstance(child, ctk.CTkLabel) and "Dados da Ordem" in str(child.cget("text")):
                                child.configure(text=f"Dados da Ordem  #{i + 1}")
                                break

    # -----------------------------------------------------------------
    #  ENVIO DE ORDENS: Ações
    # -----------------------------------------------------------------
    def _on_eo_browse_lista(self):
        path = filedialog.askopenfilename(
            title="Selecionar planilha de emails",
            initialdir=os.path.join(BASE_DIR, "BASE"),
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]

            # Detectar colunas pela header (row 1)
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(1, col).value
                if val:
                    headers[str(val).strip().lower()] = col

            # Tentar encontrar colunas de nome e email
            col_nome = None
            col_email = None
            for key, col in headers.items():
                if not col_nome and any(k in key for k in ["nome", "assessor", "name"]):
                    col_nome = col
                if not col_email and any(k in key for k in ["email", "e-mail", "mail"]):
                    col_email = col

            if not col_email:
                # Fallback: primeira coluna texto, segunda com @
                for r in range(2, min(10, ws.max_row + 1)):
                    for c in range(1, ws.max_column + 1):
                        v = str(ws.cell(r, c).value or "")
                        if "@" in v and not col_email:
                            col_email = c
                        elif v and not col_nome and "@" not in v:
                            col_nome = c

            if not col_email:
                wb.close()
                messagebox.showwarning(
                    "Coluna não encontrada",
                    "Não foi possível encontrar uma coluna de e-mail na planilha.\n\n"
                    "A planilha deve ter uma coluna com header 'Email' ou conter enderecos de e-mail."
                )
                return

            destinatarios = []
            for r in range(2, ws.max_row + 1):
                email = str(ws.cell(r, col_email).value or "").strip()
                if not email or "@" not in email:
                    continue
                nome = sanitize_text(ws.cell(r, col_nome).value) if col_nome else ""
                destinatarios.append({"nome": nome, "email": email})

            wb.close()

            if not destinatarios:
                messagebox.showwarning("Lista vazia", "Nenhum e-mail válido encontrado na planilha.")
                return

            self._eo_lista_path = path
            self._eo_destinatarios = destinatarios

            # Atualizar UI
            self.eo_lista_icon.configure(text_color=ACCENT_GREEN)
            self.eo_lista_info.configure(
                text=f"  {len(destinatarios)} destinatarios  -  {os.path.basename(path)}",
                text_color=ACCENT_GREEN
            )

            # Preview
            self.eo_preview_frame.pack(fill="x", pady=(0, 16))
            self.eo_preview_text.configure(state="normal")
            self.eo_preview_text.delete("1.0", "end")

            header_line = f"{'Nome':<35} {'Email':<40}\n"
            sep_line = f"{'-'*35} {'-'*40}\n"
            self.eo_preview_text.insert("end", header_line)
            self.eo_preview_text.insert("end", sep_line)
            for d in destinatarios[:20]:
                line = f"{d['nome']:<35} {d['email']:<40}\n"
                self.eo_preview_text.insert("end", line)
            if len(destinatarios) > 20:
                self.eo_preview_text.insert("end", f"\n... e mais {len(destinatarios) - 20} destinatarios")
            self.eo_preview_text.configure(state="disabled")

            self.eo_status_dot.configure(text_color=ACCENT_GREEN)
            self.eo_status_text.configure(text=f"  Lista carregada - {len(destinatarios)} emails prontos")

        except Exception as e:
            messagebox.showerror("Erro ao ler planilha", str(e))

    def _on_eo_browse_anexo(self):
        path = filedialog.askopenfilename(
            title="Selecionar anexo",
            filetypes=[("Todos", "*.*"), ("PDF", "*.pdf"), ("Excel", "*.xlsx"), ("Imagem", "*.png *.jpg")]
        )
        if path:
            self._eo_anexo_path = path
            self.eo_anexo_label.configure(text=os.path.basename(path), text_color=ACCENT_GREEN)

    def _on_eo_clear_anexo(self):
        self._eo_anexo_path = None
        self.eo_anexo_label.configure(text="Nenhum arquivo", text_color=TEXT_TERTIARY)

    def _on_eo_tipo_ordem_change(self, value):
        if value == "OFERTA P\u00daBLICA - RF":
            self._eo_normal_frame.pack_forget()
            self._eo_oferta_frame.pack(fill="x")
        else:
            self._eo_oferta_frame.pack_forget()
            self._eo_normal_frame.pack(fill="x")

    def _on_eo_limpar(self):
        self._eo_lista_path = None
        self._eo_destinatarios = []
        self._eo_anexo_path = None
        self.eo_lista_icon.configure(text_color=ACCENT_BLUE)
        self.eo_lista_info.configure(text="  Nenhuma lista carregada", text_color=TEXT_SECONDARY)
        self.eo_preview_frame.pack_forget()
        # Limpar campos manuais e CC
        self.eo_manual_nome.delete(0, "end")
        self.eo_manual_email.delete(0, "end")
        self.eo_cc.delete(0, "end")
        # Remover blocos de ordem extras e limpar o primeiro
        while len(self._eo_ordens_blocks) > 1:
            b = self._eo_ordens_blocks[-1]
            b["frame"].destroy()
            self._eo_ordens_blocks.pop()
        if self._eo_ordens_blocks:
            b0 = self._eo_ordens_blocks[0]
            b0["ativo"].delete(0, "end")
            b0["quantidade"].delete(0, "end")
            b0["financeiro"].delete(0, "end")
            b0["cotacao"].delete(0, "end")
        self.eo_venda_ativo.delete(0, "end")
        self.eo_venda_quantidade.delete(0, "end")
        self.eo_venda_financeiro.delete(0, "end")
        self.eo_venda_cotacao.delete(0, "end")
        self.eo_of_oferta.delete(0, "end")
        self.eo_of_prospecto.delete(0, "end")
        self.eo_of_serie.delete(0, "end")
        self.eo_of_valor.delete(0, "end")
        self.eo_of_taxa.delete(0, "end")
        self.eo_of_duration.delete(0, "end")
        self.eo_of_codigo.delete(0, "end")
        self.eo_of_vinculado.set("N\u00c3O")
        self.eo_assunto.delete(0, "end")
        self.eo_corpo.delete("1.0", "end")
        self.eo_anexo_label.configure(text="Nenhum arquivo", text_color=TEXT_TERTIARY)
        self.eo_modo.set("Individual")
        self.eo_tipo_ativo.set("Fundos")
        self.eo_tipo_ordem.set("Normal")
        self._on_eo_tipo_ordem_change("Normal")
        self.eo_status_dot.configure(text_color=TEXT_TERTIARY)
        self.eo_status_text.configure(text="  Carregue a lista de emails e configure o envio")

    def _on_eo_gerar_emails(self):
        # Resolver destinatários: lista carregada OU email manual
        destinatarios = self._eo_destinatarios
        if not destinatarios:
            manual_email = self.eo_manual_email.get().strip()
            if not manual_email or "@" not in manual_email:
                messagebox.showwarning("Sem destinatario", "Insira a lista de emails ou digite um e-mail manualmente.")
                return
            manual_nome = self.eo_manual_nome.get().strip()
            destinatarios = [{"nome": manual_nome, "email": manual_email}]

        # CC
        cc_text = self.eo_cc.get().strip()

        is_oferta = self.eo_tipo_ordem.get() == "OFERTA P\u00daBLICA - RF"

        if is_oferta:
            # Modo oferta publica — gerar assunto e corpo automaticamente
            of_oferta = self.eo_of_oferta.get().strip()
            of_prospecto = self.eo_of_prospecto.get().strip()
            of_serie = self.eo_of_serie.get().strip()
            of_valor = self.eo_of_valor.get().strip()
            of_taxa = self.eo_of_taxa.get().strip()
            of_duration = self.eo_of_duration.get().strip()
            of_vinculado = self.eo_of_vinculado.get()
            of_codigo = self.eo_of_codigo.get().strip()

            campos_faltando = []
            if not of_oferta:
                campos_faltando.append("Oferta")
            if not of_valor:
                campos_faltando.append("Valor da Reserva")
            if campos_faltando:
                messagebox.showwarning("Campos obrigat\u00f3rios", "Preencha:\n\n- " + "\n- ".join(campos_faltando))
                return

            assunto = self.eo_assunto.get().strip()
            if not assunto:
                assunto = f"Oferta P\u00fablica {of_oferta}"

            corpo_manual = self.eo_corpo.get("1.0", "end").strip()

            ordens_lista = []
        else:
            # Coletar dados de todos os blocos de ordem
            ordens_lista = []
            for b in self._eo_ordens_blocks:
                od = {
                    "ativo": b["ativo"].get().strip(),
                    "quantidade": b["quantidade"].get().strip(),
                    "financeiro": b["financeiro"].get().strip(),
                    "cotação": b["cotacao"].get().strip(),
                }
                # Incluir apenas blocos com pelo menos o ativo preenchido
                if od["ativo"]:
                    ordens_lista.append(od)

            assunto = self.eo_assunto.get().strip()
            corpo_manual = self.eo_corpo.get("1.0", "end").strip()

            if not assunto:
                messagebox.showwarning("Campo obrigat\u00f3rio", "Preencha o assunto do e-mail.")
                return
            if not corpo_manual:
                messagebox.showwarning("Campo obrigat\u00f3rio", "Preencha o corpo do e-mail.")
                return

        modo = self.eo_modo.get()
        tipo = self.eo_tipo_ativo.get()
        n = len(destinatarios)

        if modo == "Grupo":
            msg = f"Ser\u00e1 criado 1 rascunho com {n} destinat\u00e1rios em c\u00f3pia.\nOutlook precisa estar aberto. Continuar?"
        else:
            msg = f"Ser\u00e3o criados {n} rascunhos individuais.\nOutlook precisa estar aberto. Continuar?"

        if not messagebox.askyesno("Gerar Rascunhos", msg):
            return

        self.eo_enviar_btn.configure(state="disabled")
        self.eo_status_dot.configure(text_color=ACCENT_ORANGE)
        self.eo_status_text.configure(text="  Gerando rascunhos...")

        if is_oferta:
            oferta_dados = {
                "oferta": of_oferta,
                "prospecto": of_prospecto,
                "serie": of_serie,
                "valor": of_valor,
                "taxa": of_taxa,
                "duration": of_duration,
                "vinculado": of_vinculado,
                "codigo": of_codigo,
            }
            # Manter compatibilidade: ordem_dados vazio
            ordem_dados = {"ativo": "", "quantidade": "", "financeiro": "", "cotação": ""}
            venda_dados = {}
            threading.Thread(
                target=self._run_eo_emails,
                args=(assunto, corpo_manual, modo, tipo, ordem_dados, venda_dados, oferta_dados),
                kwargs={"destinatarios_override": destinatarios, "cc_text": cc_text},
                daemon=True
            ).start()
        else:
            # Para compatibilidade, ordem_dados = primeiro bloco (ou vazio)
            if ordens_lista:
                ordem_dados = ordens_lista[0]
            else:
                ordem_dados = {"ativo": "", "quantidade": "", "financeiro": "", "cotação": ""}
            venda_dados = {
                "ativo": self.eo_venda_ativo.get().strip(),
                "quantidade": self.eo_venda_quantidade.get().strip(),
                "financeiro": self.eo_venda_financeiro.get().strip(),
                "cotacao": self.eo_venda_cotacao.get().strip(),
            }
            # Ordens extras (a partir do 2o bloco)
            ordens_extras = ordens_lista[1:] if len(ordens_lista) > 1 else []
            threading.Thread(
                target=self._run_eo_emails,
                args=(assunto, corpo_manual, modo, tipo, ordem_dados, venda_dados),
                kwargs={"ordens_extras": ordens_extras, "destinatarios_override": destinatarios, "cc_text": cc_text},
                daemon=True
            ).start()

    def _run_eo_emails(self, assunto, corpo, modo, tipo, ordem_dados, venda_dados=None, oferta_dados=None, ordens_extras=None, destinatarios_override=None, cc_text=None):
        try:
            import win32com.client as win32
            import pythoncom
            pythoncom.CoInitialize()

            outlook = win32.Dispatch("Outlook.Application")

            # Usar destinatários passados (podem vir do manual ou da lista)
            destinatarios_envio = destinatarios_override if destinatarios_override else self._eo_destinatarios

            hoje = datetime.now().strftime("%d/%m/%Y")
            logo_tag = LOGO_TAG_CID if os.path.exists(LOGO_PATH) else ""

            def _get_primeiro(d):
                """Extrai primeiro nome do campo nome ou, como fallback, do email."""
                if d.get("nome"):
                    return d["nome"].split()[0].capitalize()
                local = d["email"].split("@")[0]        # nicolas.kersul
                parte = local.split(".")[0]             # nicolas
                return parte.capitalize()               # Nicolas

            _hdr_style = 'padding:5px 12px;color:#00785a;font-weight:bold;border-bottom:1.5px solid #00785a;font-size:8.5pt;'
            _val_style = 'padding:5px 12px;color:#1a1a2e;font-size:9.5pt;border-bottom:1px solid #eef1ef;'
            _val_bold = 'padding:5px 12px;color:#1a1a2e;font-size:9.5pt;font-weight:bold;border-bottom:1px solid #eef1ef;'

            # Montar lista completa de ordens (primeiro bloco + extras)
            todas_ordens = []
            if ordem_dados.get("ativo"):
                todas_ordens.append(ordem_dados)
            if ordens_extras:
                todas_ordens.extend(ordens_extras)

            # Gerar tabela HTML com todos os ativos (múltiplas linhas)
            def _build_ordem_table(ordens_list):
                if not ordens_list:
                    return ""
                # Determinar quais colunas existem em pelo menos uma ordem
                has_ativo = any(o.get("ativo") for o in ordens_list)
                has_qtd = any(o.get("quantidade") for o in ordens_list)
                has_fin = any(o.get("financeiro") for o in ordens_list)
                has_cot = any(o.get("cotação") for o in ordens_list)

                cols_def = []
                if has_ativo:
                    cols_def.append(("Ativo", "ativo", True))
                if has_qtd:
                    cols_def.append(("Quantidade", "quantidade", False))
                if has_fin:
                    cols_def.append(("Financeiro", "financeiro", False))
                if has_cot:
                    cols_def.append(("Cota\u00e7\u00e3o", "cotação", False))

                if not cols_def:
                    return ""

                hdr_cells = "".join(f'<td style="{_hdr_style}">{c[0]}</td>' for c in cols_def)
                rows_html = ""
                for ordem in ordens_list:
                    val_cells = "".join(
                        f'<td style="{_val_bold if c[2] else _val_style}">{ordem.get(c[1], "")}</td>' for c in cols_def
                    )
                    rows_html += f'<tr style="background:#ffffff;">{val_cells}</tr>'

                return (
                    f'<table cellpadding="0" cellspacing="0" border="0" '
                    f'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;margin:10px 0 6px 0;">'
                    f'<tr style="background:#f7faf9;">{hdr_cells}</tr>'
                    f'{rows_html}'
                    f'</table>'
                )

            ordem_table = _build_ordem_table(todas_ordens)

            # Tabela de venda (opcional) — só colunas preenchidas
            venda_table = ""
            if venda_dados:
                _vhdr_style = 'padding:5px 12px;color:#b85c00;font-weight:bold;border-bottom:1.5px solid #e6832a;font-size:8.5pt;'
                _venda_cols = []
                v_ativo = venda_dados.get("ativo", "")
                v_qtd = venda_dados.get("quantidade", "")
                v_fin = venda_dados.get("financeiro", "")
                v_cot = venda_dados.get("cotacao", "")
                if v_ativo:
                    _venda_cols.append(("Ativo", v_ativo, True))
                if v_qtd:
                    _venda_cols.append(("Quantidade", v_qtd, False))
                if v_fin:
                    _venda_cols.append(("Financeiro", v_fin, False))
                if v_cot:
                    _venda_cols.append(("Cota\u00e7\u00e3o", v_cot, False))

                if _venda_cols:
                    vhdr_cells = "".join(f'<td style="{_vhdr_style}">{c[0]}</td>' for c in _venda_cols)
                    vval_cells = "".join(
                        f'<td style="{_val_bold if c[2] else _val_style}">{c[1]}</td>' for c in _venda_cols
                    )
                    venda_table = (
                        f'<table cellpadding="0" cellspacing="0" border="0" '
                        f'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;margin:10px 0 6px 0;">'
                        f'<tr style="background:#fff3e8;">{vhdr_cells}</tr>'
                        f'<tr style="background:#ffffff;">{vval_cells}</tr>'
                        f'</table>'
                    )

            # === OFERTA PUBLICA: montar corpo e tabelas especificas ===
            oferta_section = ""
            if oferta_dados and oferta_dados.get("oferta"):
                od = oferta_dados

                prosp_line = ""
                if od.get("prospecto"):
                    prosp_line = (
                        f'<tr><td style="padding:4px 14px;font-size:10pt;color:#1a1a2e;">'
                        f'- Prospecto: <a href="{od["prospecto"]}" '
                        f'style="color:#1863DC;text-decoration:underline;">{od["prospecto"]}</a></td></tr>'
                    )

                serie_header = ""
                if od.get("serie"):
                    serie_header = (
                        f'<tr><td style="padding:12px 14px 4px 14px;font-size:10.5pt;color:#004d33;'
                        f'font-weight:bold;border-top:1.5px solid #00b876;">{od["serie"]}:</td></tr>'
                    )

                dur_line = ""
                if od.get("duration"):
                    dur_line = f'<tr><td style="padding:4px 14px;font-size:10pt;color:#1a1a2e;">- Duration: {od["duration"]}</td></tr>'

                cod_html = ""
                if od.get("codigo"):
                    cod_html = f' no c\u00f3digo <b>{od["codigo"]}</b> na XP'

                oferta_section = f"""
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding='0' cellspacing='0' border='0'><tr>
      <td style='background-color:#00b876;width:4px;border-radius:4px;'>&nbsp;</td>
      <td style='padding-left:12px;'>
        <span style='font-size:12.5pt;color:#004d33;font-weight:bold;letter-spacing:0.3px;'>Reserva</span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:0 4px;">
    <table cellpadding="0" cellspacing="0" border="0"
     style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;margin:10px 0 6px 0;">
      <tr><td style="padding:4px 14px;font-size:10pt;color:#1a1a2e;">- Oferta: {od["oferta"]}</td></tr>
      {prosp_line}
      {serie_header}
      <tr><td style="padding:4px 14px;font-size:10pt;color:#1a1a2e;">- Valor da Reserva: <b>{od["valor"]}</b></td></tr>
      <tr><td style="padding:4px 14px;font-size:10pt;color:#1a1a2e;">- Taxa: {od.get("taxa", "-")}</td></tr>
      {dur_line}
      <tr><td style="padding:4px 14px;font-size:10pt;color:#1a1a2e;">- Investidor Vinculado (S/N): <b>{od.get("vinculado", "N\u00c3O")}</b></td></tr>
    </table>
  </td></tr>"""

                # Se não há corpo manual, gerar texto padrão
                if not corpo:
                    corpo = (
                        f"Gostaria de confirmar a altera\u00e7\u00e3o da reserva de compra{cod_html}:"
                    )

                # Forçar ordem/venda vazias no modo oferta
                ordem_table = ""
                venda_table = ""

            def _build_html(primeiro_nome, corpo_html):
                return f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">
    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#4b5563;margin-top:8px;">
      {corpo_html}
    </p>
  </td></tr>

  {"" if not ordem_table else f"""
  <!-- Dados da Ordem (Compra) -->
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding='0' cellspacing='0' border='0'><tr>
      <td style='background-color:#00b876;width:4px;border-radius:4px;'>&nbsp;</td>
      <td style='padding-left:12px;'>
        <span style='font-size:12.5pt;color:#004d33;font-weight:bold;letter-spacing:0.3px;'>Dados da Ordem</span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style='padding:0 4px;'>
    {ordem_table}
  </td></tr>"""}
  {"" if not venda_table else f"""
  <!-- Dados da Venda (Saída) -->
  <tr><td style="padding:18px 0 8px 0;">
    <table cellpadding='0' cellspacing='0' border='0'><tr>
      <td style='background-color:#e6832a;width:4px;border-radius:4px;'>&nbsp;</td>
      <td style='padding-left:12px;'>
        <span style='font-size:12.5pt;color:#b85c00;font-weight:bold;letter-spacing:0.3px;'>Ativo de Sa&iacute;da (Venda)</span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style='padding:0 4px;'>
    {venda_table}
  </td></tr>"""}
  {oferta_section}
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">Somus Capital | Produtos &middot; {hoje}</span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">E-mail gerado automaticamente</span>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</div>
"""

            criados = 0
            erros = 0

            if modo == "Grupo":
                # Um único email com todos em To
                try:
                    primeiro = "Prezados"
                    corpo_html = corpo.replace("{nome}", primeiro).replace("\n", "<br>")
                    mail = outlook.CreateItem(0)
                    mail.To = "; ".join(d["email"] for d in destinatarios_envio)
                    if cc_text:
                        mail.CC = cc_text
                    mail.Subject = assunto
                    mail.HTMLBody = _build_html(primeiro, corpo_html)
                    _attach_logo_cid(mail)
                    if self._eo_anexo_path and os.path.exists(self._eo_anexo_path):
                        mail.Attachments.Add(os.path.abspath(self._eo_anexo_path))
                    mail.Save()
                    criados = 1
                except Exception:
                    erros = 1
            else:
                # Individual
                for d in destinatarios_envio:
                    primeiro = _get_primeiro(d)
                    corpo_html = corpo.replace("{nome}", primeiro).replace("\n", "<br>")
                    try:
                        mail = outlook.CreateItem(0)
                        mail.To = d["email"]
                        if cc_text:
                            mail.CC = cc_text
                        mail.Subject = assunto
                        mail.HTMLBody = _build_html(primeiro, corpo_html)
                        _attach_logo_cid(mail)
                        if self._eo_anexo_path and os.path.exists(self._eo_anexo_path):
                            mail.Attachments.Add(os.path.abspath(self._eo_anexo_path))
                        mail.Save()
                        criados += 1
                    except Exception:
                        erros += 1

            pythoncom.CoUninitialize()

            def _done():
                self.eo_enviar_btn.configure(state="normal")
                self.eo_status_dot.configure(text_color="#00a86b")
                self.eo_status_text.configure(
                    text=f"  Concluído - {criados} rascunho(s), {erros} erro(s)"
                )
                messagebox.showinfo("Rascunhos Criados",
                    f"{criados} rascunho(s) criado(s) no Outlook!\n{erros} erro(s).")
            self.after(0, _done)

        except Exception as e:
            def _err():
                self.eo_enviar_btn.configure(state="normal")
                self.eo_status_dot.configure(text_color=ACCENT_RED)
                self.eo_status_text.configure(text=f"  Erro: {e}")
                messagebox.showerror("Erro", str(e))
            self.after(0, _err)

    # -----------------------------------------------------------------
    #  PAGE: ORDEM MASSA
    # -----------------------------------------------------------------
    def _build_ordem_massa_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Ordem MASSA", subtitle="Mesa de Produtos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Banner
        banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=44)
        banner.pack(fill="x", pady=(0, 18))
        banner.pack_propagate(False)

        ctk.CTkLabel(
            banner, text="  \u21c8  Envio de ordens em massa a partir de planilha",
            font=("Segoe UI", 12), text_color=TEXT_WHITE, anchor="w"
        ).pack(side="left", padx=18, pady=10)

        # ======== PLANILHA ========
        ctk.CTkLabel(
            content, text="Planilha de Opera\u00e7\u00f5es",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 6))

        self.om_lista_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=12,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.om_lista_card.pack(fill="x", pady=(0, 16))

        lista_inner = ctk.CTkFrame(self.om_lista_card, fg_color="transparent")
        lista_inner.pack(fill="x", padx=20, pady=16)

        lista_left = ctk.CTkFrame(lista_inner, fg_color="transparent")
        lista_left.pack(side="left", fill="x", expand=True)

        self.om_lista_icon = ctk.CTkLabel(
            lista_left, text="\u2709",
            font=("Segoe UI", 20), text_color=ACCENT_BLUE
        )
        self.om_lista_icon.pack(side="left")

        self.om_lista_info = ctk.CTkLabel(
            lista_left, text="  Nenhuma planilha carregada",
            font=("Segoe UI", 12), text_color=TEXT_SECONDARY, anchor="w"
        )
        self.om_lista_info.pack(side="left", padx=(6, 0))

        ctk.CTkButton(
            lista_inner, text="  Carregar Planilha",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=38, corner_radius=8, width=200,
            command=self._on_om_browse_planilha,
        ).pack(side="right")

        # Preview da planilha carregada
        self.om_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.om_preview_frame.pack(fill="x", pady=(0, 16))
        self.om_preview_frame.pack_forget()

        self.om_preview_text = ctk.CTkTextbox(
            self.om_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=200, wrap="none"
        )
        self.om_preview_text.pack(fill="x", padx=4, pady=4)
        self.om_preview_text.configure(state="disabled")

        # ======== CONFIGURA\u00c7\u00d5ES ========
        ctk.CTkLabel(
            content, text="Configura\u00e7\u00f5es do Envio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        config_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                   border_width=1, border_color=BORDER_CARD)
        config_card.pack(fill="x", pady=(0, 16))

        ci = ctk.CTkFrame(config_card, fg_color="transparent")
        ci.pack(fill="x", padx=20, pady=16)

        # Assessor
        ctk.CTkLabel(
            ci, text="Nome do Assessor",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.om_assessor = ctk.CTkEntry(
            ci, font=("Segoe UI", 12), height=40,
            corner_radius=8, border_color=BORDER_CARD,
            placeholder_text="Ex: Bruno Torres"
        )
        self.om_assessor.pack(fill="x", pady=(0, 14))

        # Assunto
        ctk.CTkLabel(
            ci, text="Assunto do E-mail",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.om_assunto = ctk.CTkEntry(
            ci, font=("Segoe UI", 12), height=40,
            corner_radius=8, border_color=BORDER_CARD,
            placeholder_text="Ex: Confirma\u00e7\u00e3o de Ordem - Renda Fixa"
        )
        self.om_assunto.pack(fill="x", pady=(0, 14))

        # Corpo
        ctk.CTkLabel(
            ci, text="Corpo do E-mail",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        ctk.CTkLabel(
            ci, text="Use {assessor} para inserir o nome do assessor automaticamente",
            font=("Segoe UI", 9), text_color=TEXT_TERTIARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.om_corpo = ctk.CTkTextbox(
            ci, font=("Segoe UI", 11), height=120,
            corner_radius=8, border_width=1, border_color=BORDER_CARD,
            fg_color=BG_PRIMARY, text_color=TEXT_PRIMARY, wrap="word"
        )
        self.om_corpo.pack(fill="x", pady=(0, 14))
        self.om_corpo.insert("1.0", "Conforme alinhado com o assessor {assessor}, pe\u00e7o a confirma\u00e7\u00e3o de aplica\u00e7\u00e3o do ativo:")

        # ======== A\u00c7\u00d5ES ========
        act_frame = ctk.CTkFrame(content, fg_color="transparent")
        act_frame.pack(fill="x", pady=(6, 14))

        self.om_enviar_btn = ctk.CTkButton(
            act_frame, text="\u21c8  Gerar Rascunhos no Outlook",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            height=44, corner_radius=10,
            command=self._on_om_gerar_emails,
        )
        self.om_enviar_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            act_frame, text="Limpar Tudo",
            font=("Segoe UI", 11),
            fg_color="#6b7280", hover_color="#555d6a",
            height=44, corner_radius=10, width=120,
            command=self._on_om_limpar,
        ).pack(side="left")

        # ======== STATUS ========
        self.om_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.om_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.om_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.om_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.om_status_dot.pack(side="left")

        self.om_status_text = ctk.CTkLabel(
            si, text="  Carregue a planilha de opera\u00e7\u00f5es e configure o envio",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.om_status_text.pack(side="left")

        # Internal state
        self._om_planilha_path = None
        self._om_clientes = []  # Lista de dicts: {codigo, email, nome, ordens: [{ativo, cotação, financeiro}]}

        return page

    # -----------------------------------------------------------------
    #  ORDEM MASSA: Handlers
    # -----------------------------------------------------------------
    def _on_om_browse_planilha(self):
        path = filedialog.askopenfilename(
            title="Selecionar planilha de opera\u00e7\u00f5es",
            initialdir=os.path.expanduser("~/Downloads"),
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]

            # Detectar colunas pela header (row 1)
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(1, col).value
                if val:
                    headers[str(val).strip().lower()] = col

            # Mapear colunas
            col_codigo = None
            col_nome = None
            col_email = None
            col_ativo = None
            col_taxa = None
            col_valor = None

            for key, col in headers.items():
                if not col_codigo and any(k in key for k in ["c\u00f3digo", "codigo"]):
                    col_codigo = col
                if not col_nome and any(k in key for k in ["cliente", "nome"]):
                    col_nome = col
                if not col_email and any(k in key for k in ["email", "e-mail", "mail"]):
                    col_email = col
                if not col_ativo and any(k in key for k in ["ativo", "produto"]):
                    col_ativo = col
                if not col_taxa and any(k in key for k in ["taxa", "cota\u00e7\u00e3o", "cotacao"]):
                    col_taxa = col
                if not col_valor and any(k in key for k in ["valor", "financeiro"]):
                    col_valor = col

            if not col_email:
                wb.close()
                messagebox.showwarning(
                    "Coluna n\u00e3o encontrada",
                    "N\u00e3o foi poss\u00edvel encontrar a coluna de e-mail na planilha.\n\n"
                    "A planilha deve ter colunas: C\u00f3digo, Cliente, EMAIL, ATIVO, TAXA, Valor."
                )
                return

            clientes = {}
            for r in range(2, ws.max_row + 1):
                email_val = ws.cell(r, col_email).value if col_email else None
                if not email_val or "@" not in str(email_val):
                    continue

                email = str(email_val).strip()
                codigo_raw = ws.cell(r, col_codigo).value if col_codigo else ""
                if isinstance(codigo_raw, (int, float)):
                    codigo = str(int(codigo_raw))
                else:
                    codigo = str(codigo_raw or "").strip()

                nome_raw = sanitize_text(ws.cell(r, col_nome).value) if col_nome else ""
                ativo = str(ws.cell(r, col_ativo).value or "").strip() if col_ativo else ""
                taxa = str(ws.cell(r, col_taxa).value or "").strip() if col_taxa else ""
                valor_raw = ws.cell(r, col_valor).value if col_valor else 0

                valor_num = float(valor_raw) if isinstance(valor_raw, (int, float)) else 0
                valor_fmt = f"R$ {valor_num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

                key = (codigo, email)
                if key not in clientes:
                    primeiro_nome = nome_raw.split()[0].capitalize() if nome_raw else email.split("@")[0].split(".")[0].capitalize()
                    clientes[key] = {
                        "codigo": codigo,
                        "email": email,
                        "nome": primeiro_nome,
                        "nome_completo": nome_raw,
                        "ordens": [],
                    }

                ordem = {"ativo": ativo, "financeiro": valor_fmt}
                if taxa:
                    ordem["cota\u00e7\u00e3o"] = taxa
                clientes[key]["ordens"].append(ordem)

            wb.close()

            clientes_list = list(clientes.values())
            if not clientes_list:
                messagebox.showwarning("Planilha vazia", "Nenhuma opera\u00e7\u00e3o v\u00e1lida encontrada na planilha.")
                return

            self._om_planilha_path = path
            self._om_clientes = clientes_list

            total_ordens = sum(len(c["ordens"]) for c in clientes_list)

            # Atualizar UI
            self.om_lista_icon.configure(text_color=ACCENT_GREEN)
            self.om_lista_info.configure(
                text=f"  {len(clientes_list)} clientes, {total_ordens} ordens  -  {os.path.basename(path)}",
                text_color=ACCENT_GREEN
            )

            # Preview
            self.om_preview_frame.pack(fill="x", pady=(0, 16))
            self.om_preview_text.configure(state="normal")
            self.om_preview_text.delete("1.0", "end")

            header_line = f"{'C\u00f3digo':<12} {'Cliente':<35} {'Email':<40} {'Ativo':<25} {'Taxa':<18} {'Financeiro':<16}\n"
            sep_line = f"{'-'*12} {'-'*35} {'-'*40} {'-'*25} {'-'*18} {'-'*16}\n"
            self.om_preview_text.insert("end", header_line)
            self.om_preview_text.insert("end", sep_line)

            count = 0
            for c in clientes_list:
                for o in c["ordens"]:
                    line = f"{c['codigo']:<12} {c['nome_completo'][:34]:<35} {c['email'][:39]:<40} {o.get('ativo', '')[:24]:<25} {o.get('cota\u00e7\u00e3o', '')[:17]:<18} {o.get('financeiro', ''):<16}\n"
                    self.om_preview_text.insert("end", line)
                    count += 1
                    if count >= 30:
                        break
                if count >= 30:
                    break

            remaining = total_ordens - count
            if remaining > 0:
                self.om_preview_text.insert("end", f"\n... e mais {remaining} opera\u00e7\u00f5es")
            self.om_preview_text.configure(state="disabled")

            self.om_status_dot.configure(text_color=ACCENT_GREEN)
            self.om_status_text.configure(text=f"  Planilha carregada - {len(clientes_list)} clientes, {total_ordens} ordens prontas")

        except Exception as e:
            messagebox.showerror("Erro ao ler planilha", str(e))

    def _on_om_limpar(self):
        self._om_planilha_path = None
        self._om_clientes = []
        self.om_lista_icon.configure(text_color=ACCENT_BLUE)
        self.om_lista_info.configure(text="  Nenhuma planilha carregada", text_color=TEXT_SECONDARY)
        self.om_preview_frame.pack_forget()
        self.om_assessor.delete(0, "end")
        self.om_assunto.delete(0, "end")
        self.om_corpo.delete("1.0", "end")
        self.om_corpo.insert("1.0", "Conforme alinhado com o assessor {assessor}, pe\u00e7o a confirma\u00e7\u00e3o de aplica\u00e7\u00e3o do ativo:")
        self.om_status_dot.configure(text_color=TEXT_TERTIARY)
        self.om_status_text.configure(text="  Carregue a planilha de opera\u00e7\u00f5es e configure o envio")

    def _on_om_gerar_emails(self):
        if not self._om_clientes:
            messagebox.showwarning("Sem dados", "Carregue uma planilha de opera\u00e7\u00f5es primeiro.")
            return

        assessor = self.om_assessor.get().strip()
        assunto = self.om_assunto.get().strip()
        corpo_template = self.om_corpo.get("1.0", "end").strip()

        if not assessor:
            messagebox.showwarning("Campo obrigat\u00f3rio", "Preencha o nome do assessor.")
            return
        if not assunto:
            messagebox.showwarning("Campo obrigat\u00f3rio", "Preencha o assunto do e-mail.")
            return
        if not corpo_template:
            messagebox.showwarning("Campo obrigat\u00f3rio", "Preencha o corpo do e-mail.")
            return

        n = len(self._om_clientes)
        msg = f"Ser\u00e3o criados {n} rascunhos individuais (um por cliente).\nOutlook precisa estar aberto. Continuar?"
        if not messagebox.askyesno("Gerar Rascunhos", msg):
            return

        self.om_enviar_btn.configure(state="disabled")
        self.om_status_dot.configure(text_color=ACCENT_ORANGE)
        self.om_status_text.configure(text="  Gerando rascunhos...")

        threading.Thread(
            target=self._run_om_emails,
            args=(assunto, corpo_template, assessor),
            daemon=True,
        ).start()

    def _run_om_emails(self, assunto, corpo_template, assessor):
        try:
            import win32com.client as win32
            import pythoncom
            pythoncom.CoInitialize()

            outlook = win32.Dispatch("Outlook.Application")

            hoje = datetime.now().strftime("%d/%m/%Y")
            logo_tag = LOGO_TAG_CID if os.path.exists(LOGO_PATH) else ""

            _hdr_style = 'padding:5px 12px;color:#00785a;font-weight:bold;border-bottom:1.5px solid #00785a;font-size:8.5pt;'
            _val_style = 'padding:5px 12px;color:#1a1a2e;font-size:9.5pt;border-bottom:1px solid #eef1ef;'
            _val_bold = 'padding:5px 12px;color:#1a1a2e;font-size:9.5pt;font-weight:bold;border-bottom:1px solid #eef1ef;'

            def _build_ordem_table(ordens_list):
                if not ordens_list:
                    return ""
                has_ativo = any(o.get("ativo") for o in ordens_list)
                has_cot = any(o.get("cota\u00e7\u00e3o") for o in ordens_list)
                has_fin = any(o.get("financeiro") for o in ordens_list)

                cols_def = []
                if has_ativo:
                    cols_def.append(("Ativo", "ativo", True))
                if has_cot:
                    cols_def.append(("Cota\u00e7\u00e3o", "cota\u00e7\u00e3o", False))
                if has_fin:
                    cols_def.append(("Financeiro", "financeiro", False))

                if not cols_def:
                    return ""

                hdr_cells = "".join(f'<td style="{_hdr_style}">{c[0]}</td>' for c in cols_def)
                rows_html = ""
                for ordem in ordens_list:
                    val_cells = "".join(
                        f'<td style="{_val_bold if c[2] else _val_style}">{ordem.get(c[1], "")}</td>' for c in cols_def
                    )
                    rows_html += f'<tr style="background:#ffffff;">{val_cells}</tr>'

                return (
                    f'<table cellpadding="0" cellspacing="0" border="0" '
                    f'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;margin:10px 0 6px 0;">'
                    f'<tr style="background:#f7faf9;">{hdr_cells}</tr>'
                    f'{rows_html}'
                    f'</table>'
                )

            def _build_html(primeiro_nome, corpo_html, ordem_table):
                return f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">
    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#4b5563;margin-top:8px;">
      {corpo_html}
    </p>
  </td></tr>

  {"" if not ordem_table else f'''
  <!-- Dados da Ordem (Compra) -->
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="background-color:#00b876;width:4px;border-radius:4px;">&nbsp;</td>
      <td style="padding-left:12px;">
        <span style="font-size:12.5pt;color:#004d33;font-weight:bold;letter-spacing:0.3px;">Dados da Ordem</span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:0 4px;">
    {ordem_table}
  </td></tr>'''}

  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">Somus Capital | Produtos &middot; {hoje}</span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">E-mail gerado automaticamente</span>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</div>
"""

            criados = 0
            erros = 0

            for cliente in self._om_clientes:
                try:
                    primeiro = cliente["nome"]
                    corpo_final = corpo_template.replace("{assessor}", assessor).replace("{nome}", primeiro)
                    corpo_html = corpo_final.replace("\n", "<br>")

                    ordem_table = _build_ordem_table(cliente["ordens"])

                    html = _build_html(primeiro, corpo_html, ordem_table)

                    mail = outlook.CreateItem(0)
                    mail.To = cliente["email"]
                    mail.Subject = f"{assunto} - C\u00f3d. {cliente['codigo']}"
                    mail.HTMLBody = html
                    _attach_logo_cid(mail)
                    mail.Save()

                    criados += 1
                except Exception:
                    erros += 1

            pythoncom.CoUninitialize()

            def _done():
                self.om_enviar_btn.configure(state="normal")
                self.om_status_dot.configure(text_color="#00a86b")
                self.om_status_text.configure(
                    text=f"  Conclu\u00eddo - {criados} rascunho(s), {erros} erro(s)"
                )
                messagebox.showinfo("Rascunhos Criados",
                    f"{criados} rascunho(s) criado(s) no Outlook!\n{erros} erro(s).")
            self.after(0, _done)

        except Exception as e:
            def _err():
                self.om_enviar_btn.configure(state="normal")
                self.om_status_dot.configure(text_color=ACCENT_RED)
                self.om_status_text.configure(text=f"  Erro: {e}")
                messagebox.showerror("Erro", str(e))
            self.after(0, _err)

    # -----------------------------------------------------------------
    #  PAGE: ENVIO MESA
    # -----------------------------------------------------------------
    def _build_envio_mesa_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Envio Mesa", subtitle="Mesa de Produtos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ======== DROP ZONE ========
        self.em_drop_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=14,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.em_drop_card.pack(fill="x", pady=(0, 16))
        self.em_drop_card.bind("<Button-1>", lambda e: self._on_em_browse())

        drop_inner = ctk.CTkFrame(self.em_drop_card, fg_color="transparent")
        drop_inner.pack(fill="x", padx=30, pady=36)
        drop_inner.bind("<Button-1>", lambda e: self._on_em_browse())

        self.em_icon_label = ctk.CTkLabel(
            drop_inner, text="\u2709",
            font=("Segoe UI", 44), text_color=ACCENT_BLUE
        )
        self.em_icon_label.pack()
        self.em_icon_label.bind("<Button-1>", lambda e: self._on_em_browse())

        self.em_drop_title = ctk.CTkLabel(
            drop_inner, text="Clique para selecionar a planilha de dados",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        )
        self.em_drop_title.pack(pady=(8, 2))
        self.em_drop_title.bind("<Button-1>", lambda e: self._on_em_browse())

        self.em_drop_sub = ctk.CTkLabel(
            drop_inner, text="Detecta automaticamente: assessor, datas, valores (R$), taxas (%) e quantidades",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.em_drop_sub.pack()
        self.em_drop_sub.bind("<Button-1>", lambda e: self._on_em_browse())

        self.em_browse_btn = ctk.CTkButton(
            drop_inner, text="  Selecionar Arquivo",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=40, corner_radius=8, width=200,
            command=self._on_em_browse,
        )
        self.em_browse_btn.pack(pady=(14, 0))

        # ======== FILE INFO BAR ========
        self.em_file_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.em_file_frame.pack(fill="x", pady=(0, 12))
        self.em_file_frame.pack_forget()

        fi = ctk.CTkFrame(self.em_file_frame, fg_color="transparent")
        fi.pack(fill="x", padx=16, pady=10)

        self.em_file_icon = ctk.CTkLabel(
            fi, text="\u25cf", font=("Segoe UI", 12), text_color=ACCENT_GREEN
        )
        self.em_file_icon.pack(side="left")

        self.em_file_name = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY
        )
        self.em_file_name.pack(side="left", padx=(6, 0))

        self.em_file_info = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 10), text_color=TEXT_SECONDARY
        )
        self.em_file_info.pack(side="left", padx=(12, 0))

        ctk.CTkButton(
            fi, text="Trocar", font=("Segoe UI", 10),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=28, corner_radius=6, width=70,
            command=self._on_em_browse,
        ).pack(side="right")

        # ======== PREVIEW ========
        self.em_preview_label = ctk.CTkLabel(
            content, text="Preview dos Dados",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        self.em_preview_label.pack(fill="x", pady=(4, 8))

        self.em_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        self.em_preview_frame.pack(fill="x", pady=(0, 14))

        self.em_preview_text = ctk.CTkTextbox(
            self.em_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=200, wrap="none"
        )
        self.em_preview_text.pack(fill="x", padx=4, pady=4)
        self.em_preview_text.insert("1.0", "  Nenhum arquivo carregado...")
        self.em_preview_text.configure(state="disabled")

        # ======== CONFIGURAÇÕES ========
        ctk.CTkLabel(
            content, text="Configuracoes do Envio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        config_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                   border_width=1, border_color=BORDER_CARD)
        config_card.pack(fill="x", pady=(0, 16))

        ci = ctk.CTkFrame(config_card, fg_color="transparent")
        ci.pack(fill="x", padx=20, pady=16)

        # Assunto
        ctk.CTkLabel(
            ci, text="Assunto do E-mail",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.em_assunto = ctk.CTkEntry(
            ci, font=("Segoe UI", 12), height=40,
            corner_radius=8, border_color=BORDER_CARD,
            placeholder_text="Ex: Sugestao de troca CDI11 - JURO11"
        )
        self.em_assunto.pack(fill="x", pady=(0, 14))

        # Corpo
        ctk.CTkLabel(
            ci, text="Corpo do E-mail (texto antes da tabela)",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.em_corpo = ctk.CTkTextbox(
            ci, font=("Segoe UI", 11), height=120,
            corner_radius=8, border_width=1, border_color=BORDER_CARD,
            fg_color=BG_PRIMARY, text_color=TEXT_PRIMARY, wrap="word"
        )
        self.em_corpo.pack(fill="x", pady=(0, 14))

        # Remetente
        ctk.CTkLabel(
            ci, text="Enviar em nome de",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        self.em_remetente = ctk.CTkEntry(
            ci, font=("Segoe UI", 12), height=40,
            corner_radius=8, border_color=BORDER_CARD,
        )
        self.em_remetente.pack(fill="x", pady=(0, 0))
        self.em_remetente.insert(0, "oportunidades@somuscapital.com.br")

        # ======== ACTION BUTTONS ========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(12, 12))

        self.em_process_btn = ctk.CTkButton(
            btn_frame, text="\u2191  Gerar Rascunhos no Outlook",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=self._darken(ACCENT_GREEN),
            height=44, corner_radius=10, state="disabled",
            command=self._on_em_gerar,
        )
        self.em_process_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="Limpar Tudo",
            font=("Segoe UI", 11),
            fg_color="#6b7280", hover_color="#555d6a",
            height=44, corner_radius=10, width=120,
            command=self._on_em_limpar,
        ).pack(side="left")

        # ======== STATUS BAR ========
        self.em_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.em_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.em_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.em_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.em_status_dot.pack(side="left")

        self.em_status_text = ctk.CTkLabel(
            si, text="  Aguardando arquivo...",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.em_status_text.pack(side="left")

        # Internal state
        self._em_input_path = None
        self._em_headers = []
        self._em_data_rows = []
        self._em_assessor_groups = {}

        return page

    # -----------------------------------------------------------------
    #  ENVIO MESA: Ações
    # -----------------------------------------------------------------
    def _on_em_browse(self):
        path = filedialog.askopenfilename(
            title="Selecionar Planilha de Dados",
            initialdir=os.path.join(BASE_DIR, "Mesa Produtos", "Envio Mesa"),
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")]
        )
        if path:
            self._em_load_file(path)

    def _em_load_file(self, path):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]

            headers = []
            data_rows = []

            # Ler primeira linha para detectar headers
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                headers = [str(c).strip() if c else f"Col{i+1}" for i, c in enumerate(first_row)]

            # Se a primeira célula parece ser um código de assessor (ex: A41072), não tem header
            has_header = False
            if first_row and first_row[0]:
                val = str(first_row[0]).strip()
                # Se começa com letra e contem texto descritivo (header)
                if any(kw in val.lower() for kw in ["codigo", "código", "assessor", "cod"]):
                    has_header = True

            start_row = 2 if has_header else 1
            if not has_header:
                # Sem header, gerar nomes genéricos
                ncols = len(first_row) if first_row else 0
                headers = ["Cod Assessor", "Nome Assessor"] + [f"Col{i+1}" for i in range(2, ncols)]

            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
                if not row or not row[0]:
                    continue
                data_rows.append(list(row))

            wb.close()

            if not data_rows:
                messagebox.showwarning("Planilha vazia", "Nenhum dado encontrado na planilha.")
                return

            self._em_input_path = path
            self._em_headers = headers
            self._em_data_rows = data_rows

            # Agrupar por assessor (col 0 = código, col 1 = nome)
            groups = {}
            for row in data_rows:
                code = str(row[0]).strip() if row[0] else ""
                nome = sanitize_text(row[1]) if len(row) > 1 and row[1] else ""
                key = code
                if key not in groups:
                    groups[key] = {"nome": nome, "rows": []}
                groups[key]["rows"].append(row)
            self._em_assessor_groups = groups

            # Update UI
            fname = os.path.basename(path)
            self.em_drop_title.configure(text=fname)
            self.em_drop_sub.configure(text="Arquivo carregado - clique para trocar")
            self.em_icon_label.configure(text_color=ACCENT_GREEN, text="\u2713")
            self.em_drop_card.configure(border_color=ACCENT_GREEN, fg_color="#f0fff4")

            self.em_file_name.configure(text=fname)
            info = f"{len(data_rows)} linhas  |  {len(headers)} colunas  |  {len(groups)} assessores"
            self.em_file_info.configure(text=info)
            self.em_file_frame.pack(fill="x", pady=(0, 12))

            self._em_show_preview()
            self.em_process_btn.configure(state="normal")

            self.em_status_dot.configure(text_color=ACCENT_BLUE)
            self.em_status_text.configure(text=f"  {len(groups)} assessores identificados. Configure e gere os rascunhos.")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler arquivo:\n{e}")

    def _em_show_preview(self):
        self.em_preview_text.configure(state="normal")
        self.em_preview_text.delete("1.0", "end")

        # Header
        header_line = "  |  ".join(str(h)[:20] for h in self._em_headers)
        self.em_preview_text.insert("end", header_line + "\n")
        self.em_preview_text.insert("end", "-" * len(header_line) + "\n")

        # Data rows (max 20) — formata datas e valores para preview legivel
        def _preview_cell(val):
            if val is None:
                return ""
            if isinstance(val, datetime):
                return val.strftime("%d/%m/%Y")
            if hasattr(val, "strftime") and not isinstance(val, (int, float)):
                return val.strftime("%d/%m/%Y")
            if isinstance(val, float):
                if abs(val) >= 100:
                    return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                return f"{val:.4f}".replace(".", ",")
            return str(val)[:20]
        for row in self._em_data_rows[:20]:
            cells = [_preview_cell(c) for c in row]
            self.em_preview_text.insert("end", "  |  ".join(cells) + "\n")
        if len(self._em_data_rows) > 20:
            self.em_preview_text.insert("end", f"\n... e mais {len(self._em_data_rows) - 20} linhas")

        self.em_preview_text.configure(state="disabled")

    def _on_em_limpar(self):
        self._em_input_path = None
        self._em_headers = []
        self._em_data_rows = []
        self._em_assessor_groups = {}
        self.em_icon_label.configure(text_color=ACCENT_BLUE, text="\u2709")
        self.em_drop_title.configure(text="Clique para selecionar a planilha de dados")
        self.em_drop_sub.configure(text="Detecta automaticamente: assessor, datas, valores (R$), taxas (%) e quantidades")
        self.em_drop_card.configure(border_color=ACCENT_BLUE, fg_color="#f0f4ff")
        self.em_file_frame.pack_forget()
        self.em_preview_text.configure(state="normal")
        self.em_preview_text.delete("1.0", "end")
        self.em_preview_text.insert("1.0", "  Nenhum arquivo carregado...")
        self.em_preview_text.configure(state="disabled")
        self.em_assunto.delete(0, "end")
        self.em_corpo.delete("1.0", "end")
        self.em_process_btn.configure(state="disabled")
        self.em_status_dot.configure(text_color=TEXT_TERTIARY)
        self.em_status_text.configure(text="  Aguardando arquivo...")

    def _on_em_gerar(self):
        if not self._em_assessor_groups:
            messagebox.showwarning("Aviso", "Carregue uma planilha primeiro.")
            return

        assunto = self.em_assunto.get().strip()
        corpo = self.em_corpo.get("1.0", "end").strip()

        if not assunto:
            messagebox.showwarning("Campo obrigatorio", "Preencha o assunto do e-mail.")
            return

        n = len(self._em_assessor_groups)
        msg = (f"Serao criados ate {n} rascunhos (1 por assessor).\n\n"
               f"Assunto: {assunto}\n"
               f"Outlook precisa estar aberto. Continuar?")

        if not messagebox.askyesno("Gerar Rascunhos", msg):
            return

        self.em_process_btn.configure(state="disabled")
        self.em_status_dot.configure(text_color=ACCENT_ORANGE)
        self.em_status_text.configure(text="  Gerando rascunhos...")

        # Limpa preview para usar como log
        self.em_preview_text.configure(state="normal")
        self.em_preview_text.delete("1.0", "end")
        self.em_preview_text.configure(state="disabled")

        remetente = self.em_remetente.get().strip()
        threading.Thread(
            target=self._run_em_emails,
            args=(assunto, corpo, remetente),
            daemon=True
        ).start()

    def _em_append_log(self, msg):
        self.em_preview_text.configure(state="normal")
        self.em_preview_text.insert("end", msg + "\n")
        self.em_preview_text.see("end")
        self.em_preview_text.configure(state="disabled")

    def _run_em_emails(self, assunto, corpo, remetente):
        try:
            import win32com.client as win32
            import pythoncom
            pythoncom.CoInitialize()

            outlook = win32.Dispatch("Outlook.Application")

            # Carregar base de emails
            self.after(0, self._em_append_log, "Carregando base de emails...")
            emails_map = {}
            if os.path.exists(BASE_FILE):
                wb = openpyxl.load_workbook(BASE_FILE, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                    if row[0]:
                        code = str(row[0]).strip()
                        code_a = f"A{code}" if not code.startswith("A") else code
                        code_num = code[1:] if code.startswith("A") else code
                        info = {
                            "nome": sanitize_text(row[1]) if row[1] else "",
                            "email": str(row[2]).strip() if row[2] else "",
                        }
                        emails_map[code] = info
                        emails_map[code_a] = info
                        emails_map[code_num] = info
                wb.close()
                self.after(0, self._em_append_log, f"  {len(emails_map)} registros de email carregados.")
            else:
                self.after(0, self._em_append_log, f"  BASE EMAILS nao encontrado em {BASE_FILE}")

            # CC fixos
            cc_fixos = "artur.brito@somuscapital.com.br;leonardo.dellatorre@somuscapital.com.br"

            hoje = datetime.now().strftime("%d/%m/%Y")
            logo_b64 = ""
            if os.path.exists(LOGO_PATH):
                with open(LOGO_PATH, "rb") as f:
                    logo_b64 = base64.b64encode(f.read()).decode()
            logo_tag = (
                f'<img src="data:image/png;base64,{logo_b64}" width="155" height="38"'
                f' style="vertical-align:middle;margin-right:16px;" alt="Somus Capital">'
            ) if logo_b64 else ""

            corpo_html = corpo.replace("\n", "<br>") if corpo else ""

            # Headers para tabela (excluir col 0=código e col 1=nome assessor)
            table_headers = self._em_headers[2:] if len(self._em_headers) > 2 else self._em_headers

            criados = 0
            sem_email = 0
            erros = 0

            self.after(0, self._em_append_log, f"\nProcessando {len(self._em_assessor_groups)} assessores...")
            self.after(0, self._em_append_log, "-" * 60)

            for code, group in sorted(self._em_assessor_groups.items()):
                nome = group["nome"]
                rows = group["rows"]

                # Buscar email
                email_info = emails_map.get(code, {})
                if not email_info:
                    code_num = code[1:] if code.startswith("A") else code
                    email_info = emails_map.get(code_num, {})
                email_to = email_info.get("email", "")

                if not email_to or "@" not in email_to:
                    self.after(0, self._em_append_log, f"  {code}  {nome}:  Sem email cadastrado")
                    sem_email += 1
                    continue

                primeiro_nome = nome.split()[0].capitalize() if nome else "Prezado(a)"

                # Montar tabela HTML (esconder colunas vazias)
                visible_cols = []
                for ci_idx, h in enumerate(table_headers):
                    has_data = any(
                        r[ci_idx + 2] is not None and str(r[ci_idx + 2]).strip() != ""
                        for r in rows if len(r) > ci_idx + 2
                    )
                    if has_data:
                        visible_cols.append((ci_idx + 2, h))

                table_html = '<table cellpadding="0" cellspacing="0" border="0" '
                table_html += 'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;font-size:9pt;margin-bottom:6px;">\n'

                # Header
                table_html += '<tr>'
                for col_idx, label in visible_cols:
                    table_html += (
                        f'<td style="padding:3px 8px 4px 8px;font-weight:bold;color:#00785a;'
                        f'font-size:8.5pt;border-bottom:1.5px solid #00785a;'
                        f'white-space:nowrap;">{sanitize_text(str(label))}</td>'
                    )
                table_html += '</tr>\n'

                # Detectar tipo de cada coluna visivel pelo header
                _MONEY_KW = ("valor", "posic", "posi\u00e7", "financ", "pago", "total", "preco", "pre\u00e7o", "saldo", "montante", "aporte", "resgate")
                _RATE_KW = ("taxa", "rate", "spread", "cupom", "juros", "yield")
                _DATE_KW = ("data ", "data_", "vencimento", "venc", "emissao", "emiss\u00e3o", "aplicacao", "aplica\u00e7")
                _QTY_KW = ("qtd", "quantidade", "quant", "lote")

                def _col_type(lbl):
                    """Detecta tipo da coluna: 'money', 'rate', 'date', 'qty', 'text'."""
                    lo = str(lbl).lower().strip()
                    if any(k in lo for k in _DATE_KW):
                        return "date"
                    if any(k in lo for k in _RATE_KW):
                        return "rate"
                    if any(k in lo for k in _MONEY_KW):
                        return "money"
                    if any(k in lo for k in _QTY_KW):
                        return "qty"
                    return "text"

                col_types = {ci: _col_type(lb) for ci, lb in visible_cols}

                def _fmt_cell(val, ctype):
                    """Formata valor de celula de acordo com o tipo da coluna."""
                    if val is None or (isinstance(val, str) and val.strip() == ""):
                        return ("-", "#1a1a2e", "left")
                    # datetime → DD/MM/YYYY
                    if isinstance(val, datetime):
                        return (val.strftime("%d/%m/%Y"), "#1a1a2e", "center")
                    # date sem hora
                    if hasattr(val, "strftime") and not isinstance(val, (int, float)):
                        return (val.strftime("%d/%m/%Y"), "#1a1a2e", "center")
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        if ctype == "rate":
                            formatted = f"{val:.2f}".replace(".", ",") + "%"
                            return (formatted, "#1a1a2e", "right")
                        if ctype == "money":
                            color = "#c0392b" if val < 0 else "#1a1a2e"
                            neg = val < 0
                            v = abs(round(val, 2))
                            inteiro = int(v)
                            cents = round((v - inteiro) * 100)
                            if cents >= 100:
                                inteiro += 1
                                cents = 0
                            s_int = f"{inteiro:,}".replace(",", ".")
                            formatted = f"R$ {s_int},{cents:02d}"
                            if neg:
                                formatted = f"-{formatted}"
                            return (formatted, color, "right")
                        if ctype == "qty":
                            if val == int(val):
                                return (f"{int(val):,}".replace(",", "."), "#1a1a2e", "right")
                            return (f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."), "#1a1a2e", "right")
                        # Tipo 'text' ou 'date' com numero — detectar automaticamente
                        if ctype == "date":
                            return (str(val), "#1a1a2e", "center")
                        # Numero generico grande → BRL, pequeno → decimal
                        if abs(val) >= 100:
                            color = "#c0392b" if val < 0 else "#1a1a2e"
                            neg = val < 0
                            v = abs(round(val, 2))
                            inteiro = int(v)
                            cents = round((v - inteiro) * 100)
                            if cents >= 100:
                                inteiro += 1
                                cents = 0
                            s_int = f"{inteiro:,}".replace(",", ".")
                            formatted = f"R$ {s_int},{cents:02d}"
                            if neg:
                                formatted = f"-{formatted}"
                            return (formatted, color, "right")
                        return (f"{val:,.4f}".replace(",", "X").replace(".", ",").replace("X", "."), "#1a1a2e", "right")
                    return (sanitize_text(str(val)), "#1a1a2e", "left")

                # Data rows
                for ri, row in enumerate(rows):
                    bg = "#f7faf9" if ri % 2 == 0 else "#ffffff"
                    table_html += f'<tr style="background:{bg};">'
                    for col_idx, label in visible_cols:
                        val = row[col_idx] if len(row) > col_idx else ""
                        ctype = col_types.get(col_idx, "text")
                        formatted, color, align = _fmt_cell(val, ctype)
                        table_html += (
                            f'<td style="padding:2px 8px;text-align:{align};color:{color};'
                            f'border-bottom:1px solid #eef1ef;font-size:8.5pt;'
                            f'white-space:nowrap;">{formatted}</td>'
                        )
                    table_html += '</tr>\n'
                table_html += '</table>\n'

                # Montar HTML do email
                email_html = f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">
    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    {f'<p style="font-size:10.5pt;color:#4b5563;margin-top:8px;">{corpo_html}</p>' if corpo_html else ''}
  </td></tr>
  <tr><td style="padding:12px 4px 0 4px;">
    {table_html}
  </td></tr>
  <tr><td style="padding:16px 4px 0 4px;">
    <p style="font-size:10.5pt;color:#1a1a2e;">Qualquer d&uacute;vida estou &agrave; disposi&ccedil;&atilde;o!</p>
    <p style="font-size:10.5pt;color:#1a1a2e;">Atenciosamente,</p>
  </td></tr>
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">Envio Mesa &middot; {hoje}</span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">E-mail gerado automaticamente</span>
        </td>
      </tr>
    </table>
  </td></tr>
</table>
</div>
"""

                try:
                    mail = outlook.CreateItem(0)
                    if remetente:
                        mail.SentOnBehalfOfName = remetente
                    mail.To = email_to
                    mail.CC = cc_fixos
                    mail.Subject = assunto
                    mail.HTMLBody = email_html
                    mail.Save()
                    criados += 1
                    self.after(0, self._em_append_log,
                        f"  {code}  {nome}  ->  {email_to}  [{len(rows)} linhas]")
                except Exception as e:
                    erros += 1
                    self.after(0, self._em_append_log,
                        f"  {code}  {nome}:  ERRO - {e}")

            pythoncom.CoUninitialize()

            self.after(0, self._em_append_log, "-" * 60)
            self.after(0, self._em_append_log,
                f"\nFinalizado!  {criados} rascunhos, {sem_email} sem email, {erros} erros.")

            def _done():
                self.em_process_btn.configure(state="normal")
                self.em_status_dot.configure(text_color=ACCENT_GREEN)
                self.em_status_text.configure(
                    text=f"  Concluido - {criados} rascunho(s), {sem_email} sem email, {erros} erro(s)"
                )
                messagebox.showinfo("Rascunhos Criados",
                    f"{criados} rascunho(s) criado(s) no Outlook!\n"
                    f"{sem_email} assessor(es) sem email\n{erros} erro(s).")
            self.after(0, _done)

        except Exception as e:
            def _err(msg=str(e)):
                self.em_process_btn.configure(state="normal")
                self.em_status_dot.configure(text_color=ACCENT_RED)
                self.em_status_text.configure(text=f"  Erro: {msg}")
                messagebox.showerror("Erro", msg)
            self.after(0, _err)

    # -----------------------------------------------------------------
    #  PAGE: TAREFAS (Organizador de Tarefas)
    # -----------------------------------------------------------------
    _TAREFAS_FILE = os.path.join(BASE_DIR, "BASE", "tarefas.json")

    _PRIORIDADE_CORES = {
        "Alta": "#dc3545",
        "Media": "#e6832a",
        "Baixa": "#1863DC",
    }
    _STATUS_CORES = {
        "Pendente": "#6b7280",
        "Em Andamento": "#e6832a",
        "Concluida": "#00a86b",
    }

    def _load_tarefas(self):
        import json
        if os.path.exists(self._TAREFAS_FILE):
            try:
                with open(self._TAREFAS_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return []

    def _save_tarefas(self, tarefas):
        import json
        os.makedirs(os.path.dirname(self._TAREFAS_FILE), exist_ok=True)
        with open(self._TAREFAS_FILE, "w", encoding="utf-8") as f:
            json.dump(tarefas, f, ensure_ascii=False, indent=2)

    # -----------------------------------------------------------------
    #  PAGE: ENVIO ANIVERSÁRIOS
    # -----------------------------------------------------------------
    def _build_envio_aniversarios_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Envio Aniversários", subtitle="Mesa de Produtos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Placeholder card
        card = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=14,
            border_width=1, border_color=BORDER_CARD
        )
        card.pack(fill="x", pady=(0, 16))

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=30, pady=36)

        ctk.CTkLabel(
            inner, text="\u2605",
            font=("Segoe UI", 44), text_color=ACCENT_GREEN
        ).pack()

        ctk.CTkLabel(
            inner, text="Envio de Aniversários",
            font=("Segoe UI", 20, "bold"), text_color=TEXT_PRIMARY
        ).pack(pady=(10, 4))

        ctk.CTkLabel(
            inner, text="Módulo em construção — funcionalidade será implementada em breve.",
            font=("Segoe UI", 13), text_color=TEXT_SECONDARY
        ).pack()

        return page

    def _build_tarefas_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Tarefas", subtitle="Organizador")

        # ---- Toolbar ----
        toolbar = ctk.CTkFrame(page, fg_color=BG_CARD, height=56, corner_radius=0,
                               border_width=0)
        toolbar.pack(fill="x", padx=0, pady=0)
        toolbar.pack_propagate(False)

        tb_inner = ctk.CTkFrame(toolbar, fg_color="transparent")
        tb_inner.pack(fill="x", padx=20, pady=10)

        ctk.CTkButton(
            tb_inner, text="+  Nova Tarefa",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            height=36, corner_radius=8, width=150,
            command=self._on_tf_nova,
        ).pack(side="left", padx=(0, 12))

        # Filtro status
        ctk.CTkLabel(
            tb_inner, text="Filtro:",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        ).pack(side="left", padx=(12, 4))

        self.tf_filtro_status = ctk.CTkSegmentedButton(
            tb_inner,
            values=["Todas", "Pendente", "Em Andamento", "Concluida"],
            font=("Segoe UI", 10, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_GREEN,
            selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY,
            corner_radius=8,
            height=32,
            command=lambda v: self._tf_refresh(),
        )
        self.tf_filtro_status.pack(side="left", padx=(0, 12))
        self.tf_filtro_status.set("Todas")

        # Filtro prioridade
        self.tf_filtro_prio = ctk.CTkSegmentedButton(
            tb_inner,
            values=["Todas", "Alta", "Media", "Baixa"],
            font=("Segoe UI", 10, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_BLUE,
            selected_hover_color="#1555bb",
            unselected_color=BG_INPUT,
            unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY,
            corner_radius=8,
            height=32,
            command=lambda v: self._tf_refresh(),
        )
        self.tf_filtro_prio.pack(side="left", padx=(0, 12))
        self.tf_filtro_prio.set("Todas")

        # Contador
        self.tf_counter = ctk.CTkLabel(
            tb_inner, text="0 tarefas",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY
        )
        self.tf_counter.pack(side="right")

        # ---- KPI cards ----
        kpi_frame = ctk.CTkFrame(page, fg_color="transparent")
        kpi_frame.pack(fill="x", padx=20, pady=(12, 6))
        kpi_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self.tf_kpi_total = self._tf_make_kpi(kpi_frame, "Total", "0", TEXT_PRIMARY, 0)
        self.tf_kpi_pend = self._tf_make_kpi(kpi_frame, "Pendentes", "0", "#6b7280", 1)
        self.tf_kpi_andamento = self._tf_make_kpi(kpi_frame, "Em Andamento", "0", ACCENT_ORANGE, 2)
        self.tf_kpi_concl = self._tf_make_kpi(kpi_frame, "Concluidas", "0", ACCENT_GREEN, 3)

        # ---- Scroll area com tarefas ----
        self.tf_scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        self.tf_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        self.tf_list_frame = ctk.CTkFrame(self.tf_scroll, fg_color="transparent")
        self.tf_list_frame.pack(fill="x", padx=20, pady=10)

        # Carregar e renderizar
        self._tf_tarefas = self._load_tarefas()
        self._tf_refresh()

        return page

    def _tf_make_kpi(self, parent, label, value, color, col):
        card = ctk.CTkFrame(parent, fg_color=BG_CARD, corner_radius=10,
                            border_width=1, border_color=BORDER_CARD, height=70)
        card.grid(row=0, column=col, sticky="nsew", padx=4, pady=0)
        card.grid_propagate(False)

        ctk.CTkLabel(
            card, text=label,
            font=("Segoe UI", 9, "bold"), text_color=TEXT_TERTIARY
        ).place(relx=0.5, rely=0.3, anchor="center")

        val_label = ctk.CTkLabel(
            card, text=value,
            font=("Segoe UI", 20, "bold"), text_color=color
        )
        val_label.place(relx=0.5, rely=0.68, anchor="center")

        return val_label

    def _tf_refresh(self):
        """Reconstroi a lista de tarefas na UI."""
        # Limpar
        for w in self.tf_list_frame.winfo_children():
            w.destroy()

        filtro_st = self.tf_filtro_status.get()
        filtro_pr = self.tf_filtro_prio.get()

        tarefas = self._tf_tarefas

        # KPIs
        total = len(tarefas)
        pend = sum(1 for t in tarefas if t.get("status") == "Pendente")
        andamento = sum(1 for t in tarefas if t.get("status") == "Em Andamento")
        concl = sum(1 for t in tarefas if t.get("status") == "Concluida")
        self.tf_kpi_total.configure(text=str(total))
        self.tf_kpi_pend.configure(text=str(pend))
        self.tf_kpi_andamento.configure(text=str(andamento))
        self.tf_kpi_concl.configure(text=str(concl))

        # Filtrar
        filtered = tarefas
        if filtro_st != "Todas":
            filtered = [t for t in filtered if t.get("status") == filtro_st]
        if filtro_pr != "Todas":
            filtered = [t for t in filtered if t.get("prioridade") == filtro_pr]

        # Ordenar: Pendente > Em Andamento > Concluida, depois Alta > Media > Baixa
        _st_order = {"Pendente": 0, "Em Andamento": 1, "Concluida": 2}
        _pr_order = {"Alta": 0, "Media": 1, "Baixa": 2}
        filtered.sort(key=lambda t: (
            _st_order.get(t.get("status", ""), 9),
            _pr_order.get(t.get("prioridade", ""), 9),
        ))

        self.tf_counter.configure(text=f"{len(filtered)} tarefa(s)")

        if not filtered:
            ctk.CTkLabel(
                self.tf_list_frame,
                text="Nenhuma tarefa encontrada.\nClique em '+ Nova Tarefa' para criar.",
                font=("Segoe UI", 13), text_color=TEXT_TERTIARY,
                justify="center"
            ).pack(pady=40)
            return

        for idx, tarefa in enumerate(filtered):
            self._tf_render_card(tarefa, idx)

    def _tf_render_card(self, tarefa, idx):
        """Renderiza um card de tarefa."""
        tid = tarefa.get("id", "")
        status = tarefa.get("status", "Pendente")
        prio = tarefa.get("prioridade", "Media")
        titulo = tarefa.get("titulo", "Sem titulo")
        descricao = tarefa.get("descricao", "")
        responsavel = tarefa.get("responsavel", "")
        prazo = tarefa.get("prazo", "")
        categoria = tarefa.get("categoria", "")

        is_done = status == "Concluida"

        # Card com borda esquerda colorida pela prioridade
        prio_color = self._PRIORIDADE_CORES.get(prio, ACCENT_BLUE)
        st_color = self._STATUS_CORES.get(status, "#6b7280")

        card = ctk.CTkFrame(
            self.tf_list_frame, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD if not is_done else "#d1fae5"
        )
        card.pack(fill="x", pady=(0, 6))

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=14, pady=10)

        # Row 1: Prioridade pill + Título + Status pill + Ações
        row1 = ctk.CTkFrame(inner, fg_color="transparent")
        row1.pack(fill="x")

        # Barra prioridade
        prio_bar = ctk.CTkFrame(row1, fg_color=prio_color, width=4, corner_radius=2)
        prio_bar.pack(side="left", fill="y", padx=(0, 10), pady=2)

        # Prioridade pill
        ctk.CTkLabel(
            row1, text=f" {prio} ",
            font=("Segoe UI", 8, "bold"), text_color="white",
            fg_color=prio_color, corner_radius=4, height=18
        ).pack(side="left", padx=(0, 8))

        # Titulo
        title_font = ("Segoe UI", 13, "bold") if not is_done else ("Segoe UI", 13)
        title_color = TEXT_PRIMARY if not is_done else TEXT_TERTIARY
        ctk.CTkLabel(
            row1, text=titulo,
            font=title_font, text_color=title_color, anchor="w"
        ).pack(side="left", fill="x", expand=True)

        # Status pill
        ctk.CTkLabel(
            row1, text=f" {status} ",
            font=("Segoe UI", 8, "bold"), text_color="white",
            fg_color=st_color, corner_radius=4, height=18
        ).pack(side="right", padx=(8, 0))

        # Botões de ação
        btn_frame = ctk.CTkFrame(row1, fg_color="transparent")
        btn_frame.pack(side="right")

        if status != "Concluida":
            next_status = "Em Andamento" if status == "Pendente" else "Concluida"
            next_icon = "\u25b6" if status == "Pendente" else "\u2713"
            ctk.CTkButton(
                btn_frame, text=next_icon, width=28, height=28,
                font=("Segoe UI", 12), fg_color=BG_INPUT,
                hover_color=BORDER_CARD, text_color=st_color,
                corner_radius=6,
                command=lambda t=tarefa, ns=next_status: self._on_tf_change_status(t, ns),
            ).pack(side="left", padx=2)

        ctk.CTkButton(
            btn_frame, text="\u270e", width=28, height=28,
            font=("Segoe UI", 12), fg_color=BG_INPUT,
            hover_color=BORDER_CARD, text_color=ACCENT_BLUE,
            corner_radius=6,
            command=lambda t=tarefa: self._on_tf_editar(t),
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            btn_frame, text="\u2715", width=28, height=28,
            font=("Segoe UI", 12), fg_color=BG_INPUT,
            hover_color="#fde8e8", text_color=ACCENT_RED,
            corner_radius=6,
            command=lambda t=tarefa: self._on_tf_excluir(t),
        ).pack(side="left", padx=2)

        # Row 2: Meta info
        if descricao or responsavel or prazo or categoria:
            row2 = ctk.CTkFrame(inner, fg_color="transparent")
            row2.pack(fill="x", pady=(6, 0), padx=(14, 0))

            if descricao:
                desc_text = descricao if len(descricao) <= 120 else descricao[:120] + "..."
                ctk.CTkLabel(
                    row2, text=desc_text,
                    font=("Segoe UI", 10), text_color=TEXT_SECONDARY, anchor="w",
                    wraplength=600
                ).pack(fill="x", pady=(0, 4))

            meta_row = ctk.CTkFrame(row2, fg_color="transparent")
            meta_row.pack(fill="x")

            if responsavel:
                ctk.CTkLabel(
                    meta_row, text=f"\u2022 {responsavel}",
                    font=("Segoe UI", 9), text_color=TEXT_TERTIARY
                ).pack(side="left", padx=(0, 12))

            if categoria:
                ctk.CTkLabel(
                    meta_row, text=f"\u2022 {categoria}",
                    font=("Segoe UI", 9), text_color=TEXT_TERTIARY
                ).pack(side="left", padx=(0, 12))

            if prazo:
                prazo_color = TEXT_TERTIARY
                # Highlight se vencida
                try:
                    from datetime import datetime as _dt
                    dt_prazo = _dt.strptime(prazo, "%d/%m/%Y")
                    if dt_prazo.date() < _dt.now().date() and not is_done:
                        prazo_color = ACCENT_RED
                except Exception:
                    pass
                ctk.CTkLabel(
                    meta_row, text=f"\u2022 Prazo: {prazo}",
                    font=("Segoe UI", 9, "bold"), text_color=prazo_color
                ).pack(side="left", padx=(0, 12))

    def _on_tf_nova(self):
        self._tf_open_dialog()

    def _on_tf_editar(self, tarefa):
        self._tf_open_dialog(tarefa)

    def _tf_open_dialog(self, tarefa=None):
        """Abre dialog para criar/editar tarefa."""
        is_edit = tarefa is not None

        dialog = ctk.CTkToplevel(self)
        dialog.title("Editar Tarefa" if is_edit else "Nova Tarefa")
        dialog.geometry("520x560")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 520) // 2
        y = (dialog.winfo_screenheight() - 560) // 2
        dialog.geometry(f"520x560+{x}+{y}")

        dialog.configure(fg_color=BG_PRIMARY)

        # Header
        hdr = ctk.CTkFrame(dialog, fg_color=ACCENT_GREEN, height=50, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(
            hdr, text="  Editar Tarefa" if is_edit else "  Nova Tarefa",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_WHITE, anchor="w"
        ).pack(side="left", padx=20)

        content = ctk.CTkFrame(dialog, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=24, pady=16)

        # Titulo
        ctk.CTkLabel(content, text="Titulo *", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        titulo_entry = ctk.CTkEntry(content, font=("Segoe UI", 12), height=38,
                                    corner_radius=8, border_color=BORDER_CARD)
        titulo_entry.pack(fill="x", pady=(0, 12))
        if is_edit:
            titulo_entry.insert(0, tarefa.get("titulo", ""))

        # Descricao
        ctk.CTkLabel(content, text="Descricao", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        desc_text = ctk.CTkTextbox(content, font=("Segoe UI", 11), height=80,
                                   corner_radius=8, border_width=1, border_color=BORDER_CARD,
                                   fg_color=BG_CARD, text_color=TEXT_PRIMARY, wrap="word")
        desc_text.pack(fill="x", pady=(0, 12))
        if is_edit and tarefa.get("descricao"):
            desc_text.insert("1.0", tarefa["descricao"])

        # Row: Prioridade + Status
        row1 = ctk.CTkFrame(content, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 12))
        row1.columnconfigure((0, 1), weight=1)

        prio_frame = ctk.CTkFrame(row1, fg_color="transparent")
        prio_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        ctk.CTkLabel(prio_frame, text="Prioridade", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        prio_seg = ctk.CTkSegmentedButton(
            prio_frame, values=["Alta", "Media", "Baixa"],
            font=("Segoe UI", 10, "bold"), fg_color=BG_INPUT,
            selected_color=ACCENT_BLUE, selected_hover_color="#1555bb",
            unselected_color=BG_INPUT, unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY, corner_radius=8, height=34,
        )
        prio_seg.pack(fill="x")
        prio_seg.set(tarefa.get("prioridade", "Media") if is_edit else "Media")

        status_frame = ctk.CTkFrame(row1, fg_color="transparent")
        status_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        ctk.CTkLabel(status_frame, text="Status", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        status_seg = ctk.CTkSegmentedButton(
            status_frame, values=["Pendente", "Em Andamento", "Concluida"],
            font=("Segoe UI", 10, "bold"), fg_color=BG_INPUT,
            selected_color=ACCENT_GREEN, selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color=BG_INPUT, unselected_hover_color=BORDER_CARD,
            text_color=TEXT_PRIMARY, corner_radius=8, height=34,
        )
        status_seg.pack(fill="x")
        status_seg.set(tarefa.get("status", "Pendente") if is_edit else "Pendente")

        # Row: Responsavel + Categoria
        row2 = ctk.CTkFrame(content, fg_color="transparent")
        row2.pack(fill="x", pady=(0, 12))
        row2.columnconfigure((0, 1), weight=1)

        resp_frame = ctk.CTkFrame(row2, fg_color="transparent")
        resp_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        ctk.CTkLabel(resp_frame, text="Responsavel", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        resp_entry = ctk.CTkEntry(resp_frame, font=("Segoe UI", 12), height=38,
                                  corner_radius=8, border_color=BORDER_CARD,
                                  placeholder_text="Nome")
        resp_entry.pack(fill="x")
        if is_edit and tarefa.get("responsavel"):
            resp_entry.insert(0, tarefa["responsavel"])

        cat_frame = ctk.CTkFrame(row2, fg_color="transparent")
        cat_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))

        ctk.CTkLabel(cat_frame, text="Categoria", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        cat_entry = ctk.CTkEntry(cat_frame, font=("Segoe UI", 12), height=38,
                                 corner_radius=8, border_color=BORDER_CARD,
                                 placeholder_text="Ex: Produtos, RV, RF")
        cat_entry.pack(fill="x")
        if is_edit and tarefa.get("categoria"):
            cat_entry.insert(0, tarefa["categoria"])

        # Prazo
        ctk.CTkLabel(content, text="Prazo (DD/MM/AAAA)", font=("Segoe UI", 11, "bold"),
                      text_color=TEXT_SECONDARY, anchor="w").pack(fill="x", pady=(0, 4))
        prazo_entry = ctk.CTkEntry(content, font=("Segoe UI", 12), height=38,
                                   corner_radius=8, border_color=BORDER_CARD,
                                   placeholder_text="Ex: 15/03/2026")
        prazo_entry.pack(fill="x", pady=(0, 16))
        if is_edit and tarefa.get("prazo"):
            prazo_entry.insert(0, tarefa["prazo"])

        # Botões
        btn_row = ctk.CTkFrame(content, fg_color="transparent")
        btn_row.pack(fill="x")

        def _salvar():
            titulo = titulo_entry.get().strip()
            if not titulo:
                messagebox.showwarning("Campo obrigatorio", "Preencha o titulo da tarefa.")
                return

            import uuid
            data = {
                "id": tarefa.get("id", str(uuid.uuid4())[:8]) if is_edit else str(uuid.uuid4())[:8],
                "titulo": titulo,
                "descricao": desc_text.get("1.0", "end").strip(),
                "prioridade": prio_seg.get(),
                "status": status_seg.get(),
                "responsavel": resp_entry.get().strip(),
                "categoria": cat_entry.get().strip(),
                "prazo": prazo_entry.get().strip(),
                "criado_em": tarefa.get("criado_em", datetime.now().strftime("%d/%m/%Y %H:%M")) if is_edit else datetime.now().strftime("%d/%m/%Y %H:%M"),
            }

            if is_edit:
                for i, t in enumerate(self._tf_tarefas):
                    if t.get("id") == tarefa.get("id"):
                        self._tf_tarefas[i] = data
                        break
            else:
                self._tf_tarefas.append(data)

            self._save_tarefas(self._tf_tarefas)
            self._tf_refresh()
            dialog.destroy()

        ctk.CTkButton(
            btn_row, text="Salvar",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            height=40, corner_radius=8, width=140,
            command=_salvar,
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_row, text="Cancelar",
            font=("Segoe UI", 11),
            fg_color="#6b7280", hover_color="#555d6a",
            height=40, corner_radius=8, width=100,
            command=dialog.destroy,
        ).pack(side="left")

    def _on_tf_change_status(self, tarefa, new_status):
        for t in self._tf_tarefas:
            if t.get("id") == tarefa.get("id"):
                t["status"] = new_status
                break
        self._save_tarefas(self._tf_tarefas)
        self._tf_refresh()

    def _on_tf_excluir(self, tarefa):
        if not messagebox.askyesno("Excluir Tarefa",
            f"Excluir '{tarefa.get('titulo', '')}'?\n\nEssa acao nao pode ser desfeita."):
            return
        self._tf_tarefas = [t for t in self._tf_tarefas if t.get("id") != tarefa.get("id")]
        self._save_tarefas(self._tf_tarefas)
        self._tf_refresh()

    # -----------------------------------------------------------------
    #  PAGE: CORPORATE - DASHBOARD
    # -----------------------------------------------------------------
    def _build_corp_dashboard_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Dashboard", subtitle="Corporate")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Banner
        banner = ctk.CTkFrame(content, fg_color=ACCENT_GREEN, corner_radius=10, height=50)
        banner.pack(fill="x", pady=(0, 18))
        banner.pack_propagate(False)

        ctk.CTkLabel(
            banner, text="  Corporate - Motor NASA HD",
            font=("Segoe UI", 13, "bold"), text_color=TEXT_WHITE, anchor="w"
        ).pack(side="left", padx=18, pady=10)

        ctk.CTkLabel(
            banner, text=f"v{VERSION}",
            font=("Segoe UI", 10), text_color="#b0e0c8", anchor="e"
        ).pack(side="right", padx=18, pady=10)

        # Cards de ferramentas disponiveis
        tools_frame = ctk.CTkFrame(content, fg_color="transparent")
        tools_frame.pack(fill="x", pady=(0, 18))
        tools_frame.columnconfigure((0, 1, 2, 3), weight=1)

        tools = [
            ("Simulador de Consórcio", "Calcule parcelas, fases, lances e custo efetivo",
             ACCENT_GREEN, "\u2630", lambda: self._on_nav("consorcio")),
            ("Comparativo de VPL", "Análise NASA HD: Delta VPL, break-even, TIR/CET",
             ACCENT_BLUE, "\u25b2", lambda: self._on_nav("comparativo_vpl")),
            ("Cons. vs Financiamento", "Compare consórcio e financiamento lado a lado",
             ACCENT_ORANGE, "\u21c4", lambda: self._on_nav("consorcio_vs_financ")),
            ("Fluxo de Receitas", "Acompanhe operações, pagamentos e receitas",
             ACCENT_PURPLE, "\u25b6", lambda: self._on_nav("fluxo_receitas")),
        ]

        for col, (title, desc, color, icon, cmd) in enumerate(tools):
            card = ctk.CTkFrame(tools_frame, fg_color=BG_CARD, corner_radius=12,
                                border_width=1, border_color=BORDER_CARD)
            card.grid(row=0, column=col, sticky="nsew", padx=6, pady=4)

            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="both", padx=18, pady=18)

            icon_bar = ctk.CTkFrame(inner, fg_color=color, corner_radius=8, width=40, height=40)
            icon_bar.pack(anchor="w", pady=(0, 10))
            icon_bar.pack_propagate(False)
            ctk.CTkLabel(icon_bar, text=icon, font=("Segoe UI", 18),
                         text_color=TEXT_WHITE).pack(expand=True)

            ctk.CTkLabel(inner, text=title, font=("Segoe UI", 13, "bold"),
                         text_color=TEXT_PRIMARY, anchor="w").pack(fill="x")
            ctk.CTkLabel(inner, text=desc, font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY, anchor="w", wraplength=200).pack(fill="x", pady=(4, 10))

            ctk.CTkButton(
                inner, text="Abrir", font=("Segoe UI", 11, "bold"),
                fg_color=color, hover_color=self._darken(color),
                height=32, corner_radius=8, command=cmd,
            ).pack(anchor="w")

        # Conceitos rapidos
        info_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                 border_width=1, border_color=BORDER_CARD)
        info_card.pack(fill="x", pady=(0, 18))

        info_inner = ctk.CTkFrame(info_card, fg_color="transparent")
        info_inner.pack(fill="x", padx=24, pady=18)

        ctk.CTkLabel(info_inner, text="Conceitos NASA HD",
                     font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w").pack(fill="x", pady=(0, 10))

        conceitos = [
            ("Delta VPL", "Diferença entre o valor presente do crédito e o valor presente de todos os pagamentos. Se positivo, a operação cria valor."),
            ("ALM (CDI Líquido)", "Custo de oportunidade — taxa usada para descontar fluxos pré-contemplação."),
            ("Hurdle Rate", "Retorno mínimo exigido — taxa usada para descontar parcelas pós-contemplação."),
            ("Break-even Lance", "Percentual de lance livre que zera o Delta VPL. Acima desse valor, a operação passa a não criar valor."),
            ("CET (Custo Efetivo Total)", "Taxa equivalente anual que representa o custo real da operação, incluindo todas as taxas e encargos."),
            ("TIR", "Taxa Interna de Retorno — taxa que zera o VPL do fluxo de caixa completo."),
        ]
        for title, desc in conceitos:
            row = ctk.CTkFrame(info_inner, fg_color="transparent")
            row.pack(fill="x", pady=3)
            ctk.CTkLabel(row, text=f"\u25cf  {title}:", font=("Segoe UI", 10, "bold"),
                         text_color=ACCENT_GREEN).pack(side="left")
            ctk.CTkLabel(row, text=f"  {desc}", font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY, wraplength=600, anchor="w").pack(side="left", fill="x", expand=True)

        return page

    # -----------------------------------------------------------------
    #  PAGE: CORPORATE - SIMULADOR
    # -----------------------------------------------------------------
    def _build_consorcio_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        # Top bar profissional
        self._make_topbar(
            page, "Simulador de Consórcio", subtitle="Corporate"
        )

        # Scroll area
        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ========== DADOS DO CLIENTE ==========
        ctk.CTkLabel(
            content, text="Dados do Cliente",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        client_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                   border_width=1, border_color=BORDER_CARD)
        client_card.pack(fill="x", pady=(0, 16))

        ci = ctk.CTkFrame(client_card, fg_color="transparent")
        ci.pack(fill="x", padx=20, pady=16)
        ci.columnconfigure((0, 1), weight=1)

        ctk.CTkLabel(ci, text="Nome do Cliente", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.cs_nome = ctk.CTkEntry(ci, placeholder_text="Nome completo do cliente",
                                    height=38, corner_radius=8, fg_color=BG_INPUT,
                                    border_width=1, border_color=BORDER_LIGHT)
        self.cs_nome.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(2, 0))

        ctk.CTkLabel(ci, text="Assessor Responsavel", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=1, sticky="w", padx=(10, 0))
        self.cs_assessor = ctk.CTkEntry(ci, placeholder_text="Nome do assessor",
                                        height=38, corner_radius=8, fg_color=BG_INPUT,
                                        border_width=1, border_color=BORDER_LIGHT)
        self.cs_assessor.grid(row=1, column=1, sticky="ew", padx=(10, 0), pady=(2, 0))

        # ========== PARAMETROS DO CONSÓRCIO ==========
        ctk.CTkLabel(
            content, text="Parâmetros do Consórcio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        params_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                   border_width=1, border_color=BORDER_CARD)
        params_card.pack(fill="x", pady=(0, 16))

        pi = ctk.CTkFrame(params_card, fg_color="transparent")
        pi.pack(fill="x", padx=20, pady=16)
        pi.columnconfigure((0, 1, 2), weight=1)

        # Row 0-1: Tipo + Administradora + Prazo
        ctk.CTkLabel(pi, text="Tipo do Bem", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.cs_tipo_bem = ctk.CTkOptionMenu(
            pi, values=["Imovel", "Automovel", "Moto", "Servicos", "Outros"],
            height=38, corner_radius=8, fg_color=BG_INPUT,
            button_color=ACCENT_GREEN, button_hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_PRIMARY, font=("Segoe UI", 11),
        )
        self.cs_tipo_bem.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(2, 8))

        ctk.CTkLabel(pi, text="Administradora", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=1, sticky="w", padx=(10, 10))
        self.cs_administradora = ctk.CTkEntry(
            pi, placeholder_text="Ex: Porto Seguro, Embracon...",
            height=38, corner_radius=8, fg_color=BG_INPUT,
            border_width=1, border_color=BORDER_LIGHT)
        self.cs_administradora.grid(row=1, column=1, sticky="ew", padx=(10, 10), pady=(2, 8))

        ctk.CTkLabel(pi, text="Prazo (meses)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=2, sticky="w", padx=(10, 0))
        self.cs_prazo = ctk.CTkOptionMenu(
            pi, values=["36", "48", "60", "72", "80", "100", "120", "150", "180", "200", "240"],
            height=38, corner_radius=8, fg_color=BG_INPUT,
            button_color=ACCENT_GREEN, button_hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_PRIMARY, font=("Segoe UI", 11),
        )
        self.cs_prazo.set("120")
        self.cs_prazo.grid(row=1, column=2, sticky="ew", padx=(10, 0), pady=(2, 8))

        # Row 2-3: Valor da Carta (full width)
        ctk.CTkLabel(pi, text="Valor da Carta de Crédito (R$)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=2, column=0, columnspan=3, sticky="w", pady=(4, 0))
        self.cs_valor_carta = ctk.CTkEntry(
            pi, placeholder_text="Ex: 300000", height=42, corner_radius=8,
            fg_color=BG_INPUT, border_width=1, border_color=BORDER_LIGHT,
            font=("Segoe UI", 14))
        self.cs_valor_carta.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(2, 8))

        # Row 4-5: Taxas
        ctk.CTkLabel(pi, text="Taxa de Administração (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=4, column=0, sticky="w", padx=(0, 10))
        self.cs_taxa_adm = ctk.CTkEntry(pi, placeholder_text="Ex: 18", height=38,
                                        corner_radius=8, fg_color=BG_INPUT,
                                        border_width=1, border_color=BORDER_LIGHT)
        self.cs_taxa_adm.grid(row=5, column=0, sticky="ew", padx=(0, 10), pady=(2, 0))
        self.cs_taxa_adm.insert(0, "18")

        ctk.CTkLabel(pi, text="Fundo de Reserva (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=4, column=1, sticky="w", padx=(10, 10))
        self.cs_fundo_reserva = ctk.CTkEntry(pi, placeholder_text="Ex: 2", height=38,
                                             corner_radius=8, fg_color=BG_INPUT,
                                             border_width=1, border_color=BORDER_LIGHT)
        self.cs_fundo_reserva.grid(row=5, column=1, sticky="ew", padx=(10, 10), pady=(2, 0))
        self.cs_fundo_reserva.insert(0, "2")

        ctk.CTkLabel(pi, text="Seguro Prestamista (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=4, column=2, sticky="w", padx=(10, 0))
        self.cs_seguro = ctk.CTkEntry(pi, placeholder_text="Ex: 0", height=38,
                                      corner_radius=8, fg_color=BG_INPUT,
                                      border_width=1, border_color=BORDER_LIGHT)
        self.cs_seguro.grid(row=5, column=2, sticky="ew", padx=(10, 0), pady=(2, 0))
        self.cs_seguro.insert(0, "0")

        # Row 6-7: Correção anual + Tipo
        ctk.CTkLabel(pi, text="Correção Anual (% a.a.)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=6, column=0, sticky="w", padx=(0, 10), pady=(6, 0))
        self.cs_correção_anual = ctk.CTkEntry(pi, placeholder_text="Ex: 5.5", height=38,
                                              corner_radius=8, fg_color=BG_INPUT,
                                              border_width=1, border_color=BORDER_LIGHT)
        self.cs_correção_anual.grid(row=7, column=0, sticky="ew", padx=(0, 10), pady=(2, 2))
        self.cs_correção_anual.insert(0, "0")

        ctk.CTkLabel(pi, text="Tipo de Correção", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=6, column=1, sticky="w", padx=(10, 10), pady=(6, 0))
        self.cs_tipo_correção = ctk.CTkSegmentedButton(
            pi, values=["Pre-fixado", "Pós-fixado"],
            font=("Segoe UI", 10, "bold"),
            fg_color=BG_INPUT,
            selected_color=ACCENT_BLUE,
            selected_hover_color="#1555bb",
            unselected_color="#e0e3e8",
            unselected_hover_color="#d0d3d8",
            text_color=TEXT_PRIMARY,
            corner_radius=8, height=38,
        )
        self.cs_tipo_correção.set("Pós-fixado")
        self.cs_tipo_correção.grid(row=7, column=1, sticky="ew", padx=(10, 10), pady=(2, 2))

        ctk.CTkLabel(pi, text="Indice (INCC, IPCA...)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=6, column=2, sticky="w", padx=(10, 0), pady=(6, 0))
        self.cs_índice_correção = ctk.CTkOptionMenu(
            pi, values=["INCC", "IPCA", "IGP-M", "Outro"],
            height=38, corner_radius=8, fg_color=BG_INPUT,
            button_color=ACCENT_GREEN, button_hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_PRIMARY, font=("Segoe UI", 11),
        )
        self.cs_índice_correção.grid(row=7, column=2, sticky="ew", padx=(10, 0), pady=(2, 2))

        ctk.CTkLabel(pi, text="INCC p/ imovel, IPCA p/ veiculo. 0 = sem correção",
                     font=("Segoe UI", 8), text_color=TEXT_TERTIARY
                     ).grid(row=8, column=0, columnspan=3, sticky="w", pady=(0, 0))

        # ========== CONTEMPLAÇÃO E LANCES ==========
        ctk.CTkLabel(
            content, text="Contemplação e Lances",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(8, 8))

        cl_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                               border_width=1, border_color=BORDER_CARD)
        cl_card.pack(fill="x", pady=(0, 16))

        cli = ctk.CTkFrame(cl_card, fg_color="transparent")
        cli.pack(fill="x", padx=20, pady=16)
        cli.columnconfigure((0, 1), weight=1)

        # Row 0-1: Prazo contemplação + Parcela reduzida
        ctk.CTkLabel(cli, text="Prazo de Contemplação (mes)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.cs_prazo_contemplação = ctk.CTkEntry(
            cli, placeholder_text="Ex: 60", height=38, corner_radius=8,
            fg_color=BG_INPUT, border_width=1, border_color=BORDER_LIGHT)
        self.cs_prazo_contemplação.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(2, 2))
        self.cs_prazo_contemplação.insert(0, "60")

        ctk.CTkLabel(cli, text="Parcela Reduzida (% do Fundo Comum)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=1, sticky="w", padx=(10, 0))
        self.cs_parcela_reduzida = ctk.CTkOptionMenu(
            cli, values=["100% (Integral)", "70% (Reducao 30%)", "50% (Plano 50)"],
            height=38, corner_radius=8, fg_color=BG_INPUT,
            button_color=ACCENT_GREEN, button_hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_PRIMARY, font=("Segoe UI", 11),
        )
        self.cs_parcela_reduzida.grid(row=1, column=1, sticky="ew", padx=(10, 0), pady=(2, 2))

        ctk.CTkLabel(cli, text="Mês estimado para contemplação (sorteio ou lance)",
                     font=("Segoe UI", 8), text_color=TEXT_TERTIARY
                     ).grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(0, 8))
        ctk.CTkLabel(cli, text="% do fundo comum pago antes da contemplação",
                     font=("Segoe UI", 8), text_color=TEXT_TERTIARY
                     ).grid(row=2, column=1, sticky="w", padx=(10, 0), pady=(0, 8))

        # Row 3-4: Lance livre + Lance embutido
        ctk.CTkLabel(cli, text="Lance Livre Ofertado (% da carta)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=3, column=0, sticky="w", padx=(0, 10))
        self.cs_lance_livre = ctk.CTkEntry(
            cli, placeholder_text="Ex: 20", height=38, corner_radius=8,
            fg_color=BG_INPUT, border_width=1, border_color=BORDER_LIGHT)
        self.cs_lance_livre.grid(row=4, column=0, sticky="ew", padx=(0, 10), pady=(2, 2))
        self.cs_lance_livre.insert(0, "0")

        ctk.CTkLabel(cli, text="Lance Embutido Ofertado (% da carta)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=3, column=1, sticky="w", padx=(10, 0))
        self.cs_lance_embutido = ctk.CTkEntry(
            cli, placeholder_text="Ex: 15 (max ~30%)", height=38, corner_radius=8,
            fg_color=BG_INPUT, border_width=1, border_color=BORDER_LIGHT)
        self.cs_lance_embutido.grid(row=4, column=1, sticky="ew", padx=(10, 0), pady=(2, 2))
        self.cs_lance_embutido.insert(0, "0")

        ctk.CTkLabel(cli, text="Recursos próprios para antecipar contemplação",
                     font=("Segoe UI", 8), text_color=TEXT_TERTIARY
                     ).grid(row=5, column=0, sticky="w", padx=(0, 10))
        ctk.CTkLabel(cli, text="Descontado do crédito recebido (max ~25-30%)",
                     font=("Segoe UI", 8), text_color=TEXT_TERTIARY
                     ).grid(row=5, column=1, sticky="w", padx=(10, 0))

        # ========== BOTOES DE ACAO ==========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(4, 16))

        ctk.CTkButton(
            btn_frame, text="  Calcular Simulação",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_GREEN,
            hover_color=self._darken(ACCENT_GREEN), height=44, corner_radius=10,
            command=self._on_calcular_consorcio,
        ).pack(side="left", padx=(0, 10))

        self.cs_btn_pdf = ctk.CTkButton(
            btn_frame, text="  Gerar PDF",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_BLUE,
            hover_color=self._darken(ACCENT_BLUE), height=44, corner_radius=10,
            command=self._on_gerar_pdf_consorcio, state="disabled",
        )
        self.cs_btn_pdf.pack(side="left", padx=(10, 10))

        self.cs_btn_relatório = ctk.CTkButton(
            btn_frame, text="  Gerar Relatório",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_PURPLE,
            hover_color="#6a4daa", height=44, corner_radius=10,
            command=self._on_gerar_relatorio_consorcio, state="disabled",
        )
        self.cs_btn_relatório.pack(side="left", padx=(0, 10))

        self.cs_btn_email = ctk.CTkButton(
            btn_frame, text="  Enviar Email",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_ORANGE,
            hover_color="#c96f1f", height=44, corner_radius=10,
            command=self._on_abrir_email_consorcio, state="disabled",
        )
        self.cs_btn_email.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="  Limpar",
            font=("Segoe UI", 12), fg_color="#6b7280",
            hover_color="#555d6a", height=44, corner_radius=10,
            command=self._on_limpar_consorcio,
        ).pack(side="left", padx=(10, 0))

        # ========== RESULTADO ==========
        ctk.CTkLabel(
            content, text="Resultado da Simulação",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        self.cs_result_frame = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                            border_width=1, border_color=BORDER_CARD)
        self.cs_result_frame.pack(fill="x", pady=(0, 16))

        self.cs_result_inner = ctk.CTkFrame(self.cs_result_frame, fg_color="transparent")
        self.cs_result_inner.pack(fill="x", padx=20, pady=16)

        self.cs_result_placeholder = ctk.CTkLabel(
            self.cs_result_inner,
            text="Preencha os campos acima e clique em 'Calcular Simulação'",
            font=("Segoe UI", 12), text_color=TEXT_TERTIARY
        )
        self.cs_result_placeholder.pack(pady=20)

        return page

    # -----------------------------------------------------------------
    #  PAGE: CORPORATE - COMPARATIVO DE VPL (NASA HD)
    # -----------------------------------------------------------------
    def _build_comparativo_vpl_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Comparativo de VPL", subtitle="Corporate - Análise NASA HD")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ========== PARAMETROS DO CONSORCIO ==========
        ctk.CTkLabel(
            content, text="Parâmetros do Consórcio",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        p1_card = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                               border_width=1, border_color=BORDER_CARD)
        p1_card.pack(fill="x", pady=(0, 16))

        p1 = ctk.CTkFrame(p1_card, fg_color="transparent")
        p1.pack(fill="x", padx=20, pady=16)
        p1.columnconfigure((0, 1, 2), weight=1)

        # Row 0-1: Valor da Carta + Prazo + Contemplacao
        ctk.CTkLabel(p1, text="Valor da Carta (R$)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.vpl_valor_carta = ctk.CTkEntry(p1, placeholder_text="Ex: 3600000", height=38,
                                            corner_radius=8, fg_color=BG_INPUT,
                                            border_width=1, border_color=BORDER_LIGHT)
        self.vpl_valor_carta.grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(2, 8))

        ctk.CTkLabel(p1, text="Prazo (meses)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=1, sticky="w", padx=(10, 10))
        self.vpl_prazo = ctk.CTkEntry(p1, placeholder_text="Ex: 200", height=38,
                                      corner_radius=8, fg_color=BG_INPUT,
                                      border_width=1, border_color=BORDER_LIGHT)
        self.vpl_prazo.grid(row=1, column=1, sticky="ew", padx=(10, 10), pady=(2, 8))
        self.vpl_prazo.insert(0, "200")

        ctk.CTkLabel(p1, text="Mês Contemplação", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=0, column=2, sticky="w", padx=(10, 0))
        self.vpl_contemp = ctk.CTkEntry(p1, placeholder_text="Ex: 11", height=38,
                                        corner_radius=8, fg_color=BG_INPUT,
                                        border_width=1, border_color=BORDER_LIGHT)
        self.vpl_contemp.grid(row=1, column=2, sticky="ew", padx=(10, 0), pady=(2, 8))
        self.vpl_contemp.insert(0, "11")

        # Row 2-3: Taxas
        ctk.CTkLabel(p1, text="Taxa Admin. (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=2, column=0, sticky="w", padx=(0, 10))
        self.vpl_taxa_adm = ctk.CTkEntry(p1, placeholder_text="20", height=38,
                                         corner_radius=8, fg_color=BG_INPUT,
                                         border_width=1, border_color=BORDER_LIGHT)
        self.vpl_taxa_adm.grid(row=3, column=0, sticky="ew", padx=(0, 10), pady=(2, 8))
        self.vpl_taxa_adm.insert(0, "20")

        ctk.CTkLabel(p1, text="Fundo Reserva (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=2, column=1, sticky="w", padx=(10, 10))
        self.vpl_fundo_res = ctk.CTkEntry(p1, placeholder_text="3", height=38,
                                          corner_radius=8, fg_color=BG_INPUT,
                                          border_width=1, border_color=BORDER_LIGHT)
        self.vpl_fundo_res.grid(row=3, column=1, sticky="ew", padx=(10, 10), pady=(2, 8))
        self.vpl_fundo_res.insert(0, "3")

        ctk.CTkLabel(p1, text="Seguro (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=2, column=2, sticky="w", padx=(10, 0))
        self.vpl_seguro = ctk.CTkEntry(p1, placeholder_text="0", height=38,
                                       corner_radius=8, fg_color=BG_INPUT,
                                       border_width=1, border_color=BORDER_LIGHT)
        self.vpl_seguro.grid(row=3, column=2, sticky="ew", padx=(10, 0), pady=(2, 8))
        self.vpl_seguro.insert(0, "0")

        # Row 4-5: Lances + Reducao
        ctk.CTkLabel(p1, text="Lance Embutido (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=4, column=0, sticky="w", padx=(0, 10))
        self.vpl_lance_emb = ctk.CTkEntry(p1, placeholder_text="30", height=38,
                                          corner_radius=8, fg_color=BG_INPUT,
                                          border_width=1, border_color=BORDER_LIGHT)
        self.vpl_lance_emb.grid(row=5, column=0, sticky="ew", padx=(0, 10), pady=(2, 8))
        self.vpl_lance_emb.insert(0, "30")

        ctk.CTkLabel(p1, text="Lance Livre (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=4, column=1, sticky="w", padx=(10, 10))
        self.vpl_lance_livre = ctk.CTkEntry(p1, placeholder_text="47", height=38,
                                            corner_radius=8, fg_color=BG_INPUT,
                                            border_width=1, border_color=BORDER_LIGHT)
        self.vpl_lance_livre.grid(row=5, column=1, sticky="ew", padx=(10, 10), pady=(2, 8))
        self.vpl_lance_livre.insert(0, "0")

        ctk.CTkLabel(p1, text="Parcela Reduzida (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=4, column=2, sticky="w", padx=(10, 0))
        self.vpl_red_pct = ctk.CTkEntry(p1, placeholder_text="70", height=38,
                                        corner_radius=8, fg_color=BG_INPUT,
                                        border_width=1, border_color=BORDER_LIGHT)
        self.vpl_red_pct.grid(row=5, column=2, sticky="ew", padx=(10, 0), pady=(2, 8))
        self.vpl_red_pct.insert(0, "70")

        # Row 6-7: Correcao + ALM + Hurdle
        ctk.CTkLabel(p1, text="Correção Anual (%)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=6, column=0, sticky="w", padx=(0, 10))
        self.vpl_correcao = ctk.CTkEntry(p1, placeholder_text="3", height=38,
                                         corner_radius=8, fg_color=BG_INPUT,
                                         border_width=1, border_color=BORDER_LIGHT)
        self.vpl_correcao.grid(row=7, column=0, sticky="ew", padx=(0, 10), pady=(2, 8))
        self.vpl_correcao.insert(0, "3")

        ctk.CTkLabel(p1, text="CDI Líquido / ALM (% a.a.)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=6, column=1, sticky="w", padx=(10, 10))
        self.vpl_alm = ctk.CTkEntry(p1, placeholder_text="12", height=38,
                                    corner_radius=8, fg_color=BG_INPUT,
                                    border_width=1, border_color=BORDER_LIGHT)
        self.vpl_alm.grid(row=7, column=1, sticky="ew", padx=(10, 10), pady=(2, 8))
        self.vpl_alm.insert(0, "12")

        ctk.CTkLabel(p1, text="Hurdle Rate (% a.a.)", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).grid(row=6, column=2, sticky="w", padx=(10, 0))
        self.vpl_hurdle = ctk.CTkEntry(p1, placeholder_text="12", height=38,
                                       corner_radius=8, fg_color=BG_INPUT,
                                       border_width=1, border_color=BORDER_LIGHT)
        self.vpl_hurdle.grid(row=7, column=2, sticky="ew", padx=(10, 0), pady=(2, 8))
        self.vpl_hurdle.insert(0, "12")

        ctk.CTkLabel(p1, text="ALM: custo de oportunidade (CDI líquido de IR)  |  Hurdle: retorno mínimo exigido pós-contemplação",
                     font=("Segoe UI", 8), text_color=TEXT_TERTIARY
                     ).grid(row=8, column=0, columnspan=3, sticky="w", pady=(0, 0))

        # ========== BOTAO ==========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(4, 16))

        ctk.CTkButton(
            btn_frame, text="  Analisar VPL",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_GREEN,
            hover_color=self._darken(ACCENT_GREEN), height=44, corner_radius=10,
            command=self._on_calcular_vpl,
        ).pack(side="left", padx=(0, 10))

        self.vpl_btn_pdf = ctk.CTkButton(
            btn_frame, text="  Gerar PDF VPL",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_BLUE,
            hover_color=self._darken(ACCENT_BLUE), height=44, corner_radius=10,
            command=self._on_gerar_pdf_vpl, state="disabled",
        )
        self.vpl_btn_pdf.pack(side="left", padx=(10, 10))

        ctk.CTkButton(
            btn_frame, text="  Limpar",
            font=("Segoe UI", 12), fg_color="#6b7280",
            hover_color="#555d6a", height=44, corner_radius=10,
            command=lambda: self._on_limpar_vpl(),
        ).pack(side="left", padx=(10, 0))

        # ========== RESULTADO ==========
        ctk.CTkLabel(
            content, text="Resultado da Análise",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        self.vpl_result_frame = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                             border_width=1, border_color=BORDER_CARD)
        self.vpl_result_frame.pack(fill="x", pady=(0, 16))

        self.vpl_result_inner = ctk.CTkFrame(self.vpl_result_frame, fg_color="transparent")
        self.vpl_result_inner.pack(fill="x", padx=20, pady=16)

        self.vpl_result_placeholder = ctk.CTkLabel(
            self.vpl_result_inner,
            text="Preencha os parâmetros e clique em 'Analisar VPL'",
            font=("Segoe UI", 12), text_color=TEXT_TERTIARY
        )
        self.vpl_result_placeholder.pack(pady=20)

        return page

    # -----------------------------------------------------------------
    #  PAGE: CORPORATE - CONSÓRCIO VS FINANCIAMENTO
    # -----------------------------------------------------------------
    def _build_consorcio_vs_financ_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Consórcio vs Financiamento", subtitle="Corporate - Comparativo")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # Layout: dois cards lado a lado
        cols_frame = ctk.CTkFrame(content, fg_color="transparent")
        cols_frame.pack(fill="x", pady=(0, 16))
        cols_frame.columnconfigure((0, 1), weight=1)

        # ========== COLUNA ESQUERDA: CONSORCIO ==========
        ctk.CTkLabel(
            cols_frame, text="Consórcio",
            font=("Segoe UI", 14, "bold"), text_color=ACCENT_GREEN, anchor="w"
        ).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=(0, 8))

        c_card = ctk.CTkFrame(cols_frame, fg_color=BG_CARD, corner_radius=12,
                              border_width=1, border_color=BORDER_CARD)
        c_card.grid(row=1, column=0, sticky="nsew", padx=(0, 8))

        ci = ctk.CTkFrame(c_card, fg_color="transparent")
        ci.pack(fill="x", padx=16, pady=14)
        ci.columnconfigure((0, 1), weight=1)

        # Consorcio inputs
        _cf_labels = [
            ("Valor da Carta (R$)", "cf_c_valor", "300000"),
            ("Prazo (meses)", "cf_c_prazo", "120"),
            ("Taxa Admin. (%)", "cf_c_taxa", "18"),
            ("Fundo Reserva (%)", "cf_c_fres", "2"),
            ("Seguro (%)", "cf_c_seg", "0"),
            ("Mês Contemplação", "cf_c_contemp", "60"),
            ("Lance Embutido (%)", "cf_c_lemb", "0"),
            ("Lance Livre (%)", "cf_c_lliv", "0"),
            ("Parcela Reduzida (%)", "cf_c_red", "100"),
            ("Correção Anual (%)", "cf_c_corr", "0"),
            ("CDI Líq./ALM (% a.a.)", "cf_c_alm", "12"),
        ]
        for i, (label, attr, default) in enumerate(_cf_labels):
            col = i % 2
            row = (i // 2) * 2
            ctk.CTkLabel(ci, text=label, font=("Segoe UI", 9, "bold"),
                         text_color=TEXT_SECONDARY).grid(row=row, column=col, sticky="w", padx=4)
            e = ctk.CTkEntry(ci, placeholder_text=default, height=34,
                             corner_radius=8, fg_color=BG_INPUT,
                             border_width=1, border_color=BORDER_LIGHT,
                             font=("Segoe UI", 10))
            e.grid(row=row + 1, column=col, sticky="ew", padx=4, pady=(2, 6))
            e.insert(0, default)
            setattr(self, attr, e)

        # ========== COLUNA DIREITA: FINANCIAMENTO ==========
        ctk.CTkLabel(
            cols_frame, text="Financiamento",
            font=("Segoe UI", 14, "bold"), text_color=ACCENT_BLUE, anchor="w"
        ).grid(row=0, column=1, sticky="w", padx=(10, 0), pady=(0, 8))

        f_card = ctk.CTkFrame(cols_frame, fg_color=BG_CARD, corner_radius=12,
                              border_width=1, border_color=BORDER_CARD)
        f_card.grid(row=1, column=1, sticky="nsew", padx=(8, 0))

        fi = ctk.CTkFrame(f_card, fg_color="transparent")
        fi.pack(fill="x", padx=16, pady=14)
        fi.columnconfigure((0, 1), weight=1)

        _ff_labels = [
            ("Valor Financiado (R$)", "cf_f_valor", "300000"),
            ("Prazo (meses)", "cf_f_prazo", "120"),
            ("Taxa Mensal (%)", "cf_f_taxa", "1.0"),
            ("Método", "cf_f_metodo", None),
        ]
        for i, (label, attr, default) in enumerate(_ff_labels):
            col = i % 2
            row = (i // 2) * 2
            ctk.CTkLabel(fi, text=label, font=("Segoe UI", 9, "bold"),
                         text_color=TEXT_SECONDARY).grid(row=row, column=col, sticky="w", padx=4)
            if attr == "cf_f_metodo":
                e = ctk.CTkOptionMenu(
                    fi, values=["Price", "SAC"],
                    height=34, corner_radius=8, fg_color=BG_INPUT,
                    button_color=ACCENT_BLUE, button_hover_color="#1555bb",
                    text_color=TEXT_PRIMARY, font=("Segoe UI", 10),
                )
                e.set("Price")
            else:
                e = ctk.CTkEntry(fi, placeholder_text=default, height=34,
                                 corner_radius=8, fg_color=BG_INPUT,
                                 border_width=1, border_color=BORDER_LIGHT,
                                 font=("Segoe UI", 10))
                e.insert(0, default)
            e.grid(row=row + 1, column=col, sticky="ew", padx=4, pady=(2, 6))
            setattr(self, attr, e)

        # ========== BOTAO ==========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(4, 16))

        ctk.CTkButton(
            btn_frame, text="  Comparar",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_GREEN,
            hover_color=self._darken(ACCENT_GREEN), height=44, corner_radius=10,
            command=self._on_comparar_consorcio_financ,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="  Limpar",
            font=("Segoe UI", 12), fg_color="#6b7280",
            hover_color="#555d6a", height=44, corner_radius=10,
            command=lambda: self._on_limpar_cf(),
        ).pack(side="left", padx=(10, 0))

        # ========== RESULTADO ==========
        ctk.CTkLabel(
            content, text="Resultado Comparativo",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        self.cf_result_frame = ctk.CTkFrame(content, fg_color=BG_CARD, corner_radius=12,
                                            border_width=1, border_color=BORDER_CARD)
        self.cf_result_frame.pack(fill="x", pady=(0, 16))

        self.cf_result_inner = ctk.CTkFrame(self.cf_result_frame, fg_color="transparent")
        self.cf_result_inner.pack(fill="x", padx=20, pady=16)

        ctk.CTkLabel(
            self.cf_result_inner,
            text="Preencha ambos os lados e clique em 'Comparar'",
            font=("Segoe UI", 12), text_color=TEXT_TERTIARY
        ).pack(pady=20)

        return page

    # -----------------------------------------------------------------
    #  PAGE: CORPORATE - FLUXO DE RECEITAS
    # -----------------------------------------------------------------
    _FR_FILE = os.path.join(BASE_DIR, "Corporate", "operacoes.json")

    def _fr_load(self):
        import json
        if os.path.exists(self._FR_FILE):
            try:
                with open(self._FR_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return []

    def _fr_save(self, ops):
        import json
        os.makedirs(os.path.dirname(self._FR_FILE), exist_ok=True)
        with open(self._FR_FILE, "w", encoding="utf-8") as f:
            json.dump(ops, f, ensure_ascii=False, indent=2)

    def _build_fluxo_receitas_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Fluxo de Receitas", subtitle="Corporate - Operações e Pagamentos")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ========== KPIs ==========
        kpi_frame = ctk.CTkFrame(content, fg_color="transparent")
        kpi_frame.pack(fill="x", pady=(0, 18))
        kpi_frame.columnconfigure((0, 1, 2, 3), weight=1)

        self._fr_kpi_ops = self._make_kpi_card(kpi_frame, "Operações Ativas", "--", ACCENT_GREEN, 0)
        self._fr_kpi_receita = self._make_kpi_card(kpi_frame, "Receita Total", "--", ACCENT_BLUE, 1)
        self._fr_kpi_recebido = self._make_kpi_card(kpi_frame, "Recebido", "--", ACCENT_TEAL, 2)
        self._fr_kpi_pendente = self._make_kpi_card(kpi_frame, "Pendente", "--", ACCENT_ORANGE, 3)

        # ========== BOTOES ==========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 16))

        ctk.CTkButton(
            btn_frame, text="  + Nova Operação",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_GREEN,
            hover_color=self._darken(ACCENT_GREEN), height=42, corner_radius=10,
            command=self._fr_nova_operacao,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="  Exportar Excel",
            font=("Segoe UI", 12), fg_color=ACCENT_BLUE,
            hover_color=self._darken(ACCENT_BLUE), height=42, corner_radius=10,
            command=self._fr_exportar_excel,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="  Importar Excel",
            font=("Segoe UI", 12), fg_color=ACCENT_ORANGE,
            hover_color="#c96f1f", height=42, corner_radius=10,
            command=self._fr_importar_excel,
        ).pack(side="left", padx=(0, 10))

        # Filtro de status
        self._fr_filtro = ctk.CTkSegmentedButton(
            btn_frame, values=["Todos", "Ativo", "Contemplado", "Encerrado"],
            font=("Segoe UI", 10, "bold"), fg_color=BG_INPUT,
            selected_color=ACCENT_GREEN, selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color="#e0e3e8", unselected_hover_color="#d0d3d8",
            text_color=TEXT_PRIMARY, corner_radius=8, height=38,
            command=lambda v: self._fr_refresh(),
        )
        self._fr_filtro.set("Todos")
        self._fr_filtro.pack(side="right")

        # ========== LISTA DE OPERAÇÕES ==========
        ctk.CTkLabel(
            content, text="Operações",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 8))

        self._fr_list_frame = ctk.CTkFrame(content, fg_color="transparent")
        self._fr_list_frame.pack(fill="x", pady=(0, 16))

        # Carregar dados
        self.after(200, self._fr_refresh)

        return page

    def _fr_refresh(self):
        ops = self._fr_load()
        filtro = self._fr_filtro.get() if hasattr(self, '_fr_filtro') else "Todos"

        # Filtrar
        if filtro != "Todos":
            ops_vis = [o for o in ops if o.get("status") == filtro]
        else:
            ops_vis = ops

        # KPIs
        ativos = sum(1 for o in ops if o.get("status") in ("Ativo", "Contemplado"))
        receita_total = sum(p.get("valor", 0) for o in ops for p in o.get("pagamentos", []))
        recebido = sum(p.get("valor", 0) for o in ops for p in o.get("pagamentos", []) if p.get("pago"))
        pendente = receita_total - recebido

        self._fr_kpi_ops.configure(text=str(ativos))
        self._fr_kpi_receita.configure(text=fmt_currency(receita_total))
        self._fr_kpi_recebido.configure(text=fmt_currency(recebido))
        self._fr_kpi_pendente.configure(text=fmt_currency(pendente))

        # Limpar lista
        for w in self._fr_list_frame.winfo_children():
            w.destroy()

        if not ops_vis:
            ctk.CTkLabel(
                self._fr_list_frame,
                text="Nenhuma operação encontrada. Clique em '+ Nova Operação' para adicionar.",
                font=("Segoe UI", 12), text_color=TEXT_TERTIARY
            ).pack(pady=20)
            return

        # Renderizar cards
        _STATUS_CORES = {
            "Ativo": ACCENT_GREEN,
            "Contemplado": ACCENT_BLUE,
            "Encerrado": "#6b7280",
        }

        for op in ops_vis:
            self._fr_render_card(op, _STATUS_CORES)

    def _fr_render_card(self, op, status_cores):
        oid = op.get("id", "")
        status = op.get("status", "Ativo")
        cor = status_cores.get(status, ACCENT_GREEN)
        pagamentos = op.get("pagamentos", [])
        total_parcelas = len(pagamentos)
        pagas = sum(1 for p in pagamentos if p.get("pago"))
        pendentes = total_parcelas - pagas
        valor_recebido = sum(p.get("valor", 0) for p in pagamentos if p.get("pago"))
        valor_pendente = sum(p.get("valor", 0) for p in pagamentos if not p.get("pago"))

        # Proximo vencimento
        from datetime import datetime as dt
        hoje = dt.now()
        prox_venc = None
        prox_valor = 0
        vencidas = 0
        for p in pagamentos:
            if not p.get("pago"):
                try:
                    dv = dt.strptime(p["data_venc"], "%Y-%m-%d")
                    if dv < hoje:
                        vencidas += 1
                    if prox_venc is None or dv < prox_venc:
                        prox_venc = dv
                        prox_valor = p.get("valor", 0)
                except Exception:
                    pass

        # Card
        card = ctk.CTkFrame(self._fr_list_frame, fg_color=BG_CARD, corner_radius=12,
                            border_width=1, border_color=BORDER_CARD)
        card.pack(fill="x", pady=4)

        # Barra lateral colorida
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=0, pady=0)

        color_bar = ctk.CTkFrame(inner, fg_color=cor, width=5, corner_radius=0)
        color_bar.pack(side="left", fill="y")

        body = ctk.CTkFrame(inner, fg_color="transparent")
        body.pack(side="left", fill="x", expand=True, padx=14, pady=12)

        # Header row
        hdr = ctk.CTkFrame(body, fg_color="transparent")
        hdr.pack(fill="x")

        ctk.CTkLabel(hdr, text=op.get("cliente", "—"),
                     font=("Segoe UI", 13, "bold"), text_color=TEXT_PRIMARY).pack(side="left")

        # Status badge
        badge = ctk.CTkFrame(hdr, fg_color=cor, corner_radius=6)
        badge.pack(side="left", padx=(10, 0))
        ctk.CTkLabel(badge, text=f" {status} ", font=("Segoe UI", 9, "bold"),
                     text_color=TEXT_WHITE).pack(padx=6, pady=2)

        if vencidas > 0:
            venc_badge = ctk.CTkFrame(hdr, fg_color=ACCENT_RED, corner_radius=6)
            venc_badge.pack(side="left", padx=(6, 0))
            ctk.CTkLabel(venc_badge, text=f" {vencidas} vencida{'s' if vencidas > 1 else ''} ",
                         font=("Segoe UI", 9, "bold"), text_color=TEXT_WHITE).pack(padx=6, pady=2)

        # Botoes (direita)
        btn_box = ctk.CTkFrame(hdr, fg_color="transparent")
        btn_box.pack(side="right")

        ctk.CTkButton(btn_box, text="\u270e", width=32, height=28, fg_color=BG_INPUT,
                      hover_color=BORDER_LIGHT, text_color=TEXT_PRIMARY,
                      font=("Segoe UI", 13), corner_radius=6,
                      command=lambda o=op: self._fr_editar_operacao(o)).pack(side="left", padx=2)
        ctk.CTkButton(btn_box, text="\u2715", width=32, height=28, fg_color=BG_INPUT,
                      hover_color="#ffdddd", text_color=ACCENT_RED,
                      font=("Segoe UI", 13), corner_radius=6,
                      command=lambda o=op: self._fr_excluir_operacao(o)).pack(side="left", padx=2)

        # Info row
        info = ctk.CTkFrame(body, fg_color="transparent")
        info.pack(fill="x", pady=(4, 0))

        details = [
            f"Carta: {fmt_currency(op.get('valor_carta', 0))}",
            f"Adm: {op.get('administradora', '—')}",
            f"Prazo: {op.get('prazo_meses', 0)}m",
            f"Início: {op.get('data_inicio', '—')}",
        ]
        ctk.CTkLabel(info, text="  |  ".join(details), font=("Segoe UI", 10),
                     text_color=TEXT_SECONDARY).pack(side="left")

        # Progress row
        prog = ctk.CTkFrame(body, fg_color="transparent")
        prog.pack(fill="x", pady=(6, 0))

        pct = pagas / total_parcelas if total_parcelas > 0 else 0
        pbar = ctk.CTkProgressBar(prog, width=200, height=8, fg_color=BG_INPUT,
                                  progress_color=cor, corner_radius=4)
        pbar.set(pct)
        pbar.pack(side="left", padx=(0, 10))

        ctk.CTkLabel(prog, text=f"{pagas}/{total_parcelas} parcelas pagas ({pct:.0%})",
                     font=("Segoe UI", 10), text_color=TEXT_SECONDARY).pack(side="left")

        ctk.CTkLabel(prog, text=f"Recebido: {fmt_currency(valor_recebido)}",
                     font=("Segoe UI", 10, "bold"), text_color=ACCENT_TEAL).pack(side="right")

        if prox_venc and status != "Encerrado":
            prox_str = prox_venc.strftime("%d/%m/%Y")
            prox_color = ACCENT_RED if prox_venc < hoje else ACCENT_ORANGE
            ctk.CTkLabel(prog, text=f"Próx: {prox_str} ({fmt_currency(prox_valor)})",
                         font=("Segoe UI", 10, "bold"), text_color=prox_color).pack(side="right", padx=(0, 16))

        # Botao expandir pagamentos
        expand_btn = ctk.CTkButton(
            body, text="  Ver Pagamentos ▼",
            font=("Segoe UI", 10), fg_color="transparent",
            hover_color=BG_INPUT, text_color=ACCENT_GREEN,
            height=28, corner_radius=6, anchor="w",
            command=lambda c=card, o=op, b=body: self._fr_toggle_pagamentos(c, o, b),
        )
        expand_btn.pack(anchor="w", pady=(4, 0))

    def _fr_toggle_pagamentos(self, card, op, body):
        # Se ja tem frame de pagamentos, remove (toggle)
        tag = f"_fr_pag_{op['id']}"
        existing = getattr(self, tag, None)
        if existing and existing.winfo_exists():
            existing.destroy()
            setattr(self, tag, None)
            return

        # Criar frame de pagamentos
        pag_frame = ctk.CTkFrame(body, fg_color=BG_INPUT, corner_radius=8)
        pag_frame.pack(fill="x", pady=(6, 0))
        setattr(self, tag, pag_frame)

        pagamentos = op.get("pagamentos", [])
        if not pagamentos:
            ctk.CTkLabel(pag_frame, text="Nenhum pagamento registrado.",
                         font=("Segoe UI", 10), text_color=TEXT_TERTIARY).pack(pady=10)
            return

        # Header
        hdr = ctk.CTkFrame(pag_frame, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(8, 4))
        for txt, w in [("Mês", 50), ("Vencimento", 100), ("Valor", 110), ("Status", 80), ("Data Pgto", 100), ("Ação", 60)]:
            ctk.CTkLabel(hdr, text=txt, font=("Segoe UI", 9, "bold"),
                         text_color=TEXT_SECONDARY, width=w).pack(side="left", padx=2)

        ctk.CTkFrame(pag_frame, fg_color=BORDER_LIGHT, height=1).pack(fill="x", padx=12)

        from datetime import datetime as dt
        hoje = dt.now()

        for p in pagamentos:
            row = ctk.CTkFrame(pag_frame, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=1)

            pago = p.get("pago", False)
            try:
                dv = dt.strptime(p["data_venc"], "%Y-%m-%d")
                venc_str = dv.strftime("%d/%m/%Y")
                vencido = not pago and dv < hoje
            except Exception:
                venc_str = p.get("data_venc", "—")
                vencido = False

            # Mes
            ctk.CTkLabel(row, text=str(p.get("mes", "")), font=("Segoe UI", 10),
                         text_color=TEXT_PRIMARY, width=50).pack(side="left", padx=2)
            # Vencimento
            venc_color = ACCENT_RED if vencido else TEXT_PRIMARY
            ctk.CTkLabel(row, text=venc_str, font=("Segoe UI", 10),
                         text_color=venc_color, width=100).pack(side="left", padx=2)
            # Valor
            ctk.CTkLabel(row, text=fmt_currency(p.get("valor", 0)), font=("Segoe UI", 10),
                         text_color=TEXT_PRIMARY, width=110).pack(side="left", padx=2)
            # Status
            if pago:
                st_text, st_color = "Pago", ACCENT_TEAL
            elif vencido:
                st_text, st_color = "Vencido", ACCENT_RED
            else:
                st_text, st_color = "Pendente", ACCENT_ORANGE
            ctk.CTkLabel(row, text=st_text, font=("Segoe UI", 10, "bold"),
                         text_color=st_color, width=80).pack(side="left", padx=2)
            # Data pagamento
            dp = p.get("data_pag", "—") or "—"
            if dp != "—":
                try:
                    dp = dt.strptime(dp, "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    pass
            ctk.CTkLabel(row, text=dp, font=("Segoe UI", 10),
                         text_color=TEXT_PRIMARY, width=100).pack(side="left", padx=2)
            # Ação
            if not pago:
                ctk.CTkButton(
                    row, text="✓", width=30, height=24, fg_color=ACCENT_TEAL,
                    hover_color=self._darken(ACCENT_TEAL), text_color=TEXT_WHITE,
                    font=("Segoe UI", 12, "bold"), corner_radius=4,
                    command=lambda o=op, pm=p: self._fr_marcar_pago(o, pm),
                ).pack(side="left", padx=2)
            else:
                ctk.CTkButton(
                    row, text="↩", width=30, height=24, fg_color="#6b7280",
                    hover_color="#555d6a", text_color=TEXT_WHITE,
                    font=("Segoe UI", 12), corner_radius=4,
                    command=lambda o=op, pm=p: self._fr_desmarcar_pago(o, pm),
                ).pack(side="left", padx=2)

        # Botao marcar proximas como pagas
        pendentes = [p for p in pagamentos if not p.get("pago")]
        if pendentes:
            btn_bar = ctk.CTkFrame(pag_frame, fg_color="transparent")
            btn_bar.pack(fill="x", padx=12, pady=(6, 8))
            ctk.CTkButton(
                btn_bar, text="  Pagar Próxima Parcela",
                font=("Segoe UI", 10, "bold"), fg_color=ACCENT_GREEN,
                hover_color=self._darken(ACCENT_GREEN), height=30, corner_radius=6,
                command=lambda o=op, pn=pendentes[0]: self._fr_marcar_pago(o, pn),
            ).pack(side="left", padx=(0, 8))

            ctk.CTkButton(
                btn_bar, text="  Pagar Todas Vencidas",
                font=("Segoe UI", 10, "bold"), fg_color=ACCENT_ORANGE,
                hover_color="#c96f1f", height=30, corner_radius=6,
                command=lambda o=op: self._fr_pagar_vencidas(o),
            ).pack(side="left")

    def _fr_marcar_pago(self, op, pagamento):
        from datetime import datetime as dt
        ops = self._fr_load()
        for o in ops:
            if o["id"] == op["id"]:
                for p in o.get("pagamentos", []):
                    if p["mes"] == pagamento["mes"]:
                        p["pago"] = True
                        p["data_pag"] = dt.now().strftime("%Y-%m-%d")
                        break
                break
        self._fr_save(ops)
        self._fr_refresh()

    def _fr_desmarcar_pago(self, op, pagamento):
        ops = self._fr_load()
        for o in ops:
            if o["id"] == op["id"]:
                for p in o.get("pagamentos", []):
                    if p["mes"] == pagamento["mes"]:
                        p["pago"] = False
                        p["data_pag"] = None
                        break
                break
        self._fr_save(ops)
        self._fr_refresh()

    def _fr_pagar_vencidas(self, op):
        from datetime import datetime as dt
        hoje = dt.now()
        ops = self._fr_load()
        count = 0
        for o in ops:
            if o["id"] == op["id"]:
                for p in o.get("pagamentos", []):
                    if not p.get("pago"):
                        try:
                            dv = dt.strptime(p["data_venc"], "%Y-%m-%d")
                            if dv < hoje:
                                p["pago"] = True
                                p["data_pag"] = hoje.strftime("%Y-%m-%d")
                                count += 1
                        except Exception:
                            pass
                break
        self._fr_save(ops)
        self._fr_refresh()
        if count > 0:
            messagebox.showinfo("Pagamentos", f"{count} parcela(s) marcada(s) como paga(s).")

    def _fr_nova_operacao(self, edit_op=None):
        from datetime import datetime as dt
        import uuid

        win = ctk.CTkToplevel(self)
        win.title("Editar Operação" if edit_op else "Nova Operação")
        win.geometry("560x620")
        win.resizable(False, False)
        win.grab_set()
        win.configure(fg_color=BG_SECONDARY)
        win.after(200, win.lift)

        scroll = ctk.CTkScrollableFrame(win, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=10)

        fields = {}

        def _add_field(label, key, default="", placeholder=""):
            ctk.CTkLabel(scroll, text=label, font=("Segoe UI", 10, "bold"),
                         text_color=TEXT_SECONDARY).pack(anchor="w", pady=(6, 2))
            e = ctk.CTkEntry(scroll, placeholder_text=placeholder, height=36,
                             corner_radius=8, fg_color=BG_INPUT,
                             border_width=1, border_color=BORDER_LIGHT)
            e.pack(fill="x", pady=(0, 4))
            if default:
                e.insert(0, str(default))
            fields[key] = e

        _add_field("Nome do Cliente *", "cliente",
                   edit_op.get("cliente", "") if edit_op else "", "Ex: João Silva")
        _add_field("Assessor", "assessor",
                   edit_op.get("assessor", "") if edit_op else "", "Ex: Maria Santos")
        _add_field("Administradora", "administradora",
                   edit_op.get("administradora", "") if edit_op else "", "Ex: Porto Seguro")
        _add_field("Valor da Carta (R$) *", "valor_carta",
                   edit_op.get("valor_carta", "") if edit_op else "", "Ex: 300000")
        _add_field("Prazo (meses) *", "prazo_meses",
                   edit_op.get("prazo_meses", "") if edit_op else "", "Ex: 120")
        _add_field("Mês Contemplação *", "prazo_contemp",
                   edit_op.get("prazo_contemp", "") if edit_op else "", "Ex: 60")
        _add_field("Parcela Fase 1 (R$) *", "parcela_f1",
                   edit_op.get("parcela_f1", "") if edit_op else "", "Use Simulador para calcular")
        _add_field("Parcela Fase 2 (R$)", "parcela_f2",
                   edit_op.get("parcela_f2", "") if edit_op else "", "Parcela pós-contemplação")
        _add_field("Data Início (DD/MM/AAAA) *", "data_inicio",
                   edit_op.get("data_inicio", "") if edit_op else dt.now().strftime("%d/%m/%Y"),
                   "Ex: 15/01/2026")

        # Status
        ctk.CTkLabel(scroll, text="Status", font=("Segoe UI", 10, "bold"),
                     text_color=TEXT_SECONDARY).pack(anchor="w", pady=(6, 2))
        status_var = ctk.CTkSegmentedButton(
            scroll, values=["Ativo", "Contemplado", "Encerrado"],
            font=("Segoe UI", 10, "bold"), fg_color=BG_INPUT,
            selected_color=ACCENT_GREEN, selected_hover_color=BG_SIDEBAR_HOVER,
            unselected_color="#e0e3e8", unselected_hover_color="#d0d3d8",
            text_color=TEXT_PRIMARY, corner_radius=8, height=36,
        )
        status_var.set(edit_op.get("status", "Ativo") if edit_op else "Ativo")
        status_var.pack(fill="x", pady=(0, 4))

        # Botoes
        btn_frame = ctk.CTkFrame(scroll, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(12, 4))

        def _salvar():
            try:
                cliente = fields["cliente"].get().strip()
                if not cliente:
                    messagebox.showerror("Erro", "Nome do cliente é obrigatório.", parent=win)
                    return

                vc = self._parse_number(fields["valor_carta"].get())
                prazo = int(self._parse_number(fields["prazo_meses"].get()))
                contemp = int(self._parse_number(fields["prazo_contemp"].get()))
                p_f1 = self._parse_number(fields["parcela_f1"].get())
                p_f2_str = fields["parcela_f2"].get().strip()
                p_f2 = self._parse_number(p_f2_str) if p_f2_str else p_f1
                data_str = fields["data_inicio"].get().strip()

                # Parse data
                try:
                    data_dt = dt.strptime(data_str, "%d/%m/%Y")
                except Exception:
                    messagebox.showerror("Erro", "Data inválida. Use DD/MM/AAAA.", parent=win)
                    return

                ops = self._fr_load()

                if edit_op:
                    # Atualizar existente
                    for o in ops:
                        if o["id"] == edit_op["id"]:
                            o["cliente"] = cliente
                            o["assessor"] = fields["assessor"].get().strip()
                            o["administradora"] = fields["administradora"].get().strip()
                            o["valor_carta"] = vc
                            o["prazo_meses"] = prazo
                            o["prazo_contemp"] = contemp
                            o["parcela_f1"] = p_f1
                            o["parcela_f2"] = p_f2
                            o["data_inicio"] = data_str
                            o["status"] = status_var.get()
                            # Regenerar pagamentos se prazo mudou
                            if len(o.get("pagamentos", [])) != prazo:
                                o["pagamentos"] = self._fr_gerar_pagamentos(
                                    prazo, contemp, p_f1, p_f2, data_dt)
                            break
                else:
                    # Nova operacao
                    pagamentos = self._fr_gerar_pagamentos(prazo, contemp, p_f1, p_f2, data_dt)
                    nova = {
                        "id": str(uuid.uuid4())[:8],
                        "cliente": cliente,
                        "assessor": fields["assessor"].get().strip(),
                        "administradora": fields["administradora"].get().strip(),
                        "valor_carta": vc,
                        "prazo_meses": prazo,
                        "prazo_contemp": contemp,
                        "parcela_f1": p_f1,
                        "parcela_f2": p_f2,
                        "data_inicio": data_str,
                        "status": status_var.get(),
                        "pagamentos": pagamentos,
                    }
                    ops.append(nova)

                self._fr_save(ops)
                win.destroy()
                self._fr_refresh()

            except (ValueError, AttributeError) as e:
                messagebox.showerror("Erro", f"Preencha os campos corretamente.\n{e}", parent=win)

        ctk.CTkButton(
            btn_frame, text="  Salvar",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_GREEN,
            hover_color=self._darken(ACCENT_GREEN), height=42, corner_radius=10,
            command=_salvar,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="  Cancelar",
            font=("Segoe UI", 12), fg_color="#6b7280",
            hover_color="#555d6a", height=42, corner_radius=10,
            command=win.destroy,
        ).pack(side="left")

    def _fr_gerar_pagamentos(self, prazo, contemp, parcela_f1, parcela_f2, data_inicio):
        from datetime import datetime as dt
        from datetime import timedelta
        pagamentos = []
        for mes in range(1, prazo + 1):
            # Data de vencimento: data_inicio + mes meses
            ano = data_inicio.year + (data_inicio.month + mes - 1) // 12
            m = (data_inicio.month + mes - 1) % 12 + 1
            dia = min(data_inicio.day, 28)  # Seguro para fevereiro
            data_venc = dt(ano, m, dia)

            valor = parcela_f1 if mes <= contemp else parcela_f2
            pagamentos.append({
                "mes": mes,
                "data_venc": data_venc.strftime("%Y-%m-%d"),
                "valor": round(valor, 2),
                "pago": False,
                "data_pag": None,
            })
        return pagamentos

    def _fr_editar_operacao(self, op):
        self._fr_nova_operacao(edit_op=op)

    def _fr_excluir_operacao(self, op):
        if not messagebox.askyesno("Confirmar", f"Excluir operação de {op.get('cliente', '—')}?"):
            return
        ops = self._fr_load()
        ops = [o for o in ops if o["id"] != op["id"]]
        self._fr_save(ops)
        self._fr_refresh()

    def _fr_exportar_excel(self):
        from datetime import datetime as dt
        ops = self._fr_load()
        if not ops:
            messagebox.showinfo("Exportar", "Nenhuma operação para exportar.")
            return

        try:
            wb = openpyxl.Workbook()

            # Aba 1: Resumo das operacoes
            ws = wb.active
            ws.title = "Operacoes"
            headers = ["Cliente", "Assessor", "Administradora", "Valor Carta",
                       "Prazo", "Contemplacao", "Parcela F1", "Parcela F2",
                       "Data Inicio", "Status", "Parcelas Pagas", "Total Pago", "Total Pendente"]
            for c, h in enumerate(headers, 1):
                ws.cell(1, c, h)
                ws.cell(1, c).font = openpyxl.styles.Font(bold=True)

            for r, op in enumerate(ops, 2):
                pagamentos = op.get("pagamentos", [])
                pagas = sum(1 for p in pagamentos if p.get("pago"))
                pago = sum(p.get("valor", 0) for p in pagamentos if p.get("pago"))
                pend = sum(p.get("valor", 0) for p in pagamentos if not p.get("pago"))
                vals = [
                    op.get("cliente", ""), op.get("assessor", ""),
                    op.get("administradora", ""), op.get("valor_carta", 0),
                    op.get("prazo_meses", 0), op.get("prazo_contemp", 0),
                    op.get("parcela_f1", 0), op.get("parcela_f2", 0),
                    op.get("data_inicio", ""), op.get("status", ""),
                    f"{pagas}/{len(pagamentos)}", pago, pend,
                ]
                for c, v in enumerate(vals, 1):
                    ws.cell(r, c, v)

            # Aba 2: Todos os pagamentos detalhados
            ws2 = wb.create_sheet("Pagamentos")
            hdrs2 = ["Cliente", "Mes", "Vencimento", "Valor", "Status", "Data Pagamento"]
            for c, h in enumerate(hdrs2, 1):
                ws2.cell(1, c, h)
                ws2.cell(1, c).font = openpyxl.styles.Font(bold=True)

            row = 2
            for op in ops:
                for p in op.get("pagamentos", []):
                    ws2.cell(row, 1, op.get("cliente", ""))
                    ws2.cell(row, 2, p.get("mes", 0))
                    ws2.cell(row, 3, p.get("data_venc", ""))
                    ws2.cell(row, 4, p.get("valor", 0))
                    ws2.cell(row, 5, "Pago" if p.get("pago") else "Pendente")
                    ws2.cell(row, 6, p.get("data_pag", "") or "")
                    row += 1

            # Aba 3: Fluxo mensal (receitas por mes)
            ws3 = wb.create_sheet("Fluxo Mensal")
            ws3.cell(1, 1, "Mes/Ano")
            ws3.cell(1, 2, "Receita Prevista")
            ws3.cell(1, 3, "Receita Recebida")
            ws3.cell(1, 1).font = openpyxl.styles.Font(bold=True)
            ws3.cell(1, 2).font = openpyxl.styles.Font(bold=True)
            ws3.cell(1, 3).font = openpyxl.styles.Font(bold=True)

            # Agrupar por mes
            fluxo = {}
            for op in ops:
                for p in op.get("pagamentos", []):
                    try:
                        dv = dt.strptime(p["data_venc"], "%Y-%m-%d")
                        key = dv.strftime("%Y-%m")
                    except Exception:
                        continue
                    if key not in fluxo:
                        fluxo[key] = {"previsto": 0, "recebido": 0}
                    fluxo[key]["previsto"] += p.get("valor", 0)
                    if p.get("pago"):
                        fluxo[key]["recebido"] += p.get("valor", 0)

            for r, (mes, vals) in enumerate(sorted(fluxo.items()), 2):
                ws3.cell(r, 1, mes)
                ws3.cell(r, 2, vals["previsto"])
                ws3.cell(r, 3, vals["recebido"])

            # Salvar
            os.makedirs(CONSÓRCIO_OUTPUT_DIR, exist_ok=True)
            nome = f"Fluxo_Receitas_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = os.path.join(CONSÓRCIO_OUTPUT_DIR, nome)
            wb.save(path)
            wb.close()
            messagebox.showinfo("Exportado", f"Planilha salva!\n\n{nome}")
            os.startfile(path)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar:\n{e}")

    def _fr_importar_excel(self):
        from datetime import datetime as dt
        import uuid

        path = filedialog.askopenfilename(
            title="Selecionar planilha de operações",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")],
        )
        if not path:
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o arquivo:\n{e}")
            return

        # Verificar se tem as abas esperadas
        tem_operacoes = "Operacoes" in wb.sheetnames
        tem_pagamentos = "Pagamentos" in wb.sheetnames

        if not tem_operacoes:
            messagebox.showerror("Erro",
                "Aba 'Operacoes' não encontrada na planilha.\n"
                "Use o mesmo formato gerado por 'Exportar Excel'.")
            wb.close()
            return

        # --- Ler aba Operacoes ---
        ws_ops = wb["Operacoes"]
        ops_importadas = []
        clientes_map = {}  # cliente -> op dict (para vincular pagamentos)

        for row in ws_ops.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            # Colunas: Cliente, Assessor, Administradora, Valor Carta,
            #          Prazo, Contemplacao, Parcela F1, Parcela F2,
            #          Data Inicio, Status, Parcelas Pagas, Total Pago, Total Pendente
            cliente = str(row[0] or "").strip()
            if not cliente:
                continue

            assessor = str(row[1] or "").strip()
            administradora = str(row[2] or "").strip()
            valor_carta = float(row[3] or 0)
            prazo = int(row[4] or 0)
            contemp = int(row[5] or 0)
            parcela_f1 = float(row[6] or 0)
            parcela_f2 = float(row[7] or 0)
            data_inicio = str(row[8] or "").strip()
            status = str(row[9] or "Ativo").strip()

            if status not in ("Ativo", "Contemplado", "Encerrado"):
                status = "Ativo"

            op = {
                "id": str(uuid.uuid4())[:8],
                "cliente": cliente,
                "assessor": assessor,
                "administradora": administradora,
                "valor_carta": valor_carta,
                "prazo_meses": prazo,
                "prazo_contemp": contemp,
                "parcela_f1": parcela_f1,
                "parcela_f2": parcela_f2,
                "data_inicio": data_inicio,
                "status": status,
                "pagamentos": [],
            }
            ops_importadas.append(op)
            clientes_map[cliente] = op

        if not ops_importadas:
            messagebox.showwarning("Aviso", "Nenhuma operação encontrada na planilha.")
            wb.close()
            return

        # --- Ler aba Pagamentos (se existir) ---
        if tem_pagamentos:
            ws_pag = wb["Pagamentos"]
            for row in ws_pag.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                # Colunas: Cliente, Mes, Vencimento, Valor, Status, Data Pagamento
                cliente = str(row[0] or "").strip()
                mes = int(row[1] or 0)

                # Vencimento: pode vir como string YYYY-MM-DD ou datetime
                venc_raw = row[2]
                if isinstance(venc_raw, dt):
                    data_venc = venc_raw.strftime("%Y-%m-%d")
                else:
                    data_venc = str(venc_raw or "").strip()
                    # Tentar converter DD/MM/YYYY -> YYYY-MM-DD
                    if "/" in data_venc and len(data_venc) == 10 and data_venc[2] == "/":
                        try:
                            data_venc = dt.strptime(data_venc, "%d/%m/%Y").strftime("%Y-%m-%d")
                        except Exception:
                            pass

                valor = float(row[3] or 0)
                status_pag = str(row[4] or "").strip()
                pago = status_pag.lower() in ("pago", "sim", "true", "1")

                # Data pagamento
                dp_raw = row[5]
                if isinstance(dp_raw, dt):
                    data_pag = dp_raw.strftime("%Y-%m-%d")
                elif dp_raw:
                    data_pag = str(dp_raw).strip()
                    if "/" in data_pag and len(data_pag) == 10 and data_pag[2] == "/":
                        try:
                            data_pag = dt.strptime(data_pag, "%d/%m/%Y").strftime("%Y-%m-%d")
                        except Exception:
                            pass
                else:
                    data_pag = None

                if not data_pag or data_pag in ("", "—", "None"):
                    data_pag = None

                pagamento = {
                    "mes": mes,
                    "data_venc": data_venc,
                    "valor": round(valor, 2),
                    "pago": pago,
                    "data_pag": data_pag,
                }

                # Vincular ao cliente
                if cliente in clientes_map:
                    clientes_map[cliente]["pagamentos"].append(pagamento)

        wb.close()

        # Para operacoes sem pagamentos importados, gerar automaticamente
        for op in ops_importadas:
            if not op["pagamentos"] and op["prazo_meses"] > 0 and op["parcela_f1"] > 0:
                try:
                    data_str = op["data_inicio"]
                    # Tentar DD/MM/YYYY
                    try:
                        data_dt = dt.strptime(data_str, "%d/%m/%Y")
                    except Exception:
                        data_dt = dt.strptime(data_str, "%Y-%m-%d")
                    op["pagamentos"] = self._fr_gerar_pagamentos(
                        op["prazo_meses"], op["prazo_contemp"],
                        op["parcela_f1"], op["parcela_f2"], data_dt
                    )
                except Exception:
                    pass

        # Perguntar se quer substituir ou adicionar
        ops_existentes = self._fr_load()
        if ops_existentes:
            resp = messagebox.askyesnocancel(
                "Importar",
                f"Encontradas {len(ops_importadas)} operação(ões) na planilha.\n\n"
                f"Você já tem {len(ops_existentes)} operação(ões) cadastrada(s).\n\n"
                "SIM = Substituir tudo\n"
                "NÃO = Adicionar às existentes\n"
                "CANCELAR = Cancelar importação",
            )
            if resp is None:
                return
            if resp:
                # Substituir
                self._fr_save(ops_importadas)
            else:
                # Adicionar
                ops_existentes.extend(ops_importadas)
                self._fr_save(ops_existentes)
        else:
            self._fr_save(ops_importadas)

        self._fr_refresh()
        messagebox.showinfo("Importado",
            f"{len(ops_importadas)} operação(ões) importada(s) com sucesso!\n"
            f"Total de parcelas: {sum(len(o['pagamentos']) for o in ops_importadas)}")

    # -----------------------------------------------------------------
    #  PAGE: CONSOLIDADOR DE CARTEIRAS
    # -----------------------------------------------------------------
    def _build_consolidador_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Consolidador", subtitle="Carteiras Multi-Instituicao")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ======== DROP ZONE ========
        self.cons_drop_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=14,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.cons_drop_card.pack(fill="x", pady=(0, 16))
        self.cons_drop_card.bind("<Button-1>", lambda e: self._on_cons_browse())

        drop_inner = ctk.CTkFrame(self.cons_drop_card, fg_color="transparent")
        drop_inner.pack(fill="x", padx=30, pady=36)
        drop_inner.bind("<Button-1>", lambda e: self._on_cons_browse())

        self.cons_icon_label = ctk.CTkLabel(
            drop_inner, text="\u229a",
            font=("Segoe UI", 44), text_color=ACCENT_BLUE
        )
        self.cons_icon_label.pack()
        self.cons_icon_label.bind("<Button-1>", lambda e: self._on_cons_browse())

        self.cons_drop_title = ctk.CTkLabel(
            drop_inner, text="Clique para selecionar relatorios de carteira",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        )
        self.cons_drop_title.pack(pady=(8, 2))
        self.cons_drop_title.bind("<Button-1>", lambda e: self._on_cons_browse())

        self.cons_drop_sub = ctk.CTkLabel(
            drop_inner, text="Formatos aceitos:  PDF (XP, Itau, Safra)  |  XLSX (BTG)",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.cons_drop_sub.pack()
        self.cons_drop_sub.bind("<Button-1>", lambda e: self._on_cons_browse())

        self.cons_browse_btn = ctk.CTkButton(
            drop_inner, text="  Selecionar Arquivos",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=40, corner_radius=8, width=220,
            command=self._on_cons_browse,
        )
        self.cons_browse_btn.pack(pady=(14, 0))

        # ======== FILES LIST (hidden initially) ========
        self.cons_files_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.cons_files_frame.pack(fill="x", pady=(0, 12))
        self.cons_files_frame.pack_forget()

        self.cons_files_inner = ctk.CTkFrame(self.cons_files_frame, fg_color="transparent")
        self.cons_files_inner.pack(fill="x", padx=16, pady=10)

        self.cons_files_label = ctk.CTkLabel(
            self.cons_files_inner, text="",
            font=("Segoe UI", 11), text_color=TEXT_PRIMARY, anchor="w",
            justify="left"
        )
        self.cons_files_label.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(
            self.cons_files_inner, text="Trocar", font=("Segoe UI", 10),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=28, corner_radius=6, width=70,
            command=self._on_cons_browse,
        ).pack(side="right")

        # ======== ACTION BUTTONS ========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 12))

        self.cons_process_btn = ctk.CTkButton(
            btn_frame, text="  Consolidar e Gerar PDF",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=self._darken(ACCENT_GREEN),
            height=44, corner_radius=10, state="disabled",
            command=self._on_cons_process,
        )
        self.cons_process_btn.pack(side="left", padx=(0, 10))

        self.cons_open_btn = ctk.CTkButton(
            btn_frame, text="  Abrir PDF",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=44, corner_radius=10, state="disabled",
            command=self._on_cons_open_pdf,
        )
        self.cons_open_btn.pack(side="left", padx=(0, 10))

        self.cons_open_folder_btn = ctk.CTkButton(
            btn_frame, text="  Abrir Pasta",
            font=("Segoe UI", 12),
            fg_color="#6b7280", hover_color="#555d6a",
            height=44, corner_radius=10,
            command=self._on_cons_open_folder,
        )
        self.cons_open_folder_btn.pack(side="left")

        # ======== RESULTS AREA (hidden initially) ========
        self.cons_results_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        self.cons_results_frame.pack(fill="x", pady=(0, 14))
        self.cons_results_frame.pack_forget()

        self.cons_results_inner = ctk.CTkFrame(self.cons_results_frame, fg_color="transparent")
        self.cons_results_inner.pack(fill="x", padx=16, pady=12)

        # ======== LOG ========
        self.cons_log_label = ctk.CTkLabel(
            content, text="Log de Processamento",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        self.cons_log_label.pack(fill="x", pady=(4, 8))

        self.cons_log_frame = ctk.CTkFrame(
            content, fg_color=BG_LOG, corner_radius=10,
        )
        self.cons_log_frame.pack(fill="x", pady=(0, 8))

        self.cons_log_text = ctk.CTkTextbox(
            self.cons_log_frame, font=("Consolas", 9),
            fg_color=BG_LOG, text_color=TEXT_LOG,
            corner_radius=10, height=180, wrap="word"
        )
        self.cons_log_text.pack(fill="x", padx=4, pady=4)
        self.cons_log_text.insert("1.0", "  Aguardando arquivos...\n")
        self.cons_log_text.configure(state="disabled")

        # ======== STATUS BAR ========
        self.cons_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.cons_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.cons_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.cons_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.cons_status_dot.pack(side="left")

        self.cons_status_text = ctk.CTkLabel(
            si, text="  Aguardando arquivos...",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.cons_status_text.pack(side="left")

        # Internal state
        self._cons_file_paths = []
        self._cons_pdf_path = None
        self._cons_portfolio = None

        return page

    # -----------------------------------------------------------------
    #  CONSOLIDADOR: Acoes
    # -----------------------------------------------------------------
    def _cons_log(self, msg):
        """Append message to consolidador log."""
        self.cons_log_text.configure(state="normal")
        self.cons_log_text.insert("end", f"  {msg}\n")
        self.cons_log_text.see("end")
        self.cons_log_text.configure(state="disabled")

    def _on_cons_browse(self):
        paths = filedialog.askopenfilenames(
            title="Selecionar Relatorios de Carteira",
            initialdir=os.path.join(BASE_DIR, "Mesa Produtos", "Consolidador"),
            filetypes=[
                ("Relatorios", "*.pdf *.xlsx *.xls"),
                ("PDF files", "*.pdf"),
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*"),
            ]
        )
        if paths:
            self._cons_file_paths = list(paths)
            self._cons_pdf_path = None
            self._cons_portfolio = None

            # Update UI
            names = [os.path.basename(p) for p in self._cons_file_paths]
            label_text = f"{len(names)} arquivo(s):  " + ",  ".join(names)
            self.cons_files_label.configure(text=label_text)
            self.cons_files_frame.pack(fill="x", pady=(0, 12))

            self.cons_drop_title.configure(text=f"{len(names)} arquivo(s) selecionado(s)")
            self.cons_drop_sub.configure(text="Clique para trocar")

            self.cons_process_btn.configure(state="normal")
            self.cons_open_btn.configure(state="disabled")

            # Clear previous results
            for w in self.cons_results_inner.winfo_children():
                w.destroy()
            self.cons_results_frame.pack_forget()

            # Clear log
            self.cons_log_text.configure(state="normal")
            self.cons_log_text.delete("1.0", "end")
            self.cons_log_text.insert("1.0", "  Arquivos carregados. Clique em 'Consolidar e Gerar PDF'.\n")
            self.cons_log_text.configure(state="disabled")

            self.cons_status_dot.configure(text_color=ACCENT_BLUE)
            self.cons_status_text.configure(text=f"  {len(names)} arquivo(s) pronto(s) para consolidar")

    def _cons_detect_and_parse(self, file_path):
        """Detecta instituicao e parseia o arquivo."""
        import pdfplumber
        import warnings
        warnings.filterwarnings("ignore")

        ext = os.path.splitext(file_path)[1].lower()
        fname = os.path.basename(file_path)

        if ext == ".pdf":
            with pdfplumber.open(file_path) as pdf:
                pages_text = ""
                for p in pdf.pages[:3]:
                    pages_text += (p.extract_text() or "") + "\n"
                pages_lower = pages_text.lower()

                # Detecta relatorio consolidado gerado pelo Somus (tem dados uteis)
                if "somus capital" in pages_lower and "consolidado multi" in pages_lower:
                    from agente_investimentos.consolidador.somus_parser import parse_somus_pdf
                    self._cons_log(f"[{fname}] Detectado: Relatorio Consolidado Somus")
                    return parse_somus_pdf(file_path)

                # Detecta XP
                if ("xp investimentos" in pages_lower or "assessor" in pages_lower
                        or "posicao detalhada" in pages_lower.replace("\xe7", "c").replace("\xe3", "a")
                        or "patrimonio investimento" in pages_lower.replace("\xf4", "o").replace("\xea", "e")):
                    from agente_investimentos.consolidador.xp_parser import parse_xp_pdf
                    self._cons_log(f"[{fname}] Detectado: XP Investimentos")
                    return parse_xp_pdf(file_path)

                # Detecta Itau
                if "itau" in pages_lower or "personnalite" in pages_lower or "total investido" in pages_lower:
                    from agente_investimentos.consolidador.itau_parser import parse_itau_pdf
                    self._cons_log(f"[{fname}] Detectado: Itau")
                    return parse_itau_pdf(file_path)

                # Detecta Safra
                if "safra" in pages_lower or "safrabm" in pages_lower:
                    from agente_investimentos.consolidador.safra_parser import parse_safra_pdf
                    self._cons_log(f"[{fname}] Detectado: Safra")
                    return parse_safra_pdf(file_path)

            # Fallback Safra
            from agente_investimentos.consolidador.safra_parser import parse_safra_pdf
            self._cons_log(f"[{fname}] Fallback: tentando parser Safra")
            return parse_safra_pdf(file_path)

        elif ext in (".xlsx", ".xls"):
            from agente_investimentos.consolidador.btg_parser import parse_btg_xlsx
            self._cons_log(f"[{fname}] Detectado: BTG (Excel)")
            return parse_btg_xlsx(file_path)

        return None

    def _on_cons_process(self):
        if not self._cons_file_paths:
            return
        self.cons_process_btn.configure(state="disabled", text="  Processando...")
        self.cons_status_dot.configure(text_color=ACCENT_ORANGE)
        self.cons_status_text.configure(text="  Processando relatorios...")

        # Clear log
        self.cons_log_text.configure(state="normal")
        self.cons_log_text.delete("1.0", "end")
        self.cons_log_text.configure(state="disabled")

        threading.Thread(target=self._cons_process_thread, daemon=True).start()

    def _cons_process_thread(self):
        from agente_investimentos.consolidador.models import ConsolidatedPortfolio
        try:
            instituicoes = []
            errors = []

            for path in self._cons_file_paths:
                fname = os.path.basename(path)
                self._cons_log(f"Processando: {fname}")
                try:
                    result = self._cons_detect_and_parse(path)
                    if result and result.patrimonio_bruto > 0:
                        instituicoes.append(result)
                        self._cons_log(
                            f"  OK: {result.nome} - {result.num_ativos} ativos - "
                            f"{fmt_currency(result.patrimonio_bruto)}"
                        )
                    elif result:
                        self._cons_log(f"  AVISO: {fname} parseado mas sem patrimonio")
                    else:
                        errors.append(fname)
                        self._cons_log(f"  ERRO: {fname} - formato nao reconhecido")
                except Exception as e:
                    errors.append(fname)
                    self._cons_log(f"  ERRO: {fname} - {str(e)[:120]}")

            if not instituicoes:
                self._cons_log("\nNenhuma instituicao valida encontrada.")
                self.after(0, lambda: self._cons_finish(False, "Nenhum arquivo valido"))
                return

            # Consolida
            cp = ConsolidatedPortfolio(instituicoes=instituicoes)
            self._cons_portfolio = cp

            self._cons_log(f"\nConsolidado: {cp.num_instituicoes} instituicoes, "
                          f"{cp.num_ativos_total} ativos, {fmt_currency(cp.patrimonio_bruto_total)}")

            # Gera PDF
            self._cons_log("Gerando PDF consolidado...")
            os.makedirs(CONSOLIDADOR_OUTPUT_DIR, exist_ok=True)

            from agente_investimentos.consolidador.pdf_builder import ConsolidadorPDFBuilder
            builder = ConsolidadorPDFBuilder(cp)
            pdf_path = builder.build()

            # Copia para pasta local do app
            import shutil
            local_pdf = os.path.join(CONSOLIDADOR_OUTPUT_DIR, "Relatorio_Consolidado.pdf")
            shutil.copy2(str(pdf_path), local_pdf)
            self._cons_pdf_path = local_pdf

            self._cons_log(f"PDF gerado: {local_pdf}")
            self._cons_log("\nProcessamento concluido com sucesso!")

            self.after(0, lambda: self._cons_finish(True, ""))

        except Exception as e:
            self._cons_log(f"\nERRO GERAL: {str(e)[:200]}")
            self.after(0, lambda: self._cons_finish(False, str(e)[:100]))

    def _cons_finish(self, success, error_msg):
        self.cons_process_btn.configure(state="normal", text="  Consolidar e Gerar PDF")

        if success:
            self.cons_status_dot.configure(text_color=ACCENT_GREEN)
            self.cons_status_text.configure(text="  Consolidacao concluida com sucesso!")
            self.cons_open_btn.configure(state="normal")
            self._cons_show_results()
        else:
            self.cons_status_dot.configure(text_color=ACCENT_RED)
            self.cons_status_text.configure(text=f"  Erro: {error_msg}")

    def _cons_show_results(self):
        """Mostra KPIs e resumo da consolidacao."""
        cp = self._cons_portfolio
        if not cp:
            return

        # Clear previous results
        for w in self.cons_results_inner.winfo_children():
            w.destroy()

        self.cons_results_frame.pack(fill="x", pady=(0, 14))

        # Title
        ctk.CTkLabel(
            self.cons_results_inner, text="Resultado da Consolidacao",
            font=("Segoe UI", 15, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        ).pack(fill="x", pady=(0, 12))

        # KPI cards row
        kpi_row = ctk.CTkFrame(self.cons_results_inner, fg_color="transparent")
        kpi_row.pack(fill="x", pady=(0, 12))

        kpis = [
            ("Patrimonio Bruto", fmt_currency(cp.patrimonio_bruto_total), ACCENT_GREEN),
            ("Patrimonio Liquido", fmt_currency(cp.patrimonio_liquido_total), ACCENT_BLUE),
            ("Instituicoes", str(cp.num_instituicoes), ACCENT_PURPLE),
            ("Total Ativos", str(cp.num_ativos_total), ACCENT_ORANGE),
            ("Rent. Ano", fmt_pct(cp.rent_ano_ponderada()), ACCENT_GREEN if cp.rent_ano_ponderada() >= 0 else ACCENT_RED),
        ]

        for i, (label, value, color) in enumerate(kpis):
            card = ctk.CTkFrame(kpi_row, fg_color=BG_CARD, corner_radius=10,
                               border_width=1, border_color=BORDER_CARD)
            card.pack(side="left", fill="x", expand=True, padx=(0 if i == 0 else 6, 0))

            ctk.CTkLabel(
                card, text=label,
                font=("Segoe UI", 9), text_color=TEXT_TERTIARY, anchor="w"
            ).pack(fill="x", padx=12, pady=(10, 0))

            ctk.CTkLabel(
                card, text=value,
                font=("Segoe UI", 14, "bold"), text_color=color, anchor="w"
            ).pack(fill="x", padx=12, pady=(2, 10))

        # Per-institution summary
        for inst in cp.instituicoes:
            inst_frame = ctk.CTkFrame(self.cons_results_inner, fg_color=BG_INPUT, corner_radius=8)
            inst_frame.pack(fill="x", pady=(4, 4))

            row = ctk.CTkFrame(inst_frame, fg_color="transparent")
            row.pack(fill="x", padx=12, pady=8)

            ctk.CTkLabel(
                row, text=f"\u25cf  {inst.nome}",
                font=("Segoe UI", 12, "bold"), text_color=TEXT_PRIMARY
            ).pack(side="left")

            ctk.CTkLabel(
                row, text=f"{inst.num_ativos} ativos",
                font=("Segoe UI", 10), text_color=TEXT_SECONDARY
            ).pack(side="left", padx=(16, 0))

            ctk.CTkLabel(
                row, text=fmt_currency(inst.patrimonio_bruto),
                font=("Segoe UI", 11, "bold"), text_color=ACCENT_GREEN
            ).pack(side="right", padx=(0, 8))

            if inst.rent_carteira_ano:
                ctk.CTkLabel(
                    row, text=fmt_pct(inst.rent_carteira_ano),
                    font=("Segoe UI", 10), text_color=TEXT_SECONDARY
                ).pack(side="right", padx=(0, 16))

    def _on_cons_open_pdf(self):
        if self._cons_pdf_path and os.path.exists(self._cons_pdf_path):
            os.startfile(self._cons_pdf_path)

    def _on_cons_open_folder(self):
        os.makedirs(CONSOLIDADOR_OUTPUT_DIR, exist_ok=True)
        os.startfile(CONSOLIDADOR_OUTPUT_DIR)

    # -----------------------------------------------------------------
    #  PAGE: ORGANIZADOR DE PLANILHAS
    # -----------------------------------------------------------------
    def _build_organizador_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Organizador", subtitle="Planilhas Excel")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ======== DROP ZONE ========
        self.org_drop_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=14,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.org_drop_card.pack(fill="x", pady=(0, 16))
        self.org_drop_card.bind("<Button-1>", lambda e: self._on_browse_excel())

        drop_inner = ctk.CTkFrame(self.org_drop_card, fg_color="transparent")
        drop_inner.pack(fill="x", padx=30, pady=36)
        drop_inner.bind("<Button-1>", lambda e: self._on_browse_excel())

        self.org_icon_label = ctk.CTkLabel(
            drop_inner, text="\u25a6",
            font=("Segoe UI", 44), text_color=ACCENT_BLUE
        )
        self.org_icon_label.pack()
        self.org_icon_label.bind("<Button-1>", lambda e: self._on_browse_excel())

        self.org_drop_title = ctk.CTkLabel(
            drop_inner, text="Clique para selecionar uma planilha Excel",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        )
        self.org_drop_title.pack(pady=(8, 2))
        self.org_drop_title.bind("<Button-1>", lambda e: self._on_browse_excel())

        self.org_drop_sub = ctk.CTkLabel(
            drop_inner, text="Formatos aceitos:  .xlsx  .xls",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.org_drop_sub.pack()
        self.org_drop_sub.bind("<Button-1>", lambda e: self._on_browse_excel())

        self.org_browse_btn = ctk.CTkButton(
            drop_inner, text="  Selecionar Arquivo",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=40, corner_radius=8, width=200,
            command=self._on_browse_excel,
        )
        self.org_browse_btn.pack(pady=(14, 0))

        # ======== FILE INFO BAR (hidden initially) ========
        self.org_file_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.org_file_frame.pack(fill="x", pady=(0, 12))
        self.org_file_frame.pack_forget()

        fi = ctk.CTkFrame(self.org_file_frame, fg_color="transparent")
        fi.pack(fill="x", padx=16, pady=10)

        self.org_file_icon = ctk.CTkLabel(
            fi, text="\u25cf", font=("Segoe UI", 12), text_color=ACCENT_GREEN
        )
        self.org_file_icon.pack(side="left")

        self.org_file_name = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY
        )
        self.org_file_name.pack(side="left", padx=(6, 0))

        self.org_file_info = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 10), text_color=TEXT_SECONDARY
        )
        self.org_file_info.pack(side="left", padx=(12, 0))

        ctk.CTkButton(
            fi, text="Trocar", font=("Segoe UI", 10),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=28, corner_radius=6, width=70,
            command=self._on_browse_excel,
        ).pack(side="right")

        # ======== PREVIEW ========
        self.org_preview_label = ctk.CTkLabel(
            content, text="Preview dos Dados",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        self.org_preview_label.pack(fill="x", pady=(4, 8))

        self.org_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        self.org_preview_frame.pack(fill="x", pady=(0, 14))

        self.org_preview_text = ctk.CTkTextbox(
            self.org_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=200, wrap="none"
        )
        self.org_preview_text.pack(fill="x", padx=4, pady=4)
        self.org_preview_text.insert("1.0", "  Nenhum arquivo carregado...")
        self.org_preview_text.configure(state="disabled")

        # ======== ACTION BUTTONS ========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 12))

        self.org_process_btn = ctk.CTkButton(
            btn_frame, text="  Processar e Organizar",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=self._darken(ACCENT_GREEN),
            height=44, corner_radius=10, state="disabled",
            command=self._on_process_excel,
        )
        self.org_process_btn.pack(side="left", padx=(0, 10))

        self.org_open_btn = ctk.CTkButton(
            btn_frame, text="  Abrir Resultado",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=44, corner_radius=10, state="disabled",
            command=self._on_open_organized,
        )
        self.org_open_btn.pack(side="left", padx=(0, 10))

        self.org_open_folder_btn = ctk.CTkButton(
            btn_frame, text="  Abrir Pasta",
            font=("Segoe UI", 12),
            fg_color="#6b7280", hover_color="#555d6a",
            height=44, corner_radius=10,
            command=self._on_open_org_folder,
        )
        self.org_open_folder_btn.pack(side="left")

        # ======== STATUS BAR ========
        self.org_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.org_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.org_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.org_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.org_status_dot.pack(side="left")

        self.org_status_text = ctk.CTkLabel(
            si, text="  Aguardando arquivo...",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.org_status_text.pack(side="left")

        # Internal state
        self._org_input_path = None
        self._org_output_path = None

        return page

    # -----------------------------------------------------------------
    #  ORGANIZADOR: Ações
    # -----------------------------------------------------------------
    def _on_browse_excel(self):
        path = filedialog.askopenfilename(
            title="Selecionar Planilha Excel",
            initialdir=os.path.join(BASE_DIR, "Mesa Produtos", "Organizador"),
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self._org_load_file(path)

    def _org_load_file(self, path):
        self._org_input_path = path
        self._org_output_path = None

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            total_rows = 0
            total_cols = 0
            for ws in wb:
                total_rows += ws.max_row or 0
                total_cols = max(total_cols, ws.max_column or 0)
            wb.close()

            fname = os.path.basename(path)
            self.org_file_name.configure(text=fname)
            info = f"{len(sheets)} aba(s)  |  ~{total_rows} linhas  |  {total_cols} colunas"
            self.org_file_info.configure(text=info)
            self.org_file_frame.pack(fill="x", pady=(0, 12))

            # Update drop zone visual
            self.org_drop_title.configure(text=fname)
            self.org_drop_sub.configure(text="Arquivo carregado - clique para trocar")
            self.org_icon_label.configure(text_color=ACCENT_GREEN, text="\u2713")
            self.org_drop_card.configure(border_color=ACCENT_GREEN, fg_color="#f0fff4")

            self._org_show_preview(path)

            self.org_process_btn.configure(state="normal")
            self.org_open_btn.configure(state="disabled")

            self.org_status_dot.configure(text_color=ACCENT_BLUE)
            self.org_status_text.configure(
                text=f"  Arquivo carregado. Clique em 'Processar e Organizar'."
            )

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler arquivo:\n{e}")

    def _org_show_preview(self, path, records=None):
        try:
            self.org_preview_text.configure(state="normal")
            self.org_preview_text.delete("1.0", "end")

            if records:
                # Detect which format by checking keys
                if "produto" in records[0]:
                    # 4-row format
                    headers = ["CONTA", "ATIVO", "PRODUTO", "TIPO", "DT INICIO", "DT VENC", "QTD", "VALOR", "STATUS"]
                    col_widths = [10, 8, 16, 18, 12, 12, 8, 14, 18]
                    header_line = "".join(h.ljust(w) for h, w in zip(headers, col_widths))
                    self.org_preview_text.insert("end", header_line + "\n")
                    self.org_preview_text.insert("end", "-" * sum(col_widths) + "\n")

                    for rec in records[:20]:
                        vals = [
                            rec["conta"][:9], rec["ativo"][:7],
                            rec["produto"][:15], rec["tipo"][:17],
                            rec["data_inicio"][:11], rec["data_vencimento"][:11],
                            rec["qtd_str"][:7], rec["valor_str"][:13],
                            rec["status"][:17],
                        ]
                        line = "".join(v.ljust(w) for v, w in zip(vals, col_widths))
                        self.org_preview_text.insert("end", line + "\n")
                else:
                    # Condensed XP format
                    headers = [
                        "CLIENTE", "CONTA", "ATIVO", "DATA", "HORA",
                        "TIPO", "ORIGEM", "QTD", "VALOR",
                        "QTD EXEC", "VALOR EXEC", "STATUS"
                    ]
                    col_widths = [25, 10, 35, 12, 10, 8, 16, 12, 16, 12, 16, 12]
                    header_line = "".join(h.ljust(w) for h, w in zip(headers, col_widths))
                    self.org_preview_text.insert("end", header_line + "\n")
                    self.org_preview_text.insert("end", "-" * sum(col_widths) + "\n")

                    for rec in records[:20]:
                        vals = [
                            rec["cliente"][:24], rec["conta"][:9],
                            rec["ativo"][:34], rec["data_str"][:11], rec["hora"][:9],
                            rec["tipo"][:7], rec["origem"][:15],
                            rec["qtd_str"][:11], rec["valor_str"][:15],
                            rec["qtd_exec_str"][:11], rec["valor_exec_str"][:15],
                            rec["status"][:11],
                        ]
                        line = "".join(v.ljust(w) for v, w in zip(vals, col_widths))
                        self.org_preview_text.insert("end", line + "\n")

                if len(records) > 20:
                    self.org_preview_text.insert("end", f"\n  ... e mais {len(records) - 20} registros")
            else:
                # Show raw file preview
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                for i, row in enumerate(ws.iter_rows(max_row=min(18, ws.max_row or 18), values_only=True)):
                    cells = [str(c)[:50] if c is not None else "" for c in row]
                    self.org_preview_text.insert("end", "  |  ".join(cells) + "\n")
                wb.close()

            self.org_preview_text.configure(state="disabled")
        except Exception:
            pass

    def _detect_4row_format(self, ws):
        """Detect the 4-row grouped format: conta / ativo+produto / detalhes / status."""
        if (ws.max_column or 0) > 2:
            return False
        max_row = ws.max_row or 0
        if max_row < 4:
            return False
        # Row 1 = numeric (conta), Row 3 has date pattern DD/MM/YYYY, Row 4 = status text
        r1 = ws.cell(row=1, column=1).value
        r3 = sanitize_text(str(ws.cell(row=3, column=1).value or ""))
        r4 = sanitize_text(str(ws.cell(row=4, column=1).value or ""))
        if r1 is not None and re.search(r'\d{2}/\d{2}/\d{4}', r3):
            status_keywords = ["Pendente", "Cancelado", "Executad", "Total", "Parcial", "Registro"]
            if any(kw in r4 for kw in status_keywords):
                return True
        return False

    def _parse_4row_records(self, ws):
        """Parse the 4-row grouped format into structured records."""
        records = []
        max_row = ws.max_row or 0
        # Only iterate complete groups of 4
        usable_rows = (max_row // 4) * 4

        for i in range(0, usable_rows, 4):
            r1 = str(ws.cell(row=i + 1, column=1).value or "").strip()
            r2 = str(ws.cell(row=i + 2, column=1).value or "").strip()
            r3 = sanitize_text(str(ws.cell(row=i + 3, column=1).value or "")).strip()
            r4 = sanitize_text(str(ws.cell(row=i + 4, column=1).value or "")).strip()

            # Row 1: Conta
            conta = r1

            # Row 2: Ativo + Produto (e.g. "AURA33Smart Coupon")
            m_ativo = re.match(r'^([A-Z]{2,6}\d{1,4})(.*)', r2)
            if m_ativo:
                ativo = m_ativo.group(1)
                produto = m_ativo.group(2).strip()
            else:
                ativo = r2
                produto = ""

            # Row 3: Tipo + DataInicio + DataVencimento + Quantidade + Valor
            # e.g. "Nova Contratacao04/03/202605/08/2026310R$ 895,03"
            tipo = ""
            data_inicio = ""
            data_vencimento = ""
            qtd_str = ""
            qtd_num = 0.0
            valor_str = ""
            valor_num = 0.0

            dates = list(re.finditer(r'\d{2}/\d{2}/\d{4}', r3))
            if len(dates) >= 2:
                tipo = r3[:dates[0].start()].strip()
                data_inicio = dates[0].group()
                data_vencimento = dates[1].group()
                after_dates = r3[dates[1].end():]
                # after_dates e.g. "310R$ 895,03" or "2310R$ 6.669,43"
                m_val = re.match(r'^(\d[\d.]*)R\$\s*([\d.,]+)', after_dates)
                if m_val:
                    qtd_str = m_val.group(1)
                    qtd_num = float(qtd_str.replace('.', '').replace(',', '.'))
                    valor_str = m_val.group(2)
                    valor_num = float(valor_str.replace('.', '').replace(',', '.'))
            elif len(dates) == 1:
                tipo = r3[:dates[0].start()].strip()
                data_inicio = dates[0].group()
                after_date = r3[dates[0].end():]
                m_val = re.match(r'^(\d[\d.]*)R\$\s*([\d.,]+)', after_date)
                if m_val:
                    qtd_str = m_val.group(1)
                    qtd_num = float(qtd_str.replace('.', '').replace(',', '.'))
                    valor_str = m_val.group(2)
                    valor_num = float(valor_str.replace('.', '').replace(',', '.'))

            # Row 4: Status
            status = r4

            records.append({
                "conta": conta,
                "ativo": ativo,
                "produto": produto,
                "tipo": tipo,
                "data_inicio": data_inicio,
                "data_vencimento": data_vencimento,
                "qtd_str": qtd_str,
                "qtd_num": qtd_num,
                "valor_str": f"R$ {valor_str}" if valor_str else "",
                "valor_num": valor_num,
                "status": status,
            })

        return records

    def _detect_condensed_format(self, ws):
        """Detect if the sheet uses the condensed XP single-column format."""
        if (ws.max_column or 0) > 2:
            return False
        max_row = ws.max_row or 0
        if max_row < 6 or max_row % 6 != 0:
            return False
        # Check row 6 pattern = "Ver Nota"
        r6 = ws.cell(row=6, column=1).value
        if r6 and "Ver Nota" in str(r6):
            return True
        # Check row 1 has "Conta: "
        r1 = str(ws.cell(row=1, column=1).value or "")
        if "Conta:" in r1:
            return True
        return False

    def _parse_condensed_records(self, ws):
        """Parse the condensed XP format into structured records.
        Uses 'Conta:' markers to find each block instead of fixed 6-row stride,
        to handle injected header rows in the middle of the data."""
        records = []
        max_row = ws.max_row or 0

        # Find all rows that contain "Conta:" — these are block starts
        conta_rows = []
        for r in range(1, max_row + 1):
            v = str(ws.cell(r, 1).value or "")
            if "Conta:" in v:
                conta_rows.append(r)

        for start in conta_rows:
            # Each block: row+0=client, row+1=asset, row+2=date, row+3=data, row+4=status
            if start + 4 > max_row:
                break
            r1 = str(ws.cell(row=start, column=1).value or "")
            r2 = str(ws.cell(row=start + 1, column=1).value or "")
            r3 = str(ws.cell(row=start + 2, column=1).value or "")
            r4 = str(ws.cell(row=start + 3, column=1).value or "")
            r5 = str(ws.cell(row=start + 4, column=1).value or "")

            # --- Row 1: Cliente + Conta ---
            if "Conta:" in r1:
                parts = r1.split("Conta:")
                cliente = sanitize_text(parts[0]).strip()
                conta = parts[1].strip()
            else:
                cliente = sanitize_text(r1).strip()
                conta = ""

            # --- Row 2: Ativo ---
            ativo = sanitize_text(r2).strip()

            # --- Row 3: Data + Hora (M/D/YYYYhh:mm:ss) ---
            m_dt = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})(\d{2}:\d{2}:\d{2})$', r3.strip())
            if m_dt:
                # Convert M/D/YYYY to DD/MM/YYYY
                try:
                    dt = datetime.strptime(m_dt.group(1), "%m/%d/%Y")
                    data_str = dt.strftime("%d/%m/%Y")
                    data_dt = dt
                except ValueError:
                    data_str = m_dt.group(1)
                    data_dt = None
                hora_str = m_dt.group(2)
            else:
                data_str = r3.strip()
                data_dt = None
                hora_str = ""

            # --- Row 4: Tipo + Origem + Qtd + Valor + QtdExec + ValorExec ---
            r4_clean = r4.replace('\xa0', ' ')
            parts_r4 = r4_clean.split('R$ ')

            tipo = ""
            origem = ""
            qtd_str = ""
            valor_str = ""
            qtd_exec_str = ""
            valor_exec_str = ""
            qtd_num = 0.0
            valor_num = 0.0
            qtd_exec_num = 0.0
            valor_exec_num = 0.0

            if len(parts_r4) >= 3:
                # Part 0: "CompraEstoque XP300" or "VendaEstoque clientes2250"
                p0 = parts_r4[0]
                m_p0 = re.match(r'^(Compra|Venda)(.*?)(\d[\d.,]*)$', p0)
                if m_p0:
                    tipo = m_p0.group(1)
                    origem = m_p0.group(2).strip()
                    qtd_str = m_p0.group(3)
                    qtd_num = float(qtd_str.replace('.', '').replace(',', '.'))

                # Part 1: "30.690,99300" -> valor,XX + qtd_exec
                p1 = parts_r4[1]
                m_p1 = re.match(r'^([\d.]+,\d{2})(.+)$', p1)
                if m_p1:
                    valor_str = m_p1.group(1)
                    qtd_exec_str = m_p1.group(2)
                    valor_num = float(valor_str.replace('.', '').replace(',', '.'))
                    qtd_exec_num = float(qtd_exec_str.replace('.', '').replace(',', '.'))

                # Part 2: "30.690,99" -> valor_exec
                valor_exec_str = parts_r4[2].strip()
                valor_exec_num = float(valor_exec_str.replace('.', '').replace(',', '.'))
            elif len(parts_r4) == 2:
                p0 = parts_r4[0]
                m_p0 = re.match(r'^(Compra|Venda)(.*?)(\d[\d.,]*)$', p0)
                if m_p0:
                    tipo = m_p0.group(1)
                    origem = m_p0.group(2).strip()
                    qtd_str = m_p0.group(3)
                    qtd_num = float(qtd_str.replace('.', '').replace(',', '.'))
                valor_str = parts_r4[1].strip()
                valor_num = float(valor_str.replace('.', '').replace(',', '.'))

            # --- Row 5: Status ---
            status = sanitize_text(r5).strip()

            records.append({
                "cliente": cliente,
                "conta": conta,
                "ativo": ativo,
                "data_str": data_str,
                "data_dt": data_dt,
                "hora": hora_str,
                "tipo": tipo,
                "origem": origem,
                "qtd_str": qtd_str,
                "qtd_num": qtd_num,
                "valor_str": valor_str,
                "valor_num": valor_num,
                "qtd_exec_str": qtd_exec_str,
                "qtd_exec_num": qtd_exec_num,
                "valor_exec_str": valor_exec_str,
                "valor_exec_num": valor_exec_num,
                "status": status,
            })

        return records

    def _on_process_excel(self):
        if not self._org_input_path:
            return

        self.org_process_btn.configure(state="disabled")
        self.org_status_dot.configure(text_color=ACCENT_ORANGE)
        self.org_status_text.configure(text="  Processando...")

        threading.Thread(target=self._run_process_excel, daemon=True).start()

    def _run_process_excel(self):
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter

            path = self._org_input_path
            wb_in = openpyxl.load_workbook(path, data_only=True)
            ws_in = wb_in[wb_in.sheetnames[0]]

            is_4row = self._detect_4row_format(ws_in)
            is_condensed = self._detect_condensed_format(ws_in) if not is_4row else False

            if is_4row:
                records_4row = self._parse_4row_records(ws_in)
                records = None
            elif is_condensed:
                records_4row = None
                records = self._parse_condensed_records(ws_in)
            else:
                records_4row = None
                records = None

            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = "Dados Organizados"

            # Styles - Somus Capital branding
            header_fill = PatternFill(start_color="004d33", end_color="004d33", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            data_font = Font(name="Calibri", size=10)
            data_align_left = Alignment(horizontal="left", vertical="center")
            data_align_center = Alignment(horizontal="center", vertical="center")
            data_align_right = Alignment(horizontal="right", vertical="center")

            alt_fill = PatternFill(start_color="f0f7f4", end_color="f0f7f4", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            thin_border = Border(
                left=Side(style="thin", color="d0d0d0"),
                right=Side(style="thin", color="d0d0d0"),
                top=Side(style="thin", color="d0d0d0"),
                bottom=Side(style="thin", color="d0d0d0"),
            )

            if records_4row:
                # ===== 4-ROW GROUPED FORMAT (Smart Coupon etc.) =====
                headers = [
                    "CONTA", "ATIVO", "PRODUTO", "TIPO",
                    "DATA INICIO", "DATA VENCIMENTO",
                    "QUANTIDADE", "VALOR (R$)", "STATUS"
                ]
                col_widths = [14, 12, 20, 22, 16, 18, 14, 18, 20]

                for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
                    cell = ws_out.cell(row=1, column=col_idx)
                    cell.value = hdr
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border
                    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

                for row_idx, rec in enumerate(records_4row):
                    out_row = row_idx + 2
                    fill = alt_fill if row_idx % 2 == 0 else white_fill

                    row_values = [
                        (int(rec["conta"]) if rec["conta"].isdigit() else rec["conta"], data_align_center, None),
                        (rec["ativo"], data_align_center, None),
                        (rec["produto"], data_align_left, None),
                        (rec["tipo"], data_align_left, None),
                        (rec["data_inicio"], data_align_center, None),
                        (rec["data_vencimento"], data_align_center, None),
                        (rec["qtd_num"], data_align_right, '#,##0'),
                        (rec["valor_num"], data_align_right, '#,##0.00'),
                        (rec["status"], data_align_center, None),
                    ]

                    for col_idx, (val, align, num_fmt) in enumerate(row_values, 1):
                        cell = ws_out.cell(row=out_row, column=col_idx)
                        cell.value = val if val else ""
                        cell.font = data_font
                        cell.fill = fill
                        cell.border = thin_border
                        cell.alignment = align
                        if num_fmt and val:
                            cell.number_format = num_fmt

                total_processed = len(records_4row)
                records = records_4row  # for preview

            elif records:
                # ===== CONDENSED XP FORMAT =====
                headers = [
                    "CLIENTE", "CONTA", "ATIVO", "DATA", "HORA",
                    "TIPO", "ORIGEM", "QUANTIDADE", "VALOR (R$)",
                    "QTD EXECUTADA", "VALOR EXECUTADO (R$)", "STATUS"
                ]
                col_widths = [35, 14, 42, 14, 12, 12, 20, 16, 20, 16, 22, 14]

                # Write headers
                for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), 1):
                    cell = ws_out.cell(row=1, column=col_idx)
                    cell.value = hdr
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border
                    ws_out.column_dimensions[get_column_letter(col_idx)].width = width

                # Write data
                for row_idx, rec in enumerate(records):
                    out_row = row_idx + 2
                    fill = alt_fill if row_idx % 2 == 0 else white_fill

                    row_values = [
                        (rec["cliente"], data_align_left, None),
                        (int(rec["conta"]) if rec["conta"].isdigit() else rec["conta"], data_align_center, None),
                        (rec["ativo"], data_align_left, None),
                        (rec["data_dt"], data_align_center, "DD/MM/YYYY"),
                        (rec["hora"], data_align_center, None),
                        (rec["tipo"], data_align_center, None),
                        (rec["origem"], data_align_left, None),
                        (rec["qtd_num"], data_align_right, '#,##0.000' if rec["qtd_num"] != int(rec["qtd_num"]) else '#,##0'),
                        (rec["valor_num"], data_align_right, '#,##0.00'),
                        (rec["qtd_exec_num"], data_align_right, '#,##0.000' if rec["qtd_exec_num"] != int(rec["qtd_exec_num"]) else '#,##0'),
                        (rec["valor_exec_num"], data_align_right, '#,##0.00'),
                        (rec["status"], data_align_center, None),
                    ]

                    for col_idx, (val, align, num_fmt) in enumerate(row_values, 1):
                        cell = ws_out.cell(row=out_row, column=col_idx)
                        cell.value = val if val else ""
                        cell.font = data_font
                        cell.fill = fill
                        cell.border = thin_border
                        cell.alignment = align
                        if num_fmt and val:
                            cell.number_format = num_fmt

                total_processed = len(records)

            else:
                # ===== GENERIC FORMAT =====
                data = []
                for row in ws_in.iter_rows(values_only=True):
                    data.append(list(row))

                if not data:
                    raise ValueError("Planilha vazia")

                header_idx = 0
                max_filled = 0
                for i, row in enumerate(data[:10]):
                    filled = sum(1 for c in row if c is not None and str(c).strip())
                    if filled > max_filled:
                        max_filled = filled
                        header_idx = i

                headers_row = data[header_idx]
                data_rows = data[header_idx + 1:]
                data_rows = [
                    r for r in data_rows
                    if any(c is not None and str(c).strip() for c in r)
                ]

                n_cols = len(headers_row)
                used_cols = []
                for c in range(n_cols):
                    hv = headers_row[c]
                    if hv is not None and str(hv).strip():
                        used_cols.append(c)
                    elif any(r[c] is not None and str(r[c]).strip() for r in data_rows if c < len(r)):
                        used_cols.append(c)

                if not used_cols:
                    raise ValueError("Nenhuma coluna com dados encontrada")

                for out_col, src_col in enumerate(used_cols, 1):
                    val = headers_row[src_col]
                    cell = ws_out.cell(row=1, column=out_col)
                    cell.value = sanitize_text(val).strip().upper() if val else f"COLUNA {out_col}"
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

                for row_idx, row_data in enumerate(data_rows):
                    out_row = row_idx + 2
                    fill = alt_fill if row_idx % 2 == 0 else white_fill
                    for out_col, src_col in enumerate(used_cols, 1):
                        val = row_data[src_col] if src_col < len(row_data) else None
                        cell = ws_out.cell(row=out_row, column=out_col)
                        if isinstance(val, str):
                            val = sanitize_text(val).strip()
                        cell.value = val
                        cell.font = data_font
                        cell.fill = fill
                        cell.border = thin_border
                        if isinstance(val, datetime):
                            cell.number_format = "DD/MM/YYYY"
                            cell.alignment = data_align_center
                        elif isinstance(val, (int, float)):
                            cell.alignment = data_align_right
                            cell.number_format = '#,##0.00'
                        else:
                            cell.alignment = data_align_left

                for out_col, src_col in enumerate(used_cols, 1):
                    max_len = len(str(headers_row[src_col] or ""))
                    for rd in data_rows[:200]:
                        val = rd[src_col] if src_col < len(rd) else None
                        if val:
                            max_len = max(max_len, min(len(str(val)), 50))
                    ws_out.column_dimensions[get_column_letter(out_col)].width = min(max_len + 4, 55)

                total_processed = len(data_rows)

            # Autofilter + freeze
            last_col_letter = get_column_letter(ws_out.max_column)
            ws_out.auto_filter.ref = f"A1:{last_col_letter}{ws_out.max_row}"
            ws_out.freeze_panes = "A2"

            # Save
            os.makedirs(ORGANIZADOR_OUTPUT_DIR, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(path))[0]
            output_name = f"{base_name}_ORGANIZADO.xlsx"
            output_path = os.path.join(ORGANIZADOR_OUTPUT_DIR, output_name)

            wb_out.save(output_path)
            wb_in.close()

            self._org_output_path = output_path

            def _update_ok():
                self.org_process_btn.configure(state="normal")
                self.org_open_btn.configure(state="normal")
                self.org_status_dot.configure(text_color=ACCENT_GREEN)
                self.org_status_text.configure(
                    text=f"  Concluído! {total_processed} registros organizados  -  {output_name}"
                )
                self.org_drop_card.configure(border_color=ACCENT_GREEN)
                self._org_show_preview(output_path, records=records)
                self.org_preview_label.configure(text="Preview do Resultado")

            self.after(0, _update_ok)

        except Exception as e:
            def _update_err():
                self.org_process_btn.configure(state="normal")
                self.org_status_dot.configure(text_color=ACCENT_RED)
                self.org_status_text.configure(text=f"  Erro: {e}")

            self.after(0, _update_err)

    def _on_open_organized(self):
        if self._org_output_path and os.path.exists(self._org_output_path):
            os.startfile(self._org_output_path)

    def _on_open_org_folder(self):
        os.makedirs(ORGANIZADOR_OUTPUT_DIR, exist_ok=True)
        os.startfile(ORGANIZADOR_OUTPUT_DIR)

    # -----------------------------------------------------------------
    #  PAGE: ENVIO SALDOS
    # -----------------------------------------------------------------
    def _build_saldos_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Envio Saldos", subtitle="Saldos Diários")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ======== DROP ZONE ========
        self.sld_drop_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=14,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.sld_drop_card.pack(fill="x", pady=(0, 16))
        self.sld_drop_card.bind("<Button-1>", lambda e: self._on_browse_saldos())

        drop_inner = ctk.CTkFrame(self.sld_drop_card, fg_color="transparent")
        drop_inner.pack(fill="x", padx=30, pady=36)
        drop_inner.bind("<Button-1>", lambda e: self._on_browse_saldos())

        self.sld_icon_label = ctk.CTkLabel(
            drop_inner, text="\u21c6",
            font=("Segoe UI", 44), text_color=ACCENT_BLUE
        )
        self.sld_icon_label.pack()
        self.sld_icon_label.bind("<Button-1>", lambda e: self._on_browse_saldos())

        self.sld_drop_title = ctk.CTkLabel(
            drop_inner, text="Clique para selecionar a planilha de saldos",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        )
        self.sld_drop_title.pack(pady=(8, 2))
        self.sld_drop_title.bind("<Button-1>", lambda e: self._on_browse_saldos())

        self.sld_drop_sub = ctk.CTkLabel(
            drop_inner, text="Formatos aceitos:  .xlsx  .xls",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.sld_drop_sub.pack()
        self.sld_drop_sub.bind("<Button-1>", lambda e: self._on_browse_saldos())

        self.sld_browse_btn = ctk.CTkButton(
            drop_inner, text="  Selecionar Arquivo",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=40, corner_radius=8, width=200,
            command=self._on_browse_saldos,
        )
        self.sld_browse_btn.pack(pady=(14, 0))

        # ======== FILE INFO BAR (hidden initially) ========
        self.sld_file_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.sld_file_frame.pack(fill="x", pady=(0, 12))
        self.sld_file_frame.pack_forget()

        fi = ctk.CTkFrame(self.sld_file_frame, fg_color="transparent")
        fi.pack(fill="x", padx=16, pady=10)

        self.sld_file_icon = ctk.CTkLabel(
            fi, text="\u25cf", font=("Segoe UI", 12), text_color=ACCENT_GREEN
        )
        self.sld_file_icon.pack(side="left")

        self.sld_file_name = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY
        )
        self.sld_file_name.pack(side="left", padx=(6, 0))

        self.sld_file_info = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 10), text_color=TEXT_SECONDARY
        )
        self.sld_file_info.pack(side="left", padx=(12, 0))

        ctk.CTkButton(
            fi, text="Trocar", font=("Segoe UI", 10),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=28, corner_radius=6, width=70,
            command=self._on_browse_saldos,
        ).pack(side="right")

        # ======== PREVIEW ========
        self.sld_preview_label = ctk.CTkLabel(
            content, text="Preview dos Dados",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        self.sld_preview_label.pack(fill="x", pady=(4, 8))

        self.sld_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        self.sld_preview_frame.pack(fill="x", pady=(0, 14))

        self.sld_preview_text = ctk.CTkTextbox(
            self.sld_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=200, wrap="none"
        )
        self.sld_preview_text.pack(fill="x", padx=4, pady=4)
        self.sld_preview_text.insert("1.0", "  Nenhum arquivo carregado...")
        self.sld_preview_text.configure(state="disabled")

        # ======== ACTION BUTTONS ========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 12))

        self.sld_process_btn = ctk.CTkButton(
            btn_frame, text="  Processar Saldos",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=self._darken(ACCENT_GREEN),
            height=44, corner_radius=10, state="disabled",
            command=self._on_process_saldos,
        )
        self.sld_process_btn.pack(side="left", padx=(0, 10))

        self.sld_open_btn = ctk.CTkButton(
            btn_frame, text="  Abrir Resultado",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=44, corner_radius=10, state="disabled",
            command=self._on_open_saldos_result,
        )
        self.sld_open_btn.pack(side="left", padx=(0, 10))

        # ======== STATUS BAR ========
        self.sld_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.sld_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.sld_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.sld_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.sld_status_dot.pack(side="left")

        self.sld_status_text = ctk.CTkLabel(
            si, text="  Aguardando arquivo...",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.sld_status_text.pack(side="left")

        # Botao Atualizar
        self._make_atualizar_btn(content, os.path.join(SALDOS_DIR, "BASE", "BASE ENVIAR.xlsx"))

        # Internal state
        self._sld_input_path = None
        self._sld_output_path = None

        return page

    # -----------------------------------------------------------------
    #  ENVIO SALDOS: Ações
    # -----------------------------------------------------------------
    def _on_browse_saldos(self):
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        path = filedialog.askopenfilename(
            title="Selecionar Planilha de Saldos",
            initialdir=downloads_dir if os.path.isdir(downloads_dir) else os.path.join(SALDOS_DIR, "BASE"),
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self._sld_load_file(path)

    def _sld_load_file(self, path):
        self._sld_input_path = path
        self._sld_output_path = None

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            total_rows = 0
            total_cols = 0
            for ws in wb:
                total_rows += ws.max_row or 0
                total_cols = max(total_cols, ws.max_column or 0)
            wb.close()

            fname = os.path.basename(path)
            self.sld_file_name.configure(text=fname)

            # Verifica se o arquivo foi modificado hoje
            from datetime import datetime
            file_date = datetime.fromtimestamp(os.path.getmtime(path)).date()
            hoje = datetime.now().date()
            date_warning = ""
            if file_date < hoje:
                date_warning = f"  |  ATENCAO: arquivo de {file_date.strftime('%d/%m/%Y')}"
                messagebox.showwarning(
                    "Arquivo Desatualizado",
                    f"Este arquivo foi modificado em {file_date.strftime('%d/%m/%Y')}.\n"
                    f"Hoje e {hoje.strftime('%d/%m/%Y')}.\n\n"
                    f"Certifique-se de que esta usando o relatorio de saldos mais recente."
                )

            info = f"{len(sheets)} aba(s)  |  ~{total_rows} linhas  |  {total_cols} colunas{date_warning}"
            self.sld_file_info.configure(text=info)
            self.sld_file_frame.pack(fill="x", pady=(0, 12))

            # Update drop zone visual
            if file_date < hoje:
                self.sld_drop_title.configure(text=fname)
                self.sld_drop_sub.configure(text=f"Arquivo de {file_date.strftime('%d/%m/%Y')} - clique para trocar")
                self.sld_icon_label.configure(text_color=ACCENT_ORANGE, text="\u26a0")
                self.sld_drop_card.configure(border_color=ACCENT_ORANGE, fg_color="#fffbeb")
            else:
                self.sld_drop_title.configure(text=fname)
                self.sld_drop_sub.configure(text="Arquivo carregado - clique para trocar")
                self.sld_icon_label.configure(text_color=ACCENT_GREEN, text="\u2713")
                self.sld_drop_card.configure(border_color=ACCENT_GREEN, fg_color="#f0fff4")

            self._sld_show_preview(path)

            self.sld_process_btn.configure(state="normal")
            self.sld_open_btn.configure(state="disabled")

            self.sld_status_dot.configure(text_color=ACCENT_BLUE)
            self.sld_status_text.configure(
                text=f"  Arquivo carregado. Clique em 'Processar Saldos'."
            )

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler arquivo:\n{e}")

    def _sld_show_preview(self, path):
        try:
            self.sld_preview_text.configure(state="normal")
            self.sld_preview_text.delete("1.0", "end")

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(max_row=min(18, ws.max_row or 18), values_only=True)):
                cells = [str(c)[:50] if c is not None else "" for c in row]
                self.sld_preview_text.insert("end", "  |  ".join(cells) + "\n")
            wb.close()

            self.sld_preview_text.configure(state="disabled")
        except Exception:
            pass

    def _on_process_saldos(self):
        if not self._sld_input_path:
            messagebox.showwarning("Aviso", "Selecione um arquivo primeiro.")
            return

        confirm = messagebox.askyesno(
            "Gerar Rascunhos no Outlook",
            "Isso vai criar rascunhos de email no Outlook para cada assessor "
            "com os saldos dos seus clientes.\n\n"
            "Os emails NÃO serão enviados, apenas salvos como rascunho.\n"
            "O Outlook precisa estar aberto. Deseja continuar?"
        )
        if not confirm:
            return

        self.sld_process_btn.configure(state="disabled")
        self.sld_status_dot.configure(text_color=ACCENT_ORANGE)
        self.sld_status_text.configure(text="  Processando saldos...")

        # Atualiza preview com log
        self.sld_preview_text.configure(state="normal")
        self.sld_preview_text.delete("1.0", "end")
        self.sld_preview_text.configure(state="disabled")

        threading.Thread(target=self._run_process_saldos, daemon=True).start()

    def _run_process_saldos(self):
        try:
            from processar_saldos import processar_saldos

            def log_callback(msg, tipo="info"):
                self.after(0, self._sld_append_log, msg, tipo)

            stats = processar_saldos(self._sld_input_path, callback=log_callback)

            if stats["criados"] > 0:
                self.after(0, self.sld_status_dot.configure, {"text_color": ACCENT_GREEN})
                self.after(0, self.sld_status_text.configure, {
                    "text": f"  Concluído - {stats['criados']} rascunhos criados no Outlook"
                })
                self.after(0, lambda: messagebox.showinfo(
                    "Rascunhos Criados",
                    f"{stats['criados']} rascunhos criados nos Rascunhos do Outlook!\n\n"
                    f"Erros: {stats['erros']}  |  Sem email: {stats['sem_email']}"
                ))
            elif stats["erros"] > 0:
                self.after(0, self.sld_status_dot.configure, {"text_color": ACCENT_RED})
                self.after(0, self.sld_status_text.configure, {
                    "text": f"  Erros no processamento - verifique o log"
                })
            else:
                self.after(0, self.sld_status_dot.configure, {"text_color": ACCENT_ORANGE})
                self.after(0, self.sld_status_text.configure, {
                    "text": f"  Nenhum rascunho criado (sem assessores com email)"
                })

        except Exception as e:
            err_msg = str(e)
            self.after(0, self.sld_status_dot.configure, {"text_color": ACCENT_RED})
            self.after(0, self.sld_status_text.configure, {"text": f"  Erro: {err_msg}"})
            self.after(0, lambda m=err_msg: messagebox.showerror("Erro", m))
        finally:
            self.after(0, self.sld_process_btn.configure, {"state": "normal"})

    def _sld_append_log(self, msg, tipo="info"):
        color_map = {"ok": ACCENT_GREEN, "error": ACCENT_RED, "skip": ACCENT_ORANGE, "info": TEXT_PRIMARY}
        self.sld_preview_text.configure(state="normal")
        self.sld_preview_text.insert("end", msg + "\n")
        self.sld_preview_text.see("end")
        self.sld_preview_text.configure(state="disabled")

    def _on_open_saldos_result(self):
        if self._sld_output_path and os.path.exists(self._sld_output_path):
            os.startfile(self._sld_output_path)

    def _on_open_saldos_folder(self):
        os.makedirs(SALDOS_DIR, exist_ok=True)
        os.startfile(SALDOS_DIR)

    # -----------------------------------------------------------------
    #  SEGUROS: Renovações Anuais
    # -----------------------------------------------------------------
    def _build_renovacoes_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Renovações Anuais", subtitle="Seguros")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ======== DROP ZONE ========
        self.ren_drop_card = ctk.CTkFrame(
            content, fg_color="#f0f4ff", corner_radius=14,
            border_width=2, border_color=ACCENT_BLUE
        )
        self.ren_drop_card.pack(fill="x", pady=(0, 16))
        self.ren_drop_card.bind("<Button-1>", lambda e: self._on_browse_renovacoes())

        drop_inner = ctk.CTkFrame(self.ren_drop_card, fg_color="transparent")
        drop_inner.pack(fill="x", padx=30, pady=36)
        drop_inner.bind("<Button-1>", lambda e: self._on_browse_renovacoes())

        self.ren_icon_label = ctk.CTkLabel(
            drop_inner, text="\u26e8",
            font=("Segoe UI", 44), text_color=ACCENT_BLUE
        )
        self.ren_icon_label.pack()
        self.ren_icon_label.bind("<Button-1>", lambda e: self._on_browse_renovacoes())

        self.ren_drop_title = ctk.CTkLabel(
            drop_inner, text="Clique para selecionar a planilha de renovações",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        )
        self.ren_drop_title.pack(pady=(8, 2))
        self.ren_drop_title.bind("<Button-1>", lambda e: self._on_browse_renovacoes())

        self.ren_drop_sub = ctk.CTkLabel(
            drop_inner, text="Formatos aceitos:  .xlsx  .xls",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.ren_drop_sub.pack()
        self.ren_drop_sub.bind("<Button-1>", lambda e: self._on_browse_renovacoes())

        ren_btn_row = ctk.CTkFrame(drop_inner, fg_color="transparent")
        ren_btn_row.pack(pady=(14, 0))

        self.ren_browse_btn = ctk.CTkButton(
            ren_btn_row, text="  Selecionar Arquivo",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=40, corner_radius=8, width=200,
            command=self._on_browse_renovacoes,
        )
        self.ren_browse_btn.pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            ren_btn_row, text="  \u2913  Baixar do SharePoint",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            height=40, corner_radius=8, width=220,
            command=self._on_baixar_renovacoes,
        ).pack(side="left")

        # ======== FILE INFO BAR (hidden initially) ========
        self.ren_file_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.ren_file_frame.pack(fill="x", pady=(0, 12))
        self.ren_file_frame.pack_forget()

        fi = ctk.CTkFrame(self.ren_file_frame, fg_color="transparent")
        fi.pack(fill="x", padx=16, pady=10)

        self.ren_file_icon = ctk.CTkLabel(
            fi, text="\u25cf", font=("Segoe UI", 12), text_color=ACCENT_GREEN
        )
        self.ren_file_icon.pack(side="left")

        self.ren_file_name = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY
        )
        self.ren_file_name.pack(side="left", padx=(6, 0))

        self.ren_file_info = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 10), text_color=TEXT_SECONDARY
        )
        self.ren_file_info.pack(side="left", padx=(12, 0))

        ctk.CTkButton(
            fi, text="Trocar", font=("Segoe UI", 10),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=28, corner_radius=6, width=70,
            command=self._on_browse_renovacoes,
        ).pack(side="right")

        # ======== PREVIEW ========
        self.ren_preview_label = ctk.CTkLabel(
            content, text="Preview dos Dados",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        self.ren_preview_label.pack(fill="x", pady=(4, 8))

        self.ren_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        self.ren_preview_frame.pack(fill="x", pady=(0, 14))

        self.ren_preview_text = ctk.CTkTextbox(
            self.ren_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=200, wrap="none"
        )
        self.ren_preview_text.pack(fill="x", padx=4, pady=4)
        self.ren_preview_text.insert("1.0", "  Nenhum arquivo carregado...")
        self.ren_preview_text.configure(state="disabled")

        # ======== ACTION BUTTONS ========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 12))

        self.ren_process_btn = ctk.CTkButton(
            btn_frame, text="  Processar Renovações",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=self._darken(ACCENT_GREEN),
            height=44, corner_radius=10, state="disabled",
            command=self._on_process_renovacoes,
        )
        self.ren_process_btn.pack(side="left", padx=(0, 10))

        # ======== STATUS BAR ========
        self.ren_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.ren_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.ren_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.ren_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.ren_status_dot.pack(side="left")

        self.ren_status_text = ctk.CTkLabel(
            si, text="  Aguardando arquivo...",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.ren_status_text.pack(side="left")

        # Botao Atualizar — abre pasta Seguros no SharePoint
        atualizar_frame = ctk.CTkFrame(content, fg_color="transparent")
        atualizar_frame.pack(fill="x", pady=(24, 8))
        ctk.CTkButton(
            atualizar_frame, text="\u21bb  Atualizar",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_WHITE, height=42, corner_radius=10,
            command=lambda: self._safe_open_folder(SEGUROS_DIR)
        ).pack(fill="x")

        # Internal state
        self._ren_input_path = None
        self._ren_output_path = None

        return page

    def _safe_open_folder(self, path):
        try:
            os.makedirs(path, exist_ok=True)
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a pasta:\n{e}")

    def _on_browse_renovacoes(self):
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        path = filedialog.askopenfilename(
            title="Selecionar Planilha de Renovações",
            initialdir=downloads_dir,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self._ren_load_file(path)

    def _on_baixar_renovacoes(self):
        """Copia arquivos de renovações do SharePoint para Downloads e carrega."""
        import shutil
        downloads_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        if not os.path.isdir(SEGUROS_DIR):
            messagebox.showwarning("Aviso", f"Pasta de Seguros não encontrada:\n{SEGUROS_DIR}")
            return

        found = []
        try:
            for f in os.listdir(SEGUROS_DIR):
                if f.lower().endswith(('.xlsx', '.xls')) and 'renov' in f.lower():
                    found.append(f)
        except (PermissionError, OSError) as e:
            messagebox.showerror("Erro", f"Sem acesso à pasta do SharePoint:\n{e}")
            return

        if not found:
            messagebox.showinfo("Info", "Nenhuma planilha de renovações encontrada na pasta do SharePoint.")
            return

        copied = []
        for fname in found:
            src = os.path.join(SEGUROS_DIR, fname)
            dst = os.path.join(downloads_dir, fname)
            try:
                shutil.copy2(src, dst)
                copied.append(dst)
            except (PermissionError, OSError) as e:
                messagebox.showwarning("Aviso", f"Não foi possível copiar {fname}:\n{e}\n\nTente abrir o arquivo no SharePoint primeiro para baixá-lo.")

        if not copied:
            return

        if len(copied) == 1:
            self._ren_load_file(copied[0])
        else:
            messagebox.showinfo(
                "Arquivos Copiados",
                f"{len(copied)} planilhas copiadas para Downloads.\nSelecione qual deseja processar."
            )
            path = filedialog.askopenfilename(
                title="Selecionar Planilha de Renovações",
                initialdir=downloads_dir,
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if path:
                self._ren_load_file(path)

    def _ren_load_file(self, path):
        self._ren_input_path = path
        self._ren_output_path = None

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames
            total_rows = 0
            total_cols = 0
            for ws in wb:
                total_rows += ws.max_row or 0
                total_cols = max(total_cols, ws.max_column or 0)
            wb.close()

            fname = os.path.basename(path)
            self.ren_file_name.configure(text=fname)

            info = f"{len(sheets)} aba(s)  |  ~{total_rows} linhas  |  {total_cols} colunas"
            self.ren_file_info.configure(text=info)
            self.ren_file_frame.pack(fill="x", pady=(0, 12))

            self.ren_drop_title.configure(text=fname)
            self.ren_drop_sub.configure(text="Arquivo carregado - clique para trocar")
            self.ren_icon_label.configure(text_color=ACCENT_GREEN, text="\u2713")
            self.ren_drop_card.configure(border_color=ACCENT_GREEN, fg_color="#f0fff4")

            self._ren_show_preview(path)

            self.ren_process_btn.configure(state="normal")

            self.ren_status_dot.configure(text_color=ACCENT_BLUE)
            self.ren_status_text.configure(
                text=f"  Arquivo carregado. Clique em 'Processar Renovações'."
            )

        except PermissionError:
            messagebox.showerror(
                "Permissão Negada",
                f"Sem permissão para ler o arquivo:\n{path}\n\n"
                "Se o arquivo está no SharePoint/OneDrive, use o botão "
                "'Baixar do SharePoint' para copiá-lo para Downloads primeiro."
            )
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler arquivo:\n{e}")

    def _ren_show_preview(self, path):
        try:
            self.ren_preview_text.configure(state="normal")
            self.ren_preview_text.delete("1.0", "end")

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for i, row in enumerate(ws.iter_rows(max_row=min(18, ws.max_row or 18), values_only=True)):
                cells = [str(c)[:50] if c is not None else "" for c in row]
                self.ren_preview_text.insert("end", "  |  ".join(cells) + "\n")
            wb.close()

            self.ren_preview_text.configure(state="disabled")
        except Exception:
            pass

    def _on_process_renovacoes(self):
        if not self._ren_input_path:
            messagebox.showwarning("Aviso", "Selecione um arquivo primeiro.")
            return

        confirm = messagebox.askyesno(
            "Gerar Rascunhos no Outlook",
            "Isso vai criar rascunhos de email no Outlook para cada assessor "
            "com as renovações anuais de seguros dos seus clientes.\n\n"
            "Os emails NÃO serão enviados, apenas salvos como rascunho.\n"
            "O Outlook precisa estar aberto. Deseja continuar?"
        )
        if not confirm:
            return

        self.ren_process_btn.configure(state="disabled")
        self.ren_status_dot.configure(text_color=ACCENT_ORANGE)
        self.ren_status_text.configure(text="  Processando renovações...")

        self.ren_preview_text.configure(state="normal")
        self.ren_preview_text.delete("1.0", "end")
        self.ren_preview_text.configure(state="disabled")

        threading.Thread(target=self._run_process_renovacoes, daemon=True).start()

    def _run_process_renovacoes(self):
        try:
            from processar_renovacoes import processar_renovacoes

            def log_callback(msg, tipo="info"):
                self.after(0, self._ren_append_log, msg, tipo)

            stats = processar_renovacoes(self._ren_input_path, callback=log_callback)

            if stats["criados"] > 0:
                self.after(0, self.ren_status_dot.configure, {"text_color": ACCENT_GREEN})
                self.after(0, self.ren_status_text.configure, {
                    "text": f"  Concluído - {stats['criados']} rascunhos criados no Outlook"
                })
                self.after(0, lambda: messagebox.showinfo(
                    "Rascunhos Criados",
                    f"{stats['criados']} rascunhos criados nos Rascunhos do Outlook!\n\n"
                    f"Erros: {stats['erros']}  |  Sem email: {stats['sem_email']}"
                ))
            elif stats["erros"] > 0:
                self.after(0, self.ren_status_dot.configure, {"text_color": ACCENT_RED})
                self.after(0, self.ren_status_text.configure, {
                    "text": f"  Erros no processamento - verifique o log"
                })
            else:
                self.after(0, self.ren_status_dot.configure, {"text_color": ACCENT_ORANGE})
                self.after(0, self.ren_status_text.configure, {
                    "text": f"  Nenhum rascunho criado (sem assessores com email)"
                })

        except Exception as e:
            err_msg = str(e)
            self.after(0, self.ren_status_dot.configure, {"text_color": ACCENT_RED})
            self.after(0, self.ren_status_text.configure, {"text": f"  Erro: {err_msg}"})
            self.after(0, lambda m=err_msg: messagebox.showerror("Erro", m))
        finally:
            self.after(0, self.ren_process_btn.configure, {"state": "normal"})

    def _ren_append_log(self, msg, tipo="info"):
        self.ren_preview_text.configure(state="normal")
        self.ren_preview_text.insert("end", msg + "\n")
        self.ren_preview_text.see("end")
        self.ren_preview_text.configure(state="disabled")

    # -----------------------------------------------------------------
    #  CONSÓRCIO: Ações
    # -----------------------------------------------------------------
    def _parse_number(self, text):
        s = text.strip().replace("%", "").replace("R$", "").replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)

    def _on_calcular_consorcio(self):
        try:
            nome = self.cs_nome.get().strip()
            assessor = self.cs_assessor.get().strip()
            tipo_bem = self.cs_tipo_bem.get()
            administradora = self.cs_administradora.get().strip()
            valor_carta = self._parse_number(self.cs_valor_carta.get())
            prazo = int(self.cs_prazo.get())
            taxa_adm = self._parse_number(self.cs_taxa_adm.get())
            fundo_reserva = self._parse_number(self.cs_fundo_reserva.get())
            seguro = self._parse_number(self.cs_seguro.get())
            prazo_contemp = int(self._parse_number(self.cs_prazo_contemplação.get()))
            lance_livre_pct = self._parse_number(self.cs_lance_livre.get())
            lance_embutido_pct = self._parse_number(self.cs_lance_embutido.get())
            correção_anual = self._parse_number(self.cs_correção_anual.get())
            tipo_correção = self.cs_tipo_correção.get()
            índice_correção = self.cs_índice_correção.get()
        except (ValueError, AttributeError):
            messagebox.showerror("Erro", "Preencha todos os campos numericos corretamente.")
            return

        # Parcela reduzida
        pr_text = self.cs_parcela_reduzida.get()
        if "50%" in pr_text:
            parcela_red_pct = 50
        elif "70%" in pr_text:
            parcela_red_pct = 70
        else:
            parcela_red_pct = 100

        # Validações
        if valor_carta <= 0 or prazo <= 0:
            messagebox.showerror("Erro", "Valor da carta e prazo devem ser maiores que zero.")
            return
        if prazo_contemp <= 0 or prazo_contemp > prazo:
            messagebox.showerror("Erro", f"Prazo de contemplação deve ser entre 1 e {prazo}.")
            return
        if lance_embutido_pct > 30:
            messagebox.showwarning("Aviso",
                "Lance embutido acima de 30% e incomum.\nVerifique com a administradora.")

        # === CÁLCULOS BASE (mensais fixos sobre prazo total) ===
        fc_integral = valor_carta / prazo
        ta_mensal = (valor_carta * taxa_adm / 100) / prazo
        fr_mensal = (valor_carta * fundo_reserva / 100) / prazo
        sg_mensal = (valor_carta * seguro / 100) / prazo
        taxas_mensais = ta_mensal + fr_mensal + sg_mensal

        # === FASE 1: Antes da contemplação ===
        fc_reduzido = fc_integral * (parcela_red_pct / 100)
        parcela_fase1 = fc_reduzido + taxas_mensais
        fundo_pago_fase1 = fc_reduzido * prazo_contemp

        # === LANCES (no momento da contemplação) ===
        lance_livre_valor = valor_carta * (lance_livre_pct / 100)
        lance_embutido_valor = valor_carta * (lance_embutido_pct / 100)
        lance_total = lance_livre_valor + lance_embutido_valor
        carta_líquida = valor_carta - lance_embutido_valor

        # === FASE 2: Após contemplação ===
        meses_restantes = prazo - prazo_contemp
        fundo_restante = max(0, valor_carta - fundo_pago_fase1 - lance_total)
        fc_fase2 = fundo_restante / meses_restantes if meses_restantes > 0 else 0
        parcela_fase2 = fc_fase2 + taxas_mensais

        # === CORREÇÃO ANUAL ===
        taxa_corr = correção_anual / 100  # ex: 5.5% -> 0.055

        # Calcular parcelas corrigidas mês a mês (fator composto anual)
        desemb_fase1_corr = 0.0
        desemb_fase2_corr = 0.0
        parcela_f1_final = parcela_fase1  # última parcela fase 1 (corrigida)
        parcela_f2_final = parcela_fase2  # última parcela fase 2 (corrigida)

        for mes in range(1, prazo + 1):
            fator = (1 + taxa_corr) ** ((mes - 1) // 12)
            if mes <= prazo_contemp:
                p_corr = parcela_fase1 * fator
                desemb_fase1_corr += p_corr
                parcela_f1_final = p_corr
            else:
                p_corr = parcela_fase2 * fator
                desemb_fase2_corr += p_corr
                parcela_f2_final = p_corr

        # === TOTAIS ===
        desemb_fase1 = parcela_fase1 * prazo_contemp
        desemb_fase2 = parcela_fase2 * meses_restantes if meses_restantes > 0 else 0
        total_desembolsado = desemb_fase1 + lance_livre_valor + desemb_fase2
        total_taxas = taxas_mensais * prazo

        # Totais com correção
        total_desemb_corr = desemb_fase1_corr + lance_livre_valor + desemb_fase2_corr
        tem_correção = correção_anual > 0

        # === CUSTO EFETIVO ===
        from gerar_pdf_consorcio import _calc_custo_efetivo
        total_ref = total_desemb_corr if tem_correção else total_desembolsado
        ce_mensal, ce_anual, ce_pct = _calc_custo_efetivo(carta_líquida, total_ref, prazo)
        custo_total = total_ref - carta_líquida

        # Salvar dados para o PDF
        self._consórcio_dados = {
            "cliente_nome": nome or "Cliente",
            "assessor": assessor,
            "tipo_bem": tipo_bem,
            "administradora": administradora,
            "valor_carta": valor_carta,
            "prazo_meses": prazo,
            "taxa_adm": taxa_adm,
            "fundo_reserva": fundo_reserva,
            "seguro": seguro,
            "prazo_contemplação": prazo_contemp,
            "parcela_reduzida_pct": parcela_red_pct,
            "lance_livre_pct": lance_livre_pct,
            "lance_embutido_pct": lance_embutido_pct,
            "lance_livre_valor": lance_livre_valor,
            "lance_embutido_valor": lance_embutido_valor,
            "carta_líquida": carta_líquida,
            "parcela_fase1": parcela_fase1,
            "parcela_fase2": parcela_fase2,
            "fc_integral": fc_integral,
            "fc_reduzido": fc_reduzido,
            "fc_fase2": fc_fase2,
            "ta_mensal": ta_mensal,
            "fr_mensal": fr_mensal,
            "sg_mensal": sg_mensal,
            "desemb_fase1": desemb_fase1,
            "desemb_fase2": desemb_fase2,
            "total_desembolsado": total_desembolsado,
            "total_taxas": total_taxas,
            "correção_anual": correção_anual,
            "tipo_correção": tipo_correção,
            "índice_correção": índice_correção,
        }

        # ======== ATUALIZAR UI DE RESULTADO ========
        for w in self.cs_result_inner.winfo_children():
            w.destroy()

        def _result_row(parent, label, value, color=TEXT_PRIMARY, bold=False):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text=label, font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY).pack(side="left")
            f = ("Segoe UI", 11, "bold") if bold else ("Segoe UI", 11)
            ctk.CTkLabel(r, text=value, font=f,
                         text_color=color).pack(side="right")

        def _dot_row(parent, label, value, color):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text="\u25cf", font=("Segoe UI", 9),
                         text_color=color).pack(side="left")
            ctk.CTkLabel(r, text=f"  {label}", font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY).pack(side="left")
            ctk.CTkLabel(r, text=value, font=("Segoe UI", 11, "bold"),
                         text_color=TEXT_PRIMARY).pack(side="right")

        def _separator(parent):
            ctk.CTkFrame(parent, fg_color=BORDER_LIGHT, height=1).pack(fill="x", pady=8)

        # ------ FASE 1 ------
        red_label = f"  (reduzida {parcela_red_pct}%)" if parcela_red_pct < 100 else ""
        f1 = ctk.CTkFrame(self.cs_result_inner, fg_color=ACCENT_GREEN, corner_radius=10)
        f1.pack(fill="x", pady=(0, 6))
        f1i = ctk.CTkFrame(f1, fg_color="transparent")
        f1i.pack(fill="x", padx=20, pady=12)
        ctk.CTkLabel(f1i, text=f"FASE 1  -  Antes da Contemplação{red_label}",
                     font=("Segoe UI", 9), text_color="#80c0a0").pack(anchor="w")
        ctk.CTkLabel(f1i, text=f"{fmt_currency(parcela_fase1)} /mes   x   {prazo_contemp} meses",
                     font=("Segoe UI", 22, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(2, 0))
        if tem_correção:
            ctk.CTkLabel(f1i, text=f"Parcela final estimada (corrigida): {fmt_currency(parcela_f1_final)}",
                         font=("Segoe UI", 9), text_color="#b0e0c8").pack(anchor="w")

        fc_lbl = "Fundo Comum" + (f" ({parcela_red_pct}%)" if parcela_red_pct < 100 else "")
        _dot_row(self.cs_result_inner, fc_lbl, fmt_currency(fc_reduzido), ACCENT_BLUE)
        _dot_row(self.cs_result_inner, "Taxa de Administração", fmt_currency(ta_mensal), ACCENT_ORANGE)
        _dot_row(self.cs_result_inner, "Fundo de Reserva", fmt_currency(fr_mensal), ACCENT_TEAL)
        _dot_row(self.cs_result_inner, "Seguro Prestamista", fmt_currency(sg_mensal), ACCENT_PURPLE)

        _separator(self.cs_result_inner)

        # ------ CONTEMPLAÇÃO / LANCES ------
        if lance_livre_valor > 0 or lance_embutido_valor > 0:
            cf = ctk.CTkFrame(self.cs_result_inner, fg_color=ACCENT_BLUE, corner_radius=10)
            cf.pack(fill="x", pady=(0, 6))
            ci = ctk.CTkFrame(cf, fg_color="transparent")
            ci.pack(fill="x", padx=20, pady=12)
            ctk.CTkLabel(ci, text=f"CONTEMPLAÇÃO  -  Mês {prazo_contemp}",
                         font=("Segoe UI", 9), text_color="#a0c0f0").pack(anchor="w")

            if lance_livre_valor > 0:
                r = ctk.CTkFrame(ci, fg_color="transparent")
                r.pack(fill="x", pady=1)
                ctk.CTkLabel(r, text="Lance Livre (recursos próprios)",
                             font=("Segoe UI", 10), text_color="#c0d8f8").pack(side="left")
                ctk.CTkLabel(r, text=fmt_currency(lance_livre_valor),
                             font=("Segoe UI", 12, "bold"), text_color=TEXT_WHITE).pack(side="right")

            if lance_embutido_valor > 0:
                r = ctk.CTkFrame(ci, fg_color="transparent")
                r.pack(fill="x", pady=1)
                ctk.CTkLabel(r, text="Lance Embutido (descontado da carta)",
                             font=("Segoe UI", 10), text_color="#c0d8f8").pack(side="left")
                ctk.CTkLabel(r, text=fmt_currency(lance_embutido_valor),
                             font=("Segoe UI", 12, "bold"), text_color=TEXT_WHITE).pack(side="right")

            r = ctk.CTkFrame(ci, fg_color="transparent")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text="Carta de Crédito Líquida",
                         font=("Segoe UI", 10, "bold"), text_color="#c0d8f8").pack(side="left")
            ctk.CTkLabel(r, text=fmt_currency(carta_líquida),
                         font=("Segoe UI", 14, "bold"), text_color=TEXT_WHITE).pack(side="right")

            _separator(self.cs_result_inner)

        # ------ FASE 2 ------
        if meses_restantes > 0:
            f2 = ctk.CTkFrame(self.cs_result_inner, fg_color="#005c3d", corner_radius=10)
            f2.pack(fill="x", pady=(0, 6))
            f2i = ctk.CTkFrame(f2, fg_color="transparent")
            f2i.pack(fill="x", padx=20, pady=12)
            ctk.CTkLabel(f2i, text="FASE 2  -  Após Contemplação  (parcela reajustada)",
                         font=("Segoe UI", 9), text_color="#80c0a0").pack(anchor="w")
            ctk.CTkLabel(f2i, text=f"{fmt_currency(parcela_fase2)} /mes   x   {meses_restantes} meses",
                         font=("Segoe UI", 22, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(2, 0))
            if tem_correção:
                ctk.CTkLabel(f2i, text=f"Parcela final estimada (corrigida): {fmt_currency(parcela_f2_final)}",
                             font=("Segoe UI", 9), text_color="#b0e0c8").pack(anchor="w")

            _dot_row(self.cs_result_inner, "Fundo Comum (reajustado)", fmt_currency(fc_fase2), ACCENT_BLUE)
            _dot_row(self.cs_result_inner, "Taxa de Administração", fmt_currency(ta_mensal), ACCENT_ORANGE)
            _dot_row(self.cs_result_inner, "Fundo de Reserva", fmt_currency(fr_mensal), ACCENT_TEAL)
            _dot_row(self.cs_result_inner, "Seguro Prestamista", fmt_currency(sg_mensal), ACCENT_PURPLE)

        _separator(self.cs_result_inner)

        # ------ TOTAIS ------
        _result_row(self.cs_result_inner, "Valor da Carta de Crédito", fmt_currency(valor_carta), bold=True)
        if lance_embutido_valor > 0:
            _result_row(self.cs_result_inner, "Carta Líquida (apos lance embutido)", fmt_currency(carta_líquida), ACCENT_RED, True)
        _result_row(self.cs_result_inner, f"Desembolso Fase 1 ({prazo_contemp} meses)", fmt_currency(desemb_fase1))
        if lance_livre_valor > 0:
            _result_row(self.cs_result_inner, "Lance Livre (desembolso próprio)", fmt_currency(lance_livre_valor))
        if meses_restantes > 0:
            _result_row(self.cs_result_inner, f"Desembolso Fase 2 ({meses_restantes} meses)", fmt_currency(desemb_fase2))
        _result_row(self.cs_result_inner, "Total Desembolsado (sem correção)", fmt_currency(total_desembolsado), ACCENT_GREEN, True)
        if tem_correção:
            corr_label = f"{tipo_correção} {fmt_pct(correção_anual)} a.a."
            if tipo_correção == "Pós-fixado":
                corr_label += f" ({índice_correção})"
            _result_row(self.cs_result_inner, f"Total Corrigido ({corr_label})", fmt_currency(total_desemb_corr), ACCENT_RED, True)
        _result_row(self.cs_result_inner, "Total em Taxas e Encargos", fmt_currency(total_taxas))

        _separator(self.cs_result_inner)

        # ------ RESUMO DA OPERAÇÃO ------
        ro = ctk.CTkFrame(self.cs_result_inner, fg_color=ACCENT_GREEN, corner_radius=10)
        ro.pack(fill="x", pady=(0, 6))
        roi = ctk.CTkFrame(ro, fg_color="transparent")
        roi.pack(fill="x", padx=20, pady=12)
        ctk.CTkLabel(roi, text="CUSTO EFETIVO DO CONSÓRCIO",
                     font=("Segoe UI", 9), text_color="#80c0a0").pack(anchor="w")
        ctk.CTkLabel(roi, text=f"{fmt_pct(ce_pct)} sobre o crédito",
                     font=("Segoe UI", 20, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(2, 0))
        ctk.CTkLabel(roi, text=f"{fmt_pct(ce_anual * 100)} a.a.  |  {fmt_pct(ce_mensal * 100)} a.m.",
                     font=("Segoe UI", 12, "bold"), text_color="#b0e0c8").pack(anchor="w")

        _separator(self.cs_result_inner)

        _result_row(self.cs_result_inner, "Crédito Recebido (carta líquida)", fmt_currency(carta_líquida), ACCENT_BLUE, True)
        _result_row(self.cs_result_inner, "Custo Total da Operação", fmt_currency(custo_total), ACCENT_RED, True)
        relação = total_ref / carta_líquida if carta_líquida > 0 else 0
        _result_row(self.cs_result_inner, "Relação Custo / Crédito", f"{relação:.2f}x".replace(".", ","))
        _result_row(self.cs_result_inner, "Parcela Media", fmt_currency(total_ref / prazo if prazo > 0 else 0))

        # Habilitar PDF, Relatório e Email
        self.cs_btn_pdf.configure(state="normal")
        self.cs_btn_relatório.configure(state="normal")
        self.cs_btn_email.configure(state="normal")

    def _on_gerar_pdf_consorcio(self):
        if not hasattr(self, '_consórcio_dados') or not self._consórcio_dados:
            messagebox.showwarning("Aviso", "Calcule a simulação primeiro.")
            return
        try:
            filepath = generate_consorcio_pdf(self._consórcio_dados, CONSÓRCIO_OUTPUT_DIR)
            messagebox.showinfo("PDF Gerado",
                                f"PDF salvo com sucesso!\n\n{os.path.basename(filepath)}")
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF:\n{e}")

    def _on_gerar_relatorio_consorcio(self):
        if not hasattr(self, '_consórcio_dados') or not self._consórcio_dados:
            messagebox.showwarning("Aviso", "Calcule a simulação primeiro.")
            return
        try:
            filepath = generate_relatorio_consorcio(self._consórcio_dados, CONSÓRCIO_OUTPUT_DIR)
            messagebox.showinfo("Relatório Gerado",
                                f"Relatório explicativo salvo!\n\n{os.path.basename(filepath)}")
            os.startfile(filepath)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{e}")

    def _on_abrir_email_consorcio(self):
        if not hasattr(self, '_consórcio_dados') or not self._consórcio_dados:
            messagebox.showwarning("Aviso", "Calcule a simulação primeiro.")
            return

        d = self._consórcio_dados
        from gerar_pdf_consorcio import _calc, fmt_currency as fc2, fmt_pct as fp2
        c = _calc(d)

        # Janela de composição de email
        win = ctk.CTkToplevel(self)
        win.title("Enviar Email - Consórcio")
        win.geometry("680x720")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()
        win.configure(fg_color=BG_PRIMARY)

        # Header
        hdr = ctk.CTkFrame(win, fg_color=ACCENT_GREEN, height=50, corner_radius=0)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="  Enviar Simulação por Email",
                     font=("Segoe UI", 16, "bold"), text_color=TEXT_WHITE).pack(side="left", padx=20)

        body = ctk.CTkFrame(win, fg_color=BG_PRIMARY)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # -- Destinatario --
        ctk.CTkLabel(body, text="Para (email do cliente)", font=("Segoe UI", 11, "bold"),
                     text_color=TEXT_SECONDARY).pack(anchor="w")
        email_to = ctk.CTkEntry(body, placeholder_text="cliente@email.com", height=38,
                                corner_radius=8, fg_color=BG_INPUT,
                                border_width=1, border_color=BORDER_LIGHT,
                                font=("Segoe UI", 11))
        email_to.pack(fill="x", pady=(2, 10))

        # -- CC --
        ctk.CTkLabel(body, text="CC (opcional)", font=("Segoe UI", 11, "bold"),
                     text_color=TEXT_SECONDARY).pack(anchor="w")
        email_cc = ctk.CTkEntry(body, placeholder_text="copia@email.com", height=38,
                                corner_radius=8, fg_color=BG_INPUT,
                                border_width=1, border_color=BORDER_LIGHT,
                                font=("Segoe UI", 11))
        email_cc.pack(fill="x", pady=(2, 10))

        # -- Assunto --
        nome_cli = d.get("cliente_nome", "Cliente")
        tipo = d.get("tipo_bem", "")
        assunto_default = f"Simulação de Consórcio - {nome_cli} - {tipo}"

        ctk.CTkLabel(body, text="Assunto", font=("Segoe UI", 11, "bold"),
                     text_color=TEXT_SECONDARY).pack(anchor="w")
        email_assunto = ctk.CTkEntry(body, height=38, corner_radius=8, fg_color=BG_INPUT,
                                     border_width=1, border_color=BORDER_LIGHT,
                                     font=("Segoe UI", 11))
        email_assunto.pack(fill="x", pady=(2, 10))
        email_assunto.insert(0, assunto_default)

        # -- Anexos --
        ctk.CTkLabel(body, text="Anexos", font=("Segoe UI", 11, "bold"),
                     text_color=TEXT_SECONDARY).pack(anchor="w", pady=(4, 4))

        anexo_frame = ctk.CTkFrame(body, fg_color=BG_CARD, corner_radius=10,
                                   border_width=1, border_color=BORDER_CARD)
        anexo_frame.pack(fill="x", pady=(0, 10))
        af = ctk.CTkFrame(anexo_frame, fg_color="transparent")
        af.pack(fill="x", padx=16, pady=12)

        chk_proposta_var = ctk.BooleanVar(value=True)
        chk_proposta = ctk.CTkCheckBox(
            af, text="  Proposta de Simulação (PDF)",
            variable=chk_proposta_var, font=("Segoe UI", 11),
            fg_color=ACCENT_GREEN, hover_color=BG_SIDEBAR_HOVER,
            text_color=TEXT_PRIMARY, corner_radius=4,
        )
        chk_proposta.pack(anchor="w", pady=2)

        chk_relatório_var = ctk.BooleanVar(value=True)
        chk_relatório = ctk.CTkCheckBox(
            af, text="  Relatório Explicativo (PDF)",
            variable=chk_relatório_var, font=("Segoe UI", 11),
            fg_color=ACCENT_PURPLE, hover_color="#6a4daa",
            text_color=TEXT_PRIMARY, corner_radius=4,
        )
        chk_relatório.pack(anchor="w", pady=2)

        ctk.CTkLabel(af, text="Os PDFs serão gerados automaticamente ao enviar",
                     font=("Segoe UI", 9), text_color=TEXT_TERTIARY).pack(anchor="w", pady=(4, 0))

        # -- Preview resumo --
        ctk.CTkLabel(body, text="Resumo no corpo do email", font=("Segoe UI", 11, "bold"),
                     text_color=TEXT_SECONDARY).pack(anchor="w", pady=(4, 4))

        prev = ctk.CTkFrame(body, fg_color=BG_CARD, corner_radius=10,
                            border_width=1, border_color=BORDER_CARD)
        prev.pack(fill="x", pady=(0, 10))
        pi = ctk.CTkFrame(prev, fg_color="transparent")
        pi.pack(fill="x", padx=16, pady=12)

        preview_items = [
            ("Carta de Crédito", fc2(d.get("valor_carta", 0))),
            ("Administradora", d.get("administradora", "-")),
            ("Prazo", f"{d.get('prazo_meses', 0)} meses"),
            ("Parcela Fase 1", f"{fc2(d.get('parcela_fase1', 0))}/mes"),
            ("Custo Efetivo a.a.", fp2(c["ce_anual"] * 100)),
        ]
        for label, val in preview_items:
            r = ctk.CTkFrame(pi, fg_color="transparent")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text=label, font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY).pack(side="left")
            ctk.CTkLabel(r, text=val, font=("Segoe UI", 10, "bold"),
                         text_color=TEXT_PRIMARY).pack(side="right")

        # -- Botoes --
        btn_f = ctk.CTkFrame(body, fg_color="transparent")
        btn_f.pack(fill="x", pady=(8, 0))

        def _do_enviar():
            to = email_to.get().strip()
            if not to:
                messagebox.showwarning("Aviso", "Preencha o email do destinatario.", parent=win)
                return
            cc = email_cc.get().strip()
            assunto = email_assunto.get().strip()
            anexar_prop = chk_proposta_var.get()
            anexar_rel = chk_relatório_var.get()

            if not anexar_prop and not anexar_rel:
                messagebox.showwarning("Aviso", "Selecione ao menos um anexo.", parent=win)
                return

            try:
                import win32com.client as win32
                import pythoncom
                pythoncom.CoInitialize()

                outlook = win32.Dispatch("Outlook.Application")
            except Exception as e:
                messagebox.showerror("Erro",
                    f"Não foi possível conectar ao Outlook.\nVerifique se está aberto.\n\n{e}",
                    parent=win)
                return

            try:
                # Gerar PDFs necessários
                anexos = []
                if anexar_prop:
                    prop_path = generate_consorcio_pdf(d, CONSÓRCIO_OUTPUT_DIR)
                    anexos.append(prop_path)
                if anexar_rel:
                    rel_path = generate_relatorio_consorcio(d, CONSÓRCIO_OUTPUT_DIR)
                    anexos.append(rel_path)

                # Montar email
                dados_resumo = {
                    "tipo_bem": d.get("tipo_bem", ""),
                    "valor_carta_fmt": fc2(d.get("valor_carta", 0)),
                    "administradora": d.get("administradora", "-"),
                    "prazo": str(d.get("prazo_meses", 0)),
                    "parcela_f1": fc2(d.get("parcela_fase1", 0)),
                    "ce_anual": fp2(c["ce_anual"] * 100),
                }
                html_body = build_email_consorcio(nome_cli, dados_resumo)

                mail = outlook.CreateItem(0)
                mail.To = to
                if cc:
                    mail.CC = cc
                mail.Subject = assunto
                mail.HTMLBody = html_body

                for a in anexos:
                    mail.Attachments.Add(os.path.abspath(a))

                mail.Save()

                n_anexos = len(anexos)
                pythoncom.CoUninitialize()
                win.destroy()
                messagebox.showinfo("Email Criado",
                    f"Rascunho salvo no Outlook!\n\n"
                    f"Para: {to}\n"
                    f"{n_anexos} anexo(s) incluido(s).\n\n"
                    f"Abra o Outlook e revise na pasta Rascunhos.")

            except Exception as e:
                pythoncom.CoUninitialize()
                messagebox.showerror("Erro", f"Erro ao criar email:\n{e}", parent=win)

        ctk.CTkButton(
            btn_f, text="  Enviar",
            font=("Segoe UI", 13, "bold"), fg_color=ACCENT_GREEN,
            hover_color=BG_SIDEBAR_HOVER, height=44, corner_radius=10,
            command=_do_enviar,
        ).pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_f, text="  Cancelar",
            font=("Segoe UI", 12), fg_color="#6b7280",
            hover_color="#555d6a", height=44, corner_radius=10,
            command=win.destroy,
        ).pack(side="left")

    def _on_limpar_consorcio(self):
        self.cs_nome.delete(0, "end")
        self.cs_assessor.delete(0, "end")
        self.cs_tipo_bem.set("Imovel")
        self.cs_administradora.delete(0, "end")
        self.cs_valor_carta.delete(0, "end")
        self.cs_prazo.set("120")
        self.cs_taxa_adm.delete(0, "end")
        self.cs_taxa_adm.insert(0, "18")
        self.cs_fundo_reserva.delete(0, "end")
        self.cs_fundo_reserva.insert(0, "2")
        self.cs_seguro.delete(0, "end")
        self.cs_seguro.insert(0, "0")
        self.cs_prazo_contemplação.delete(0, "end")
        self.cs_prazo_contemplação.insert(0, "60")
        self.cs_parcela_reduzida.set("100% (Integral)")
        self.cs_lance_livre.delete(0, "end")
        self.cs_lance_livre.insert(0, "0")
        self.cs_lance_embutido.delete(0, "end")
        self.cs_lance_embutido.insert(0, "0")
        self.cs_correção_anual.delete(0, "end")
        self.cs_correção_anual.insert(0, "0")
        self.cs_tipo_correção.set("Pós-fixado")
        self.cs_índice_correção.set("INCC")

        for w in self.cs_result_inner.winfo_children():
            w.destroy()
        self.cs_result_placeholder = ctk.CTkLabel(
            self.cs_result_inner,
            text="Preencha os campos acima e clique em 'Calcular Simulação'",
            font=("Segoe UI", 12), text_color=TEXT_TERTIARY
        )
        self.cs_result_placeholder.pack(pady=20)
        self.cs_btn_pdf.configure(state="disabled")
        self.cs_btn_relatório.configure(state="disabled")
        self.cs_btn_email.configure(state="disabled")
        if hasattr(self, '_consórcio_dados'):
            self._consórcio_dados = None

    def _on_abrir_pasta_consorcio(self):
        os.makedirs(CONSÓRCIO_OUTPUT_DIR, exist_ok=True)
        os.startfile(CONSÓRCIO_OUTPUT_DIR)

    # =================================================================
    #  HANDLERS: COMPARATIVO DE VPL (NASA HD)
    # =================================================================
    def _on_calcular_vpl(self):
        try:
            params = {
                "valor_carta": self._parse_number(self.vpl_valor_carta.get()),
                "prazo_meses": int(self._parse_number(self.vpl_prazo.get())),
                "taxa_adm": self._parse_number(self.vpl_taxa_adm.get()),
                "fundo_reserva": self._parse_number(self.vpl_fundo_res.get()),
                "seguro": self._parse_number(self.vpl_seguro.get()),
                "prazo_contemp": int(self._parse_number(self.vpl_contemp.get())),
                "lance_embutido_pct": self._parse_number(self.vpl_lance_emb.get()),
                "lance_livre_pct": self._parse_number(self.vpl_lance_livre.get()),
                "parcela_red_pct": self._parse_number(self.vpl_red_pct.get()),
                "correcao_anual": self._parse_number(self.vpl_correcao.get()),
                "alm_anual": self._parse_number(self.vpl_alm.get()),
                "hurdle_anual": self._parse_number(self.vpl_hurdle.get()),
            }
        except (ValueError, AttributeError):
            messagebox.showerror("Erro", "Preencha todos os campos corretamente.")
            return

        if params["valor_carta"] <= 0 or params["prazo_meses"] <= 0:
            messagebox.showerror("Erro", "Valor e prazo devem ser maiores que zero.")
            return

        # Calcular
        fluxo_r = calcular_fluxo_consorcio(params)
        vpl_r = calcular_vpl_hd(params, fluxo_r)

        # Salvar para PDF
        self._vpl_result = {**params, **vpl_r, **fluxo_r}

        # ======== Mostrar resultado ========
        for w in self.vpl_result_inner.winfo_children():
            w.destroy()

        def _row(parent, label, value, color=TEXT_PRIMARY, bold=False):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text=label, font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY).pack(side="left")
            f = ("Segoe UI", 11, "bold") if bold else ("Segoe UI", 11)
            ctk.CTkLabel(r, text=value, font=f, text_color=color).pack(side="right")

        def _sep(parent):
            ctk.CTkFrame(parent, fg_color=BORDER_LIGHT, height=1).pack(fill="x", pady=8)

        vc = params["valor_carta"]
        delta = vpl_r["delta_vpl"]
        cria = vpl_r["cria_valor"]

        # Banner principal
        banner_color = ACCENT_GREEN if cria else ACCENT_RED
        banner_text = "CRIA VALOR" if cria else "NÃO CRIA VALOR"
        banner_icon = "\u2714" if cria else "\u2718"

        banner = ctk.CTkFrame(self.vpl_result_inner, fg_color=banner_color, corner_radius=12)
        banner.pack(fill="x", pady=(0, 10))
        bi = ctk.CTkFrame(banner, fg_color="transparent")
        bi.pack(fill="x", padx=24, pady=16)
        ctk.CTkLabel(bi, text=f"{banner_icon}  {banner_text}",
                     font=("Segoe UI", 22, "bold"), text_color=TEXT_WHITE).pack(anchor="w")
        ctk.CTkLabel(bi, text=f"Delta VPL: {fmt_currency(delta)}",
                     font=("Segoe UI", 14, "bold"), text_color="#ffffffcc").pack(anchor="w", pady=(4, 0))
        ctk.CTkLabel(bi, text=f"Break-even lance livre: {fmt_pct(vpl_r['break_even_lance'])}",
                     font=("Segoe UI", 11), text_color="#ffffffaa").pack(anchor="w", pady=(2, 0))

        # Detalhes
        _sep(self.vpl_result_inner)

        ctk.CTkLabel(self.vpl_result_inner, text="DECOMPOSIÇÃO DO VPL",
                     font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY, anchor="w").pack(fill="x")

        _row(self.vpl_result_inner, "B0 — PV do Crédito (a taxa ALM)", fmt_currency(vpl_r["b0"]), ACCENT_GREEN, True)
        _row(self.vpl_result_inner, "H0 — PV dos Pagamentos pré-T", fmt_currency(vpl_r["h0"]), ACCENT_ORANGE, True)
        _row(self.vpl_result_inner, "D0 — Valor criado (B0 − H0)", fmt_currency(vpl_r["d0"]),
             ACCENT_GREEN if vpl_r["d0"] >= 0 else ACCENT_RED, True)
        _row(self.vpl_result_inner, "PV Parcelas pós-T (a taxa Hurdle)", fmt_currency(vpl_r["pv_pos_t"]), ACCENT_BLUE, True)
        _row(self.vpl_result_inner, "Delta VPL (D0 − PV pós-T)", fmt_currency(delta),
             ACCENT_GREEN if cria else ACCENT_RED, True)

        _sep(self.vpl_result_inner)

        ctk.CTkLabel(self.vpl_result_inner, text="CUSTOS E TAXAS",
                     font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY, anchor="w").pack(fill="x")

        tir_a = vpl_r["tir_anual"] * 100
        tir_m = vpl_r["tir_mensal"] * 100
        cet_a = vpl_r["cet_anual"] * 100

        _row(self.vpl_result_inner, "TIR Mensal", f"{tir_m:.3f}%".replace(".", ","))
        _row(self.vpl_result_inner, "TIR Anual", f"{tir_a:.2f}%".replace(".", ","))
        _row(self.vpl_result_inner, "CET Anual", f"{cet_a:.2f}%".replace(".", ","))
        _row(self.vpl_result_inner, "VPL Total (a taxa ALM)", fmt_currency(vpl_r["vpl_total"]))

        _sep(self.vpl_result_inner)

        ctk.CTkLabel(self.vpl_result_inner, text="DADOS DA OPERAÇÃO",
                     font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY, anchor="w").pack(fill="x")

        _row(self.vpl_result_inner, "Carta de Crédito", fmt_currency(vc))
        _row(self.vpl_result_inner, "Carta Líquida", fmt_currency(fluxo_r["carta_liquida"]), ACCENT_GREEN, True)
        _row(self.vpl_result_inner, "Total Desembolsado", fmt_currency(fluxo_r["total_pago"]))
        _row(self.vpl_result_inner, "Parcela Fase 1 (base)", fmt_currency(fluxo_r["parcela_f1_base"]))
        _row(self.vpl_result_inner, "Parcela Fase 2 (base)", fmt_currency(fluxo_r["parcela_f2_base"]))
        _row(self.vpl_result_inner, "Lance Embutido", fmt_currency(fluxo_r["lance_embutido_valor"]))
        _row(self.vpl_result_inner, "Lance Livre", fmt_currency(fluxo_r["lance_livre_valor"]))

        # Habilitar PDF
        self.vpl_btn_pdf.configure(state="normal")

    def _on_gerar_pdf_vpl(self):
        if not hasattr(self, '_vpl_result') or not self._vpl_result:
            messagebox.showwarning("Aviso", "Execute a análise VPL primeiro.")
            return
        try:
            from fpdf import FPDF
            r = self._vpl_result
            pdf = FPDF()
            pdf.add_page()
            pdf.set_auto_page_break(auto=True, margin=15)

            # Header
            pdf.set_font("Helvetica", "B", 18)
            pdf.cell(0, 12, "SOMUS CAPITAL - Analise VPL (NASA HD)", ln=True, align="C")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 8, f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
            pdf.ln(8)

            # Status
            cria = r["cria_valor"]
            status_text = "CRIA VALOR" if cria else "NAO CRIA VALOR"
            pdf.set_font("Helvetica", "B", 16)
            pdf.set_text_color(0, 128, 0) if cria else pdf.set_text_color(200, 0, 0)
            pdf.cell(0, 12, status_text, ln=True, align="C")
            pdf.set_text_color(0, 0, 0)
            pdf.ln(4)

            # Delta VPL
            pdf.set_font("Helvetica", "B", 14)
            delta_str = f"R$ {r['delta_vpl']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            pdf.cell(0, 10, f"Delta VPL: {delta_str}", ln=True, align="C")
            be_str = f"{r['break_even_lance']:.2f}%".replace(".", ",")
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 8, f"Break-even Lance Livre: {be_str}", ln=True, align="C")
            pdf.ln(6)

            # Tabela de decomposicao
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Decomposicao do VPL", ln=True)
            pdf.set_font("Helvetica", "", 10)

            def _add_row(label, val):
                val_str = f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                pdf.cell(100, 7, f"  {label}", border=0)
                pdf.cell(0, 7, val_str, ln=True, align="R")

            _add_row("B0 - PV do Credito (ALM)", r["b0"])
            _add_row("H0 - PV Pagamentos pre-T", r["h0"])
            _add_row("D0 - Valor criado (B0 - H0)", r["d0"])
            _add_row("PV Parcelas pos-T (Hurdle)", r["pv_pos_t"])
            _add_row("Delta VPL (D0 - PV pos-T)", r["delta_vpl"])
            pdf.ln(4)

            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Custos e Taxas", ln=True)
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(100, 7, "  TIR Mensal"); pdf.cell(0, 7, f"{r['tir_mensal']*100:.3f}%".replace(".", ","), ln=True, align="R")
            pdf.cell(100, 7, "  TIR Anual"); pdf.cell(0, 7, f"{r['tir_anual']*100:.2f}%".replace(".", ","), ln=True, align="R")
            pdf.cell(100, 7, "  CET Anual"); pdf.cell(0, 7, f"{r['cet_anual']*100:.2f}%".replace(".", ","), ln=True, align="R")
            pdf.ln(4)

            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "Dados da Operacao", ln=True)
            pdf.set_font("Helvetica", "", 10)
            _add_row("Carta de Credito", r["valor_carta"])
            _add_row("Carta Liquida", r["carta_liquida"])
            _add_row("Total Desembolsado", r["total_pago"])
            _add_row("Parcela Fase 1 (base)", r["parcela_f1_base"])
            _add_row("Parcela Fase 2 (base)", r["parcela_f2_base"])
            _add_row("Lance Embutido", r["lance_embutido_valor"])
            _add_row("Lance Livre", r["lance_livre_valor"])

            pdf.ln(10)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 5, "Analise gerada pelo motor NASA HD - Somus Capital", ln=True, align="C")
            pdf.cell(0, 5, "Valores sujeitos a variacao. Simulacao nao constitui oferta.", ln=True, align="C")

            os.makedirs(CONSÓRCIO_OUTPUT_DIR, exist_ok=True)
            nome = f"VPL_NASA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            path = os.path.join(CONSÓRCIO_OUTPUT_DIR, nome)
            pdf.output(path)
            messagebox.showinfo("PDF Gerado", f"PDF VPL salvo!\n\n{nome}")
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar PDF VPL:\n{e}")

    def _on_limpar_vpl(self):
        defaults = {
            "vpl_valor_carta": "", "vpl_prazo": "200", "vpl_contemp": "11",
            "vpl_taxa_adm": "20", "vpl_fundo_res": "3", "vpl_seguro": "0",
            "vpl_lance_emb": "30", "vpl_lance_livre": "0", "vpl_red_pct": "70",
            "vpl_correcao": "3", "vpl_alm": "12", "vpl_hurdle": "12",
        }
        for attr, val in defaults.items():
            w = getattr(self, attr)
            w.delete(0, "end")
            if val:
                w.insert(0, val)
        for w in self.vpl_result_inner.winfo_children():
            w.destroy()
        ctk.CTkLabel(
            self.vpl_result_inner,
            text="Preencha os parâmetros e clique em 'Analisar VPL'",
            font=("Segoe UI", 12), text_color=TEXT_TERTIARY
        ).pack(pady=20)
        self.vpl_btn_pdf.configure(state="disabled")
        self._vpl_result = None

    # =================================================================
    #  HANDLERS: CONSÓRCIO VS FINANCIAMENTO
    # =================================================================
    def _on_comparar_consorcio_financ(self):
        try:
            params_c = {
                "valor_carta": self._parse_number(self.cf_c_valor.get()),
                "prazo_meses": int(self._parse_number(self.cf_c_prazo.get())),
                "taxa_adm": self._parse_number(self.cf_c_taxa.get()),
                "fundo_reserva": self._parse_number(self.cf_c_fres.get()),
                "seguro": self._parse_number(self.cf_c_seg.get()),
                "prazo_contemp": int(self._parse_number(self.cf_c_contemp.get())),
                "lance_embutido_pct": self._parse_number(self.cf_c_lemb.get()),
                "lance_livre_pct": self._parse_number(self.cf_c_lliv.get()),
                "parcela_red_pct": self._parse_number(self.cf_c_red.get()),
                "correcao_anual": self._parse_number(self.cf_c_corr.get()),
                "alm_anual": self._parse_number(self.cf_c_alm.get()),
            }
            params_f = {
                "valor": self._parse_number(self.cf_f_valor.get()),
                "prazo_meses": int(self._parse_number(self.cf_f_prazo.get())),
                "taxa_mensal_pct": self._parse_number(self.cf_f_taxa.get()),
                "metodo": self.cf_f_metodo.get().lower(),
            }
        except (ValueError, AttributeError):
            messagebox.showerror("Erro", "Preencha todos os campos corretamente.")
            return

        if params_c["valor_carta"] <= 0 or params_f["valor"] <= 0:
            messagebox.showerror("Erro", "Valores devem ser maiores que zero.")
            return

        # Calcular
        result = comparar_consorcio_financiamento(params_c, params_f)

        # ======== Mostrar resultado ========
        for w in self.cf_result_inner.winfo_children():
            w.destroy()

        def _row(parent, label, value, color=TEXT_PRIMARY, bold=False):
            r = ctk.CTkFrame(parent, fg_color="transparent")
            r.pack(fill="x", pady=1)
            ctk.CTkLabel(r, text=label, font=("Segoe UI", 10),
                         text_color=TEXT_SECONDARY).pack(side="left")
            f = ("Segoe UI", 11, "bold") if bold else ("Segoe UI", 11)
            ctk.CTkLabel(r, text=value, font=f, text_color=color).pack(side="right")

        def _sep(parent):
            ctk.CTkFrame(parent, fg_color=BORDER_LIGHT, height=1).pack(fill="x", pady=8)

        # Cards lado a lado para resultado
        cols = ctk.CTkFrame(self.cf_result_inner, fg_color="transparent")
        cols.pack(fill="x", pady=(0, 10))
        cols.columnconfigure((0, 1), weight=1)

        # -- Card Consorcio --
        cc = ctk.CTkFrame(cols, fg_color=ACCENT_GREEN, corner_radius=12)
        cc.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        cci = ctk.CTkFrame(cc, fg_color="transparent")
        cci.pack(fill="x", padx=18, pady=14)
        ctk.CTkLabel(cci, text="CONSÓRCIO", font=("Segoe UI", 10, "bold"),
                     text_color="#80c0a0").pack(anchor="w")
        ctk.CTkLabel(cci, text=fmt_currency(result["consorcio"]["total_pago"]),
                     font=("Segoe UI", 18, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(2, 0))
        ctk.CTkLabel(cci, text="total desembolsado",
                     font=("Segoe UI", 9), text_color="#b0e0c8").pack(anchor="w")

        tir_c_a = result["tir_consorcio_anual"] * 100
        ctk.CTkLabel(cci, text=f"TIR: {tir_c_a:.2f}% a.a.".replace(".", ","),
                     font=("Segoe UI", 11, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(6, 0))
        ctk.CTkLabel(cci, text=f"Relação custo/crédito: {result['razao_vpl_consorcio']:.2f}x".replace(".", ","),
                     font=("Segoe UI", 10), text_color="#b0e0c8").pack(anchor="w")

        # -- Card Financiamento --
        fc = ctk.CTkFrame(cols, fg_color=ACCENT_BLUE, corner_radius=12)
        fc.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        fci = ctk.CTkFrame(fc, fg_color="transparent")
        fci.pack(fill="x", padx=18, pady=14)
        ctk.CTkLabel(fci, text="FINANCIAMENTO", font=("Segoe UI", 10, "bold"),
                     text_color="#a0c0f0").pack(anchor="w")
        ctk.CTkLabel(fci, text=fmt_currency(result["financiamento"]["total_pago"]),
                     font=("Segoe UI", 18, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(2, 0))
        ctk.CTkLabel(fci, text="total desembolsado",
                     font=("Segoe UI", 9), text_color="#c0d8f8").pack(anchor="w")

        tir_f_a = result["tir_financ_anual"] * 100
        ctk.CTkLabel(fci, text=f"TIR: {tir_f_a:.2f}% a.a.".replace(".", ","),
                     font=("Segoe UI", 11, "bold"), text_color=TEXT_WHITE).pack(anchor="w", pady=(6, 0))
        ctk.CTkLabel(fci, text=f"Relação custo/crédito: {result['razao_vpl_financ']:.2f}x".replace(".", ","),
                     font=("Segoe UI", 10), text_color="#c0d8f8").pack(anchor="w")

        _sep(self.cf_result_inner)

        # Resultado comparativo
        eco = result["economia_vpl"]
        vantagem = "Consórcio" if eco > 0 else "Financiamento"
        eco_color = ACCENT_GREEN if eco > 0 else ACCENT_BLUE

        result_banner = ctk.CTkFrame(self.cf_result_inner, fg_color=eco_color, corner_radius=12)
        result_banner.pack(fill="x", pady=(0, 10))
        rbi = ctk.CTkFrame(result_banner, fg_color="transparent")
        rbi.pack(fill="x", padx=24, pady=14)
        ctk.CTkLabel(rbi, text=f"Vantagem: {vantagem}",
                     font=("Segoe UI", 16, "bold"), text_color=TEXT_WHITE).pack(anchor="w")
        ctk.CTkLabel(rbi, text=f"Economia VPL: {fmt_currency(abs(eco))}",
                     font=("Segoe UI", 12, "bold"), text_color="#ffffffcc").pack(anchor="w", pady=(2, 0))

        _sep(self.cf_result_inner)

        # VPLs
        _row(self.cf_result_inner, "VPL Consórcio (a taxa ALM)", fmt_currency(result["vpl_consorcio"]), ACCENT_GREEN, True)
        _row(self.cf_result_inner, "VPL Financiamento (a taxa ALM)", fmt_currency(result["vpl_financiamento"]), ACCENT_BLUE, True)
        _row(self.cf_result_inner, "Economia VPL (Cons. − Financ.)", fmt_currency(eco), eco_color, True)

        _sep(self.cf_result_inner)

        # Detalhes
        cons = result["consorcio"]
        fin = result["financiamento"]
        _row(self.cf_result_inner, "Crédito Líquido (Consórcio)", fmt_currency(cons["carta_liquida"]))
        _row(self.cf_result_inner, "Valor Financiado", fmt_currency(fin["valor"]))
        _row(self.cf_result_inner, "Total Juros (Financiamento)", fmt_currency(fin["total_juros"]), ACCENT_RED)
        _row(self.cf_result_inner, "Parcela F1 Consórcio (base)", fmt_currency(cons["parcela_f1_base"]))
        _row(self.cf_result_inner, "Parcela F2 Consórcio (base)", fmt_currency(cons["parcela_f2_base"]))
        if fin["parcelas"]:
            _row(self.cf_result_inner, "Parcela Financiamento (1a)", fmt_currency(fin["parcelas"][0]["parcela"]))

    def _on_limpar_cf(self):
        defaults_c = {
            "cf_c_valor": "300000", "cf_c_prazo": "120", "cf_c_taxa": "18",
            "cf_c_fres": "2", "cf_c_seg": "0", "cf_c_contemp": "60",
            "cf_c_lemb": "0", "cf_c_lliv": "0", "cf_c_red": "100",
            "cf_c_corr": "0", "cf_c_alm": "12",
        }
        defaults_f = {
            "cf_f_valor": "300000", "cf_f_prazo": "120", "cf_f_taxa": "1.0",
        }
        for attr, val in {**defaults_c, **defaults_f}.items():
            w = getattr(self, attr)
            w.delete(0, "end")
            w.insert(0, val)
        self.cf_f_metodo.set("Price")
        for w in self.cf_result_inner.winfo_children():
            w.destroy()
        ctk.CTkLabel(
            self.cf_result_inner,
            text="Preencha ambos os lados e clique em 'Comparar'",
            font=("Segoe UI", 12), text_color=TEXT_TERTIARY
        ).pack(pady=20)

    # =================================================================
    #  DATA LOADING
    # =================================================================
    def _load_data_async(self):
        if self.role == "corporate":
            return  # Corporate nao carrega dados de Mesa Produtos
        self.after(500, lambda: threading.Thread(target=self._do_load_data, daemon=True).start())

    def _do_load_data(self):
        try:
            assessores = load_base_emails()
            summary = load_eventos_summary()

            def update_ui():
                self.kpi_assessores.configure(text=str(len(assessores)))
                self.kpi_eventos.configure(text=f"{summary['total_eventos']:,}".replace(",", "."))
                self.kpi_ativos.configure(text=str(summary["ativos"]))
                self.kpi_datas.configure(text=str(summary["datas_unicas"]))
                self.total_value_label.configure(text=fmt_currency(summary["valor_total"]))

                if summary["data_min"] and summary["data_max"]:
                    d1 = summary["data_min"].strftime("%d/%m/%Y")
                    d2 = summary["data_max"].strftime("%d/%m/%Y")
                    self.período_text.configure(
                        text=f"  Periodo:  {d1}   \u2192   {d2}   |   {summary['datas_unicas']} datas com eventos"
                    )
                    self.total_sub_label.configure(
                        text=f"Somatorio de {summary['total_eventos']:,} eventos de {len(assessores)} assessores".replace(",", ".")
                    )

                self._update_pdf_status()

                # Tipo breakdown cards
                tipo_colors = {
                    "PAGAMENTO DE JUROS": ACCENT_BLUE,
                    "AMORTIZACAO": ACCENT_GREEN,
                    "INCORPORACAO": ACCENT_ORANGE,
                    "PREMIO": ACCENT_PURPLE,
                }
                tipo_names = {
                    "PAGAMENTO DE JUROS": "Pagamento de Juros",
                    "AMORTIZACAO": "Amortizacao",
                    "INCORPORACAO": "Incorporacao",
                    "PREMIO": "Premio",
                }
                tipos = summary.get("tipos", {})
                idx = 0
                for tipo_key in ["PAGAMENTO DE JUROS", "AMORTIZACAO", "INCORPORACAO", "PREMIO"]:
                    if tipo_key in tipos:
                        r = idx // 2
                        c = idx % 2
                        self._make_tipo_card(
                            self.tipo_frame,
                            tipo_names.get(tipo_key, tipo_key),
                            tipos[tipo_key]["count"],
                            tipos[tipo_key]["total"],
                            tipo_colors.get(tipo_key, TEXT_SECONDARY),
                            r, c
                        )
                        idx += 1

            self.after(0, update_ui)
        except Exception as e:
            import traceback; traceback.print_exc()
            err_msg = f"  Erro ao carregar: {e}"
            self.after(0, lambda m=err_msg: self.período_text.configure(text=m))

    def _update_pdf_status(self):
        if os.path.exists(OUTPUT_DIR):
            pdfs = [f for f in os.listdir(OUTPUT_DIR) if f.endswith(".pdf")]
            if pdfs:
                self.período_pdfs.configure(text=f"\u2713 {len(pdfs)} PDFs gerados  ")
                return
        self.período_pdfs.configure(text="Nenhum PDF gerado  ")

    # =================================================================
    #  LOGGING (Operations page)
    # =================================================================
    def _log(self, msg, tag=None):
        def _do():
            self.ops_log.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            prefix = {"ok": "\u2713", "error": "\u2717", "skip": "\u2013"}.get(tag, " ")
            self.ops_log.insert("end", f"  [{ts}]  {prefix}  {msg}\n")
            self.ops_log.see("end")
            self.ops_log.configure(state="disabled")
        self.after(0, _do)

    def _clear_ops_log(self):
        self.ops_log.configure(state="normal")
        self.ops_log.delete("1.0", "end")
        self.ops_log.configure(state="disabled")

    def _set_ops_status(self, text, color=TEXT_PRIMARY):
        self.after(0, lambda: self.ops_status_text.configure(text=f"  {text}"))
        self.after(0, lambda: self.ops_status_dot.configure(text_color=color))

    def _set_ops_progress(self, value, text=""):
        self.after(0, lambda: self.ops_progress.set(value))
        self.after(0, lambda: self.ops_progress_text.configure(text=text))

    def _set_ops_counter(self, text):
        self.after(0, lambda: self.ops_counter.configure(text=text))

    def _set_sidebar_buttons_enabled(self, enabled):
        state = "normal" if enabled else "disabled"
        for btn in self.sidebar_buttons.values():
            self.after(0, lambda b=btn, s=state: b.configure(state=s))

    # =================================================================
    #  ACTIONS
    # =================================================================
    def _go_gerar_pdfs(self):
        self._show_page("operations")
        self._action_gerar_pdfs()

    def _go_gerar_emails(self):
        self._show_page("operations")
        self._action_gerar_emails()

    def _action_gerar_pdfs(self):
        self._clear_ops_log()
        self.ops_title.configure(text="Gerando PDFs")
        self._update_sidebar_active("fluxo_rf")
        self._set_sidebar_buttons_enabled(False)
        self._set_ops_status("Iniciando...", ACCENT_ORANGE)
        self._set_ops_progress(0, "Preparando...")
        self._set_ops_counter("")
        threading.Thread(target=self._run_gerar_pdfs, daemon=True).start()

    def _run_gerar_pdfs(self):
        try:
            self._log("Importando gerador de PDFs...")
            import gerar_pdfs
            import importlib
            importlib.reload(gerar_pdfs)

            self._log("Carregando base de assessores...")
            assessores = gerar_pdfs.load_base_emails(BASE_FILE)
            self._log(f"{len(assessores)} assessores na base.")

            self._log("Carregando eventos...")
            eventos = gerar_pdfs.load_eventos(ENTRADA_FILE)
            total = sum(len(v) for v in eventos.values())
            self._log(f"{total} eventos carregados.")

            os.makedirs(OUTPUT_DIR, exist_ok=True)
            self._log("Gerando PDFs...\n")
            self._set_ops_status("Gerando PDFs...", ACCENT_ORANGE)

            ok = 0
            erros = 0
            codes = sorted(assessores.keys())
            n_total = len(codes)

            for i, código in enumerate(codes):
                info = assessores[código]
                evts = eventos.get(código, [])
                progress = (i + 1) / n_total

                try:
                    gerar_pdfs.generate_pdf(código, info, evts, OUTPUT_DIR)
                    self._log(f"{código}  {info['nome']}  ({len(evts)} eventos)", "ok")
                    ok += 1
                except Exception as e:
                    self._log(f"{código}  {info['nome']}:  {e}", "error")
                    erros += 1

                self._set_ops_progress(progress, f"{i + 1} de {n_total}")
                self._set_ops_counter(f"{ok} / {n_total}")

            self._log(f"\nFinalizado!  {ok} PDFs gerados,  {erros} erros.")
            self._set_ops_progress(1.0, "Concluído")
            self._set_ops_status(f"Concluído  -  {ok} PDFs gerados", "#00a86b" if erros == 0 else ACCENT_ORANGE)
            self._set_ops_counter(f"{ok} PDFs")
            self.after(0, self._update_pdf_status)

            if erros == 0:
                messagebox.showinfo("Sucesso", f"{ok} PDFs gerados com sucesso!\n\nPasta: {OUTPUT_DIR}")
            else:
                messagebox.showwarning("Atencao", f"{ok} PDFs gerados, {erros} com erro.")

        except Exception as e:
            self._log(f"ERRO: {e}", "error")
            self._set_ops_status(f"Erro:  {e}", ACCENT_RED)
            messagebox.showerror("Erro", str(e))
        finally:
            self._set_sidebar_buttons_enabled(True)

    def _action_gerar_emails(self):
        if not os.path.exists(OUTPUT_DIR) or not any(f.endswith(".pdf") for f in os.listdir(OUTPUT_DIR)):
            messagebox.showwarning("PDFs não encontrados", "Gere os PDFs primeiro antes de criar os emails.")
            self._show_page("fluxo_rf")
            return

        resp = messagebox.askyesno(
            "Gerar Emails no Outlook",
            "Isso vai criar um rascunho de email no Outlook para cada assessor.\n\n"
            "Os emails NÃO serão enviados, apenas criados como rascunho.\n\n"
            "O Outlook precisa estar aberto. Deseja continuar?"
        )
        if not resp:
            self._show_page("fluxo_rf")
            return

        self._clear_ops_log()
        self.ops_title.configure(text="Gerando Emails")
        self._update_sidebar_active("fluxo_rf")
        self._set_sidebar_buttons_enabled(False)
        self._set_ops_status("Conectando ao Outlook...", ACCENT_ORANGE)
        self._set_ops_progress(0, "Preparando...")
        self._set_ops_counter("")
        threading.Thread(target=self._run_gerar_emails, daemon=True).start()

    def _run_gerar_emails(self):
        try:
            import win32com.client as win32
            import pythoncom
            pythoncom.CoInitialize()

            self._log("Conectando ao Outlook...")
            try:
                outlook = win32.Dispatch("Outlook.Application")
            except Exception as e:
                self._log(f"Outlook não encontrado: {e}", "error")
                self._set_ops_status("Outlook não encontrado", ACCENT_RED)
                messagebox.showerror("Erro", "Não foi possível conectar ao Outlook.\nVerifique se está aberto.")
                return

            self._log("Carregando base de assessores...")
            assessores = load_base_emails()
            self._log(f"{len(assessores)} assessores na base.\n")

            criados = 0
            erros = 0
            sem_pdf = 0
            codes = sorted(assessores.keys())
            n_total = len(codes)
            self._set_ops_status("Criando rascunhos...", ACCENT_BLUE)

            for i, código in enumerate(codes):
                info = assessores[código]
                nome = info["nome"]
                email_assessor = info["email"]
                email_assistente = info["email_assistente"]
                progress = (i + 1) / n_total

                pdf_path = find_pdf_for_assessor(código, nome)
                if not pdf_path:
                    self._log(f"{código}  {nome}:  PDF não encontrado", "skip")
                    sem_pdf += 1
                    self._set_ops_progress(progress, f"{i + 1} de {n_total}")
                    continue

                if not email_assessor or email_assessor == "-":
                    self._log(f"{código}  {nome}:  Sem email", "skip")
                    self._set_ops_progress(progress, f"{i + 1} de {n_total}")
                    continue

                try:
                    mail = outlook.CreateItem(0)
                    mail.To = email_assessor
                    if email_assistente and email_assistente != "-":
                        mail.CC = email_assistente
                    mail.Subject = f"Fluxo de Renda Fixa - {nome} ({código})"
                    mail.HTMLBody = build_email_body(nome)
                    _attach_logo_cid(mail)
                    mail.Attachments.Add(os.path.abspath(pdf_path))
                    mail.Save()

                    cc = f"  CC: {email_assistente}" if email_assistente and email_assistente != "-" else ""
                    self._log(f"{código}  {nome}  \u2192  {email_assessor}{cc}", "ok")
                    criados += 1
                except Exception as e:
                    self._log(f"{código}  {nome}:  {e}", "error")
                    erros += 1

                self._set_ops_progress(progress, f"{i + 1} de {n_total}")
                self._set_ops_counter(f"{criados} emails")

            pythoncom.CoUninitialize()

            self._log(f"\nFinalizado!  {criados} emails criados,  {erros} erros,  {sem_pdf} sem PDF.")
            self._log("Os emails estao na pasta RASCUNHOS do Outlook.")
            self._set_ops_progress(1.0, "Concluído")
            self._set_ops_status(f"Concluído  -  {criados} emails nos Rascunhos", "#00a86b" if erros == 0 else ACCENT_ORANGE)

            if erros == 0:
                messagebox.showinfo("Emails Criados", f"{criados} emails criados nos Rascunhos do Outlook!")
            else:
                messagebox.showwarning("Atencao", f"{criados} emails criados, {erros} com erro.")

        except Exception as e:
            self._log(f"ERRO: {e}", "error")
            self._set_ops_status(f"Erro:  {e}", ACCENT_RED)
            messagebox.showerror("Erro", str(e))
        finally:
            self._set_sidebar_buttons_enabled(True)

    def _on_abrir_pasta(self):
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        os.startfile(OUTPUT_DIR)

    # -----------------------------------------------------------------
    #  REPORTAR ERRO
    # -----------------------------------------------------------------
    def _on_reportar_erro(self):
        win = ctk.CTkToplevel(self)
        win.title("Reportar Erro - SomusApp")
        win.geometry("520x480")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        # Centralizar
        win.update_idletasks()
        x = (win.winfo_screenwidth() - 520) // 2
        y = (win.winfo_screenheight() - 480) // 2
        win.geometry(f"520x480+{x}+{y}")

        try:
            ico_path = os.path.join(BASE_DIR, "assets", "icon_somus.ico")
            if os.path.exists(ico_path):
                win.iconbitmap(ico_path)
        except Exception:
            pass

        # Header
        header = ctk.CTkFrame(win, fg_color=ACCENT_RED, height=52, corner_radius=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header, text="\u26a0  Reportar Erro",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_WHITE
        ).pack(side="left", padx=20)

        # Body
        body = ctk.CTkFrame(win, fg_color=BG_PRIMARY, corner_radius=0)
        body.pack(fill="both", expand=True)

        inner = ctk.CTkFrame(body, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=24, pady=20)

        # Página onde ocorreu
        ctk.CTkLabel(
            inner, text="Página onde ocorreu o erro",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        erro_página = ctk.CTkComboBox(
            inner, font=("Segoe UI", 12), height=38,
            corner_radius=8, border_color=BORDER_CARD,
            values=["Dashboard", "FLUXO - RF", "Informativo", "Envio de Ordens",
                    "Ctrl Receita", "Organizador", "Envio Saldos",
                    "Corporate - Dashboard", "Corporate - Simulador", "Outro"],
            state="readonly"
        )
        erro_página.pack(fill="x", pady=(0, 14))
        erro_página.set("Outro")

        # Descrição do erro
        ctk.CTkLabel(
            inner, text="Descreva o erro encontrado",
            font=("Segoe UI", 11, "bold"), text_color=TEXT_SECONDARY, anchor="w"
        ).pack(fill="x", pady=(0, 4))

        erro_texto = ctk.CTkTextbox(
            inner, font=("Segoe UI", 11), height=180,
            corner_radius=8, border_width=1, border_color=BORDER_CARD,
            fg_color=BG_CARD, text_color=TEXT_PRIMARY, wrap="word"
        )
        erro_texto.pack(fill="x", pady=(0, 14))

        # Status
        status_label = ctk.CTkLabel(
            inner, text="",
            font=("Segoe UI", 10), text_color=TEXT_TERTIARY, anchor="w"
        )
        status_label.pack(fill="x", pady=(0, 8))

        # Botoes
        btn_frame = ctk.CTkFrame(inner, fg_color="transparent")
        btn_frame.pack(fill="x")

        def _enviar():
            página = erro_página.get()
            descrição = erro_texto.get("1.0", "end").strip()

            if not descrição:
                messagebox.showwarning("Campo obrigatorio", "Descreva o erro antes de enviar.", parent=win)
                return

            enviar_btn.configure(state="disabled")
            status_label.configure(text="Enviando...", text_color=ACCENT_ORANGE)
            win.update()

            try:
                import win32com.client as win32
                import pythoncom
                pythoncom.CoInitialize()

                outlook = win32.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = "artur.brito@somuscapital.com.br"
                mail.Subject = f"[SomusApp - Bug Report] {página}"

                hoje = datetime.now().strftime("%d/%m/%Y %H:%M")
                mail.HTMLBody = f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:700px;margin:0 auto;">
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr><td style="background:#dc3545;padding:14px 20px;">
    <span style="font-size:15pt;font-weight:bold;color:#fff;">\u26a0 Bug Report - SomusApp</span>
  </td></tr>
</table>
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:16px;">
  <tr><td style="padding:0 4px;">
    <table cellpadding="0" cellspacing="0" border="0"
      style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;font-size:10pt;width:100%;margin-bottom:16px;">
      <tr style="background:#f7faf9;">
        <td style="padding:8px 14px;font-weight:bold;color:#004d33;border-bottom:1.5px solid #004d33;width:140px;">Página</td>
        <td style="padding:8px 14px;color:#1a1a2e;border-bottom:1.5px solid #004d33;">{página}</td>
      </tr>
      <tr>
        <td style="padding:8px 14px;font-weight:bold;color:#004d33;border-bottom:1px solid #eee;">Data/Hora</td>
        <td style="padding:8px 14px;color:#1a1a2e;border-bottom:1px solid #eee;">{hoje}</td>
      </tr>
      <tr style="background:#f7faf9;">
        <td style="padding:8px 14px;font-weight:bold;color:#004d33;border-bottom:1px solid #eee;">Modulo</td>
        <td style="padding:8px 14px;color:#1a1a2e;border-bottom:1px solid #eee;">{self.current_module}</td>
      </tr>
    </table>
  </td></tr>
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="background-color:#dc3545;width:4px;border-radius:4px;">&nbsp;</td>
      <td style="padding-left:12px;">
        <span style="font-size:12.5pt;color:#dc3545;font-weight:bold;">Descrição do Erro</span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:8px 4px;">
    <div style="background:#fdf2f2;border:1px solid #f5c6cb;border-radius:6px;padding:14px 18px;font-size:10.5pt;color:#1a1a2e;">
      {descrição.replace(chr(10), '<br>')}
    </div>
  </td></tr>
  <tr><td style="padding:24px 0 0 0;">
    <hr style="border:none;border-top:1px solid #e0e0e0;margin:0 0 8px 0;">
    <span style="font-size:8pt;color:#9ca3af;">Enviado automaticamente pelo SomusApp - BETA</span>
  </td></tr>
</table>
</div>
"""
                mail.Send()
                pythoncom.CoUninitialize()

                status_label.configure(text="Erro reportado com sucesso!", text_color="#00a86b")
                messagebox.showinfo("Enviado", "Bug report enviado com sucesso!", parent=win)
                win.after(500, win.destroy)

            except Exception as e:
                enviar_btn.configure(state="normal")
                status_label.configure(text=f"Falha: {e}", text_color=ACCENT_RED)
                messagebox.showerror("Erro", f"Não foi possível enviar:\n{e}", parent=win)

        enviar_btn = ctk.CTkButton(
            btn_frame, text="\u2709  Enviar",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_RED, hover_color="#b02a37",
            height=42, corner_radius=10,
            command=_enviar,
        )
        enviar_btn.pack(side="left", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="Cancelar",
            font=("Segoe UI", 11),
            fg_color="#6b7280", hover_color="#555d6a",
            height=42, corner_radius=10, width=100,
            command=win.destroy,
        ).pack(side="left")

    # =================================================================
    #  TOP PICKS
    # =================================================================
    def _build_top_picks_page(self):
        page = ctk.CTkFrame(self, fg_color=BG_PRIMARY, corner_radius=0)

        self._make_topbar(page, "Top Picks", subtitle="Maiores Taxas por Faixa e Indexador")

        scroll = ctk.CTkScrollableFrame(page, fg_color=BG_PRIMARY, corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        content = ctk.CTkFrame(scroll, fg_color="transparent")
        content.pack(fill="x", padx=28, pady=20)

        # ======== DROP ZONE ========
        self.tp_drop_card = ctk.CTkFrame(
            content, fg_color="#fff8f0", corner_radius=14,
            border_width=2, border_color=ACCENT_ORANGE
        )
        self.tp_drop_card.pack(fill="x", pady=(0, 16))
        self.tp_drop_card.bind("<Button-1>", lambda e: self._on_browse_top_picks())

        drop_inner = ctk.CTkFrame(self.tp_drop_card, fg_color="transparent")
        drop_inner.pack(fill="x", padx=30, pady=36)
        drop_inner.bind("<Button-1>", lambda e: self._on_browse_top_picks())

        self.tp_icon_label = ctk.CTkLabel(
            drop_inner, text="\u2605",
            font=("Segoe UI", 44), text_color=ACCENT_ORANGE
        )
        self.tp_icon_label.pack()
        self.tp_icon_label.bind("<Button-1>", lambda e: self._on_browse_top_picks())

        self.tp_drop_title = ctk.CTkLabel(
            drop_inner, text="Clique para selecionar a planilha de ativos",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        )
        self.tp_drop_title.pack(pady=(8, 2))
        self.tp_drop_title.bind("<Button-1>", lambda e: self._on_browse_top_picks())

        self.tp_drop_sub = ctk.CTkLabel(
            drop_inner, text="Formatos aceitos:  .xlsx  .xls",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.tp_drop_sub.pack()
        self.tp_drop_sub.bind("<Button-1>", lambda e: self._on_browse_top_picks())

        self.tp_browse_btn = ctk.CTkButton(
            drop_inner, text="  Selecionar Arquivo",
            font=("Segoe UI", 12, "bold"),
            fg_color=ACCENT_ORANGE, hover_color="#c96e1f",
            height=40, corner_radius=8, width=200,
            command=self._on_browse_top_picks,
        )
        self.tp_browse_btn.pack(pady=(14, 0))

        # ======== FILE INFO BAR ========
        self.tp_file_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.tp_file_frame.pack(fill="x", pady=(0, 12))
        self.tp_file_frame.pack_forget()

        fi = ctk.CTkFrame(self.tp_file_frame, fg_color="transparent")
        fi.pack(fill="x", padx=16, pady=10)

        self.tp_file_icon = ctk.CTkLabel(
            fi, text="\u25cf", font=("Segoe UI", 12), text_color=ACCENT_GREEN
        )
        self.tp_file_icon.pack(side="left")

        self.tp_file_name = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 11, "bold"), text_color=TEXT_PRIMARY
        )
        self.tp_file_name.pack(side="left", padx=(6, 0))

        self.tp_file_info = ctk.CTkLabel(
            fi, text="", font=("Segoe UI", 10), text_color=TEXT_SECONDARY
        )
        self.tp_file_info.pack(side="left", padx=(12, 0))

        ctk.CTkButton(
            fi, text="Trocar", font=("Segoe UI", 10),
            fg_color=BG_INPUT, hover_color=BORDER_CARD,
            text_color=TEXT_SECONDARY, height=28, corner_radius=6, width=70,
            command=self._on_browse_top_picks,
        ).pack(side="right")

        # ======== PREVIEW ========
        self.tp_preview_label = ctk.CTkLabel(
            content, text="Preview dos Dados",
            font=("Segoe UI", 14, "bold"), text_color=TEXT_PRIMARY, anchor="w"
        )
        self.tp_preview_label.pack(fill="x", pady=(4, 8))

        self.tp_preview_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=12,
            border_width=1, border_color=BORDER_CARD
        )
        self.tp_preview_frame.pack(fill="x", pady=(0, 14))

        self.tp_preview_text = ctk.CTkTextbox(
            self.tp_preview_frame, font=("Consolas", 9),
            fg_color=BG_CARD, text_color=TEXT_PRIMARY,
            corner_radius=10, height=200, wrap="none"
        )
        self.tp_preview_text.pack(fill="x", padx=4, pady=4)
        self.tp_preview_text.insert("1.0", "  Nenhum arquivo carregado...")
        self.tp_preview_text.configure(state="disabled")

        # ======== ACTION BUTTONS ========
        btn_frame = ctk.CTkFrame(content, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(0, 12))

        self.tp_process_btn = ctk.CTkButton(
            btn_frame, text="  Gerar Top Picks",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_GREEN, hover_color=self._darken(ACCENT_GREEN),
            height=44, corner_radius=10, state="disabled",
            command=self._on_process_top_picks,
        )
        self.tp_process_btn.pack(side="left", padx=(0, 10))

        self.tp_open_btn = ctk.CTkButton(
            btn_frame, text="  Abrir Resultado",
            font=("Segoe UI", 13, "bold"),
            fg_color=ACCENT_BLUE, hover_color="#1555bb",
            height=44, corner_radius=10, state="disabled",
            command=self._on_open_top_picks_result,
        )
        self.tp_open_btn.pack(side="left", padx=(0, 10))

        # ======== STATUS BAR ========
        self.tp_status_frame = ctk.CTkFrame(
            content, fg_color=BG_CARD, corner_radius=10,
            border_width=1, border_color=BORDER_CARD
        )
        self.tp_status_frame.pack(fill="x", pady=(0, 8))

        si = ctk.CTkFrame(self.tp_status_frame, fg_color="transparent")
        si.pack(fill="x", padx=16, pady=10)

        self.tp_status_dot = ctk.CTkLabel(
            si, text="\u25cf", font=("Segoe UI", 11), text_color=TEXT_TERTIARY
        )
        self.tp_status_dot.pack(side="left")

        self.tp_status_text = ctk.CTkLabel(
            si, text="  Aguardando arquivo...",
            font=("Segoe UI", 11), text_color=TEXT_SECONDARY
        )
        self.tp_status_text.pack(side="left")

        # Internal state
        self._tp_input_path = None
        self._tp_output_path = None

        return page

    # -----------------------------------------------------------------
    #  TOP PICKS: Ações
    # -----------------------------------------------------------------
    def _on_browse_top_picks(self):
        path = filedialog.askopenfilename(
            title="Selecionar Planilha de Ativos",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self._tp_load_file(path)

    def _tp_load_file(self, path):
        self._tp_input_path = path
        self._tp_output_path = None

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            total_rows = ws.max_row or 0
            total_cols = ws.max_column or 0
            wb.close()

            fname = os.path.basename(path)
            self.tp_file_name.configure(text=fname)
            info = f"~{total_rows} linhas  |  {total_cols} colunas"
            self.tp_file_info.configure(text=info)
            self.tp_file_frame.pack(fill="x", pady=(0, 12))

            self.tp_drop_title.configure(text=fname)
            self.tp_drop_sub.configure(text="Arquivo carregado - clique para trocar")
            self.tp_icon_label.configure(text_color=ACCENT_GREEN)
            self.tp_drop_card.configure(border_color=ACCENT_GREEN, fg_color="#f0fff4")

            # Preview
            self.tp_preview_text.configure(state="normal")
            self.tp_preview_text.delete("1.0", "end")

            wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws2 = wb2[wb2.sheetnames[0]]
            lines = []
            for i, row in enumerate(ws2.iter_rows(max_row=min(12, total_rows), values_only=True)):
                vals = [str(c) if c is not None else "" for c in row]
                lines.append("  ".join(f"{v:<18s}" for v in vals[:10]))
            wb2.close()

            self.tp_preview_text.insert("1.0", "\n".join(lines))
            self.tp_preview_text.configure(state="disabled")

            self.tp_process_btn.configure(state="normal")
            self.tp_open_btn.configure(state="disabled")
            self.tp_status_dot.configure(text_color=ACCENT_GREEN)
            self.tp_status_text.configure(text=f"  Arquivo carregado: {fname}")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao ler arquivo:\n{e}")

    def _on_process_top_picks(self):
        self.tp_process_btn.configure(state="disabled")
        self.tp_status_dot.configure(text_color=ACCENT_ORANGE)
        self.tp_status_text.configure(text="  Processando Top Picks...")
        threading.Thread(target=self._run_top_picks, daemon=True).start()

    def _on_open_top_picks_result(self):
        if self._tp_output_path and os.path.exists(self._tp_output_path):
            os.startfile(self._tp_output_path)

    @staticmethod
    def _tp_extract_tax_number(tax_text):
        """Extrai valor numerico de textos como 'IPC-A + 8,50%' ou '117,50% CDI' ou '13,50%'."""
        if not tax_text:
            return 0.0
        s = str(tax_text).replace("%", "").strip()
        nums = re.findall(r'(\d+[,.]?\d*)', s)
        if not nums:
            return 0.0
        val_str = nums[-1].replace(",", ".")
        try:
            return float(val_str)
        except ValueError:
            return 0.0

    @staticmethod
    def _tp_count_business_days(start_date, end_date):
        """Conta dias uteis (seg-sex) entre duas datas."""
        from datetime import timedelta
        if not start_date or not end_date:
            return 0
        if isinstance(start_date, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    start_date = datetime.strptime(start_date, fmt)
                    break
                except ValueError:
                    continue
        if isinstance(end_date, str):
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    end_date = datetime.strptime(end_date, fmt)
                    break
                except ValueError:
                    continue
        if not isinstance(start_date, datetime) or not isinstance(end_date, datetime):
            return 0
        if end_date <= start_date:
            return 0
        total_days = (end_date - start_date).days
        full_weeks = total_days // 7
        remaining = total_days % 7
        count = full_weeks * 5
        d = start_date.weekday()
        for i in range(1, remaining + 1):
            if (d + i) % 7 < 5:
                count += 1
        return count

    @staticmethod
    def _tp_classify_indexador(indexador_text):
        """Classifica indexador em '% CDI', 'IPC-A', 'PRE FIXADO' ou None."""
        if not indexador_text:
            return None
        s = indexador_text.upper()
        if "CDI" in s and "CDI+" not in s.replace(" ", ""):
            return "% CDI"
        if "IPC" in s or "IPCA" in s:
            return "IPC-A"
        if "PR\u00c9" in s or "PRE" in s:
            return "PR\u00c9 FIXADO"
        return None

    @staticmethod
    def _tp_extract_ativo_name(ativo_text):
        """Extrai nome do emissor: 'CDB FIBRA - JUL/2031' -> 'FIBRA'."""
        if not ativo_text:
            return ""
        s = str(ativo_text)
        pos = s.find("-")
        if pos > 4:
            return s[3:pos].strip()
        return s[3:].strip() if len(s) > 3 else s

    HIGH_GRADE_RATINGS = {
        "AA-", "AA", "AA+", "AAA",
        "brAA-", "brAA", "brAA+", "brAAA",
        "AA-.br", "AA.br", "AA+.br", "AAA.br",
    }

    def _tp_find_best_per_faixa(self, rows_data, col_map, idx_dias_uteis, filter_ratings=None):
        """Encontra o melhor ativo por faixa x indexador.
        Se filter_ratings for um set, filtra apenas esses ratings."""
        faixas = [(90, 199), (200, 400), (600, 800)]
        result_rows = []

        for faixa_min, faixa_max in faixas:
            best = {}
            for row in rows_data:
                dias = row[idx_dias_uteis]
                if not isinstance(dias, (int, float)) or dias < faixa_min or dias > faixa_max:
                    continue

                # Filtro de rating (high grade)
                if filter_ratings is not None:
                    rating_val = str(row[col_map["rating"]]).strip() if row[col_map["rating"]] else ""
                    if rating_val not in filter_ratings:
                        continue

                tag = self._tp_classify_indexador(
                    str(row[col_map["indexador"]]) if row[col_map["indexador"]] else ""
                )
                if not tag:
                    continue

                tax_text = str(row[col_map["tax_max"]]) if row[col_map["tax_max"]] else ""
                tax_val = self._tp_extract_tax_number(tax_text)

                if tag not in best or tax_val > best[tag][1]:
                    best[tag] = (row, tax_val)

            for tag in ["% CDI", "PR\u00c9 FIXADO", "IPC-A"]:
                if tag in best:
                    row = best[tag][0]
                    ativo_raw = str(row[col_map["ativo"]]) if row[col_map["ativo"]] else ""
                    # Tipo = primeiros 3 caracteres (CDB, LCA, LCI, etc.)
                    tipo = ativo_raw[:3].strip() if len(ativo_raw) >= 3 else ativo_raw
                    emissor = self._tp_extract_ativo_name(ativo_raw)
                    dias_val = row[idx_dias_uteis]
                    venc = row[col_map["vencimento"]]
                    indexador = str(row[col_map["indexador"]]) if row[col_map["indexador"]] else ""
                    tax_min_txt = str(row[col_map["tax_min"]]) if row[col_map["tax_min"]] else ""
                    tax_max_txt = str(row[col_map["tax_max"]]) if row[col_map["tax_max"]] else ""
                    rating_val = str(row[col_map["rating"]]).strip() if row[col_map["rating"]] else "Sem Rating"
                    if not rating_val:
                        rating_val = "Sem Rating"
                    roa_val = row[col_map["roa"]] if row[col_map["roa"]] else ""

                    result_rows.append([tipo, emissor, dias_val, venc,
                                        indexador, tax_min_txt, tax_max_txt, rating_val, roa_val])

        return result_rows

    def _run_top_picks(self):
        try:
            wb = openpyxl.load_workbook(self._tp_input_path, data_only=True)
            ws = wb[wb.sheetnames[0]]

            headers = []
            rows_data = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c else "" for c in row]
                    continue
                rows_data.append(list(row))
            wb.close()

            # Mapear colunas pelo header
            col_map = {}
            header_keys = {
                "Ativo": "ativo", "Ticker": "ticker", "Indexador": "indexador",
                "Rating": "rating", "Vencimento": "vencimento",
            }
            for idx, h in enumerate(headers):
                clean = h.strip()
                if clean in header_keys:
                    col_map[header_keys[clean]] = idx
                elif clean.startswith("Tax.M\u00edn") or clean == "Tax.Min":
                    col_map["tax_min"] = idx
                elif clean == "Tax.M\u00e1x" or clean == "Tax.Max":
                    col_map["tax_max"] = idx
                elif clean == "ROA E. Aprox." or clean.startswith("ROA"):
                    col_map["roa"] = idx

            col_map.setdefault("ativo", 0)
            col_map.setdefault("ticker", 1)
            col_map.setdefault("indexador", 3)
            col_map.setdefault("rating", 9)
            col_map.setdefault("vencimento", 11)
            col_map.setdefault("tax_min", 12)
            col_map.setdefault("tax_max", 13)
            col_map.setdefault("roa", 19)

            today = datetime.now()

            for row in rows_data:
                venc = row[col_map["vencimento"]]
                dias = self._tp_count_business_days(today, venc)
                row.append(dias)
            idx_dias_uteis = len(headers)

            # Gerar output
            wb_out = openpyxl.Workbook()

            # Aba MAIORES TAXAS
            ws_taxas = wb_out.active
            ws_taxas.title = "MAIORES TAXAS"
            out_headers = headers + ["Dias \u00dateis Restantes"]
            ws_taxas.append(out_headers)
            for row in rows_data:
                ws_taxas.append(row)

            # Aba HIGH GRADE
            ws_hg = wb_out.create_sheet("HIGH GRADE")
            ws_hg.append(out_headers)
            for row in rows_data:
                rating_val = str(row[col_map["rating"]]) if row[col_map["rating"]] else ""
                if rating_val.strip() in self.HIGH_GRADE_RATINGS:
                    ws_hg.append(row)

            # Aba CONSOLIDADO — High Yield (todos) + High Grade (filtrado)
            ws_cons = wb_out.create_sheet("CONSOLIDADO")
            cons_headers = ["Tipo", "Emissor", "Dias \u00dateis", "Vencimento",
                            "Indexador", "Taxa M\u00edn", "Taxa M\u00e1x", "Rating", "ROA"]

            # Secao 1: HIGH YIELD (todos os ratings)
            ws_cons.append(["TOP PICKS - HIGH YIELD"])
            ws_cons.append(cons_headers)
            hy_rows = self._tp_find_best_per_faixa(rows_data, col_map, idx_dias_uteis)
            for cr in hy_rows:
                ws_cons.append(cr)

            # Linha em branco de separacao
            ws_cons.append([])

            # Secao 2: HIGH GRADE (apenas ratings AA+)
            ws_cons.append(["TOP PICKS - HIGH GRADE"])
            ws_cons.append(cons_headers)
            hg_rows = self._tp_find_best_per_faixa(rows_data, col_map, idx_dias_uteis,
                                                    filter_ratings=self.HIGH_GRADE_RATINGS)
            for cr in hg_rows:
                ws_cons.append(cr)

            all_cons_rows = hy_rows + hg_rows

            # Formatar datas no CONSOLIDADO
            for r in range(1, ws_cons.max_row + 1):
                cell = ws_cons.cell(row=r, column=4)
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD/MM/YYYY"

            # Formatar datas no MAIORES TAXAS
            venc_col_out = col_map["vencimento"] + 1
            for r in range(2, ws_taxas.max_row + 1):
                cell = ws_taxas.cell(row=r, column=venc_col_out)
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD/MM/YYYY"

            # Salvar
            output_dir = os.path.join(BASE_DIR, "Mesa Produtos", "Top Picks")
            os.makedirs(output_dir, exist_ok=True)
            output_name = f"TOP_PICKS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = os.path.join(output_dir, output_name)
            wb_out.save(output_path)
            self._tp_output_path = output_path

            n_taxas = len(rows_data)
            n_hg = ws_hg.max_row - 1
            n_hy = len(hy_rows)
            n_hg_picks = len(hg_rows)

            def update_ok():
                self.tp_status_dot.configure(text_color=ACCENT_GREEN)
                self.tp_status_text.configure(
                    text=f"  Conclu\u00eddo! {n_taxas} ativos | {n_hy} high yield | {n_hg_picks} high grade"
                )
                self.tp_process_btn.configure(state="normal")
                self.tp_open_btn.configure(state="normal")

                self.tp_preview_text.configure(state="normal")
                self.tp_preview_text.delete("1.0", "end")
                preview_lines = []
                preview_lines.append("  === TOP PICKS - HIGH YIELD ===")
                preview_lines.append("  " + "  ".join(f"{h:<16s}" for h in cons_headers))
                preview_lines.append("  " + "-" * 145)
                for cr in hy_rows:
                    vals = [str(v) if v is not None else "" for v in cr]
                    preview_lines.append("  " + "  ".join(f"{v:<16s}" for v in vals))
                preview_lines.append("")
                preview_lines.append("  === TOP PICKS - HIGH GRADE ===")
                preview_lines.append("  " + "  ".join(f"{h:<16s}" for h in cons_headers))
                preview_lines.append("  " + "-" * 145)
                for cr in hg_rows:
                    vals = [str(v) if v is not None else "" for v in cr]
                    preview_lines.append("  " + "  ".join(f"{v:<16s}" for v in vals))
                if not all_cons_rows:
                    preview_lines.append("  Nenhum top pick encontrado nas faixas definidas.")
                self.tp_preview_text.insert("1.0", "\n".join(preview_lines))
                self.tp_preview_text.configure(state="disabled")

            self.after(0, update_ok)

        except Exception as e:
            import traceback
            traceback.print_exc()
            def update_err():
                self.tp_status_dot.configure(text_color=ACCENT_RED)
                self.tp_status_text.configure(text=f"  Erro: {e}")
                self.tp_process_btn.configure(state="normal")
            self.after(0, update_err)

    @staticmethod
    def _darken(hex_color):
        r = max(0, int(hex_color[1:3], 16) - 30)
        g = max(0, int(hex_color[3:5], 16) - 30)
        b = max(0, int(hex_color[5:7], 16) - 30)
        return f"#{r:02x}{g:02x}{b:02x}"


# =====================================================================
#  LOGIN
# =====================================================================

_LOGIN_BG = "#004d33"
_LOGIN_BG_LIGHT = "#005c3d"
_LOGIN_CARD = "#ffffff"
_LOGIN_BTN = "#f0e0c0"
_LOGIN_BTN_HOVER = "#e6d1a8"
_LOGIN_BTN_TEXT = "#004d33"
_LOGIN_ACCENT = "#00b876"


class LoginWindow(ctk.CTkToplevel):
    """Tela de login com senha fixa - visual premium Somus Capital."""

    PASSWORDS = {
        "@Produtos1": "admin",
        "Corporate1": "corporate",
    }

    def __init__(self, master, on_success):
        super().__init__(master)
        self.on_success = on_success
        self.title("Somus Capital")
        self.geometry("480x600")
        self.resizable(False, False)
        self.configure(fg_color=_LOGIN_BG)
        self.transient(master)
        self.grab_set()

        # Centralizar na tela
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 480) // 2
        y = (self.winfo_screenheight() - 600) // 2
        self.geometry(f"480x600+{x}+{y}")

        # Icone
        try:
            ico_path = os.path.join(BASE_DIR, "assets", "icon_somus.ico")
            if os.path.exists(ico_path):
                self.iconbitmap(ico_path)
        except Exception:
            pass

        # Fechar janela = fechar app
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self._build_ui()

    def _build_ui(self):
        # ── Topo verde com logo ──
        top_frame = ctk.CTkFrame(self, fg_color="transparent", height=180)
        top_frame.pack(fill="x", pady=(0, 0))
        top_frame.pack_propagate(False)

        # Logo (branca sobre fundo verde = visivel)
        try:
            logo_img = Image.open(LOGO_PATH)
            self._logo_ctk = ctk.CTkImage(
                light_image=logo_img, dark_image=logo_img, size=(200, 49)
            )
            ctk.CTkLabel(top_frame, image=self._logo_ctk, text="").place(
                relx=0.5, rely=0.38, anchor="center"
            )
        except Exception:
            ctk.CTkLabel(
                top_frame, text="SOMUS CAPITAL",
                font=("Segoe UI", 24, "bold"), text_color=TEXT_WHITE
            ).place(relx=0.5, rely=0.38, anchor="center")

        # Subtítulos
        ctk.CTkLabel(
            top_frame, text="Mesa de Produtos  |  Agente de Investimentos",
            font=("Segoe UI", 11), text_color="#8fbfaa"
        ).place(relx=0.5, rely=0.68, anchor="center")

        # Linha decorativa verde claro
        line = ctk.CTkFrame(self, fg_color=_LOGIN_ACCENT, height=3, corner_radius=2)
        line.pack(fill="x", padx=60, pady=(0, 0))

        # ── Card branco central ──
        card = ctk.CTkFrame(
            self, fg_color=_LOGIN_CARD, corner_radius=16,
            border_width=0, width=360, height=280
        )
        card.pack(pady=(30, 0))
        card.pack_propagate(False)

        # Icone Somus pequeno no card
        try:
            icon_path = os.path.join(BASE_DIR, "assets", "icon_somus.png")
            icon_img = Image.open(icon_path)
            self._icon_ctk = ctk.CTkImage(
                light_image=icon_img, dark_image=icon_img, size=(48, 48)
            )
            ctk.CTkLabel(card, image=self._icon_ctk, text="").pack(pady=(24, 4))
        except Exception:
            ctk.CTkLabel(
                card, text="S", font=("Segoe UI", 28, "bold"),
                text_color=_LOGIN_BG, width=48, height=48,
                fg_color="#e8f5e9", corner_radius=24
            ).pack(pady=(24, 4))

        # Título do card
        ctk.CTkLabel(
            card, text="Acesso Restrito",
            font=("Segoe UI", 16, "bold"), text_color=TEXT_PRIMARY
        ).pack(pady=(0, 16))

        # Input senha
        self.password_entry = ctk.CTkEntry(
            card, placeholder_text="Digite a senha...",
            show="\u2022", width=280, height=44,
            font=("Segoe UI", 13),
            border_color="#d1d5db", border_width=1,
            fg_color="#f9fafb", text_color=TEXT_PRIMARY,
            corner_radius=10
        )
        self.password_entry.pack(pady=(0, 6))
        self.password_entry.bind("<Return>", lambda e: self._try_login())
        self.password_entry.focus_set()

        # Mensagem de erro
        self.error_label = ctk.CTkLabel(
            card, text="", font=("Segoe UI", 10),
            text_color="#dc3545", height=18
        )
        self.error_label.pack(pady=(0, 8))

        # Botao entrar (bege/dourado)
        self.login_btn = ctk.CTkButton(
            card, text="Entrar", width=280, height=44,
            font=("Segoe UI", 14, "bold"),
            fg_color=_LOGIN_BTN, hover_color=_LOGIN_BTN_HOVER,
            text_color=_LOGIN_BTN_TEXT,
            corner_radius=10, command=self._try_login
        )
        self.login_btn.pack(pady=(0, 24))

        # ── Footer ──
        ctk.CTkLabel(
            self, text="SomusApp - BETA",
            font=("Segoe UI", 9), text_color="#5a9a7d"
        ).pack(side="bottom", pady=(0, 16))

    def _try_login(self):
        senha = self.password_entry.get()
        role = self.PASSWORDS.get(senha)
        if role:
            self.grab_release()
            self.destroy()
            self.on_success(role)
        else:
            self.error_label.configure(text="Senha incorreta. Tente novamente.")
            self.password_entry.delete(0, "end")
            self.password_entry.focus_set()
            # Shake animation
            orig_x = self.winfo_x()
            orig_y = self.winfo_y()
            import time
            for dx in [10, -10, 8, -8, 4, -4, 0]:
                self.geometry(f"+{orig_x + dx}+{orig_y}")
                self.update_idletasks()
                time.sleep(0.03)

    def _on_close(self):
        self.grab_release()
        self.destroy()
        self.master.quit()


def main():
    # Definir AppUserModelID para o Windows usar o icone do app na barra de tarefas
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("somus.capital.fluxorf")
    except Exception:
        pass

    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("dark-blue")

    # Janela root oculta (host para o login)
    root = ctk.CTk()
    root.withdraw()

    try:
        ico_path = os.path.join(BASE_DIR, "assets", "icon_somus.ico")
        if os.path.exists(ico_path):
            root.iconbitmap(ico_path)
    except Exception:
        pass

    login_ok = [False]
    login_role = [None]

    def on_login_success(role):
        login_ok[0] = True
        login_role[0] = role
        root.quit()

    login = LoginWindow(root, on_login_success)
    root.mainloop()

    # Apos login bem-sucedido, destruir root e criar o app em novo mainloop
    if login_ok[0]:
        root.destroy()
        app = App(role=login_role[0])
        app.mainloop()


if __name__ == "__main__":
    main()
