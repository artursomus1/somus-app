"""
Somus Capital - Processador de Saldos Di\u00e1rios
Gera rascunhos de e-mail no Outlook para cada assessor com os saldos dos clientes.
"""

import os
import base64
from datetime import datetime, timedelta
from collections import defaultdict

import openpyxl


# === PATHS ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Mesa Produtos/Envio Saldos/
APP_ROOT = os.path.dirname(os.path.dirname(BASE_DIR))  # APP SOMUS/
SALDOS_BASE_DIR = os.path.join(BASE_DIR, "BASE")
BASE_EMAILS_FILE = os.path.join(APP_ROOT, "BASE", "BASE EMAILS.xlsx")
BASE_XLSM_FILE = os.path.join(APP_ROOT, "BASE", "BASE.xlsm")
BASE_ENVIAR_FILE = os.path.join(SALDOS_BASE_DIR, "BASE ENVIAR.xlsx")
LOGO_PATH = os.path.join(APP_ROOT, "assets", "logo_somus.png")
SAIDA_DIR = os.path.join(BASE_DIR, "SAIDA SALDOS")


def _load_logo_b64():
    """Carrega logo Somus como base64 data URI para embedding inline."""
    if not os.path.exists(LOGO_PATH):
        return ""
    with open(LOGO_PATH, "rb") as f:
        return base64.b64encode(f.read()).decode()


def sanitize_text(text):
    if text is None:
        return ""
    text = str(text)
    replacements = {
        '\u2013': '-', '\u2014': '-', '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2026': '...', '\u2022': '-',
        '\u00ba': 'o', '\u00aa': 'a', '\u00a0': ' ',
        '\ufeff': '', '\u200b': '', '\u200e': '', '\u200f': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def fmt_brl(value):
    """Formata n\u00famero como R$ brasileiro."""
    if value is None or value == "" or value == "-":
        return "-"
    try:
        num = float(str(value).replace('.', '').replace(',', '.')) if isinstance(value, str) else float(value)
    except (ValueError, TypeError):
        return str(value)
    if num == 0:
        return "-"
    negative = num < 0
    num = abs(num)
    # Arredonda para 2 casas decimais antes de separar inteiro/centavos
    num = round(num, 2)
    inteiro = int(num)
    centavos = round((num - inteiro) * 100)
    if centavos >= 100:
        inteiro += 1
        centavos = 0
    s_inteiro = f"{inteiro:,}".replace(",", ".")
    result = f"R$ {s_inteiro},{centavos:02d}"
    if negative:
        result = f"-{result}"
    return result


def load_base_somus():
    """Carrega mapeamento Conta(C\u00f3digo Cliente) -> C\u00f3digo Assessor (A-prefix) do BASE.xlsm."""
    mapping = {}
    if not os.path.exists(BASE_XLSM_FILE):
        return mapping
    wb = openpyxl.load_workbook(BASE_XLSM_FILE, read_only=True, data_only=True)
    if "BASE SOMUS" not in wb.sheetnames:
        wb.close()
        return mapping
    ws = wb["BASE SOMUS"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] and row[2]:
            conta = str(row[0]).strip()
            assessor_code = str(row[2]).strip()  # A69519 format
            assessor_nome = sanitize_text(row[3]) if row[3] else ""
            mapping[conta] = {"code": assessor_code, "nome": assessor_nome}
    wb.close()
    return mapping


def load_emails():
    """Carrega mapeamento C\u00f3digo Assessor -> e-mail de ambas as fontes."""
    emails = {}

    # 1) BASE EMAILS.xlsx (prioridade - tem assistente tamb\u00e9m)
    if os.path.exists(BASE_EMAILS_FILE):
        wb = openpyxl.load_workbook(BASE_EMAILS_FILE, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0]:
                code = str(row[0]).strip()
                emails[code] = {
                    "nome": sanitize_text(row[1]) if row[1] else "",
                    "email": str(row[2]).strip() if row[2] else "",
                    "assistente": sanitize_text(row[3]) if row[3] else "-",
                    "email_assistente": str(row[4]).strip() if row[4] else "-",
                }
                # Tamb\u00e9m sem prefixo A
                if code.startswith("A"):
                    emails[code[1:]] = emails[code]
        wb.close()

    # 2) Lista de e-mails do BASE.xlsm (fallback)
    if os.path.exists(BASE_XLSM_FILE):
        wb = openpyxl.load_workbook(BASE_XLSM_FILE, read_only=True, data_only=True)
        if "Lista de e-mails" in wb.sheetnames:
            ws = wb["Lista de e-mails"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row[0]:
                    code = str(row[0]).strip()
                    a_code = f"A{code}" if not code.startswith("A") else code
                    if a_code not in emails and code not in emails:
                        emails[code] = {
                            "nome": sanitize_text(row[1]) if row[1] else "",
                            "email": str(row[2]).strip() if row[2] else "",
                            "assistente": sanitize_text(row[3]) if row[3] else "-",
                            "email_assistente": str(row[4]).strip() if row[4] else "-",
                        }
                        emails[a_code] = emails[code]
        wb.close()

    return emails


def parse_saldo_value(val):
    """Converte valor da planilha para float. Retorna 0.0 se vazio."""
    if val is None or val == "" or val == "-":
        return 0.0
    try:
        if isinstance(val, str):
            return float(val.replace('.', '').replace(',', '.'))
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def load_saldos(filepath):
    """L\u00ea a planilha de saldos e retorna lista de registros."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        conta = str(row[0]).strip()
        cliente = sanitize_text(row[1]) if row[1] else ""
        assessor_raw = str(row[2]).strip() if row[2] else ""
        d0 = parse_saldo_value(row[3])
        d1 = parse_saldo_value(row[4])
        d2 = parse_saldo_value(row[5])
        d3 = parse_saldo_value(row[6])
        total = parse_saldo_value(row[7])
        records.append({
            "conta": conta,
            "cliente": cliente,
            "assessor_raw": assessor_raw,
            "d0": d0,
            "d1": d1,
            "d2": d2,
            "d3": d3,
            "total": total,
        })
    wb.close()
    return records


def _parse_date_obj(val):
    """Converte valor para objeto date. Retorna None se n\u00e3o conseguir."""
    if val is None or val == "" or val == "-":
        return None
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def fmt_date(val):
    """Formata data para DD/MM/YYYY."""
    if val is None or val == "" or val == "-":
        return "-"
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    # Tenta parsear "2026-02-26 00:00:00"
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s[:10]


def load_vencimentos_rf():
    """Carrega vencimentos de renda fixa do BASE ENVIAR.xlsx, agrupados por assessor.
    Filtra apenas vencimentos dos pr\u00f3ximos 4 dias (hoje at\u00e9 hoje+4).
    Deduplicado por (assessor, cliente, ativo, vencimento) — mant\u00e9m o maior financeiro."""
    venc_por_assessor = defaultdict(list)
    if not os.path.exists(BASE_ENVIAR_FILE):
        return venc_por_assessor
    wb = openpyxl.load_workbook(BASE_ENVIAR_FILE, read_only=True, data_only=True)
    if "Vencimentos renda fixa" not in wb.sheetnames:
        wb.close()
        return venc_por_assessor
    ws = wb["Vencimentos renda fixa"]
    hoje = datetime.now().date()
    limite = hoje + timedelta(days=4)
    # Row 2 = header: _, CLIENTE, NOME, AI, ATIVO, EMISSOR, INDEXADOR, VENCIMENTO, FINANCEIRO
    # Usa dict para deduplicar por (assessor, cliente, ativo, vencimento) — mant\u00e9m maior financeiro
    dedup = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[1]:
            continue
        ai = str(row[3]).strip() if row[3] else ""
        if not ai:
            continue
        # Filtra apenas vencimentos entre hoje e hoje+4 dias
        dt_venc = _parse_date_obj(row[7])
        if not dt_venc or dt_venc < hoje or dt_venc > limite:
            continue
        cliente_cod = str(row[1]).strip()
        ativo = sanitize_text(row[4]).strip() if row[4] else ""
        vencimento_str = fmt_date(row[7])
        financeiro = parse_saldo_value(row[8])
        dedup_key = (ai, cliente_cod, ativo, vencimento_str)
        existing = dedup.get(dedup_key)
        if existing is None or abs(financeiro) > abs(existing["financeiro"]):
            dedup[dedup_key] = {
                "cliente_cod": cliente_cod,
                "nome": sanitize_text(row[2]) if row[2] else "",
                "ativo": ativo,
                "emissor": sanitize_text(row[5]) if row[5] else "",
                "indexador": sanitize_text(row[6]) if row[6] else "",
                "vencimento": vencimento_str,
                "financeiro": financeiro,
                "_ai": ai,
            }
    wb.close()
    for rec in dedup.values():
        ai = rec.pop("_ai")
        venc_por_assessor[ai].append(rec)
    return venc_por_assessor


def load_vencimentos_fundos():
    """Carrega vencimentos de fundos do BASE ENVIAR.xlsx, agrupados por assessor.
    Filtra apenas líquida\u00e7\u00f5es dos pr\u00f3ximos 4 dias (hoje at\u00e9 hoje+4)."""
    venc_por_assessor = defaultdict(list)
    if not os.path.exists(BASE_ENVIAR_FILE):
        return venc_por_assessor
    wb = openpyxl.load_workbook(BASE_ENVIAR_FILE, read_only=True, data_only=True)
    # Aceita variantes do nome da aba
    sheet_name = None
    for sn in wb.sheetnames:
        if "vencimento" in sn.lower() and "fundo" in sn.lower():
            sheet_name = sn
            break
    if not sheet_name:
        wb.close()
        return venc_por_assessor
    ws = wb[sheet_name]
    hoje = datetime.now().date()
    limite = hoje + timedelta(days=4)
    # Row 2 = header: _, CLIENTE, NOME, AI, OPERAÇÃO, FUNDO, DATA COTIZACAO, DATA LIQUIDACAO, FINANCEIRO
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        if not row[1]:
            continue
        ai = str(row[3]).strip() if row[3] else ""
        if not ai:
            continue
        # Filtra apenas líquida\u00e7\u00f5es entre hoje e hoje+4 dias
        dt_liq = _parse_date_obj(row[7])
        if not dt_liq or dt_liq < hoje or dt_liq > limite:
            continue
        venc_por_assessor[ai].append({
            "cliente_cod": str(row[1]).strip(),
            "nome": sanitize_text(row[2]) if row[2] else "",
            "operacao": sanitize_text(row[4]) if row[4] else "",
            "fundo": sanitize_text(row[5]) if row[5] else "",
            "dt_cotizacao": fmt_date(row[6]),
            "dt_liquidacao": fmt_date(row[7]),
            "financeiro": parse_saldo_value(row[8]),
        })
    wb.close()
    return venc_por_assessor


def classify_and_group(records, base_somus, emails_map):
    """
    Classifica registros em categorias e agrupa por assessor.

    Categorias (mesmo crit\u00e9rio do VBA):
    - positivos: D0 >= 100
    - negativos_d0: D0 <= -50
    - negativos_d1: D+1 <= -50

    Retorna dict: assessor_code -> {
        'nome': str, 'email': str, 'email_assistente': str,
        'primeiro_nome': str,
        'positivos': [...], 'negativos_d0': [...], 'negativos_d1': [...]
    }
    """
    assessor_data = {}

    for rec in records:
        conta = rec["conta"]

        # Mapeia conta -> assessor individual via BASE SOMUS
        somus_info = base_somus.get(conta)
        if somus_info:
            assessor_code = somus_info["code"]
        else:
            # Fallback: usa o assessor da planilha com prefixo A
            raw = rec["assessor_raw"]
            assessor_code = f"A{raw}" if raw and not raw.startswith("A") else raw

        if not assessor_code:
            continue

        # Inicializa assessor se necess\u00e1rio
        if assessor_code not in assessor_data:
            # Busca email
            email_info = emails_map.get(assessor_code, {})
            if not email_info:
                # Tenta sem A
                code_num = assessor_code[1:] if assessor_code.startswith("A") else assessor_code
                email_info = emails_map.get(code_num, {})

            nome = email_info.get("nome", somus_info["nome"] if somus_info else "")
            primeiro_nome = nome.split()[0].capitalize() if nome else ""

            assessor_data[assessor_code] = {
                "nome": nome,
                "email": email_info.get("email", ""),
                "assistente": email_info.get("assistente", "-"),
                "email_assistente": email_info.get("email_assistente", "-"),
                "primeiro_nome": primeiro_nome,
                "positivos": [],
                "negativos_d0": [],
                "negativos_d1": [],
            }

        # Classifica
        if rec["d0"] >= 100:
            assessor_data[assessor_code]["positivos"].append(rec)
        if rec["d0"] <= -50:
            assessor_data[assessor_code]["negativos_d0"].append(rec)
        if rec["d1"] <= -50:
            assessor_data[assessor_code]["negativos_d1"].append(rec)

    # Ordena cada lista por saldo D0 do maior para o menor (absoluto)
    for data in assessor_data.values():
        data["positivos"].sort(key=lambda r: r["d0"], reverse=True)
        data["negativos_d0"].sort(key=lambda r: r["d0"])  # mais negativo primeiro
        data["negativos_d1"].sort(key=lambda r: r["d1"])  # mais negativo primeiro

    return assessor_data


def build_html_table(records, columns, empty_msg=""):
    """Gera tabela HTML compacta e clean."""
    if not records:
        if empty_msg:
            return (
                f'<p style="font-family:Calibri,Arial,sans-serif;font-size:10.5pt;'
                f'color:#6b7280;font-style:italic;margin:4px 0 16px 0;">'
                f'{empty_msg}</p>\n'
            )
        return ""

    html = '<table cellpadding="0" cellspacing="0" border="0" '
    html += 'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;'
    html += 'font-size:9pt;margin-bottom:6px;">\n'

    # Header — apenas texto bold com borda inferior fina
    html += '<tr>'
    for key, label in columns:
        html += (
            f'<td style="padding:3px 8px 4px 8px;font-weight:bold;color:#00785a;'
            f'font-size:8.5pt;border-bottom:1.5px solid #00785a;'
            f'white-space:nowrap;">{label}</td>'
        )
    html += '</tr>\n'

    money_keys = ("d0", "d1", "d2", "d3", "total", "financeiro")

    for i, rec in enumerate(records):
        bg = "#f7faf9" if i % 2 == 0 else "#ffffff"
        html += f'<tr style="background:{bg};">'
        for key, label in columns:
            val = rec.get(key, "")
            if key in money_keys:
                val = fmt_brl(val) if val != 0 else "-"
                raw = rec.get(key, 0)
                color = "#c0392b" if isinstance(raw, (int, float)) and raw < 0 else "#1a1a2e"
                html += (
                    f'<td style="padding:2px 8px;text-align:right;color:{color};'
                    f'border-bottom:1px solid #eef1ef;font-size:8.5pt;'
                    f'white-space:nowrap;">{val}</td>'
                )
            else:
                html += (
                    f'<td style="padding:2px 8px;color:#1a1a2e;'
                    f'border-bottom:1px solid #eef1ef;font-size:8.5pt;">'
                    f'{sanitize_text(str(val))}</td>'
                )
        html += '</tr>\n'

    html += '</table>\n'
    return html


def _section_header(title):
    """Gera um header de se\u00e7\u00e3o estilizado com pill lateral verde."""
    return (
        f'<tr><td style="padding:22px 0 8px 0;">'
        f'<table cellpadding="0" cellspacing="0" border="0"><tr>'
        f'<td style="background-color:#00b876;width:4px;border-radius:4px;">&nbsp;</td>'
        f'<td style="padding-left:12px;">'
        f'<span style="font-family:Calibri,Arial,sans-serif;font-size:12.5pt;'
        f'color:#004d33;font-weight:bold;letter-spacing:0.3px;">{title}</span>'
        f'</td></tr></table>'
        f'</td></tr>\n'
    )


def build_email_html(assessor_code, data):
    """Monta o corpo HTML completo do email para um assessor - estilo Somus Capital."""
    primeiro_nome = data["primeiro_nome"] or "Assessor"
    hoje = datetime.now().strftime("%d/%m/%Y")

    cols_rf = [
        ("cliente_cod", "Conta"),
        ("nome", "Cliente"),
        ("ativo", "Ativo"),
        ("emissor", "Emissor"),
        ("indexador", "Indexador"),
        ("vencimento", "Vencimento"),
        ("financeiro", "Financeiro"),
    ]

    cols_fundos = [
        ("cliente_cod", "Conta"),
        ("nome", "Cliente"),
        ("operacao", u"Opera\u00e7\u00e3o"),
        ("fundo", "Fundo"),
        ("dt_cotizacao", u"Dt. Cotiza\u00e7\u00e3o"),
        ("dt_liquidacao", u"Dt. Liquida\u00e7\u00e3o"),
        ("financeiro", "Financeiro"),
    ]

    cols_saldos = [
        ("cliente", "Cliente"),
        ("conta", "Conta"),
        ("d0", "D0"),
        ("d1", "D+1"),
        ("d2", "D+2"),
        ("d3", "D+3"),
        ("total", "Total"),
    ]

    venc_rf = data.get("vencimentos_rf", [])
    venc_fundos = data.get("vencimentos_fundos", [])
    negativos_d0 = data["negativos_d0"]
    negativos_d1 = data["negativos_d1"]
    positivos = data["positivos"]

    # === HEADER SOMUS ===
    logo_b64 = _load_logo_b64()
    logo_tag = (
        f'<img src="data:image/png;base64,{logo_b64}" width="155" height="38"'
        f' style="vertical-align:middle;margin-right:16px;" alt="Somus Capital">'
    ) if logo_b64 else ""

    html = f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">

<!-- Header -->
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Produtos</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>

<!-- Corpo -->
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">

    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#4b5563;margin-top:0;">
      Segue abaixo a rela\u00e7\u00e3o de saldos di\u00e1rios e vencimentos dos seus clientes.
    </p>

  </td></tr>
"""

    # === 1. VENCIMENTOS RF ===
    html += _section_header(u"Vencimentos de Renda Fixa \u2014 Pr\u00f3ximos 4 Dias")
    html += '<tr><td style="padding:0 4px;">\n'
    html += build_html_table(
        venc_rf, cols_rf,
        empty_msg=u"N\u00e3o h\u00e1 vencimentos de renda fixa para os pr\u00f3ximos 4 dias."
    )
    html += '</td></tr>\n'

    # === 2. VENCIMENTOS FUNDOS ===
    html += _section_header(u"Líquida\u00e7\u00f5es de Fundos \u2014 Pr\u00f3ximos 4 Dias")
    html += '<tr><td style="padding:0 4px;">\n'
    html += build_html_table(
        venc_fundos, cols_fundos,
        empty_msg=u"N\u00e3o h\u00e1 líquida\u00e7\u00f5es de fundos para os pr\u00f3ximos 4 dias."
    )
    html += '</td></tr>\n'

    # === 3. NEGATIVOS D0 ===
    html += _section_header("Saldos Negativos em D0")
    html += '<tr><td style="padding:0 4px;">\n'
    html += build_html_table(
        negativos_d0, cols_saldos,
        empty_msg=u"N\u00e3o h\u00e1 saldos negativos em D0 para hoje."
    )
    html += '</td></tr>\n'

    # === 4. NEGATIVOS D+1 ===
    html += _section_header(u"D\u00e9bitos em D+1")
    html += '<tr><td style="padding:0 4px;">\n'
    html += build_html_table(
        negativos_d1, cols_saldos,
        empty_msg=u"N\u00e3o h\u00e1 d\u00e9bitos em D+1 para hoje."
    )
    html += '</td></tr>\n'

    # === 5. POSITIVOS ===
    html += _section_header("Saldos Acima de R$100")
    html += '<tr><td style="padding:0 4px;">\n'
    html += build_html_table(
        positivos, cols_saldos,
        empty_msg=u"N\u00e3o h\u00e1 saldos positivos acima de R$100 para hoje."
    )
    html += '</td></tr>\n'

    # === FOOTER SOMUS ===
    html += f"""
  <!-- Footer -->
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Produtos</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">
            Saldos Di\u00e1rios &middot; {hoje}
          </span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">
            E-mail gerado automaticamente
          </span>
        </td>
      </tr>
    </table>
  </td></tr>

</table>
</div>
"""
    return html


def processar_saldos(saldo_filepath, callback=None):
    """
    Processa a planilha de saldos e cria rascunhos no Outlook.

    Args:
        saldo_filepath: caminho da planilha de saldos
        callback: fun\u00e7\u00e3o(msg, tipo) para log. tipo = 'info', 'ok', 'skip', 'error'

    Returns:
        dict com estat\u00edsticas: criados, erros, sem_email, total_assessores
    """
    def log(msg, tipo="info"):
        if callback:
            callback(msg, tipo)

    stats = {"criados": 0, "erros": 0, "sem_email": 0, "total_assessores": 0}

    # 1. Carregar bases
    log("Carregando BASE SOMUS...")
    base_somus = load_base_somus()
    log(f"  {len(base_somus)} clientes mapeados na BASE SOMUS.")

    log("Carregando emails...")
    emails_map = load_emails()
    log(f"  {len(emails_map)} registros de email carregados.")

    # 2. Ler saldos
    log(f"Lendo planilha de saldos...")
    records = load_saldos(saldo_filepath)
    log(f"  {len(records)} registros lidos.")

    if not records:
        log("Nenhum registro encontrado na planilha.", "error")
        return stats

    # 3. Carregar vencimentos do BASE ENVIAR (pr\u00f3ximos 4 dias)
    hoje_str = datetime.now().strftime("%d/%m/%Y")
    limite_str = (datetime.now() + timedelta(days=4)).strftime("%d/%m/%Y")
    log(f"Carregando vencimentos de Renda Fixa ({hoje_str} a {limite_str})...")
    venc_rf = load_vencimentos_rf()
    total_rf = sum(len(v) for v in venc_rf.values())
    log(f"  {total_rf} vencimentos RF para {len(venc_rf)} assessores.")

    log(f"Carregando vencimentos de Fundos ({hoje_str} a {limite_str})...")
    venc_fundos = load_vencimentos_fundos()
    total_fundos = sum(len(v) for v in venc_fundos.values())
    log(f"  {total_fundos} vencimentos de Fundos para {len(venc_fundos)} assessores.")

    # 4. Classificar e agrupar por assessor
    log("Classificando e agrupando por assessor...")
    assessor_data = classify_and_group(records, base_somus, emails_map)
    stats["total_assessores"] = len(assessor_data)
    log(f"  {len(assessor_data)} assessores identificados.")

    # Injetar vencimentos nos dados de cada assessor
    for code, data in assessor_data.items():
        data["vencimentos_rf"] = venc_rf.get(code, [])
        data["vencimentos_fundos"] = venc_fundos.get(code, [])

    # Filtrar apenas assessores que t\u00eam pelo menos 1 registro em alguma categoria
    assessores_com_dados = {
        code: data for code, data in assessor_data.items()
        if data["positivos"] or data["negativos_d0"] or data["negativos_d1"]
        or data["vencimentos_rf"] or data["vencimentos_fundos"]
    }
    log(f"  {len(assessores_com_dados)} assessores com dados relevantes.")

    if not assessores_com_dados:
        log("Nenhum assessor com dados para enviar.", "skip")
        return stats

    # 4. Conectar ao Outlook
    log("Conectando ao Outlook...")
    try:
        import win32com.client as win32
        import pythoncom
        pythoncom.CoInitialize()
        outlook = win32.Dispatch("Outlook.Application")
    except Exception as e:
        log(f"Erro ao conectar ao Outlook: {e}", "error")
        log("Verifique se o Outlook est\u00e1 aberto.", "error")
        return stats

    # 4.1 Selecionar conta artur.brito@somuscapital.com.br
    SENDER_EMAIL = "artur.brito@somuscapital.com.br"
    sender_account = None
    try:
        for acc in outlook.Session.Accounts:
            if acc.SmtpAddress.lower() == SENDER_EMAIL:
                sender_account = acc
                break
        if sender_account:
            log(f"Conta selecionada: {SENDER_EMAIL}")
        else:
            log(f"Conta {SENDER_EMAIL} n\u00e3o encontrada no Outlook. Usando conta padr\u00e3o.", "skip")
    except Exception:
        log("N\u00e3o foi poss\u00edvel selecionar conta. Usando conta padr\u00e3o.", "skip")

    # 5. Criar rascunhos
    log("\nCriando rascunhos no Outlook...")
    log("-" * 60)

    for code in sorted(assessores_com_dados.keys()):
        data = assessores_com_dados[code]
        email = data["email"]
        nome = data["nome"]
        n_pos = len(data["positivos"])
        n_neg0 = len(data["negativos_d0"])
        n_neg1 = len(data["negativos_d1"])
        n_rf = len(data.get("vencimentos_rf", []))
        n_fund = len(data.get("vencimentos_fundos", []))

        if not email or email == "-":
            log(f"  {code}  {nome}:  Sem email cadastrado", "skip")
            stats["sem_email"] += 1
            continue

        try:
            mail = outlook.CreateItem(0)  # olMailItem
            if sender_account:
                mail.SendUsingAccount = sender_account
            mail.To = email

            # CC para assistente se houver
            email_assist = data.get("email_assistente", "-")
            if email_assist and email_assist != "-":
                mail.CC = email_assist

            mail.Subject = f"{code} | Saldos di\u00e1rios e vencimentos de Renda Fixa"

            mail.HTMLBody = build_email_html(code, data)

            mail.Save()  # Salva como rascunho (NÃO envia)

            cc_info = f"  CC: {email_assist}" if email_assist and email_assist != "-" else ""
            log(f"  {code}  {nome}  ->  {email}{cc_info}  [+{n_pos} -{n_neg0} d1:{n_neg1} RF:{n_rf} FD:{n_fund}]", "ok")
            stats["criados"] += 1

        except Exception as e:
            log(f"  {code}  {nome}:  ERRO - {e}", "error")
            stats["erros"] += 1

    try:
        pythoncom.CoUninitialize()
    except Exception:
        pass

    log("-" * 60)
    log(f"\nFinalizado!  {stats['criados']} rascunhos criados, "
        f"{stats['erros']} erros, {stats['sem_email']} sem email.")
    log("Os e-mails est\u00e3o na pasta RASCUNHOS do Outlook.")

    return stats


# === Para teste standalone ===
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python processar_saldos.py <planilha_saldos.xlsx>")
        sys.exit(1)

    def print_log(msg, tipo="info"):
        prefix = {"ok": "[OK]", "error": "[ERRO]", "skip": "[SKIP]", "info": "[INFO]"}
        print(f"{prefix.get(tipo, '[INFO]')} {msg}")

    result = processar_saldos(sys.argv[1], callback=print_log)
    print(f"\nResultado: {result}")
