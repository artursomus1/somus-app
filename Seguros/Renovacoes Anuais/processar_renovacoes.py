"""
Somus Capital - Processador de Renovações Anuais de Seguros
Gera rascunhos de e-mail no Outlook para cada assessor com as renovações dos seus clientes.
"""

import os
import base64
from datetime import datetime
from collections import defaultdict

import openpyxl


# === PATHS ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Seguros/Renovacoes Anuais/
APP_ROOT = os.path.dirname(os.path.dirname(BASE_DIR))  # APP SOMUS/
BASE_EMAILS_FILE = os.path.join(APP_ROOT, "BASE", "BASE EMAILS.xlsx")
BASE_XLSM_FILE = os.path.join(APP_ROOT, "BASE", "BASE.xlsm")
LOGO_PATH = os.path.join(APP_ROOT, "assets", "logo_somus.png")


def _load_logo_b64():
    """Carrega logo Somus como base64 data URI para embedding inline."""
    if not os.path.exists(LOGO_PATH):
        return ""
    with open(LOGO_PATH, "rb") as f:
        return base64.b64encode(f.read()).decode()


def sanitize_text(text):
    if text is None:
        return ""
    text = str(text)
    replacements = {
        '\u2013': '-', '\u2014': '-', '\u2018': "'", '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2026': '...', '\u2022': '-',
        '\u00ba': 'o', '\u00aa': 'a', '\u00a0': ' ',
        '\ufeff': '', '\u200b': '', '\u200e': '', '\u200f': '',
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def fmt_brl(value):
    """Formata numero como R$ brasileiro."""
    if value is None or value == "" or value == "-":
        return "-"
    try:
        num = float(str(value).replace('.', '').replace(',', '.')) if isinstance(value, str) else float(value)
    except (ValueError, TypeError):
        return str(value)
    if num == 0:
        return "-"
    negative = num < 0
    num = abs(num)
    num = round(num, 2)
    inteiro = int(num)
    centavos = round((num - inteiro) * 100)
    if centavos >= 100:
        inteiro += 1
        centavos = 0
    s_inteiro = f"{inteiro:,}".replace(",", ".")
    result = f"R$ {s_inteiro},{centavos:02d}"
    if negative:
        result = f"-{result}"
    return result


def fmt_date(val):
    """Formata data para DD/MM/YYYY."""
    if val is None or val == "" or val == "-":
        return "-"
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s[:10]


def load_emails():
    """Carrega mapeamento Codigo Assessor -> info de email."""
    emails = {}

    # 1) BASE EMAILS.xlsx (prioridade)
    if os.path.exists(BASE_EMAILS_FILE):
        wb = openpyxl.load_workbook(BASE_EMAILS_FILE, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0]:
                code = str(row[0]).strip()
                emails[code] = {
                    "nome": sanitize_text(row[1]) if row[1] else "",
                    "email": str(row[2]).strip() if row[2] else "",
                    "assistente": sanitize_text(row[3]) if row[3] else "-",
                    "email_assistente": str(row[4]).strip() if row[4] else "-",
                }
                if code.startswith("A"):
                    emails[code[1:]] = emails[code]
        wb.close()

    # 2) Lista de e-mails do BASE.xlsm (fallback)
    if os.path.exists(BASE_XLSM_FILE):
        wb = openpyxl.load_workbook(BASE_XLSM_FILE, read_only=True, data_only=True)
        if "Lista de e-mails" in wb.sheetnames:
            ws = wb["Lista de e-mails"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if row[0]:
                    code = str(row[0]).strip()
                    a_code = f"A{code}" if not code.startswith("A") else code
                    if a_code not in emails and code not in emails:
                        emails[code] = {
                            "nome": sanitize_text(row[1]) if row[1] else "",
                            "email": str(row[2]).strip() if row[2] else "",
                            "assistente": sanitize_text(row[3]) if row[3] else "-",
                            "email_assistente": str(row[4]).strip() if row[4] else "-",
                        }
                        emails[a_code] = emails[code]
        wb.close()

    return emails


def _build_nome_to_email_map(emails):
    """Constroi mapeamento reverso: nome (lowercase) -> email_info.
    Tambem mapeia primeiro nome para facilitar matching parcial."""
    nome_map = {}
    for code, info in emails.items():
        nome = info.get("nome", "").strip()
        if not nome:
            continue
        nome_lower = nome.lower()
        nome_map[nome_lower] = info
        # Primeiro nome
        primeiro = nome.split()[0].lower()
        if primeiro not in nome_map:
            nome_map[primeiro] = info
    return nome_map


def _match_aai_to_email(aai_name, nome_map):
    """Tenta encontrar o email do assessor pelo nome (case insensitive).
    Tenta matching exato, por primeiro nome, e por conteudo parcial."""
    aai_lower = aai_name.strip().lower()
    if not aai_lower:
        return None

    # 1) Match exato
    if aai_lower in nome_map:
        return nome_map[aai_lower]

    # 2) Match por primeiro nome do AAI
    primeiro_aai = aai_lower.split()[0]
    if primeiro_aai in nome_map:
        return nome_map[primeiro_aai]

    # 3) Match parcial: AAI contem o nome ou vice-versa
    for nome_key, info in nome_map.items():
        if len(nome_key) >= 3 and (nome_key in aai_lower or aai_lower in nome_key):
            return info

    return None


def _parse_premio(val):
    """Converte premio para float, lidando com formatos BR (R$ 47.716,88)."""
    if val is None or val == "" or val == "-":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _find_header_row(ws, max_search=5):
    """Encontra a linha do header procurando por palavras-chave nas primeiras linhas."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search, values_only=True), start=1):
        vals = [str(c).strip().upper() if c else "" for c in row]
        joined = " ".join(vals)
        if "CLIENTE" in joined or "NOME" in joined:
            return i
    return 1  # fallback


def load_renovacoes(filepath):
    """Le planilha de renovacoes e retorna lista de registros normalizados.

    Detecta tres formatos:
    1) Aba CONSOLIDADO: [vazio, ASSESSOR, CLIENTE, SEGURADORA, AQUISICAO, DATA RENOVACAO]
    2) Abas mensais: [vazio, AAI, CLIENTE, SEG, PA, DATA RENOVACAO] (AAI pode estar vazio)
    3) Aba "em breve": [AAI, Nome, Seguradora, Premio, Periodicidade, Data emissao, Proximo pgto]
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    records = []

    sheet_names = wb.sheetnames
    meses = {"janeiro", "fevereiro", "marco", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"}

    has_consolidado = any(s.strip().upper() == "CONSOLIDADO" for s in sheet_names)
    has_monthly = any(s.strip().lower() in meses for s in sheet_names)
    has_em_breve = any("em breve" in s.lower() for s in sheet_names)

    if has_consolidado:
        # Formato 1: aba CONSOLIDADO com ASSESSOR preenchido
        consolidado_name = next(s for s in sheet_names if s.strip().upper() == "CONSOLIDADO")
        ws = wb[consolidado_name]
        header_row = _find_header_row(ws)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 6:
                continue
            # [vazio, ASSESSOR, CLIENTE, SEGURADORA, AQUISICAO, DATA RENOVACAO]
            assessor = sanitize_text(row[1]) if row[1] else ""
            cliente = sanitize_text(row[2]) if row[2] else ""
            seguradora = sanitize_text(row[3]) if row[3] else ""
            premio = _parse_premio(row[4])
            data_renov = row[5]

            if not assessor or not cliente:
                continue

            records.append({
                "aai": assessor.strip(),
                "cliente": cliente.strip(),
                "seguradora": seguradora.strip(),
                "premio": premio,
                "data_renovacao": data_renov,
                "periodicidade": "Anual",
                "aba": consolidado_name,
            })

    elif has_monthly:
        # Formato 2: abas mensais — coluna AAI pode estar vazia em linhas de dados
        for sheet_name in sheet_names:
            if sheet_name.strip().lower() not in meses:
                continue
            ws = wb[sheet_name]
            header_row = _find_header_row(ws)
            current_aai = ""
            for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
                if not row or len(row) < 6:
                    continue
                # [vazio, AAI, CLIENTE, SEG, PA, DATA RENOVACAO]
                # AAI so aparece na primeira linha do grupo — herdar para as seguintes
                if row[1]:
                    current_aai = sanitize_text(row[1]).strip()
                cliente = sanitize_text(row[2]) if row[2] else ""
                seguradora = sanitize_text(row[3]) if row[3] else ""
                premio = _parse_premio(row[4])
                data_renov = row[5]

                if not cliente:
                    continue

                records.append({
                    "aai": current_aai,
                    "cliente": cliente.strip(),
                    "seguradora": seguradora.strip(),
                    "premio": premio,
                    "data_renovacao": data_renov,
                    "periodicidade": "Anual",
                    "aba": sheet_name,
                })

    if not records and has_em_breve:
        # Formato 3: aba "em breve"
        em_breve_name = next(s for s in sheet_names if "em breve" in s.lower())
        ws = wb[em_breve_name]
        header_row = _find_header_row(ws)
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 7:
                continue
            # [AAI, Nome, Seguradora, Premio, Periodicidade, Data emissao, Proximo pagamento]
            aai = sanitize_text(row[0]) if row[0] else ""
            cliente = sanitize_text(row[1]) if row[1] else ""
            seguradora = sanitize_text(row[2]) if row[2] else ""
            premio = _parse_premio(row[3])
            periodicidade = sanitize_text(row[4]) if row[4] else ""
            data_renov = row[6]  # Proximo pagamento

            if not aai or not cliente:
                continue

            records.append({
                "aai": aai.strip(),
                "cliente": cliente.strip(),
                "seguradora": seguradora.strip(),
                "premio": premio,
                "data_renovacao": data_renov,
                "periodicidade": periodicidade.strip() or "Anual",
                "aba": em_breve_name,
            })

    wb.close()
    return records


def classify_and_group(records, nome_map):
    """Agrupa renovacoes por assessor (AAI).

    Retorna dict: aai_name -> {
        'nome': str, 'email': str, 'primeiro_nome': str,
        'email_assistente': str, 'renovacoes': [...]
    }
    """
    assessor_data = {}

    for rec in records:
        aai = rec["aai"]
        if not aai:
            continue

        if aai not in assessor_data:
            email_info = _match_aai_to_email(aai, nome_map)
            nome_completo = email_info.get("nome", aai) if email_info else aai
            primeiro_nome = nome_completo.split()[0].capitalize() if nome_completo else aai
            assessor_data[aai] = {
                "nome": nome_completo,
                "email": email_info.get("email", "") if email_info else "",
                "assistente": email_info.get("assistente", "-") if email_info else "-",
                "email_assistente": email_info.get("email_assistente", "-") if email_info else "-",
                "primeiro_nome": primeiro_nome,
                "renovacoes": [],
            }

        assessor_data[aai]["renovacoes"].append(rec)

    # Ordena renovacoes por data
    for data in assessor_data.values():
        data["renovacoes"].sort(key=lambda r: str(r.get("data_renovacao", "")))

    return assessor_data


def build_email_html(assessor_name, data):
    """Monta o corpo HTML completo do email para um assessor - estilo Somus Capital."""
    primeiro_nome = data["primeiro_nome"] or "Assessor"
    hoje = datetime.now().strftime("%d/%m/%Y")
    renovacoes = data["renovacoes"]

    # Calcular prêmio total
    premio_total = sum(r.get("premio", 0) or 0 for r in renovacoes if isinstance(r.get("premio"), (int, float)))

    # Nome do mês atual
    mes_nome = ["", "Janeiro", "Fevereiro", "Mar\u00e7o", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"][datetime.now().month]

    logo_b64 = _load_logo_b64()
    logo_tag = (
        f'<img src="data:image/png;base64,{logo_b64}" width="155" height="38"'
        f' style="vertical-align:middle;margin-right:16px;" alt="Somus Capital">'
    ) if logo_b64 else ""

    # Tabela de renovações
    table_html = '<table cellpadding="0" cellspacing="0" border="0" '
    table_html += 'style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;'
    table_html += 'font-size:9pt;margin-bottom:6px;width:100%;">\n'

    # Header
    columns = [("cliente", "Cliente"), ("seguradora", "Seguradora"),
               ("premio", u"Pr\u00eamio Atual"), ("data_renovacao", u"Data de Renova\u00e7\u00e3o")]
    table_html += '<tr>'
    for key, label in columns:
        align = "text-align:center;" if key == "premio" else ""
        table_html += (
            f'<td style="padding:3px 8px 4px 8px;font-weight:bold;color:#00785a;'
            f'font-size:8.5pt;border-bottom:1.5px solid #00785a;'
            f'white-space:nowrap;{align}">{label}</td>'
        )
    table_html += '</tr>\n'

    for i, rec in enumerate(renovacoes):
        bg = "#f7faf9" if i % 2 == 0 else "#ffffff"
        table_html += f'<tr style="background:{bg};">'
        table_html += (
            f'<td style="padding:2px 8px;color:#1a1a2e;'
            f'border-bottom:1px solid #eef1ef;font-size:8.5pt;">'
            f'{sanitize_text(rec["cliente"])}</td>'
        )
        table_html += (
            f'<td style="padding:2px 8px;color:#1a1a2e;'
            f'border-bottom:1px solid #eef1ef;font-size:8.5pt;">'
            f'{sanitize_text(rec["seguradora"])}</td>'
        )
        table_html += (
            f'<td style="padding:2px 8px;text-align:center;color:#1a1a2e;'
            f'border-bottom:1px solid #eef1ef;font-size:8.5pt;'
            f'white-space:nowrap;">{fmt_brl(rec["premio"])}</td>'
        )
        table_html += (
            f'<td style="padding:2px 8px;color:#1a1a2e;'
            f'border-bottom:1px solid #eef1ef;font-size:8.5pt;'
            f'white-space:nowrap;">{fmt_date(rec["data_renovacao"])}</td>'
        )
        table_html += '</tr>\n'

    table_html += '</table>\n'

    html = f"""
<div style="font-family:Calibri,Arial,sans-serif;max-width:960px;margin:0 auto;">

<!-- Header -->
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding:16px 0 12px 0;vertical-align:middle;">
      {logo_tag}<!--
      --><span style="font-size:16pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:300;"> | </span><!--
      --><span style="font-size:16pt;color:#004d33;font-weight:bold;">Seguros</span>
    </td>
  </tr>
  <tr>
    <td style="padding:0 0 16px 0;">
      <hr style="border:none;border-top:2px solid #004d33;margin:0;">
    </td>
  </tr>
</table>

<!-- Corpo -->
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:18px;">
  <tr><td style="padding:0 4px;">
    <p style="font-size:11pt;color:#1a1a2e;margin-bottom:4px;">
      Prezado(a) <b>{primeiro_nome}</b>, tudo bem?
    </p>
    <p style="font-size:10.5pt;color:#4b5563;margin-top:0;margin-bottom:4px;">
      Segue abaixo a rela\u00e7\u00e3o de renova\u00e7\u00f5es anuais de seguros dos seus clientes referente ao m\u00eas de <b>{mes_nome}</b>.
    </p>
    <p style="font-size:12pt;color:#004d33;font-weight:bold;margin-top:8px;margin-bottom:0;">
      Pr\u00eamio Total: {fmt_brl(premio_total)}
    </p>
  </td></tr>

  <!-- Se\u00e7\u00e3o Renova\u00e7\u00f5es -->
  <tr><td style="padding:22px 0 8px 0;">
    <table cellpadding="0" cellspacing="0" border="0"><tr>
      <td style="background-color:#00b876;width:4px;border-radius:4px;">&nbsp;</td>
      <td style="padding-left:12px;">
        <span style="font-family:Calibri,Arial,sans-serif;font-size:12.5pt;
        color:#004d33;font-weight:bold;letter-spacing:0.3px;">Renova\u00e7\u00f5es Anuais de Seguros</span>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:0 4px;">
    {table_html}
  </td></tr>

  <!-- Footer -->
  <tr><td style="padding:28px 0 0 0;">
    <hr style="border:none;border-top:2px solid #004d33;margin:0 0 12px 0;">
    <table width="100%" cellpadding="0" cellspacing="0" border="0">
      <tr>
        <td style="vertical-align:middle;">
          <span style="font-size:10pt;font-weight:bold;color:#004d33;">Somus Capital</span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:300;"> | </span><!--
          --><span style="font-size:10pt;color:#004d33;font-weight:bold;">Seguros</span>
        </td>
        <td style="text-align:right;">
          <span style="font-size:8.5pt;color:#6b7280;">
            Renova\u00e7\u00f5es Anuais &middot; {hoje}
          </span><br>
          <span style="font-size:7.5pt;color:#9ca3af;">
            E-mail gerado automaticamente
          </span>
        </td>
      </tr>
    </table>
  </td></tr>

</table>
</div>
"""
    return html


def processar_renovacoes(filepath, callback=None):
    """
    Processa a planilha de renovacoes e cria rascunhos no Outlook.

    Args:
        filepath: caminho da planilha de renovacoes
        callback: funcao(msg, tipo) para log. tipo = 'info', 'ok', 'skip', 'error'

    Returns:
        dict com estatisticas: criados, erros, sem_email, total_assessores
    """
    def log(msg, tipo="info"):
        if callback:
            callback(msg, tipo)

    stats = {"criados": 0, "erros": 0, "sem_email": 0, "total_assessores": 0}

    # 1. Carregar emails
    log("Carregando emails...")
    emails_map = load_emails()
    log(f"  {len(emails_map)} registros de email carregados.")

    # Construir mapa nome -> email
    nome_map = _build_nome_to_email_map(emails_map)
    log(f"  {len(nome_map)} mapeamentos nome->email criados.")

    # 2. Ler renovacoes
    log(f"Lendo planilha de renovacoes...")
    all_records = load_renovacoes(filepath)
    log(f"  {len(all_records)} registros totais lidos.")

    # Filtrar apenas renovacoes do proximo mes
    hoje = datetime.now()
    if hoje.month == 12:
        mes_atual = 1
        ano_atual = hoje.year + 1
    else:
        mes_atual = hoje.month + 1
        ano_atual = hoje.year
    mes_nome = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"][mes_atual]

    records = []
    for rec in all_records:
        dt = rec.get("data_renovacao")
        if dt is None:
            continue
        if isinstance(dt, datetime):
            if dt.month == mes_atual and dt.year == ano_atual:
                records.append(rec)
        elif isinstance(dt, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    parsed = datetime.strptime(dt.strip(), fmt)
                    if parsed.month == mes_atual and parsed.year == ano_atual:
                        records.append(rec)
                    break
                except ValueError:
                    continue

    log(f"  {len(records)} renovacoes em {mes_nome}/{ano_atual} (filtradas de {len(all_records)} totais).")

    if not records:
        log(f"Nenhuma renovacao encontrada para {mes_nome}/{ano_atual}.", "error")
        return stats

    # 3. Agrupar por assessor
    log("Agrupando por assessor...")
    assessor_data = classify_and_group(records, nome_map)
    stats["total_assessores"] = len(assessor_data)
    log(f"  {len(assessor_data)} assessores identificados.")

    # Filtrar apenas assessores com renovacoes
    assessores_com_dados = {
        aai: data for aai, data in assessor_data.items()
        if data["renovacoes"]
    }
    log(f"  {len(assessores_com_dados)} assessores com renovacoes.")

    if not assessores_com_dados:
        log("Nenhum assessor com renovacoes para enviar.", "skip")
        return stats

    # 4. Conectar ao Outlook
    log("Conectando ao Outlook...")
    try:
        import win32com.client as win32
        import pythoncom
        pythoncom.CoInitialize()
        outlook = win32.Dispatch("Outlook.Application")
    except Exception as e:
        log(f"Erro ao conectar ao Outlook: {e}", "error")
        log("Verifique se o Outlook esta aberto.", "error")
        return stats

    # Selecionar conta
    SENDER_EMAIL = "artur.brito@somuscapital.com.br"
    sender_account = None
    try:
        for acc in outlook.Session.Accounts:
            if acc.SmtpAddress.lower() == SENDER_EMAIL:
                sender_account = acc
                break
        if sender_account:
            log(f"Conta selecionada: {SENDER_EMAIL}")
        else:
            log(f"Conta {SENDER_EMAIL} nao encontrada no Outlook. Usando conta padrao.", "skip")
    except Exception:
        log("Nao foi possivel selecionar conta. Usando conta padrao.", "skip")

    # 5. Criar rascunhos
    log("\nCriando rascunhos no Outlook...")
    log("-" * 60)

    for aai in sorted(assessores_com_dados.keys()):
        data = assessores_com_dados[aai]
        email = data["email"]
        nome = data["nome"]
        n_renov = len(data["renovacoes"])

        if not email or email == "-":
            log(f"  {aai}  ({nome}):  Sem email cadastrado", "skip")
            stats["sem_email"] += 1
            continue

        try:
            mail = outlook.CreateItem(0)
            if sender_account:
                mail.SendUsingAccount = sender_account
            mail.To = email

            # CC para assistente se houver
            email_assist = data.get("email_assistente", "-")
            if email_assist and email_assist != "-":
                mail.CC = email_assist

            mail.Subject = f"{aai} | Renovacoes Anuais de Seguros"

            mail.HTMLBody = build_email_html(aai, data)

            mail.Save()

            cc_info = f"  CC: {email_assist}" if email_assist and email_assist != "-" else ""
            log(f"  {aai}  ({nome})  ->  {email}{cc_info}  [{n_renov} renovacoes]", "ok")
            stats["criados"] += 1

        except Exception as e:
            log(f"  {aai}  ({nome}):  ERRO - {e}", "error")
            stats["erros"] += 1

    try:
        pythoncom.CoUninitialize()
    except Exception:
        pass

    log("-" * 60)
    log(f"\nFinalizado!  {stats['criados']} rascunhos criados, "
        f"{stats['erros']} erros, {stats['sem_email']} sem email.")
    log("Os e-mails estao na pasta RASCUNHOS do Outlook.")

    return stats


# === Para teste standalone ===
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Uso: python processar_renovacoes.py <planilha_renovacoes.xlsx>")
        sys.exit(1)

    def print_log(msg, tipo="info"):
        prefix = {"ok": "[OK]", "error": "[ERRO]", "skip": "[SKIP]", "info": "[INFO]"}
        print(f"{prefix.get(tipo, '[INFO]')} {msg}")

    result = processar_renovacoes(sys.argv[1], callback=print_log)
    print(f"\nResultado: {result}")
